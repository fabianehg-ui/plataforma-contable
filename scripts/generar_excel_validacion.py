"""
Genera un Excel con los resultados de procesar los 192 XMLs reales
para que el usuario pueda revisar visualmente qué hizo el motor.
"""
import json
import sys
import xml.etree.ElementTree as ET
import glob
from pathlib import Path
from collections import Counter

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

sys.path.insert(0, str(Path(__file__).parent.parent))

from core.procesadores.motor_mapeo_v03 import (
    CatalogoEmpresa, resolver_mapeo, normalizar_nit, formato_cc_salida,
    calcular_retencion_renta, calcular_reteiva, calcular_reteica
)
from core.procesadores.detector_tipo_doc import (
    detectar_tipo_documento, mapear_a_comprobante
)


NS = {
    'cac': 'urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2',
    'cbc': 'urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2',
}


def get_inner(root):
    tag = root.tag.split('}')[-1]
    if tag == 'AttachedDocument':
        for desc in root.iter('{urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2}Description'):
            if desc.text and ('<Invoice' in desc.text or '<CreditNote' in desc.text or '<DebitNote' in desc.text):
                try:
                    return ET.fromstring(desc.text)
                except Exception:
                    pass
    return root


def extraer_doc(path):
    try:
        with open(path, 'r', encoding='utf-8') as f:
            c = f.read()
        root = ET.fromstring(c)
    except Exception:
        return None
    inner = get_inner(root)
    sup = inner.find('.//cac:AccountingSupplierParty', NS)
    nit_emi = (sup.find('.//cbc:CompanyID', NS).text
               if sup is not None and sup.find('.//cbc:CompanyID', NS) is not None else '')
    nombre_emi = (sup.find('.//cbc:RegistrationName', NS).text
                  if sup is not None and sup.find('.//cbc:RegistrationName', NS) is not None else '')
    delivery_line = ''
    for d in inner.findall('.//cac:Delivery', NS):
        line = d.find('.//cbc:Line', NS)
        if line is not None and line.text:
            delivery_line = line.text.strip()
            break
    total = inner.find('.//cac:LegalMonetaryTotal/cbc:PayableAmount', NS)
    total_val = float(total.text) if total is not None and total.text else 0.0
    iva_total = 0.0
    for tt in inner.findall('.//cac:TaxTotal', NS):
        for ts in tt.findall('.//cac:TaxSubtotal', NS):
            cat = ts.find('.//cac:TaxCategory/cac:TaxScheme/cbc:ID', NS)
            tax_amount = ts.find('.//cbc:TaxAmount', NS)
            if cat is not None and cat.text == '01' and tax_amount is not None:
                try:
                    iva_total += float(tax_amount.text or 0)
                except Exception:
                    pass
    descs = []
    for line in inner.findall('.//cac:InvoiceLine', NS) + inner.findall('.//cac:CreditNoteLine', NS) + inner.findall('.//cac:DebitNoteLine', NS):
        for desc in line.findall('.//cac:Item/cbc:Description', NS):
            if desc.text:
                descs.append(desc.text)
    notas = []
    for note in inner.findall('.//cbc:Note', NS):
        if note.text:
            notas.append(note.text)
    return {
        'archivo': Path(path).name,
        'nit_emisor': nit_emi,
        'nombre_emisor': nombre_emi,
        'direccion_entrega': delivery_line,
        'total': total_val,
        'iva': iva_total,
        'base_imponible': total_val - iva_total if iva_total else total_val,
        'items_descripcion': ' | '.join(descs[:5]),
        'items_descripciones_lista': descs,
        'notas': notas,
        'xml_root': root  # para detector
    }


def main():
    cat = CatalogoEmpresa.cargar(
        Path(__file__).parent.parent / 'core' / 'data' / 'empresas' / '900451388_silla_tres'
    )

    base_xmls = '/home/claude/analisis/all_xmls'
    docs = []
    for tipo in ['FE', 'NC', 'ND', 'DS']:
        for x in sorted(glob.glob(f'{base_xmls}/{tipo}/*.xml')):
            d = extraer_doc(x)
            if d:
                d['tipo'] = tipo
                docs.append(d)

    print(f"Procesando {len(docs)} XMLs...")

    catalogo_comprobantes = cat.empresa_json.get('comprobantes_por_tipo_dian', None)

    filas = []
    for d in docs:
        # Detectar tipo DIAN real
        tipo_dian, fuente_tipo = detectar_tipo_documento(
            xml_root=d.get('xml_root'),
            nombre_archivo=d['archivo']
        )
        comp_info = mapear_a_comprobante(tipo_dian, catalogo_comprobantes)
        comprobante = comp_info.get('comprobante', '')
        comp_desc = comp_info.get('descripcion', '')

        # Si es ACUSE, no procesar
        if tipo_dian == 'ACUSE':
            filas.append({
                'tipo_carpeta': d['tipo'],
                'tipo_dian_real': tipo_dian,
                'comprobante': comprobante,
                'comprobante_desc': comp_desc,
                'nit_emisor': d['nit_emisor'],
                'nombre_emisor': d['nombre_emisor'],
                'direccion_xml': d['direccion_entrega'],
                'notas_xml': ' | '.join(d.get('notas', []))[:200],
                'total': d['total'],
                'base': d['base_imponible'],
                'iva': d['iva'],
                'cuenta_resuelta': '',
                'cc_resuelto': '',
                'cc_nombre': '',
                'concepto_retencion': '',
                'fuente_cuenta': 'no_aplica_acuse',
                'fuente_cc': '',
                'pista_palabra_clave': '',
                'retfuente_valor': 0,
                'retfuente_motivo_no_aplica': '',
                'reteiva_valor': 0,
                'reteiva_motivo_no_aplica': '',
                'pendiente': False,
                'advertencias': 'Acuse de recibo - no contable',
                'archivo_xml': d['archivo']
            })
            continue

        r = resolver_mapeo(
            cat,
            nit_emisor=d['nit_emisor'],
            nombre_emisor=d['nombre_emisor'],
            direccion_entrega=d['direccion_entrega'],
            items_descripcion=d['items_descripcion'],
            notas_xml=d.get('notas', []),
            items_descripciones_lista=d.get('items_descripciones_lista', [])
        )
        ret_renta = calcular_retencion_renta(
            cat, r.concepto_retencion, d['base_imponible'], r.regimen_emisor, r.autorretenedor_renta
        )
        ret_iva = calcular_reteiva(cat, d['iva'], r.regimen_emisor, d['base_imponible'])

        # Aplicar formato de CC según empresa.json
        cc_formato = cat.empresa_json.get('formato_salida', {}).get('cc_formato', 'sin_guion')
        cc_salida = formato_cc_salida(r.cc, cc_formato)

        filas.append({
            'tipo_carpeta': d['tipo'],
            'tipo_dian_real': tipo_dian,
            'comprobante': comprobante,
            'comprobante_desc': comp_desc,
            'nit_emisor': d['nit_emisor'],
            'nombre_emisor': d['nombre_emisor'],
            'direccion_xml': d['direccion_entrega'],
            'notas_xml': ' | '.join(d.get('notas', []))[:200],
            'total': d['total'],
            'base': d['base_imponible'],
            'iva': d['iva'],
            'cuenta_resuelta': r.cuenta,
            'cc_resuelto': cc_salida,
            'cc_nombre': cat.centros_costo.get(r.cc, ''),
            'concepto_retencion': r.concepto_retencion,
            'fuente_cuenta': r.fuente_cuenta,
            'fuente_cc': r.fuente_cc,
            'pista_palabra_clave': r.pista_palabra_clave[:80] if r.pista_palabra_clave else '',
            'retfuente_valor': ret_renta.valor if ret_renta and ret_renta.aplicada else 0,
            'retfuente_motivo_no_aplica': ret_renta.motivo_no_aplicada if ret_renta and not ret_renta.aplicada else '',
            'reteiva_valor': ret_iva.valor if ret_iva and ret_iva.aplicada else 0,
            'reteiva_motivo_no_aplica': ret_iva.motivo_no_aplicada if ret_iva and not ret_iva.aplicada else '',
            'pendiente': r.es_pendiente_revision,
            'advertencias': '; '.join(r.advertencias) if r.advertencias else '',
            'archivo_xml': d['archivo']
        })

    df_resultados = pd.DataFrame(filas)

    # Hoja 2: Mapeo aprendido (tabla compacta)
    cc_formato = cat.empresa_json.get('formato_salida', {}).get('cc_formato', 'sin_guion')
    mapeo_filas = []
    for nit, info in cat.mapeo_nits.items():
        cc_int = info.get('cc_default', '')
        cc_out = formato_cc_salida(cc_int, cc_formato)
        mapeo_filas.append({
            'nit': nit,
            'nombre': info.get('nombre', ''),
            'cuenta_default': info.get('cuenta_default', ''),
            'cc_default_codigo': cc_out,
            'cc_default_interno': cc_int,
            'cc_nombre': cat.centros_costo.get(cc_int, ''),
            'concepto_retencion': info.get('concepto_retencion', ''),
            'regimen': info.get('regimen', 'ordinario'),
            'autorretenedor_renta': info.get('autorretenedor_renta', False),
            'confianza_cuenta': info.get('confianza_cuenta', 0),
            'confianza_cc': info.get('confianza_cc', 0),
            'cuentas_alternas': ', '.join(info.get('cuentas_vistas', [])[1:5]),
            'ccs_alternos': ', '.join([formato_cc_salida(c, cc_formato) for c in info.get('ccs_vistos', [])[1:5]])
        })
    df_mapeo = pd.DataFrame(mapeo_filas).sort_values('nombre')

    # Hoja 3: Direcciones
    direcciones_filas = []
    for direc, info in cat.direcciones_locales.get('direcciones', {}).items():
        cc_int = info.get('cc', '')
        direcciones_filas.append({
            'direccion_normalizada': direc,
            'cc_codigo': formato_cc_salida(cc_int, cc_formato),
            'cc_interno': cc_int,
            'cc_nombre': info.get('cc_nombre', ''),
            'docs_marzo_2026': info.get('docs_observados_marzo_2026', 0),
            'variantes': ' | '.join(info.get('variantes_originales', []))
        })
    df_dirs = pd.DataFrame(direcciones_filas).sort_values('docs_marzo_2026', ascending=False)

    # Hoja 4: Resumen
    from collections import Counter as _Counter
    tipos_real = _Counter(f['tipo_dian_real'] for f in filas)
    resumen = {
        'Total XMLs procesados': [len(docs)],
        '— TIPOS DIAN DETECTADOS —': [''],
        'FE - Facturas (→ comprobante 3)': [tipos_real.get('FE', 0)],
        'NC - Notas crédito (→ comprobante 12)': [tipos_real.get('NC', 0)],
        'ND - Notas débito (→ comprobante 7)': [tipos_real.get('ND', 0)],
        'DS - Documentos soporte (→ comprobante 3)': [tipos_real.get('DS', 0)],
        'ACUSE - Acuses recibo (descartados)': [tipos_real.get('ACUSE', 0)],
        '— CLASIFICACIÓN —': [''],
        'Auto-clasificados (con cuenta)': [sum(1 for f in filas if f['fuente_cuenta'] == 'mapeo_nit')],
        'Auto-detectados como insumo': [sum(1 for f in filas if f['fuente_cuenta'] == 'auto_insumos')],
        'Pendientes catalogación': [sum(1 for f in filas if f['pendiente'])],
        'CC por palabra clave (CTS1, etc.)': [sum(1 for f in filas if f['fuente_cc'] == 'palabra_clave')],
        'CC por dirección XML': [sum(1 for f in filas if f['fuente_cc'] == 'direccion_xml')],
        'CC por NIT alta confianza': [sum(1 for f in filas if f['fuente_cc'] == 'nit_default_alta_conf')],
        'CC por NIT baja confianza': [sum(1 for f in filas if f['fuente_cc'] == 'nit_default_baja_conf')],
        'CC default (1010 GENERAL)': [sum(1 for f in filas if f['fuente_cc'] == 'empresa_default')],
        '— TOTALES —': [''],
        'Total compras': [sum(f['total'] for f in filas if f['tipo_dian_real'] != 'ACUSE')],
        'Total IVA': [sum(f['iva'] for f in filas if f['tipo_dian_real'] != 'ACUSE')],
        'Total retefuente practicada': [sum(f['retfuente_valor'] for f in filas)],
        'Total reteIVA practicada (solo a RST)': [sum(f['reteiva_valor'] for f in filas)],
        'ReteICA': ['NO se practica - empresa no obligada']
    }
    df_resumen = pd.DataFrame(resumen).T.reset_index()
    df_resumen.columns = ['Métrica', 'Valor']

    # Escribir Excel
    out = '/home/claude/v03/RESULTADOS_VALIDACION_192_XMLS.xlsx'
    with pd.ExcelWriter(out, engine='openpyxl') as w:
        df_resumen.to_excel(w, sheet_name='1_Resumen', index=False)
        df_resultados.to_excel(w, sheet_name='2_XMLs_Procesados', index=False)
        df_mapeo.to_excel(w, sheet_name='3_NITs_Aprendidos', index=False)
        df_dirs.to_excel(w, sheet_name='4_Direcciones_Mapeadas', index=False)

    # Aplicar formato
    from openpyxl import load_workbook
    wb = load_workbook(out)

    bold = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F4E78')
    pendiente_fill = PatternFill('solid', fgColor='FFE699')
    high_conf = PatternFill('solid', fgColor='C6EFCE')
    low_conf = PatternFill('solid', fgColor='FFC7CE')

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        # Header
        for cell in ws[1]:
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Auto width
        for col in ws.columns:
            max_len = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 50))
                except Exception:
                    pass
            ws.column_dimensions[column].width = min(max(max_len + 2, 10), 50)
        # Freeze
        ws.freeze_panes = 'A2'

    # Hoja XMLs: pintar pendientes y baja confianza
    ws = wb['2_XMLs_Procesados']
    # Detectar dinámicamente la columna 'pendiente'
    headers_xmls = [c.value for c in ws[1]]
    col_pendiente = headers_xmls.index('pendiente') + 1 if 'pendiente' in headers_xmls else None
    col_palabra_clave = headers_xmls.index('pista_palabra_clave') + 1 if 'pista_palabra_clave' in headers_xmls else None
    palabra_fill = PatternFill('solid', fgColor='D6EAF8')

    for row_i in range(2, ws.max_row + 1):
        if col_pendiente:
            pend = ws.cell(row=row_i, column=col_pendiente).value
            if pend is True:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=row_i, column=c).fill = pendiente_fill
        # Pintar la celda de pista en azul claro si tiene contenido
        if col_palabra_clave:
            pista = ws.cell(row=row_i, column=col_palabra_clave).value
            if pista:
                ws.cell(row=row_i, column=col_palabra_clave).fill = palabra_fill

    # Hoja NITs aprendidos: pintar baja confianza
    ws = wb['3_NITs_Aprendidos']
    for row_i in range(2, ws.max_row + 1):
        conf = ws.cell(row=row_i, column=9).value  # confianza_cuenta
        if isinstance(conf, (int, float)):
            if conf >= 0.7:
                ws.cell(row=row_i, column=9).fill = high_conf
            elif conf < 0.5:
                ws.cell(row=row_i, column=9).fill = low_conf

    # Formato número en columnas de moneda (detectar dinámicamente)
    money_headers = {'total', 'base', 'iva', 'retfuente_valor', 'reteiva_valor', 'reteica_valor'}
    ws = wb['2_XMLs_Procesados']
    for col_idx, header in enumerate(headers_xmls, 1):
        if header in money_headers:
            for row_i in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_i, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '$#,##0;($#,##0);-'

    # Resumen: columna 2 es Valor
    ws = wb['1_Resumen']
    for row_i in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_i, column=2)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '$#,##0;($#,##0);-' if cell.value > 1000 else '#,##0'

    wb.save(out)
    print(f"✅ Excel generado: {out}")


if __name__ == '__main__':
    main()
