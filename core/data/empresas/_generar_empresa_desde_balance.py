import sys, json, types, os
sys.path.insert(0,'.')
sys.modules['db']=types.ModuleType('db')
m=types.ModuleType('db.supabase_client'); m.get_supabase=lambda:None; m.get_supabase_admin=lambda:None
sys.modules['db.supabase_client']=m
from core.utils import conocimiento_balance as cb
from core.utils.nits import normalizar_nit
from openpyxl import load_workbook
from collections import defaultdict

BAL='/tmp/mapeo.xlsx'
NIT, DV, RAZON, SLUG = '900495853','4','MILAGROS GROUP SAS','900495853_milagros'
def norm_cc(x): return str(x or '').strip().replace('-','').replace(' ','')

# --- COMPRAS (proveedores) via conocimiento_balance arreglado ---
k = cb.procesar_balance(open(BAL,'rb').read())

# --- leer balance crudo para nombres y para VENTAS (clientes, clase 4) y CxC ---
wb=load_workbook(BAL, read_only=True, data_only=True); ws=wb.active
filas=[list(f) for f in ws.iter_rows(values_only=True)]; wb.close()
enc=[str(c or '').strip().lower() for c in filas[3]]
c_cta=enc.index('cuenta'); c_nom=enc.index('nombre'); c_nit=enc.index('nit')
c_nnit=enc.index('nombre nit'); c_cc=enc.index('centro de costos'); c_ncc=enc.index('nombre cc')
c_deb=enc.index('débitos'); c_cre=enc.index('créditos')
cuenta_nom={}; cc_nom={}
ing=defaultdict(lambda:defaultdict(float)); ing_cc=defaultdict(lambda:defaultdict(float))
cxc=defaultdict(lambda:defaultdict(float)); nom_cli={}
for r in filas[4:]:
    if not r or r[c_cta] is None: continue
    cta=str(r[c_cta]).strip()
    if not cta or cta=='#N/A': continue
    cuenta_nom.setdefault(cta, str(r[c_nom] or '').strip())
    if r[c_cc]:
        ccn=norm_cc(r[c_cc])
        if ccn: cc_nom.setdefault(ccn, str(r[c_ncc] or '').strip())
    nit=normalizar_nit(r[c_nit]); 
    if not nit: continue
    cre=float(r[c_cre] or 0); 
    nnit=str(r[c_nnit] or '').strip()
    # VENTAS: ingresos clase 4 por crédito
    if cta[:1]=='4' and len(cta)>=6 and cre>0:
        ing[nit][cta]+=cre
        if r[c_cc]: ing_cc[nit][norm_cc(r[c_cc])]+=cre
        if nnit: nom_cli.setdefault(nit,nnit)
    # CxC clientes (13)
    if cta.startswith('13') and len(cta)>=6:
        mov=abs(float(r[c_deb] or 0)-cre)
        if mov>0: cxc[nit][cta]+=mov

# --- mapeo COMPRAS ---
mapeo={}
for nit_norm, info in k['nits'].items():
    cta=info['cuentas_gasto'][0] if info.get('cuentas_gasto') else ''
    if not cta: continue
    cc=info.get('centro_costo_sugerido','') or '001001'
    mapeo[nit_norm]={"razon_social":info.get('nombre',''),
        "default":{"cuenta":cta,"centro_costo":cc,"concepto":cuenta_nom.get(cta,info.get('nombre',''))},
        "reglas_item":[]}

# --- mapeo VENTAS (clientes) ---
mapeo_ventas={}
for nit_c, ctas in ing.items():
    cta=max(ctas.items(),key=lambda x:x[1])[0]
    cc=max(ing_cc[nit_c].items(),key=lambda x:x[1])[0] if ing_cc.get(nit_c) else '001001'
    cobrar=max(cxc[nit_c].items(),key=lambda x:x[1])[0] if cxc.get(nit_c) else '13050505'
    mapeo_ventas[nit_c]={"razon_social":nom_cli.get(nit_c,''),
        "default":{"cuenta":cta,"centro_costo":cc,"concepto":cuenta_nom.get(cta,'Venta')},
        "cuenta_por_cobrar":cobrar,"reglas_item":[]}

mapeo_json={
  "_comentario":f"Mapeo de {RAZON} desde balance. 'mapeo'=compras (por NIT emisor/proveedor); 'mapeo_ventas'=ventas (por NIT receptor/cliente).",
  "_fallback_global":{"cuenta":"519095","centro_costo":"001001","concepto":"PENDIENTE DE MAPEO - REVISAR"},
  "_fallback_ventas":{"cuenta":"41401520","centro_costo":"001003","concepto":"VENTA PENDIENTE DE MAPEO - REVISAR"},
  "mapeo":mapeo, "mapeo_ventas":mapeo_ventas
}
centros_json={"_comentario":f"Centros de costo de {RAZON} (sin guiones).","centros_costo":dict(sorted(cc_nom.items()))}

emp=json.load(open('core/data/empresas/901038325_jiper/empresa.json'))
emp["_comentario"]=f"Config base de {RAZON} generada del balance. VERIFICAR con contador: cuentas IVA, retención, IVA generado en ventas y comprobantes."
emp["nit"]=NIT; emp["dv"]=DV; emp["razon_social"]=RAZON
emp["cc_default"]="001001" if "001001" in cc_nom else (sorted(cc_nom)[0] if cc_nom else "001001")
emp["cc_default_nombre"]=cc_nom.get(emp["cc_default"],"")
emp["cuentas_venta"]={"_comentario":"VERIFICAR con contador.","cuenta_por_cobrar_default":"13050505",
    "iva_generado":"PENDIENTE_VERIFICAR","comprobante_venta":"PENDIENTE_VERIFICAR"}

base=f'/tmp/stage/core/data/empresas/{SLUG}'; os.makedirs(base,exist_ok=True)
json.dump(emp,open(f'{base}/empresa.json','w'),ensure_ascii=False,indent=2)
json.dump(mapeo_json,open(f'{base}/mapeo_nits.json','w'),ensure_ascii=False,indent=2)
json.dump(centros_json,open(f'{base}/centros_costo.json','w'),ensure_ascii=False,indent=2)
idx=json.load(open('core/data/empresas/_empresas_index.json'))
if not any(e['nit']==NIT for e in idx['empresas']):
    idx['empresas'].append({"id":NIT,"nit":NIT,"razon_social":RAZON,"carpeta":SLUG,"activa":True})
json.dump(idx,open('/tmp/stage/core/data/empresas/_empresas_index.json','w'),ensure_ascii=False,indent=2)

print("COMPRAS (proveedores):", len(mapeo))
print("VENTAS (clientes):", len(mapeo_ventas))
for nit_c,v in mapeo_ventas.items():
    print(f"  cliente {nit_c:12s} -> ingreso {v['default']['cuenta']} CxC {v['cuenta_por_cobrar']} CC {v['default']['centro_costo']} | {v['razon_social'][:22]}")
