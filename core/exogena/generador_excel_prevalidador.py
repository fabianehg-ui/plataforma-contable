"""
Generador Excel — Idéntico al Prevalidador DIAN AG 2025
========================================================

Produce un archivo .xlsx con una hoja por formato (F1001, F1003, F1005, F1006,
F1008, F1011, F1012) que replica fielmente las hojas del prevalidador oficial
en cuanto a:
  - Orden de columnas
  - Nombres de columnas (visibles)
  - Tipos de datos por columna
  - Longitudes máximas
  - Obligatoriedad

El contador puede copy/paste estos datos directamente a las hojas del
prevalidador .xlsm para subir a MUISCA.

La metadata exacta de las hojas se carga desde `data_prevalidador.json`,
que fue extraído del prevalidador oficial v3.3.0-26.
"""

from __future__ import annotations
from pathlib import Path
from typing import Optional
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Importar generadores del v2
from generador_xml_v2 import (
    Tercero, TerceroDestino, CabeceraXML, normalizar_nit, calcular_dv, TIPO_DOC_NIT,
    PAIS_COLOMBIA,
    RegistroF1001, RegistroF1003, RegistroF1005, RegistroF1006, RegistroF1007,
    RegistroF1008, RegistroF1009, RegistroF1011, RegistroF1012,
    RegistroF1647, RegistroF2276,
    atributos_f1001, atributos_f1003, atributos_f1005, atributos_f1006,
    atributos_f1007, atributos_f1008, atributos_f1009,
    atributos_f1011, atributos_f1012, atributos_f1647, atributos_f2276,
)


# ================================================================
# Cargar metadata del prevalidador (exacta)
# ================================================================

_METADATA_PATH = Path(__file__).parent / 'data_prevalidador.json'


def cargar_metadata_prevalidador() -> dict:
    """Carga el JSON con la estructura exacta de cada hoja del prevalidador."""
    with open(_METADATA_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)


METADATA = cargar_metadata_prevalidador()


# ================================================================
# ESTILOS — Imitando el prevalidador
# ================================================================

# Color azul oscuro como el prevalidador (color "21" = azul institucional DIAN)
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name='Arial', bold=True, color="FFFFFF", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FILL = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
TITLE_FONT = Font(name='Arial', bold=True, color="FFFFFF", size=12)
TITLE_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

OBLIGATORIO_FONT = Font(name='Arial', bold=True, color="FFFFFF", size=10)
OBLIGATORIO_FILL = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")

NUMERICO_ALIGN = Alignment(horizontal="right", vertical="center")
TEXTO_ALIGN = Alignment(horizontal="left", vertical="center")

THIN = Side(border_style="thin", color="999999")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


# ================================================================
# Estructura de una hoja
# ================================================================

def _crear_hoja_formato(
    wb: Workbook,
    formato: str,
    registros_attrs: list[dict],
    cabecera_info: dict | None = None,
):
    """
    Crea una hoja del formato indicado, imitando la estructura del prevalidador.
    
    Args:
        wb: workbook openpyxl
        formato: 'F1001', 'F1003', 'F1005', etc.
        registros_attrs: lista de dicts con los atributos del XSD ya formateados
                         (las llaves son los nombres de atributo: cpt, tdoc, nid, etc.)
        cabecera_info: opcional. Dict con {ano, nit_informante, razon_informante, ...}
    
    Estructura de hoja generada:
        R1: Título grande con identidad del formato
        R2: Fila vacía
        R3: Cabeceras (nombres visibles del prevalidador)
        R4: Marca de obligatoriedad ("*" en rojo si obligatorio)
        R5: Datos (atributo XSD entre paréntesis para referencia)
        R6+: Datos reales (uno por registro)
    """
    meta = METADATA[formato]
    columnas = meta['columnas']  # lista ordenada de definiciones
    n_cols = len(columnas)
    
    ws = wb.create_sheet(formato)
    
    # ============================================================
    # R1: Título grande con identidad
    # ============================================================
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1, value=meta['identidad'])
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = TITLE_ALIGN
    ws.row_dimensions[1].height = 35
    
    # ============================================================
    # R2: Información del informante (opcional)
    # ============================================================
    if cabecera_info:
        info_text = (
            f"Año gravable: {cabecera_info.get('ano', 2025)}  |  "
            f"Informante: {cabecera_info.get('razon_informante', '')}  "
            f"(NIT: {cabecera_info.get('nit_informante', '')})  |  "
            f"Registros: {len(registros_attrs)}"
        )
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        cell = ws.cell(row=2, column=1, value=info_text)
        cell.font = Font(name='Arial', italic=True, color="555555", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 18
    
    # ============================================================
    # R3: Cabeceras (nombres visibles del prevalidador)
    # ============================================================
    for i, col_def in enumerate(columnas, start=1):
        c = ws.cell(row=3, column=i, value=col_def['nombre'])
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = BORDER
    ws.row_dimensions[3].height = 50
    
    # ============================================================
    # R4: Marca de obligatoriedad
    # ============================================================
    for i, col_def in enumerate(columnas, start=1):
        if col_def['obligatorio']:
            c = ws.cell(row=4, column=i, value="* OBLIGATORIO")
            c.fill = OBLIGATORIO_FILL
            c.font = OBLIGATORIO_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
        else:
            c = ws.cell(row=4, column=i, value="opcional")
            c.font = Font(name='Arial', italic=True, color="666666", size=8)
            c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[4].height = 18
    
    # ============================================================
    # R5: Tipo y atributo XSD (referencia técnica)
    # ============================================================
    for i, col_def in enumerate(columnas, start=1):
        ref = f"({col_def['atributo']}) {col_def['tipo']}{col_def['longitud']}"
        c = ws.cell(row=5, column=i, value=ref)
        c.font = Font(name='Consolas', color="333333", size=8)
        c.fill = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[5].height = 15
    
    # ============================================================
    # R6+: Datos reales
    # ============================================================
    for row_idx, attrs in enumerate(registros_attrs, start=6):
        for col_idx, col_def in enumerate(columnas, start=1):
            atributo = col_def['atributo']
            valor = attrs.get(atributo, '')
            
            # Convertir tipo según definición
            if col_def['tipo'] == 'N' and valor not in ('', None):
                try:
                    valor = int(valor) if isinstance(valor, str) and valor.isdigit() else valor
                except (ValueError, TypeError):
                    pass
            
            c = ws.cell(row=row_idx, column=col_idx, value=valor if valor != '' else None)
            
            # Alineación según tipo
            if col_def['tipo'] == 'N':
                c.alignment = NUMERICO_ALIGN
                # Formato numérico con separador de miles
                try:
                    longitud = int(col_def['longitud']) if col_def['longitud'] is not None else 0
                except (ValueError, TypeError):
                    longitud = 0
                if longitud >= 10:
                    c.number_format = '#,##0'
            else:
                c.alignment = TEXTO_ALIGN
            
            c.border = BORDER
            c.font = Font(name='Calibri', size=10)
    
    # ============================================================
    # ANCHOS DE COLUMNA (basados en el prevalidador)
    # ============================================================
    for i, col_def in enumerate(columnas, start=1):
        letra = get_column_letter(i)
        ws.column_dimensions[letra].width = max(col_def.get('ancho_excel', 15), 10)
    
    # ============================================================
    # CONGELAR primeras 5 filas (cabecera + obligatoriedad + tipo)
    # ============================================================
    ws.freeze_panes = 'A6'
    
    return ws


# ================================================================
# Hoja de RESUMEN
# ================================================================

def _crear_hoja_resumen(wb: Workbook, info_global: dict, formatos_data: dict):
    """
    Hoja resumen con totales por formato.
    
    Args:
        info_global: {ano, nit_informante, razon_informante, fecha_generacion}
        formatos_data: {'F1005': {'registros': N, 'valor_total': X}, ...}
    """
    ws = wb.create_sheet('Resumen', 0)
    
    # Título
    ws.merge_cells('A1:D1')
    c = ws.cell(row=1, column=1, value="INFORMACIÓN EXÓGENA — RESUMEN")
    c.fill = TITLE_FILL
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=14)
    c.alignment = TITLE_ALIGN
    ws.row_dimensions[1].height = 30
    
    # Info del informante
    fila = 3
    info_rows = [
        ('Año gravable', info_global.get('ano', 2025)),
        ('NIT informante', info_global.get('nit_informante', '')),
        ('Razón social', info_global.get('razon_informante', '')),
        ('Fecha generación', info_global.get('fecha_generacion', '')),
        ('Resolución vigente', '000227 de 2025 + modificaciones'),
    ]
    for label, value in info_rows:
        ws.cell(row=fila, column=1, value=label).font = Font(bold=True)
        ws.cell(row=fila, column=2, value=str(value))
        fila += 1
    
    fila += 2
    
    # Tabla de formatos
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
    c = ws.cell(row=fila, column=1, value="FORMATOS GENERADOS")
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = TITLE_ALIGN
    fila += 1
    
    headers = ['Formato', 'Descripción', 'Registros', 'Valor Total']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=fila, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = BORDER
    fila += 1
    
    total_general = 0
    for fmt_key in ['F1001', 'F1003', 'F1005', 'F1006', 'F1007', 'F1008', 'F1009',
                     'F1011', 'F1012', 'F1647', 'F2276']:
        if fmt_key not in formatos_data:
            continue
        data = formatos_data[fmt_key]
        meta = METADATA[fmt_key]
        
        ws.cell(row=fila, column=1, value=fmt_key).font = Font(bold=True)
        ws.cell(row=fila, column=2, value=meta['identidad'].split(' - ')[-1] if ' - ' in meta['identidad'] else meta['identidad'])
        c_n = ws.cell(row=fila, column=3, value=data.get('registros', 0))
        c_n.alignment = NUMERICO_ALIGN
        c_v = ws.cell(row=fila, column=4, value=data.get('valor_total', 0))
        c_v.number_format = '"$"#,##0'
        c_v.alignment = NUMERICO_ALIGN
        
        for col in range(1, 5):
            ws.cell(row=fila, column=col).border = BORDER
        
        total_general += data.get('valor_total', 0)
        fila += 1
    
    # Total general
    ws.cell(row=fila, column=1, value='TOTAL').font = Font(bold=True, size=11)
    c_tot = ws.cell(row=fila, column=4, value=total_general)
    c_tot.number_format = '"$"#,##0'
    c_tot.font = Font(bold=True, size=11)
    c_tot.alignment = NUMERICO_ALIGN
    
    # Anchos
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 22
    
    return ws


# ================================================================
# Función principal: generar el Excel
# ================================================================

def generar_excel_prevalidador(
    ruta_salida: str | Path,
    info_global: dict,
    registros_por_formato: dict,
) -> Path:
    """
    Genera el Excel maestro idéntico a las hojas del prevalidador DIAN.
    
    Args:
        ruta_salida: ruta donde guardar el .xlsx
        info_global: {
            'ano': 2025,
            'nit_informante': '900533491',
            'razon_informante': 'QUINTO SENTIDO SAS',
            'fecha_generacion': '2026-04-15',
        }
        registros_por_formato: {
            'F1001': {'registros': [RegistroF1001, ...]},
            'F1003': {'registros': [RegistroF1003, ...]},
            'F1005': {'registros': [RegistroF1005, ...]},
            ...
        }
    
    Returns:
        Path del archivo generado.
    """
    wb = Workbook()
    # Quitar la hoja default
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Mapeo formato → función de atributos + extractor de valor
    FUNCS_ATRIBUTOS = {
        'F1001': (atributos_f1001, lambda r: r.pago_deducible + r.pago_no_deducible),
        'F1003': (atributos_f1003, lambda r: r.retencion),
        'F1005': (atributos_f1005, lambda r: r.iva_descontable + r.iva_dev_ventas + r.iva_mayor_costo),
        'F1006': (atributos_f1006, lambda r: r.iva_generado + r.iva_dev_compras + r.impuesto_consumo),
        'F1007': (atributos_f1007, lambda r: r.ingresos_brutos),
        'F1008': (atributos_f1008, lambda r: r.saldo),
        'F1009': (atributos_f1009, lambda r: r.saldo),
        'F1011': (atributos_f1011, lambda r: r.saldo),
        'F1012': (atributos_f1012, lambda r: r.valor),
        'F1647': (atributos_f1647, lambda r: r.valor_total),
        'F2276': (atributos_f2276, lambda r: r.total_ingresos_brutos),
    }
    
    # Acumular info para el resumen
    formatos_resumen = {}
    
    # Crear hoja por formato (orden oficial de los formatos)
    ORDEN_FORMATOS = [
        'F1001', 'F1003', 'F1005', 'F1006', 'F1007', 'F1008', 'F1009',
        'F1011', 'F1012', 'F1647', 'F2276',
    ]
    for fmt_key in ORDEN_FORMATOS:
        if fmt_key not in registros_por_formato:
            continue
        
        registros = registros_por_formato[fmt_key].get('registros', [])
        if not registros:
            continue
        
        fn_attrs, fn_valor = FUNCS_ATRIBUTOS[fmt_key]
        registros_attrs = [fn_attrs(r) for r in registros]
        valor_total = sum(fn_valor(r) for r in registros)
        
        _crear_hoja_formato(wb, fmt_key, registros_attrs, info_global)
        
        formatos_resumen[fmt_key] = {
            'registros': len(registros),
            'valor_total': valor_total,
        }
    
    # Crear hoja de resumen al principio
    _crear_hoja_resumen(wb, info_global, formatos_resumen)
    
    # Guardar
    ruta = Path(ruta_salida)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)
    return ruta
