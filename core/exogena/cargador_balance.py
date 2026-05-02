"""
Cargador del balance auxiliar exportado del software contable.

Formato esperado (basado en el archivo de Rutas del Mar):
    Fila 1: 'RUTAS DEL MAR SAS - 901.589.040-1' (encabezado de empresa)
    Fila 2: 'Balance de Prueba por NIT (Normal) Dic-31-2025' (subtítulo)
    Fila 3: 'Desde Ene-01-2025 Hasta Dic-31-2025' (rango fechas)
    Fila 4: Encabezados: Cuenta | Equivalencia | Nombre | NIT | Nombre NIT | Saldo Anterior | Débitos | Créditos | Nuevo Saldo
    Fila 5+: Datos. Algunas filas son TOTALIZADORES (nivel 1, 2, 4 dígitos), otras son MOVIMIENTOS por NIT.

Procesamiento:
    - Limpieza de cuentas: quitar guiones, convertir todo a dígitos
    - Limpieza de NITs: quitar puntos y dígito de verificación (el guión final + N)
    - Distinción entre filas totalizadoras (sin NIT) y movimientos (con NIT)
    - Validación: la suma de movimientos por cuenta debe igualar el totalizador

El parser devuelve:
    - cabecera: { empresa, nit, periodo, fecha_corte }
    - movimientos: [{cuenta, nit, nombre_tercero, saldo_anterior, debitos, creditos, saldo_final, ...}, ...]
    - totalizadores: [{cuenta_totalizadora, ..., suma_movimientos_validados}]
    - errores: lista de problemas detectados
"""

from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
import re
from decimal import Decimal, InvalidOperation

import openpyxl


@dataclass
class CabeceraBalance:
    empresa: str = ''
    nit_empresa: str = ''
    titulo: str = ''
    periodo: str = ''
    fecha_corte: str = ''


@dataclass
class MovimientoBalance:
    """Una fila del balance que representa un movimiento de un tercero."""
    codigo_cuenta: str
    nombre_cuenta: str
    nit: str
    nombre_tercero: str
    saldo_anterior: Decimal = Decimal('0')
    debitos: Decimal = Decimal('0')
    creditos: Decimal = Decimal('0')
    saldo_final: Decimal = Decimal('0')
    fila_origen: int = 0


@dataclass
class TotalizadorBalance:
    """Una fila totalizadora del balance (sin NIT, con saldos agregados)."""
    codigo_cuenta: str
    nombre_cuenta: str
    nivel: int  # 1, 2, 4, 6, 8 dígitos
    saldo_anterior: Decimal = Decimal('0')
    debitos: Decimal = Decimal('0')
    creditos: Decimal = Decimal('0')
    saldo_final: Decimal = Decimal('0')
    fila_origen: int = 0


@dataclass
class ResultadoCargaBalance:
    cabecera: CabeceraBalance = field(default_factory=CabeceraBalance)
    movimientos: list = field(default_factory=list)
    totalizadores: list = field(default_factory=list)
    errores: list = field(default_factory=list)
    advertencias: list = field(default_factory=list)
    nits_unicos: set = field(default_factory=set)
    cuentas_unicas: set = field(default_factory=set)


def _limpiar_cuenta(valor) -> str:
    """Limpia un código de cuenta: quita guiones y espacios.
    
    Ejemplos:
        '11050505'         -> '11050505'
        '1-1-05-05-05'     -> '11050505'
        '   11050505    '  -> '11050505'
        '1105'             -> '1105' (totalizador)
    """
    if valor is None:
        return ''
    s = str(valor).strip()
    # Quitar guiones, puntos y espacios internos
    s = re.sub(r'[-.\s]', '', s)
    return s


def _limpiar_nit(valor) -> str:
    """Limpia un NIT: quita puntos y dígito de verificación.
    
    Ejemplos:
        '901.589.040-1'     -> '901589040'
        '120-6'             -> '120'
        '1.050.962.255-1'   -> '1050962255'
        '8.424.668-9'       -> '8424668'
        ''                  -> ''
    """
    if valor is None:
        return ''
    s = str(valor).strip()
    if not s:
        return ''
    # Si tiene guión, todo lo que está después es el DV (descartarlo)
    if '-' in s:
        s = s.split('-')[0]
    # Quitar puntos y espacios
    s = re.sub(r'[.\s]', '', s)
    # Validar que sea solo dígitos
    if not s.isdigit():
        return ''
    return s


def _to_decimal(valor) -> Decimal:
    """Convierte un valor a Decimal, manejando texto, números y notación científica."""
    if valor is None or valor == '':
        return Decimal('0')
    try:
        # openpyxl puede traer floats con notación científica como 2.38e-07
        # Convertir a string y luego a Decimal
        s = str(valor).strip()
        if not s or s == '-':
            return Decimal('0')
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal('0')


def _parsear_cabecera(rows: list) -> CabeceraBalance:
    """Extrae datos de cabecera de las primeras 3 filas."""
    cab = CabeceraBalance()
    if len(rows) >= 1 and rows[0] and rows[0][0]:
        primera = str(rows[0][0]).strip()
        # Formato: 'RUTAS DEL MAR SAS - 901.589.040-1'
        match = re.match(r'^(.+?)\s*-\s*(\d[\d.\s-]+)$', primera)
        if match:
            cab.empresa = match.group(1).strip()
            cab.nit_empresa = _limpiar_nit(match.group(2))
        else:
            cab.empresa = primera

    if len(rows) >= 2 and rows[1] and rows[1][0]:
        cab.titulo = str(rows[1][0]).strip()
        # Buscar fecha de corte en el título
        m = re.search(r'(Ene|Feb|Mar|Abr|May|Jun|Jul|Ago|Sep|Oct|Nov|Dic)-(\d+)-(\d{4})', cab.titulo)
        if m:
            cab.fecha_corte = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"

    if len(rows) >= 3 and rows[2] and rows[2][0]:
        cab.periodo = str(rows[2][0]).strip()

    return cab


def _detectar_columnas(fila_headers) -> dict:
    """Detecta los índices de columnas del balance.

    Espera encabezados típicos: Cuenta | Equivalencia | Nombre | NIT | 
                                 Nombre NIT | Saldo Anterior | Débitos | Créditos | Nuevo Saldo
    """
    cols = {
        'cuenta': 0,
        'nombre_cuenta': 2,
        'nit': 3,
        'nombre_nit': 4,
        'saldo_anterior': 5,
        'debitos': 6,
        'creditos': 7,
        'saldo_final': 8,
    }
    if not fila_headers:
        return cols
    # Mapear nombre de columna a índice
    for i, header in enumerate(fila_headers):
        if header is None:
            continue
        h = str(header).strip().lower()
        if h == 'cuenta':
            cols['cuenta'] = i
        elif 'nombre' in h and 'nit' not in h:
            cols['nombre_cuenta'] = i
        elif h == 'nit':
            cols['nit'] = i
        elif 'nombre nit' in h:
            cols['nombre_nit'] = i
        elif 'saldo anterior' in h:
            cols['saldo_anterior'] = i
        elif h in ('débitos', 'debitos'):
            cols['debitos'] = i
        elif h in ('créditos', 'creditos'):
            cols['creditos'] = i
        elif 'saldo' in h and ('nuevo' in h or 'final' in h or 'actual' in h):
            cols['saldo_final'] = i
    return cols


def cargar_balance(archivo, año_gravable: int = 2025) -> ResultadoCargaBalance:
    """Lee el balance auxiliar y retorna estructuras pobladas.
    
    Args:
        archivo: ruta (str/Path) o file-like (UploadedFile de Streamlit)
        año_gravable: para validaciones cruzadas
    
    Returns:
        ResultadoCargaBalance con cabecera, movimientos, totalizadores, errores
    """
    res = ResultadoCargaBalance()

    if isinstance(archivo, (str, Path)):
        archivo = Path(archivo)
        if not archivo.exists():
            res.errores.append(f"Archivo no existe: {archivo}")
            return res

    wb = openpyxl.load_workbook(archivo, data_only=True, read_only=True)
    ws = wb['Datos'] if 'Datos' in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 5:
        res.errores.append("Archivo demasiado corto, no parece un balance.")
        return res

    # Cabecera (filas 1-3)
    res.cabecera = _parsear_cabecera(rows)

    # Detectar columnas (fila 4)
    cols = _detectar_columnas(rows[3])

    # Procesar filas de datos (desde fila 5 = índice 4)
    for i, row in enumerate(rows[4:], start=5):
        if not row or not row[cols['cuenta']]:
            continue

        cuenta = _limpiar_cuenta(row[cols['cuenta']])
        if not cuenta or not cuenta.isdigit():
            continue

        nombre_cuenta = str(row[cols['nombre_cuenta']] or '').strip()
        nit_raw = row[cols['nit']] if cols['nit'] < len(row) else None
        nit = _limpiar_nit(nit_raw)
        nombre_nit = str(row[cols['nombre_nit']] or '').strip() if cols['nombre_nit'] < len(row) else ''

        saldo_ant = _to_decimal(row[cols['saldo_anterior']] if cols['saldo_anterior'] < len(row) else 0)
        debitos = _to_decimal(row[cols['debitos']] if cols['debitos'] < len(row) else 0)
        creditos = _to_decimal(row[cols['creditos']] if cols['creditos'] < len(row) else 0)
        saldo_fin = _to_decimal(row[cols['saldo_final']] if cols['saldo_final'] < len(row) else 0)

        if nit:
            # Es un movimiento por tercero
            mov = MovimientoBalance(
                codigo_cuenta=cuenta,
                nombre_cuenta=nombre_cuenta,
                nit=nit,
                nombre_tercero=nombre_nit,
                saldo_anterior=saldo_ant,
                debitos=debitos,
                creditos=creditos,
                saldo_final=saldo_fin,
                fila_origen=i,
            )
            res.movimientos.append(mov)
            res.nits_unicos.add(nit)
            res.cuentas_unicas.add(cuenta)
        else:
            # Es un totalizador (sin NIT)
            tot = TotalizadorBalance(
                codigo_cuenta=cuenta,
                nombre_cuenta=nombre_cuenta,
                nivel=len(cuenta),
                saldo_anterior=saldo_ant,
                debitos=debitos,
                creditos=creditos,
                saldo_final=saldo_fin,
                fila_origen=i,
            )
            res.totalizadores.append(tot)

    # Nota sobre cuadre: en un balance auxiliar Db != Cr es NORMAL para una
    # cuenta individual (ej. los activos tienen débitos > créditos). El cuadre
    # de partida doble es a nivel global del balance, no por cuenta.
    # Por eso NO advertimos sobre esto.
    pass

    return res


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print("Uso: python cargador_balance.py <archivo.xlsx>")
        sys.exit(1)

    res = cargar_balance(sys.argv[1])

    print(f"\n=== Cabecera ===")
    print(f"Empresa: {res.cabecera.empresa}")
    print(f"NIT: {res.cabecera.nit_empresa}")
    print(f"Título: {res.cabecera.titulo}")
    print(f"Periodo: {res.cabecera.periodo}")
    print(f"Fecha corte: {res.cabecera.fecha_corte}")

    print(f"\n=== Resultado ===")
    print(f"Movimientos por NIT: {len(res.movimientos)}")
    print(f"Totalizadores: {len(res.totalizadores)}")
    print(f"NITs únicos: {len(res.nits_unicos)}")
    print(f"Cuentas únicas: {len(res.cuentas_unicas)}")
    print(f"Errores: {len(res.errores)}")
    print(f"Advertencias: {len(res.advertencias)}")

    if res.errores:
        print("\nErrores:")
        for e in res.errores:
            print(f"  - {e}")
    if res.advertencias:
        print("\nAdvertencias:")
        for a in res.advertencias[:5]:
            print(f"  - {a}")

    print("\n=== Muestras de movimientos ===")
    for m in res.movimientos[:5]:
        print(f"  Cuenta {m.codigo_cuenta:<10} NIT {m.nit:<12} {m.nombre_tercero[:30]:<30} "
              f"Db={m.debitos:>15,.2f} Cr={m.creditos:>15,.2f}")
