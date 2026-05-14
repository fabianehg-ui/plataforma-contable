"""
Conciliación Enriquecida — UI Streamlit
========================================

Función pública: render_conciliacion_enriquecida(sb, empresa_id, año_gravable)

Se inyecta dentro del tab "Conciliación" de pages/7_Informacion_Exogena.py.

Muestra:
    1. Resumen ejecutivo (cuántos conceptos cuadrados / con diferencia / pendientes)
    2. Filtros: por formato, por estado de cuadre, por monto
    3. Tabla expandible — un acordeón por (formato, concepto)
       que al abrirse muestra las cuentas que aportan y la diferencia
    4. Botón para exportar a Excel
"""

from __future__ import annotations
import io

import streamlit as st
import pandas as pd

from core.exogena import conciliacion_enriquecida_1 as ce


# ============================================================================
# Punto de entrada
# ============================================================================

def render_conciliacion_enriquecida(
    sb,
    empresa_id: str,
    empresa_nombre: str,
    año_gravable: int = 2025,
):
    """Renderiza el componente completo de conciliación enriquecida."""

    st.markdown("---")
    st.subheader("📊 Conciliación detallada por concepto")
    st.caption(
        "Cada formato/concepto desagregado en sus cuentas componentes, "
        "con saldo del balance vs valor reportado y diferencias explicadas."
    )

    # Botón para construir/refrescar el dictamen
    col_btn, col_info = st.columns([1, 3])
    with col_btn:
        if st.button("🔄 Construir dictamen", key="cce_btn_construir"):
            with st.spinner("Analizando movimientos..."):
                try:
                    dictamen = ce.construir_conciliacion_enriquecida(
                        sb=sb,
                        empresa_id=empresa_id,
                        año_gravable=año_gravable,
                    )
                    st.session_state["cce_dictamen"] = dictamen
                except Exception as e:
                    st.error(f"Error construyendo dictamen: {e}")
                    return

    if "cce_dictamen" not in st.session_state:
        with col_info:
            st.info(
                "Click 'Construir dictamen' para generar el análisis. "
                "Requiere que el motor de clasificación se haya ejecutado primero."
            )
        return

    dictamen = st.session_state["cce_dictamen"]

    # ============================================================
    # Resumen ejecutivo
    # ============================================================
    resumen = ce.dictamen_a_resumen_dict(dictamen)

    st.markdown("### Resumen ejecutivo")
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total conceptos", resumen["total_conceptos"])
    col2.metric(
        "✅ Cuadrados",
        resumen["conceptos_cuadrados"],
        delta=f"{resumen['porcentaje_cuadrado']:.0f}%",
    )
    col3.metric(
        "⚠️ Con diferencia explicada",
        resumen["conceptos_con_diferencia_explicada"],
    )
    col4.metric(
        "❌ Pendientes",
        resumen["conceptos_pendientes"],
    )

    # ============================================================
    # Filtros
    # ============================================================
    st.markdown("### Filtros")
    col_f1, col_f2, col_f3 = st.columns(3)

    formatos_disponibles = sorted({b.formato_dian for b in dictamen.bloques})
    with col_f1:
        formato_filtro = st.selectbox(
            "Formato",
            options=["(todos)"] + formatos_disponibles,
            key="cce_filtro_formato",
        )

    with col_f2:
        estado_filtro = st.selectbox(
            "Estado",
            options=[
                "(todos)",
                "✅ Cuadrados",
                "⚠️ Con diferencia explicada",
                "❌ Pendientes",
            ],
            key="cce_filtro_estado",
        )

    with col_f3:
        monto_minimo = st.number_input(
            "Monto mínimo (filtra conceptos pequeños)",
            min_value=0,
            value=0,
            step=100_000,
            key="cce_filtro_monto",
        )

    # Aplicar filtros
    bloques_filtrados = []
    for b in dictamen.bloques:
        if formato_filtro != "(todos)" and b.formato_dian != formato_filtro:
            continue
        if estado_filtro != "(todos)":
            mapping = {
                "✅ Cuadrados": "cuadrado",
                "⚠️ Con diferencia explicada": "diferencia_explicada",
                "❌ Pendientes": "pendiente",
            }
            if b.estado != mapping[estado_filtro]:
                continue
        if abs(b.total_balance) < monto_minimo:
            continue
        bloques_filtrados.append(b)

    st.caption(f"Mostrando {len(bloques_filtrados)} de {len(dictamen.bloques)} bloques")

    # ============================================================
    # Botón de descarga Excel
    # ============================================================
    if bloques_filtrados:
        excel_buf = _generar_excel(dictamen, bloques_filtrados)
        st.download_button(
            "📥 Descargar conciliación detallada (Excel)",
            data=excel_buf,
            file_name=f"conciliacion_detallada_{año_gravable}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="cce_btn_excel",
        )

    # ============================================================
    # Tabla expandible — un expander por (formato, concepto)
    # ============================================================
    if not bloques_filtrados:
        st.info("No hay bloques con los filtros actuales.")
        return

    st.markdown("### Detalle por concepto")

    for bloque in bloques_filtrados:
        # Icono según estado
        icon = {
            "cuadrado": "✅",
            "diferencia_explicada": "⚠️",
            "pendiente": "❌",
            "inconsistente": "🚨",
        }.get(bloque.estado, "❓")

        # Encabezado del expander
        diff_str = (
            f" (dif ${bloque.diferencia:,.0f})"
            if bloque.estado != "cuadrado"
            else ""
        )

        with st.expander(
            f"{icon} **F{bloque.formato_dian} / Concepto {bloque.concepto_dian}** — "
            f"{bloque.descripcion_concepto[:50]} — "
            f"Balance: ${bloque.total_balance:,.0f} → "
            f"Reportado: ${bloque.total_reportado:,.0f}{diff_str}"
        ):
            # Tabla de cuentas
            df = pd.DataFrame([
                {
                    "Estado": "🚫" if c.excluida else "",
                    "Código": c.codigo_cuenta,
                    "Nombre": c.nombre_cuenta,
                    "Capa": c.capa_nombre,
                    "Saldo balance": f"${c.saldo_balance:,.0f}",
                    "Reportado": f"${c.valor_reportado:,.0f}",
                    "Diferencia": (
                        f"${c.diferencia:,.0f}" if c.diferencia != 0 else "—"
                    ),
                    "NITs": c.nits_distintos if c.nits_distintos > 0 else "",
                    "Motivo": c.motivo_diferencia or "",
                }
                for c in bloque.cuentas
            ])

            st.dataframe(df, use_container_width=True, hide_index=True)

            # Fila de totales
            col_t1, col_t2, col_t3 = st.columns(3)
            col_t1.metric("Total balance", f"${bloque.total_balance:,.0f}")
            col_t2.metric("Total reportado", f"${bloque.total_reportado:,.0f}")
            col_t3.metric(
                "Diferencia",
                f"${bloque.diferencia:,.0f}",
                delta=None if bloque.estado == "cuadrado" else "Revisar",
                delta_color="off" if bloque.estado == "cuadrado" else "inverse",
            )

            # Explicaciones
            if bloque.explicaciones:
                st.markdown("**Explicación de diferencias:**")
                for exp in bloque.explicaciones[:5]:
                    st.markdown(exp)
                if len(bloque.explicaciones) > 5:
                    st.caption(
                        f"...y {len(bloque.explicaciones) - 5} explicaciones más "
                        "(ver Excel descargable)"
                    )


# ============================================================================
# Generador de Excel
# ============================================================================

def _generar_excel(dictamen, bloques_filtrados) -> bytes:
    """Genera Excel multi-hoja con el dictamen.

    Estructura:
        - Resumen ejecutivo
        - Resumen por formato (totales y conteo de conceptos por formato)
        - Una hoja por cada formato (F1001, F1003, F1005, ...) con sus
          conceptos desglosados y filas de subtotal por concepto.
        - Conceptos con diferencia (solo si hay alguno)
        - Detalle plano completo (compatibilidad hacia atrás / auditoría)
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()

    # Filas planas (todos los bloques, no solo los filtrados, para que el
    # archivo refleje el dictamen completo y no dependa de la vista en pantalla).
    filas_planas = ce.dictamen_a_filas_excel(dictamen)

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # ────────────────────────────────────────────────────────────────
        # Hoja 1: Resumen ejecutivo
        # ────────────────────────────────────────────────────────────────
        resumen = ce.dictamen_a_resumen_dict(dictamen)
        df_resumen = pd.DataFrame([
            {"Métrica": "Año gravable", "Valor": resumen["año_gravable"]},
            {"Métrica": "Total conceptos", "Valor": resumen["total_conceptos"]},
            {"Métrica": "Cuadrados", "Valor": resumen["conceptos_cuadrados"]},
            {"Métrica": "Con diferencia explicada", "Valor": resumen["conceptos_con_diferencia_explicada"]},
            {"Métrica": "Pendientes", "Valor": resumen["conceptos_pendientes"]},
            {"Métrica": "% Cuadrado", "Valor": f"{resumen['porcentaje_cuadrado']:.1f}%"},
        ])
        df_resumen.to_excel(writer, sheet_name="Resumen", index=False)

        # ────────────────────────────────────────────────────────────────
        # Hoja 2: Resumen por formato (índice navegable)
        # ────────────────────────────────────────────────────────────────
        formatos: dict[str, dict] = {}
        for b in dictamen.bloques:
            fmt = b.formato_dian or "(sin_formato)"
            if fmt not in formatos:
                formatos[fmt] = {
                    "conceptos": 0,
                    "cuadrados": 0,
                    "con_diferencia": 0,
                    "pendientes": 0,
                    "total_balance": 0.0,
                    "total_reportado": 0.0,
                    "total_diferencia": 0.0,
                }
            f_info = formatos[fmt]
            f_info["conceptos"] += 1
            if b.estado == "cuadrado":
                f_info["cuadrados"] += 1
            elif b.estado == "diferencia_explicada":
                f_info["con_diferencia"] += 1
            else:
                f_info["pendientes"] += 1
            f_info["total_balance"] += float(b.total_balance)
            f_info["total_reportado"] += float(b.total_reportado)
            f_info["total_diferencia"] += float(b.diferencia)

        df_por_fmt = pd.DataFrame([
            {
                "Formato": f"F{fmt}" if fmt != "(sin_formato)" else fmt,
                "Conceptos": v["conceptos"],
                "Cuadrados": v["cuadrados"],
                "Con diferencia": v["con_diferencia"],
                "Pendientes": v["pendientes"],
                "Total balance": v["total_balance"],
                "Total reportado": v["total_reportado"],
                "Diferencia": v["total_diferencia"],
            }
            for fmt, v in sorted(formatos.items())
        ])
        df_por_fmt.to_excel(writer, sheet_name="Resumen por formato", index=False)

        # ────────────────────────────────────────────────────────────────
        # Hoja 3+: Una hoja por cada formato
        # ────────────────────────────────────────────────────────────────
        # Agrupar bloques por formato manteniendo el orden ya establecido
        # en dictamen.bloques (formato, concepto).
        bloques_por_fmt: dict[str, list] = {}
        for b in dictamen.bloques:
            bloques_por_fmt.setdefault(b.formato_dian or "(sin_formato)", []).append(b)

        # Estilos compartidos
        header_font = Font(bold=True, color="FFFFFF", name="Arial")
        header_fill = PatternFill("solid", start_color="305496")
        subtotal_font = Font(bold=True, name="Arial")
        subtotal_fill = PatternFill("solid", start_color="DDEBF7")
        excluida_fill = PatternFill("solid", start_color="FCE4D6")

        for fmt in sorted(bloques_por_fmt.keys()):
            sheet_name = _nombre_hoja_formato(fmt)
            bloques_fmt = bloques_por_fmt[fmt]

            filas_hoja = []
            for bloque in bloques_fmt:
                for c in bloque.cuentas:
                    filas_hoja.append({
                        "Concepto": bloque.concepto_dian,
                        "Descripción": bloque.descripcion_concepto,
                        "Cuenta": c.codigo_cuenta,
                        "Nombre cuenta": c.nombre_cuenta,
                        "Capa": c.capa_nombre,
                        "Saldo balance": float(c.saldo_balance),
                        "Valor reportado": float(c.valor_reportado),
                        "Diferencia": float(c.diferencia),
                        "NITs": c.nits_distintos,
                        "Estado": "🚫 Excluida" if c.excluida else "✓ Activa",
                        "Motivo": c.motivo_diferencia or "",
                    })
                # Subtotal por concepto
                filas_hoja.append({
                    "Concepto": bloque.concepto_dian,
                    "Descripción": bloque.descripcion_concepto,
                    "Cuenta": "TOTAL CONCEPTO",
                    "Nombre cuenta": "",
                    "Capa": "",
                    "Saldo balance": float(bloque.total_balance),
                    "Valor reportado": float(bloque.total_reportado),
                    "Diferencia": float(bloque.diferencia),
                    "NITs": "",
                    "Estado": bloque.estado.upper(),
                    "Motivo": "",
                })

            df_fmt = pd.DataFrame(filas_hoja)
            df_fmt.to_excel(writer, sheet_name=sheet_name, index=False)

            # Aplicar formato a la hoja recién creada
            ws = writer.sheets[sheet_name]

            # Header
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Anchos de columna
            anchos = {
                "A": 10, "B": 35, "C": 14, "D": 38, "E": 16,
                "F": 16, "G": 16, "H": 14, "I": 8, "J": 14, "K": 35,
            }
            for col, w in anchos.items():
                ws.column_dimensions[col].width = w

            # Formato numérico (col F=Saldo, G=Reportado, H=Diferencia)
            for row in range(2, ws.max_row + 1):
                for col_letter in ("F", "G", "H"):
                    ws[f"{col_letter}{row}"].number_format = '#,##0;(#,##0);-'

            # Resaltar filas de subtotal y de exclusiones
            for row_idx, fila in enumerate(filas_hoja, start=2):
                if fila["Cuenta"] == "TOTAL CONCEPTO":
                    for col_idx in range(1, len(df_fmt.columns) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font = subtotal_font
                        cell.fill = subtotal_fill
                elif fila.get("Estado", "").startswith("🚫"):
                    for col_idx in range(1, len(df_fmt.columns) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = excluida_fill

            # Congelar primera fila para navegación cómoda
            ws.freeze_panes = "A2"

        # ────────────────────────────────────────────────────────────────
        # Hoja: Conceptos con diferencia (solo si hay alguno)
        # ────────────────────────────────────────────────────────────────
        filas_diff = [
            f for f in filas_planas
            if f["Cuenta"] == "TOTAL CONCEPTO" and f["Diferencia"] != 0
        ]
        if filas_diff:
            df_diff = pd.DataFrame(filas_diff)
            df_diff.to_excel(writer, sheet_name="Conceptos con diferencia", index=False)

        # ────────────────────────────────────────────────────────────────
        # Hoja final: Detalle plano completo (compatibilidad / auditoría)
        # ────────────────────────────────────────────────────────────────
        df_detalle = pd.DataFrame(filas_planas)
        df_detalle.to_excel(writer, sheet_name="Detalle por concepto", index=False)

    buf.seek(0)
    return buf.getvalue()


def _nombre_hoja_formato(fmt: str) -> str:
    """Genera un nombre de hoja válido para Excel (≤31 chars, sin caracteres
    prohibidos: : \\ / ? * [ ])."""
    if not fmt or fmt == "(sin_formato)":
        return "Sin formato"
    base = f"F{fmt}" if not fmt.startswith("F") else fmt
    # Excel no permite ciertos caracteres en nombres de hoja
    for char in [":", "\\", "/", "?", "*", "[", "]"]:
        base = base.replace(char, "_")
    return base[:31]
