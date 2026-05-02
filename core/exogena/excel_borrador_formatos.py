"""
Generador de Excel borrador completo para Información Exógena DIAN.

Este Excel es el documento maestro de auditoría tributaria que muestra:
    1. Hoja Resumen: estado general de los formatos
    2. Una hoja por cada formato a reportar (1001, 1003, 1007, etc.)
       - Datos borrador: detalle por NIT (cómo va al XML)
       - Conciliación: balance vs reportado por concepto
    3. Hoja de cuentas no reportadas (agrupadas por motivo)
    4. Hoja de auditoría: reglas y conciliaciones aplicadas

El contador puede usar este Excel para:
    - Validar que los datos del XML estén correctos antes de enviar
    - Justificar las diferencias entre balance y formatos
    - Tener un soporte de auditoría completo
"""

from __future__ import annotations
from datetime import datetime
from io import BytesIO
from decimal import Decimal
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from .conciliacion import (
        detectar_tipo_conciliacion,
        construir_cuadre_balance_vs_formatos,
        PREFIJOS_EXCLUIR_SS,
    )
except ImportError:
    from conciliacion import (
        detectar_tipo_conciliacion,
        construir_cuadre_balance_vs_formatos,
        PREFIJOS_EXCLUIR_SS,
    )


# ================================================================
# ESTILOS
# ================================================================

HEADER_FILL = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill(start_color="5DADE2", end_color="5DADE2", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)

TITLE_FILL = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
TITLE_FONT = Font(bold=True, color="FFFFFF", size=14)

CONCILIACION_FILL = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="ABEBC6", end_color="ABEBC6", fill_type="solid")
WARNING_FILL = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")
ERROR_FILL = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
EXCLUIDO_FILL = PatternFill(start_color="D5DBDB", end_color="D5DBDB", fill_type="solid")

THIN = Side(border_style="thin", color="999999")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


# ================================================================
# NOMBRES DE FORMATOS (para mostrar en hojas)
# ================================================================

NOMBRES_FORMATOS = {
    '1001': '1001 - Pagos o abonos en cuenta y retenciones practicadas',
    '1003': '1003 - Retenciones en la fuente que le practicaron',
    '1004': '1004 - Descuentos tributarios solicitados',
    '1005': '1005 - Impuesto a las ventas por pagar',
    '1006': '1006 - Impuesto a las ventas por pagar - generado',
    '1007': '1007 - Ingresos recibidos',
    '1008': '1008 - Saldos de cuentas por cobrar',
    '1009': '1009 - Saldos de cuentas por pagar',
    '1010': '1010 - Información de socios o accionistas',
    '1011': '1011 - Información de declaraciones tributarias',
    '1012': '1012 - Saldos de cuentas, inversiones y otros',
    '1056': '1056 - Pagos a Establecimientos Permanentes',
    '1647': '1647 - Ingresos recibidos para terceros',
    '2275': '2275 - Costos y gastos por pagos al exterior',
    '2276': '2276 - Información de rentas de trabajo',
    '2278': '2278 - Información de cuentas en el exterior',
    '5253': '5253 - Pagos por GMF',
}

DESCRIPCIONES_CONCEPTOS = {
    5002: 'Honorarios',
    5003: 'Comisiones',
    5004: 'Servicios',
    5005: 'Arrendamientos',
    5006: 'Intereses y rendimientos financieros causados',
    5007: 'Compra de activos movibles',
    5008: 'Compra de activos fijos',
    5010: 'Aportes parafiscales SENA, Cajas, ICBF',
    5011: 'Pagos a EPS / Salud / ARL',
    5012: 'Aportes pensiones obligatorios y voluntarios',
    5013: 'Donaciones en dinero',
    5014: 'Donaciones en activos diferentes a dinero',
    5015: 'Impuestos solicitados como deducción',
    5016: 'Demás costos y deducciones',
    5044: 'Loterías, rifas, apuestas',
    5045: 'Retenciones en tarjetas débito y crédito',
    5055: 'Viáticos',
    5056: 'Gastos de representación',
    5101: 'GMF (Gravamen movimientos financieros) - 50% deducible',
}


# ================================================================
# UTILIDADES
# ================================================================

def _aplicar_titulo(ws, fila, texto, ncols=8):
    """Aplica un título grande a toda una fila."""
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ncols)
    cell = ws.cell(row=fila, column=1, value=texto)
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[fila].height = 25


def _aplicar_header(ws, fila, columnas, fill=HEADER_FILL, font=HEADER_FONT):
    """Aplica formato a una fila de encabezados."""
    for col, value in enumerate(columnas, start=1):
        cell = ws.cell(row=fila, column=col, value=value)
        cell.fill = fill
        cell.font = font
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws.row_dimensions[fila].height = 35


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _normalizar_nit(nit_raw):
    """Devuelve (nit_sin_dv, dv) — separa el dígito de verificación si existe."""
    if not nit_raw:
        return ('', '')
    s = str(nit_raw).strip().replace('.', '').replace(' ', '')
    if '-' in s:
        partes = s.split('-')
        return (partes[0], partes[1] if len(partes) > 1 else '')
    return (s, '')


def _tipo_documento_default(nit: str) -> str:
    """Determina el tipo de documento según rangos DIAN.
    
    Returns: '13' (CC), '31' (NIT), '41' (Pasaporte), etc.
    """
    if not nit:
        return ''
    try:
        n = int(str(nit).strip())
    except (ValueError, TypeError):
        return ''
    
    # Rangos oficiales DIAN
    if 10000 <= n < 100_000_000:
        return '13'   # Cédula de ciudadanía vieja o nueva
    elif 100_000_000 <= n < 600_000_000:
        return '13'   # CC nueva
    elif 600_000_000 <= n < 800_000_000:
        return '31'   # NIT residentes
    elif 800_000_000 <= n < 1_000_000_000:
        return '31'   # NIT empresarial
    elif n >= 1_000_000_000:
        return '13'   # CC nueva post-2000
    return ''


# ================================================================
# CONSTRUCCIÓN DE HOJAS POR FORMATO
# ================================================================

def _construir_hoja_formato(
    wb,
    formato_dian: str,
    movimientos: list[dict],
    cuadre_formato,  # CuadreFormato del módulo conciliacion
    terceros_dict: dict,  # nit → datos del tercero
):
    """Construye una hoja con datos borrador del formato + conciliación abajo.
    
    Args:
        wb: workbook openpyxl
        formato_dian: '1001', '1003', etc.
        movimientos: lista de dicts con datos clasificados (este formato)
        cuadre_formato: CuadreFormato del módulo conciliacion (este formato)
        terceros_dict: dict nit_sin_dv → {tipo_doc, dv, razon_social, ...}
    """
    nombre_hoja = f'F{formato_dian}'
    ws = wb.create_sheet(nombre_hoja)
    
    # ============================================================
    # TÍTULO
    # ============================================================
    _aplicar_titulo(ws, 1, NOMBRES_FORMATOS.get(formato_dian, f'Formato {formato_dian}'),
                    ncols=8)
    
    # ============================================================
    # SECCIÓN 1: DATOS BORRADOR (detalle por NIT)
    # ============================================================
    fila = 3
    ws.cell(row=fila, column=1, value="📋 DATOS BORRADOR (lo que va al XML)")
    ws.cell(row=fila, column=1).font = Font(bold=True, size=12, color="1A5276")
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=8)
    fila += 2
    
    headers = [
        "Tipo Doc", "NIT", "DV", "Razón Social",
        "Concepto", "Descripción Concepto", "Valor",
        "Cuenta Origen",
    ]
    _aplicar_header(ws, fila, headers)
    fila += 1
    
    fila_inicio_datos = fila
    
    # Agrupar movimientos por (NIT, concepto) para evitar duplicados
    consolidado = defaultdict(lambda: {
        'tipo_doc': '', 'nit': '', 'dv': '', 'razon_social': '',
        'concepto': None, 'valor': Decimal('0'), 'cuentas': set(),
    })
    
    for m in movimientos:
        nit_raw = m.get('nit', '') or ''
        concepto = m.get('concepto_dian')
        if concepto is None:
            continue
        
        # Normalizar NIT
        nit_clean, dv_extracted = _normalizar_nit(nit_raw)
        
        # Buscar tercero en el dict
        t = terceros_dict.get(nit_clean, {})
        
        # Tomar el valor según tipo de cuenta
        cuenta = str(m.get('codigo_cuenta', ''))
        debitos = Decimal(str(m.get('debitos', 0) or 0))
        creditos = Decimal(str(m.get('creditos', 0) or 0))
        
        if cuenta.startswith(('23', '25')):
            valor = abs(debitos)  # pasivos pagados: solo débitos
        else:
            valor = max(abs(debitos), abs(creditos))
        
        if valor <= 0:
            continue
        
        key = (nit_clean, concepto)
        c = consolidado[key]
        if not c['nit']:
            c['nit'] = nit_clean
            c['dv'] = t.get('dv') or dv_extracted
            c['tipo_doc'] = t.get('tipo_documento') or _tipo_documento_default(nit_clean)
            c['razon_social'] = t.get('razon_social') or t.get('nombre_completo') or m.get('nombre_tercero', '')
            c['concepto'] = concepto
        c['valor'] += valor
        c['cuentas'].add(cuenta)
    
    # Ordenar por concepto y valor descendente
    items = sorted(consolidado.values(), key=lambda x: (x['concepto'], -x['valor']))
    
    total_xml = Decimal('0')
    for item in items:
        ws.cell(row=fila, column=1, value=item['tipo_doc'])
        ws.cell(row=fila, column=2, value=item['nit'])
        ws.cell(row=fila, column=3, value=item['dv'])
        ws.cell(row=fila, column=4, value=str(item['razon_social'])[:60])
        ws.cell(row=fila, column=5, value=item['concepto'])
        ws.cell(row=fila, column=6, value=DESCRIPCIONES_CONCEPTOS.get(
            item['concepto'], f"Concepto {item['concepto']}"
        )[:50])
        cell_valor = ws.cell(row=fila, column=7, value=float(item['valor']))
        cell_valor.number_format = '"$"#,##0.00'
        ws.cell(row=fila, column=8, value=', '.join(sorted(item['cuentas']))[:40])
        
        # Bordes
        for col in range(1, 9):
            ws.cell(row=fila, column=col).border = BORDER
        
        total_xml += item['valor']
        fila += 1
    
    # Si no hay datos
    if not items:
        ws.cell(row=fila, column=1, value="(No hay movimientos para este formato)")
        ws.cell(row=fila, column=1).font = Font(italic=True, color="888888")
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=8)
        fila += 1
    else:
        # Total XML
        fila_total = fila
        ws.cell(row=fila_total, column=1, value=f"TOTAL XML ({len(items)} filas)")
        ws.merge_cells(start_row=fila_total, start_column=1, end_row=fila_total, end_column=6)
        cell_t = ws.cell(row=fila_total, column=7, value=float(total_xml))
        cell_t.number_format = '"$"#,##0.00'
        for col in range(1, 9):
            ws.cell(row=fila_total, column=col).fill = TOTAL_FILL
            ws.cell(row=fila_total, column=col).font = Font(bold=True)
            ws.cell(row=fila_total, column=col).border = BORDER
        fila += 1
    
    # ============================================================
    # SECCIÓN 2: CONCILIACIÓN (debajo del detalle)
    # ============================================================
    fila += 3
    ws.cell(row=fila, column=1, value="📐 CONCILIACIÓN: BALANCE vs VALOR REPORTADO")
    ws.cell(row=fila, column=1).font = Font(bold=True, size=12, color="1A5276")
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=8)
    fila += 2
    
    headers_concilia = [
        "Concepto", "Descripción", "Cuentas", "Movs",
        "Saldo Balance", "A Reportar", "Diferencia", "Estado / Motivo",
    ]
    _aplicar_header(ws, fila, headers_concilia, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
    fila += 1
    
    if cuadre_formato is None or not cuadre_formato.filas:
        ws.cell(row=fila, column=1, value="(No hay datos de conciliación)")
        ws.cell(row=fila, column=1).font = Font(italic=True, color="888888")
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=8)
        fila += 1
    else:
        for f in cuadre_formato.filas:
            ws.cell(row=fila, column=1, value=f.concepto_dian)
            ws.cell(row=fila, column=2, value=str(f.descripcion_concepto)[:45])
            ws.cell(row=fila, column=3, value=f.cantidad_cuentas)
            ws.cell(row=fila, column=4, value=f.cantidad_movimientos)
            
            cell_b = ws.cell(row=fila, column=5, value=float(f.saldo_balance))
            cell_b.number_format = '"$"#,##0.00'
            cell_r = ws.cell(row=fila, column=6, value=float(f.valor_reportado))
            cell_r.number_format = '"$"#,##0.00'
            cell_d = ws.cell(row=fila, column=7, value=float(f.diferencia))
            cell_d.number_format = '"$"#,##0.00'
            
            estado_motivo = f"{f.estado}: {f.motivo}" if f.motivo else f.estado
            ws.cell(row=fila, column=8, value=estado_motivo[:80])
            
            # Color según estado
            if f.estado == 'OK':
                fill = TOTAL_FILL
            elif f.estado in ('NO_DEDUCIBLE', 'GMF_50'):
                fill = WARNING_FILL
            elif f.estado == 'DIFERENCIA':
                fill = ERROR_FILL
            else:
                fill = CONCILIACION_FILL
            
            for col in range(1, 9):
                ws.cell(row=fila, column=col).fill = fill
                ws.cell(row=fila, column=col).border = BORDER
            
            fila += 1
        
        # Total formato
        fila_total = fila
        ws.cell(row=fila_total, column=1, value="TOTAL FORMATO")
        ws.merge_cells(start_row=fila_total, start_column=1, end_row=fila_total, end_column=4)
        ws.cell(row=fila_total, column=5,
                value=float(cuadre_formato.total_balance)).number_format = '"$"#,##0.00'
        ws.cell(row=fila_total, column=6,
                value=float(cuadre_formato.total_reportado)).number_format = '"$"#,##0.00'
        ws.cell(row=fila_total, column=7,
                value=float(cuadre_formato.total_diferencia)).number_format = '"$"#,##0.00'
        for col in range(1, 9):
            ws.cell(row=fila_total, column=col).fill = TOTAL_FILL
            ws.cell(row=fila_total, column=col).font = Font(bold=True)
            ws.cell(row=fila_total, column=col).border = BORDER
    
    # Anchos
    _set_col_widths(ws, [10, 15, 5, 40, 10, 35, 18, 20])
    
    # Freeze panes
    ws.freeze_panes = ws.cell(row=fila_inicio_datos, column=1)


# ================================================================
# HOJA RESUMEN
# ================================================================

def _construir_hoja_resumen(wb, cabecera, formatos_data, total_general):
    """Crea la primera hoja con resumen de todos los formatos."""
    ws = wb.create_sheet("Resumen", 0)  # insertar al principio
    
    ws['A1'] = "INFORMACIÓN EXÓGENA DIAN — BORRADOR DE FORMATOS"
    ws['A1'].font = Font(bold=True, size=16, color="1A5276")
    ws.merge_cells('A1:H1')
    
    # Datos empresa
    ws['A3'] = "Empresa:"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = cabecera.get('razon_social', '')
    
    ws['A4'] = "NIT:"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = cabecera.get('nit', '')
    
    ws['A5'] = "Año gravable:"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = cabecera.get('año_gravable', '')
    
    ws['A6'] = "Generado:"
    ws['A6'].font = Font(bold=True)
    ws['B6'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    
    # Tabla de formatos
    fila = 9
    headers = ["Formato", "Nombre", "# Filas XML", "Total Reportado",
               "Total Balance", "Diferencia", "Estado"]
    _aplicar_header(ws, fila, headers)
    fila += 1
    
    fila_inicio = fila
    
    for fmt_code in sorted(formatos_data.keys()):
        data = formatos_data[fmt_code]
        ws.cell(row=fila, column=1, value=fmt_code)
        ws.cell(row=fila, column=2, value=NOMBRES_FORMATOS.get(fmt_code, f'Formato {fmt_code}')[:60])
        ws.cell(row=fila, column=3, value=data['n_filas'])
        ws.cell(row=fila, column=4, value=float(data['total_reportado'])).number_format = '"$"#,##0.00'
        ws.cell(row=fila, column=5, value=float(data['total_balance'])).number_format = '"$"#,##0.00'
        diff = data['total_balance'] - data['total_reportado']
        ws.cell(row=fila, column=6, value=float(diff)).number_format = '"$"#,##0.00'
        
        # Estado
        if abs(diff) < 1:
            estado = '✓ Cuadra'
            fill = TOTAL_FILL
        elif diff > 0:
            estado = '⚠ Hay no deducibles'
            fill = WARNING_FILL
        else:
            estado = '❌ Reportado > Balance'
            fill = ERROR_FILL
        
        ws.cell(row=fila, column=7, value=estado)
        for col in range(1, 8):
            ws.cell(row=fila, column=col).border = BORDER
            if col == 7:
                ws.cell(row=fila, column=col).fill = fill
        fila += 1
    
    # Total general
    fila_total = fila
    ws.cell(row=fila_total, column=1, value="TOTAL GENERAL").font = Font(bold=True)
    ws.merge_cells(start_row=fila_total, start_column=1, end_row=fila_total, end_column=2)
    
    total_filas = sum(d['n_filas'] for d in formatos_data.values())
    total_reportado = sum(d['total_reportado'] for d in formatos_data.values())
    total_balance_g = sum(d['total_balance'] for d in formatos_data.values())
    total_diff_g = total_balance_g - total_reportado
    
    ws.cell(row=fila_total, column=3, value=total_filas)
    ws.cell(row=fila_total, column=4, value=float(total_reportado)).number_format = '"$"#,##0.00'
    ws.cell(row=fila_total, column=5, value=float(total_balance_g)).number_format = '"$"#,##0.00'
    ws.cell(row=fila_total, column=6, value=float(total_diff_g)).number_format = '"$"#,##0.00'
    
    for col in range(1, 8):
        ws.cell(row=fila_total, column=col).fill = TOTAL_FILL
        ws.cell(row=fila_total, column=col).font = Font(bold=True)
        ws.cell(row=fila_total, column=col).border = BORDER
    
    _set_col_widths(ws, [10, 50, 12, 18, 18, 18, 22])
    ws.freeze_panes = ws.cell(row=fila_inicio, column=1)


# ================================================================
# HOJA: CUENTAS NO REPORTADAS
# ================================================================

def _clasificar_motivo_no_reporte(codigo_cuenta, nombre_cuenta):
    """Determina por qué una cuenta no se reporta. Retorna motivo legible."""
    cuenta = str(codigo_cuenta).strip()
    nombre_norm = (nombre_cuenta or '').lower()
    
    # Excluida explícitamente (gasto PyG seguridad social)
    for excl in PREFIJOS_EXCLUIR_SS:
        if cuenta.startswith(excl):
            return ('🚫 Excluida: gasto PyG ya cubierto en pasivos 23/25', 1)
    
    # Cuentas internas (caja, bancos, etc) generalmente no se reportan en exógena
    if cuenta.startswith(('11', '12')):
        return ('ℹ️ Caja / Bancos / Inversiones (no aplica reporte)', 4)
    
    # Inventarios
    if cuenta.startswith('14'):
        return ('ℹ️ Inventarios (no aplica reporte)', 4)
    
    # Activos fijos
    if cuenta.startswith('15'):
        return ('ℹ️ Activos fijos (van al 1011 si aplica)', 4)
    
    # Patrimonio
    if cuenta.startswith('3'):
        return ('ℹ️ Cuentas de patrimonio', 4)
    
    # Cierre / orden
    if cuenta.startswith(('59', '69', '79', '89', '99')):
        return ('ℹ️ Cuentas de cierre/orden', 4)
    
    # Comisiones / intereses bancarios (cuentas 5305 que no son GMF)
    if cuenta.startswith('5305'):
        if 'comision' in nombre_norm or 'interes' in nombre_norm or 'manejo' in nombre_norm:
            return ('ℹ️ Gastos bancarios (no GMF)', 4)
    
    # Default: sin regla
    return ('❓ Sin regla en las 3 capas (revisar)', 2)


def _construir_hoja_no_reportadas(wb, todos_movimientos, movs_clasificados):
    """Hoja con cuentas que NO van a ningún formato, agrupadas por motivo."""
    ws = wb.create_sheet("Cuentas no reportadas")
    
    _aplicar_titulo(ws, 1, "📊 CUENTAS DEL BALANCE NO INCLUIDAS EN FORMATOS DIAN", ncols=6)
    
    ws['A3'] = ("Estas cuentas NO se reportan en información exógena. Algunas son por exclusión "
                "automática (gastos PyG ya cubiertos) y otras requieren revisión manual.")
    ws.merge_cells('A3:F3')
    ws['A3'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 30
    
    # Identificar cuentas no reportadas
    cuentas_reportadas = set()
    for m in movs_clasificados:
        if m.get('formato_dian') and m.get('formato_dian') != '999999':
            cuentas_reportadas.add(str(m.get('codigo_cuenta', '')).strip())
    
    # Agrupar todos los movimientos por cuenta (los que no están reportados)
    no_reportadas = defaultdict(lambda: {
        'codigo': '', 'nombre': '', 'saldo': Decimal('0'),
        'movs': 0, 'nits': set(), 'motivo': '', 'orden': 99,
    })
    
    for m in todos_movimientos:
        cuenta = str(m.get('codigo_cuenta', '')).strip()
        if not cuenta or cuenta in cuentas_reportadas:
            continue
        
        nombre = m.get('nombre_cuenta', '')
        debitos = Decimal(str(m.get('debitos', 0) or 0))
        creditos = Decimal(str(m.get('creditos', 0) or 0))
        valor = max(abs(debitos), abs(creditos))
        
        c = no_reportadas[cuenta]
        c['codigo'] = cuenta
        c['nombre'] = nombre
        c['saldo'] += valor
        c['movs'] += 1
        if m.get('nit'):
            c['nits'].add(m['nit'])
        if not c['motivo']:
            c['motivo'], c['orden'] = _clasificar_motivo_no_reporte(cuenta, nombre)
    
    # Agrupar por motivo
    grupos = defaultdict(list)
    for cuenta_data in no_reportadas.values():
        motivo = cuenta_data['motivo']
        grupos[motivo].append(cuenta_data)
    
    # Ordenar grupos por orden y mostrar
    fila = 5
    for motivo in sorted(grupos.keys(), key=lambda m: grupos[m][0]['orden']):
        cuentas_grupo = sorted(grupos[motivo], key=lambda x: x['saldo'], reverse=True)
        total_grupo = sum(c['saldo'] for c in cuentas_grupo)
        
        # Subheader del grupo
        ws.cell(row=fila, column=1, value=f"{motivo} ({len(cuentas_grupo)} cuentas, ${total_grupo:,.0f})")
        ws.cell(row=fila, column=1).font = Font(bold=True, size=11, color="1A5276")
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=6)
        fila += 1
        
        # Headers del grupo
        headers = ["Cuenta", "Nombre", "Movs", "NITs", "Saldo", "Motivo"]
        _aplicar_header(ws, fila, headers, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
        fila += 1
        
        for c in cuentas_grupo:
            ws.cell(row=fila, column=1, value=c['codigo'])
            ws.cell(row=fila, column=2, value=str(c['nombre'])[:50])
            ws.cell(row=fila, column=3, value=c['movs'])
            ws.cell(row=fila, column=4, value=len(c['nits']))
            ws.cell(row=fila, column=5, value=float(c['saldo'])).number_format = '"$"#,##0.00'
            ws.cell(row=fila, column=6, value=c['motivo'][:60])
            
            # Color según orden (gravedad)
            if c['orden'] == 1:    # Excluida explícita
                fill = EXCLUIDO_FILL
            elif c['orden'] == 2:  # Sin regla (revisar)
                fill = WARNING_FILL
            else:                  # Info
                fill = None
            
            for col in range(1, 7):
                if fill:
                    ws.cell(row=fila, column=col).fill = fill
                ws.cell(row=fila, column=col).border = BORDER
            fila += 1
        
        fila += 1  # espacio entre grupos
    
    if not no_reportadas:
        ws.cell(row=fila, column=1, value="✅ Todas las cuentas del balance están incluidas en algún formato.")
        ws.cell(row=fila, column=1).font = Font(italic=True, color="2E7D32", size=11)
    
    _set_col_widths(ws, [12, 50, 8, 8, 18, 50])
    ws.freeze_panes = ws.cell(row=6, column=1)


# ================================================================
# FUNCIÓN PRINCIPAL
# ================================================================

def generar_excel_borrador_completo(
    movimientos_clasificados: list[dict],
    todos_movimientos_balance: list[dict],
    cuadre: dict,  # de construir_cuadre_balance_vs_formatos
    terceros_dict: dict,
    cabecera: dict,
) -> BytesIO:
    """Genera el Excel borrador completo de auditoría tributaria.
    
    Args:
        movimientos_clasificados: solo los que tienen formato/concepto asignado
        todos_movimientos_balance: TODOS los movimientos del balance (para no reportadas)
        cuadre: dict {formato: CuadreFormato}
        terceros_dict: dict {nit: {tipo_doc, dv, razon_social, ...}}
        cabecera: dict con razon_social, nit, año_gravable
    
    Returns:
        BytesIO con el Excel listo para descargar
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # eliminar hoja default
    
    # Agrupar movimientos clasificados por formato
    movs_por_formato = defaultdict(list)
    for m in movimientos_clasificados:
        fmt = m.get('formato_dian')
        if fmt and fmt != '999999':
            movs_por_formato[fmt].append(m)
    
    # Datos para hoja resumen
    formatos_data = {}
    
    # Construir una hoja por cada formato (ordenado)
    for formato in sorted(movs_por_formato.keys()):
        movs = movs_por_formato[formato]
        cuadre_fmt = cuadre.get(formato)
        
        _construir_hoja_formato(wb, formato, movs, cuadre_fmt, terceros_dict)
        
        # Calcular totales para resumen
        total_balance = cuadre_fmt.total_balance if cuadre_fmt else Decimal('0')
        total_reportado = cuadre_fmt.total_reportado if cuadre_fmt else Decimal('0')
        
        # Contar filas únicas (NIT, concepto)
        filas_xml = len(set((m.get('nit'), m.get('concepto_dian')) for m in movs
                            if m.get('nit') and m.get('concepto_dian')))
        
        formatos_data[formato] = {
            'n_filas': filas_xml,
            'total_balance': total_balance,
            'total_reportado': total_reportado,
        }
    
    # Hoja resumen (al inicio)
    _construir_hoja_resumen(wb, cabecera, formatos_data, sum(d['total_balance'] for d in formatos_data.values()))
    
    # Hoja cuentas no reportadas (al final)
    _construir_hoja_no_reportadas(wb, todos_movimientos_balance, movimientos_clasificados)
    
    # Guardar
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ================================================================
# TEST
# ================================================================

if __name__ == '__main__':
    print("Test de excel_borrador_formatos.py")
    
    # Datos de prueba
    movs_clasif = [
        # 1001 - 5004 Servicios
        {'codigo_cuenta': '514005', 'nombre_cuenta': 'TRANSPORTE',
         'nit': '900555666', 'nombre_tercero': 'TRANSPORTES SA',
         'debitos': 25000000, 'creditos': 0,
         'formato_dian': '1001', 'concepto_dian': 5004},
        # 1001 - 5011 Salud
        {'codigo_cuenta': '255005', 'nombre_cuenta': 'PAGOS SALUD EPS',
         'nit': '900111222', 'nombre_tercero': 'SANITAS EPS',
         'debitos': 24000000, 'creditos': 0,
         'formato_dian': '1001', 'concepto_dian': 5011},
        # 1001 - 5012 Pensión
        {'codigo_cuenta': '255015', 'nombre_cuenta': 'AFP PORVENIR',
         'nit': '900222111', 'nombre_tercero': 'PORVENIR',
         'debitos': 36000000, 'creditos': 0,
         'formato_dian': '1001', 'concepto_dian': 5012},
        # 1007 - Ingresos
        {'codigo_cuenta': '413505', 'nombre_cuenta': 'INGRESOS',
         'nit': '901234567', 'nombre_tercero': 'CLIENTE A',
         'debitos': 0, 'creditos': 50000000,
         'formato_dian': '1007', 'concepto_dian': 4001},
    ]
    
    # Movimientos completos del balance (incluye no reportados)
    todos_movs = movs_clasif + [
        # Cuenta excluida (gasto PyG)
        {'codigo_cuenta': '510568', 'nombre_cuenta': 'APORTE RIESGOS PROFESIONALES',
         'nit': '900111222', 'debitos': 477800, 'creditos': 0,
         'formato_dian': None, 'concepto_dian': None},
        # Cuenta sin regla
        {'codigo_cuenta': '130505', 'nombre_cuenta': 'CLIENTES NACIONALES',
         'nit': '901234567', 'debitos': 5000000, 'creditos': 0,
         'formato_dian': None, 'concepto_dian': None},
        # Caja
        {'codigo_cuenta': '110505', 'nombre_cuenta': 'CAJA GENERAL',
         'nit': None, 'debitos': 1000000, 'creditos': 0,
         'formato_dian': None, 'concepto_dian': None},
    ]
    
    # Construir cuadre
    from conciliacion import construir_cuadre_balance_vs_formatos
    cuadre = construir_cuadre_balance_vs_formatos(movs_clasif)
    
    # Terceros
    terceros = {
        '900555666': {'tipo_documento': '31', 'dv': '1', 'razon_social': 'TRANSPORTES SA'},
        '900111222': {'tipo_documento': '31', 'dv': '5', 'razon_social': 'SANITAS EPS'},
        '900222111': {'tipo_documento': '31', 'dv': '8', 'razon_social': 'PORVENIR PENSIONES'},
        '901234567': {'tipo_documento': '31', 'dv': '4', 'razon_social': 'CLIENTE A SAS'},
    }
    
    cabecera = {
        'razon_social': 'RUTAS DEL MAR SAS',
        'nit': '901589040-1',
        'año_gravable': 2025,
    }
    
    excel = generar_excel_borrador_completo(
        movs_clasif, todos_movs, cuadre, terceros, cabecera
    )
    
    with open('/tmp/borrador_completo.xlsx', 'wb') as f:
        f.write(excel.getvalue())
    
    import os
    size = os.path.getsize('/tmp/borrador_completo.xlsx')
    print(f"✅ Excel generado: /tmp/borrador_completo.xlsx ({size:,} bytes)")
    
    # Verificar que tiene las hojas correctas
    wb = openpyxl.load_workbook('/tmp/borrador_completo.xlsx')
    print(f"   Hojas: {wb.sheetnames}")
