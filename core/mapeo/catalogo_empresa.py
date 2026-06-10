"""
Catálogo de mapeo por empresa.

Lee un "Balance de Prueba por NIT con CC" (el que trae las columnas de NIT del
tercero además del centro de costo) y construye, para esa empresa, el catálogo
que el procesador de XMLs usa para armar el movimiento:

    - cuentas válidas (código -> nombre)
    - centros de costo válidos (código -> nombre)
    - regla proveedor -> cuenta de gasto/costo (la cuenta donde históricamente
      se registró ese tercero), con sus centros de costo asociados
    - cuenta(s) por pagar usadas por cada tercero (clase 22/23)

Idea de uso: se carga UNA vez por empresa (NIT) y queda memorizado; antes de
procesar el token de esa empresa, el procesador consulta este catálogo para
asignar cuenta y centro de costo a cada documento, en vez de dejar todo en
"PENDIENTE DE MAPEO".

Formato de entrada (hoja "Datos"):
    Fila 1: EMPRESA - NIT
    Fila 2: título + periodo
    Fila 3: Desde ... Hasta ...
    Fila 4: Cuenta | Equivalencia | Nombre | NIT | Nombre NIT |
            Centro de Costos | Nombre CC | Saldo Anterior | Débitos |
            Créditos | Nuevo Saldo
    Fila 5+: datos
"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Dict, List, Optional, BinaryIO, Union
import re

import pandas as pd
from openpyxl import load_workbook

CLASES_GASTO_COSTO = {"5", "6", "7"}
CLASES_INGRESO = {"4"}
PREF_POR_PAGAR = ("22", "23")          # proveedores / cuentas por pagar


def _norm_nit(nit: str) -> str:
    """Normaliza un NIT a solo dígitos, sin dígito de verificación.

    Usa el guion como separador del DV ('900.495.853-4' -> '900495853').
    Si no hay guion, conserva todos los dígitos ('901304654' -> '901304654').
    """
    s = str(nit or "").strip()
    if "-" in s:
        s = s.rsplit("-", 1)[0]          # quita el DV después del guion
    return re.sub(r"[^0-9]", "", s)


def _to_float(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


@dataclass
class CatalogoEmpresa:
    nit: str                       # NIT de la empresa (normalizado, sin DV)
    nit_completo: str              # NIT tal cual viene (con DV)
    empresa: str
    periodo: str
    cuentas: Dict[str, str]        # codigo -> nombre
    centros: Dict[str, str]        # codigo CC -> nombre
    proveedor_cuenta: Dict[str, dict]   # nit_tercero -> info de mapeo
    detalle: pd.DataFrame          # filas crudas normalizadas

    def resumen(self) -> dict:
        return {
            "empresa": self.empresa,
            "nit": self.nit_completo,
            "periodo": self.periodo,
            "n_cuentas": len(self.cuentas),
            "n_centros": len(self.centros),
            "n_proveedores": len(self.proveedor_cuenta),
        }


def cargar_balance_nit_cc(fuente: Union[str, BinaryIO]) -> tuple:
    """Lee un balance 'por NIT' (con o sin centro de costo) y devuelve
    (df, empresa, nit, periodo, tiene_cc).

    Detecta las columnas por su encabezado, así sirve para empresas que NO
    manejan centro de costo (no trae 'Centro de Costos') y para distintos
    formatos, mientras tenga al menos: Cuenta, Nombre, NIT, Débitos, Créditos,
    Nuevo Saldo. La fila de encabezados se busca en las primeras filas.
    """
    wb = load_workbook(fuente, read_only=True, data_only=True)
    ws = wb["Datos"] if "Datos" in wb.sheetnames else wb.active
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(filas) < 3:
        raise ValueError("El archivo no tiene la estructura esperada.")

    # Encabezado de empresa / periodo (primeras filas de texto)
    cab = str(filas[0][0] or "").strip()
    if " - " in cab:
        empresa, nit = cab.rsplit(" - ", 1)
        empresa, nit = empresa.strip(), nit.strip()
    else:
        empresa, nit = cab, ""
    periodo = ""
    for r in filas[1:4]:
        m = re.search(r"([A-Za-z]{3}-\d{1,2}-\d{4})", str(r[0] or ""))
        if m:
            periodo = m.group(1)
            break

    # Buscar fila de encabezados (la que contiene 'Cuenta' y 'Nuevo Saldo')
    def _norm(s):
        return str(s or "").strip().lower()
    idx_hdr = None
    for i, r in enumerate(filas[:8]):
        celdas = [_norm(c) for c in r]
        if "cuenta" in celdas and any("nuevo saldo" in c for c in celdas):
            idx_hdr = i
            break
    if idx_hdr is None:
        raise ValueError("No encontré la fila de encabezados (Cuenta ... Nuevo Saldo).")

    enc = [_norm(c) for c in filas[idx_hdr]]

    def col(*nombres):
        for nm in nombres:
            for j, h in enumerate(enc):
                if h == nm:
                    return j
        return None

    c_cuenta = col("cuenta")
    c_nombre = col("nombre")
    c_nit = col("nit", "nit tercero")
    c_nnit = col("nombre nit", "nombre tercero")
    c_cc = col("centro de costos", "centro de costo", "cc")
    c_ncc = col("nombre cc", "nombre centro de costos")
    c_deb = col("débitos", "debitos")
    c_cre = col("créditos", "creditos")
    c_ns = col("nuevo saldo")
    faltan = [n for n, v in [("Cuenta", c_cuenta), ("Nombre", c_nombre),
              ("NIT", c_nit), ("Débitos", c_deb), ("Créditos", c_cre),
              ("Nuevo Saldo", c_ns)] if v is None]
    if faltan:
        raise ValueError("Faltan columnas en el balance: " + ", ".join(faltan))

    tiene_cc = c_cc is not None

    regs = []
    for r in filas[idx_hdr + 1:]:
        if c_cuenta >= len(r) or r[c_cuenta] is None:
            continue
        cuenta = str(r[c_cuenta]).strip()
        if not cuenta or cuenta == "#N/A":
            continue
        regs.append({
            "cuenta": cuenta,
            "nombre": str(r[c_nombre] or "").strip() if c_nombre is not None else "",
            "nit": str(r[c_nit] or "").strip() if c_nit is not None else "",
            "nombre_nit": str(r[c_nnit] or "").strip() if c_nnit is not None else "",
            "cc": str(r[c_cc] or "").strip() if tiene_cc else "",
            "nombre_cc": str(r[c_ncc] or "").strip() if c_ncc is not None else "",
            "debitos": _to_float(r[c_deb]),
            "creditos": _to_float(r[c_cre]),
            "nuevo_saldo": _to_float(r[c_ns]),
        })
    df = pd.DataFrame(regs)
    if df.empty:
        raise ValueError("No se encontraron filas de datos válidas.")
    df["es_numerica"] = df["cuenta"].str.fullmatch(r"\d+")
    df["clase"] = df["cuenta"].str[0].where(df["es_numerica"], "")
    df["mov"] = df["debitos"] - df["creditos"]
    return df, empresa, nit, periodo, tiene_cc


def _hojas(df: pd.DataFrame) -> pd.DataFrame:
    """Marca las cuentas hoja (las que no son prefijo de otra más larga)."""
    cods = sorted(df.loc[df["es_numerica"], "cuenta"].unique())
    padres = set()
    for i in range(len(cods) - 1):
        if cods[i + 1].startswith(cods[i]):
            padres.add(cods[i])
    return df[df["es_numerica"] & ~df["cuenta"].isin(padres)].copy()


def construir_catalogo(fuente: Union[str, BinaryIO]) -> CatalogoEmpresa:
    """Construye el catálogo de mapeo de una empresa desde su balance.

    Funciona con balances 'por NIT con CC' y 'por NIT' (sin centro de costo).
    """
    df, empresa, nit_completo, periodo, tiene_cc = cargar_balance_nit_cc(fuente)

    # catálogo de cuentas (cualquier nivel) y de centros de costo
    cuentas = {}
    for c, n in df[["cuenta", "nombre"]].drop_duplicates().itertuples(index=False):
        cuentas.setdefault(c, n)
    centros = {}
    for c, n in df[df["cc"] != ""][["cc", "nombre_cc"]].drop_duplicates().itertuples(index=False):
        centros.setdefault(c, n)

    hojas = _hojas(df)

    # regla proveedor -> cuenta de gasto/costo (cuenta dominante por |movimiento|)
    gc = hojas[(hojas["nit"] != "") & (hojas["clase"].isin(CLASES_GASTO_COSTO))].copy()
    proveedor_cuenta: Dict[str, dict] = {}
    for nit_t, grp in gc.groupby("nit"):
        # cuenta dominante por valor absoluto del movimiento
        por_cuenta = grp.groupby(["cuenta", "nombre"])["mov"].apply(
            lambda s: s.abs().sum()).sort_values(ascending=False)
        if por_cuenta.empty:
            continue
        (cta, nom), _ = list(por_cuenta.items())[0]
        # centros de costo asociados a ese tercero (con peso)
        ccs = (grp.groupby(["cc", "nombre_cc"])["mov"].apply(lambda s: s.abs().sum())
               .sort_values(ascending=False))
        centros_list = [{"cc": cc, "nombre": ncc, "peso": float(v)}
                        for (cc, ncc), v in ccs.items() if cc]
        # cuenta(s) por pagar de ese tercero
        pp = hojas[(hojas["nit"] == nit_t)
                   & hojas["cuenta"].str.startswith(PREF_POR_PAGAR)]
        por_pagar = sorted(pp["cuenta"].unique().tolist())
        proveedor_cuenta[_norm_nit(nit_t)] = {
            "nit": nit_t,
            "nombre": grp["nombre_nit"].iloc[0],
            "cuenta_gasto": cta,
            "cuenta_gasto_nombre": nom,
            "cuentas_alternas": [c for (c, _n) in list(por_cuenta.index)[1:4]],
            "centros": centros_list,
            "cuentas_por_pagar": por_pagar,
        }

    return CatalogoEmpresa(
        nit=_norm_nit(nit_completo), nit_completo=nit_completo, empresa=empresa,
        periodo=periodo, cuentas=cuentas, centros=centros,
        proveedor_cuenta=proveedor_cuenta, detalle=df,
    )


def sugerir_cuenta(catalogo: CatalogoEmpresa, nit_proveedor: str) -> Optional[dict]:
    """Devuelve la sugerencia de mapeo para un proveedor (por su NIT)."""
    return catalogo.proveedor_cuenta.get(_norm_nit(nit_proveedor))
