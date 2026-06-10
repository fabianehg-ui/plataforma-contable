"""Exporta el catálogo de mapeo de una empresa a Excel (para revisión/memoria)."""
from __future__ import annotations

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .catalogo_empresa import CatalogoEmpresa

FUENTE = "Calibri"
AZUL = "1F4E78"
GRIS = "808080"


def _hdr(ws, headers, fila=1):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=fila, column=j, value=h)
        c.font = Font(name=FUENTE, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def exportar_catalogo_excel(cat: CatalogoEmpresa) -> bytes:
    wb = Workbook()

    # Hoja 1: Proveedor -> Cuenta (el mapeo)
    ws = wb.active
    ws.title = "Proveedor-Cuenta"
    ws.cell(row=1, column=1, value=f"{cat.empresa}  ·  {cat.nit_completo}").font = Font(
        name=FUENTE, bold=True, size=12)
    ws.cell(row=2, column=1,
            value="Mapeo sugerido proveedor → cuenta (editable). Periodo " + cat.periodo
            ).font = Font(name=FUENTE, italic=True, size=9, color=GRIS)
    headers = ["NIT proveedor", "Nombre", "Cuenta gasto/costo", "Nombre cuenta",
               "Centro de costo principal", "Cuentas por pagar", "Cuentas alternas"]
    _hdr(ws, headers, fila=4)
    i = 5
    for nit, info in sorted(cat.proveedor_cuenta.items(), key=lambda x: x[1]["nombre"]):
        cc0 = info["centros"][0]["cc"] if info["centros"] else ""
        ws.cell(row=i, column=1, value=info["nit"])
        ws.cell(row=i, column=2, value=info["nombre"])
        ws.cell(row=i, column=3, value=info["cuenta_gasto"])
        ws.cell(row=i, column=4, value=info["cuenta_gasto_nombre"])
        ws.cell(row=i, column=5, value=cc0)
        ws.cell(row=i, column=6, value=", ".join(info["cuentas_por_pagar"]))
        ws.cell(row=i, column=7, value=", ".join(info["cuentas_alternas"]))
        for j in range(1, 8):
            ws.cell(row=i, column=j).font = Font(name=FUENTE, size=9)
        i += 1
    ws.freeze_panes = "A5"
    for j, w in zip(range(1, 8), [16, 32, 16, 28, 22, 22, 20]):
        ws.column_dimensions[get_column_letter(j)].width = w

    # Hoja 2: Cuentas
    ws2 = wb.create_sheet("Cuentas")
    _hdr(ws2, ["Cuenta", "Nombre"])
    for k, (c, n) in enumerate(sorted(cat.cuentas.items()), start=2):
        ws2.cell(row=k, column=1, value=c).font = Font(name=FUENTE, size=9)
        ws2.cell(row=k, column=2, value=n).font = Font(name=FUENTE, size=9)
    ws2.freeze_panes = "A2"
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 40

    # Hoja 3: Centros de costo
    ws3 = wb.create_sheet("Centros de costo")
    _hdr(ws3, ["Centro de Costo", "Nombre"])
    for k, (c, n) in enumerate(sorted(cat.centros.items()), start=2):
        ws3.cell(row=k, column=1, value=c).font = Font(name=FUENTE, size=9)
        ws3.cell(row=k, column=2, value=n).font = Font(name=FUENTE, size=9)
    ws3.freeze_panes = "A2"
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 36

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
