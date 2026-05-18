"""
procesador_acuses.py
====================
Procesador del Excel de RADIAN exportado del catálogo VPFE de la DIAN.

Estructura del archivo origen (columnas estándar DIAN AG2025):
    A: Tipo de documento          B: CUFE/CUDE          C: Folio
    D: Prefijo                    E: Divisa             F: Forma de Pago ★
    G: Medio de Pago              H: Fecha Emisión      I: Fecha Recepción
    J: NIT Emisor                 K: Nombre Emisor      L: NIT Receptor
    M: Nombre Receptor            N: IVA                O: ICA
    P: IC                         Q: INC                R: Timbre
    S: INC Bolsas                 T: IN Carbono         U: IN Combustibles
    V: IC Datos                   W: ICL                X: INPP
    Y: IBUA                       Z: ICUI               AA: Rete IVA
    AB: Rete Renta                AC: Rete ICA          AD: Total
    AE: Estado                    AF: Grupo

Códigos DIAN (Anexo Técnico FE 1.9):
    Forma de Pago:
        1 = Contado
        2 = Crédito
    Grupo:
        Emitido  = factura que YO emití (mis ventas, va a F1007/F1008)
        Recibido = factura que YO recibí (mis compras, va a F1001/F1009)
    Tipo de documento:
        - Factura electrónica
        - Nota crédito electrónica
        - Nota débito electrónica
        - Documento soporte con no obligados
        - Application response (NO es una factura, es solo el acuse de recibo)
"""

from __future__ import annotations
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional
import io

import pandas as pd


# ============================================================================
# CONSTANTES
# ============================================================================

# Columnas estándar del catálogo DIAN VPFE
COLUMNAS_ESPERADAS = [
    'Tipo de documento', 'CUFE/CUDE', 'Folio', 'Prefijo', 'Divisa',
    'Forma de Pago', 'Medio de Pago', 'Fecha Emisión', 'Fecha Recepción',
    'NIT Emisor', 'Nombre Emisor', 'NIT Receptor', 'Nombre Receptor',
    'IVA', 'ICA', 'IC', 'INC', 'Timbre', 'INC Bolsas', 'IN Carbono',
    'IN Combustibles', 'IC Datos', 'ICL', 'INPP', 'IBUA', 'ICUI',
    'Rete IVA', 'Rete Renta', 'Rete ICA', 'Total', 'Estado', 'Grupo',
]

# Tipo de documento "Application response" NO es una factura, es solo
# el acuse de recibo del receptor → siempre se filtra
TIPO_APPLICATION_RESPONSE = 'Application response'

# Catálogos para etiquetas legibles
FORMA_PAGO_LABEL = {
    1: 'Contado',
    2: 'Crédito',
}

GRUPO_LABEL = {
    'Emitido':  'Emitido (ventas)',
    'Recibido': 'Recibido (compras)',
}

# Medios de pago más comunes (Anexo Técnico FE 1.9, no exhaustivo)
MEDIO_PAGO_LABEL = {
    '1': 'No definido',
    '2': 'Tarjeta crédito',
    '10': 'Efectivo',
    '20': 'Cheque',
    '30': 'Crédito documentario',
    '42': 'Consignación bancaria',
    '45': 'Transferencia débito bancaria',
    '47': 'Transferencia',
    '48': 'Tarjeta crédito (otro)',
    '49': 'Tarjeta débito',
    '78': 'Tarjeta servicios',
    'ZZZ': 'No registrado',
}


# ============================================================================
# RESULTADOS / DATA CLASSES
# ============================================================================

@dataclass
class ResumenAcuses:
    """Resumen ejecutivo del archivo RADIAN procesado."""
    total_documentos: int = 0
    total_application_response: int = 0
    total_facturas: int = 0
    total_emitidos: int = 0
    total_recibidos: int = 0
    total_credito: int = 0
    total_contado: int = 0
    valor_total: float = 0.0
    fecha_min: Optional[str] = None
    fecha_max: Optional[str] = None
    proveedores_unicos: int = 0
    clientes_unicos: int = 0


@dataclass
class ResultadoFiltro:
    """Resultado de filtrar el dataset."""
    df: pd.DataFrame = field(default_factory=pd.DataFrame)
    filas_originales: int = 0
    filas_filtradas: int = 0
    filtros_aplicados: dict = field(default_factory=dict)
    advertencias: list = field(default_factory=list)
    
    @property
    def total_valor(self) -> float:
        if 'Total' not in self.df.columns or self.df.empty:
            return 0.0
        return float(self.df['Total'].sum())


# ============================================================================
# CARGA Y VALIDACIÓN
# ============================================================================

def cargar_excel_radian(archivo) -> pd.DataFrame:
    """
    Carga el Excel del catálogo VPFE y valida su estructura.
    
    Args:
        archivo: ruta al archivo o BytesIO (de Streamlit uploader)
    
    Returns:
        DataFrame con todas las filas y columnas esperadas
    
    Raises:
        ValueError: si el archivo no tiene la estructura esperada
    """
    try:
        df = pd.read_excel(archivo)
    except Exception as e:
        raise ValueError(f"No se pudo leer el Excel: {e}")
    
    # Validar columnas mínimas obligatorias
    columnas_min = ['Tipo de documento', 'Forma de Pago', 'Grupo', 
                     'NIT Emisor', 'Nombre Emisor', 'Total', 
                     'Fecha Emisión', 'Folio', 'Estado']
    faltantes = [c for c in columnas_min if c not in df.columns]
    if faltantes:
        raise ValueError(
            f"El archivo no parece ser del catálogo VPFE de la DIAN. "
            f"Columnas faltantes: {faltantes}"
        )
    
    # Normalizar tipos: NITs como string sin decimales
    for col_nit in ['NIT Emisor', 'NIT Receptor']:
        if col_nit in df.columns:
            df[col_nit] = df[col_nit].apply(_limpiar_nit)
    
    # Forma de Pago como int (puede venir como float si hay NaN)
    if 'Forma de Pago' in df.columns:
        df['Forma de Pago'] = pd.to_numeric(df['Forma de Pago'], errors='coerce').astype('Int64')
    
    # Total como float
    if 'Total' in df.columns:
        df['Total'] = pd.to_numeric(df['Total'], errors='coerce').fillna(0.0)
    
    return df


def _limpiar_nit(valor) -> str:
    """Convierte NIT a string limpio: '900533491.0' → '900533491'."""
    if pd.isna(valor):
        return ''
    s = str(valor).strip()
    # Quitar .0 final si viene como float
    if s.endswith('.0'):
        s = s[:-2]
    # Quitar espacios y guiones (DV)
    s = s.replace(' ', '').split('-')[0]
    return s


# ============================================================================
# RESUMEN
# ============================================================================

def resumir_acuses(df: pd.DataFrame) -> ResumenAcuses:
    """Calcula métricas generales del archivo."""
    r = ResumenAcuses()
    
    if df.empty:
        return r
    
    r.total_documentos = len(df)
    
    # Application response no son facturas reales
    mask_app_resp = df['Tipo de documento'] == TIPO_APPLICATION_RESPONSE
    r.total_application_response = int(mask_app_resp.sum())
    
    df_fact = df[~mask_app_resp]
    r.total_facturas = len(df_fact)
    
    if not df_fact.empty:
        r.total_emitidos = int((df_fact['Grupo'] == 'Emitido').sum())
        r.total_recibidos = int((df_fact['Grupo'] == 'Recibido').sum())
        r.total_credito = int((df_fact['Forma de Pago'] == 2).sum())
        r.total_contado = int((df_fact['Forma de Pago'] == 1).sum())
        r.valor_total = float(df_fact['Total'].sum())
        
        # Fechas
        fechas = pd.to_datetime(df_fact['Fecha Emisión'], errors='coerce', dayfirst=True)
        fechas = fechas.dropna()
        if not fechas.empty:
            r.fecha_min = fechas.min().strftime('%d-%m-%Y')
            r.fecha_max = fechas.max().strftime('%d-%m-%Y')
        
        # Proveedores y clientes únicos
        recibidos = df_fact[df_fact['Grupo'] == 'Recibido']
        emitidos = df_fact[df_fact['Grupo'] == 'Emitido']
        r.proveedores_unicos = int(recibidos['NIT Emisor'].nunique()) if not recibidos.empty else 0
        r.clientes_unicos = int(emitidos['NIT Receptor'].nunique()) if not emitidos.empty else 0
    
    return r


# ============================================================================
# FILTRADO
# ============================================================================

def filtrar(
    df: pd.DataFrame,
    forma_pago: Optional[int] = 2,
    grupo: Optional[str] = 'Recibido',
    excluir_application_response: bool = True,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    nit_contraparte: Optional[str] = None,
    solo_aprobados: bool = True,
) -> ResultadoFiltro:
    """
    Aplica filtros al dataset RADIAN.
    
    Args:
        df: DataFrame cargado con cargar_excel_radian()
        forma_pago: 1=Contado, 2=Crédito, None=todos
        grupo: 'Recibido', 'Emitido' o None
        excluir_application_response: descarta los acuses (no son facturas)
        fecha_desde: 'YYYY-MM-DD' o None
        fecha_hasta: 'YYYY-MM-DD' o None
        nit_contraparte: filtrar por NIT específico
        solo_aprobados: incluye solo estado "Aprobado" o "Aprobado con notificación"
    
    Returns:
        ResultadoFiltro con DataFrame filtrado y metadata
    """
    resultado = ResultadoFiltro(
        filas_originales=len(df),
        filtros_aplicados={},
        advertencias=[],
    )
    
    if df.empty:
        return resultado
    
    df_f = df.copy()
    
    # 1. Excluir Application Response (no son facturas)
    if excluir_application_response:
        antes = len(df_f)
        df_f = df_f[df_f['Tipo de documento'] != TIPO_APPLICATION_RESPONSE]
        descartados = antes - len(df_f)
        resultado.filtros_aplicados['Excluir Application Response'] = f'{descartados} descartados'
    
    # 2. Solo aprobados (descarta rechazados, en proceso, etc.)
    if solo_aprobados and 'Estado' in df_f.columns:
        antes = len(df_f)
        df_f = df_f[df_f['Estado'].str.contains('Aprobado', na=False, case=False)]
        descartados = antes - len(df_f)
        if descartados > 0:
            resultado.filtros_aplicados['Solo aprobados'] = f'{descartados} no aprobados descartados'
    
    # 3. Filtro por grupo (Emitido/Recibido)
    if grupo:
        df_f = df_f[df_f['Grupo'] == grupo]
        resultado.filtros_aplicados['Grupo'] = grupo
    
    # 4. Filtro por forma de pago
    if forma_pago is not None:
        df_f = df_f[df_f['Forma de Pago'] == forma_pago]
        resultado.filtros_aplicados['Forma de Pago'] = f'{forma_pago} - {FORMA_PAGO_LABEL.get(forma_pago, "?")}'
    
    # 5. Filtros por fecha
    if fecha_desde or fecha_hasta:
        fechas = pd.to_datetime(df_f['Fecha Emisión'], errors='coerce', dayfirst=True)
        df_f = df_f.assign(_fecha_parsed=fechas)
        if fecha_desde:
            df_f = df_f[df_f['_fecha_parsed'] >= pd.to_datetime(fecha_desde)]
            resultado.filtros_aplicados['Desde'] = fecha_desde
        if fecha_hasta:
            df_f = df_f[df_f['_fecha_parsed'] <= pd.to_datetime(fecha_hasta)]
            resultado.filtros_aplicados['Hasta'] = fecha_hasta
        df_f = df_f.drop(columns=['_fecha_parsed'])
    
    # 6. Filtro por NIT
    if nit_contraparte:
        nit_limpio = _limpiar_nit(nit_contraparte)
        col_nit = 'NIT Emisor' if grupo == 'Recibido' else 'NIT Receptor'
        df_f = df_f[df_f[col_nit] == nit_limpio]
        resultado.filtros_aplicados['NIT'] = nit_limpio
    
    resultado.df = df_f.reset_index(drop=True)
    resultado.filas_filtradas = len(df_f)
    
    if resultado.filas_filtradas == 0:
        resultado.advertencias.append('Ningún documento coincide con los filtros aplicados')
    
    return resultado


# ============================================================================
# AGRUPACIONES
# ============================================================================

def resumir_por_proveedor(df: pd.DataFrame) -> pd.DataFrame:
    """Agrupa por NIT Emisor (proveedor) con totales y conteos."""
    if df.empty:
        return pd.DataFrame()
    
    agrupado = df.groupby(['NIT Emisor', 'Nombre Emisor']).agg(
        cantidad_facturas=('Folio', 'count'),
        total_iva=('IVA', 'sum'),
        total_rete_iva=('Rete IVA', 'sum'),
        total_rete_renta=('Rete Renta', 'sum'),
        total_rete_ica=('Rete ICA', 'sum'),
        total_general=('Total', 'sum'),
        fecha_primera=('Fecha Emisión', 'min'),
        fecha_ultima=('Fecha Emisión', 'max'),
    ).reset_index().sort_values('total_general', ascending=False)
    
    return agrupado


def resumir_por_cliente(df: pd.DataFrame) -> pd.DataFrame:
    """Agrupa por NIT Receptor (cliente) con totales y conteos."""
    if df.empty:
        return pd.DataFrame()
    
    agrupado = df.groupby(['NIT Receptor', 'Nombre Receptor']).agg(
        cantidad_facturas=('Folio', 'count'),
        total_iva=('IVA', 'sum'),
        total_general=('Total', 'sum'),
        fecha_primera=('Fecha Emisión', 'min'),
        fecha_ultima=('Fecha Emisión', 'max'),
    ).reset_index().sort_values('total_general', ascending=False)
    
    return agrupado


# ============================================================================
# EXPORTACIÓN A EXCEL
# ============================================================================

def exportar_a_excel(
    resultado: ResultadoFiltro,
    incluir_resumen: bool = True,
    nombre_empresa: str = '',
) -> bytes:
    """
    Genera un Excel con varias hojas:
      - Resumen ejecutivo
      - Detalle de documentos filtrados
      - Resumen por proveedor (si grupo=Recibido)
      - Resumen por cliente (si grupo=Emitido)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Estilos
    blue_fill = PatternFill('solid', start_color='4472C4')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='808080')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # =========================================================================
    # HOJA 1: RESUMEN
    # =========================================================================
    if incluir_resumen:
        ws_r = wb.create_sheet('Resumen')
        ws_r.merge_cells('A1:C1')
        ws_r['A1'] = 'REPORTE RADIAN — Documentos filtrados'
        ws_r['A1'].font = title_font
        ws_r['A1'].fill = blue_fill
        ws_r['A1'].alignment = center
        ws_r.row_dimensions[1].height = 28
        
        if nombre_empresa:
            ws_r['A2'] = 'Empresa:'
            ws_r['B2'] = nombre_empresa
            ws_r['A2'].font = Font(bold=True)
        
        ws_r['A3'] = 'Generado:'
        ws_r['B3'] = datetime.now().strftime('%d-%m-%Y %H:%M:%S')
        ws_r['A3'].font = Font(bold=True)
        
        # Filtros aplicados
        ws_r['A5'] = 'FILTROS APLICADOS'
        ws_r['A5'].font = header_font
        ws_r['A5'].fill = blue_fill
        ws_r.merge_cells('A5:C5')
        
        row = 6
        for k, v in resultado.filtros_aplicados.items():
            ws_r.cell(row, 1, value=k).font = Font(bold=True)
            ws_r.cell(row, 2, value=str(v))
            row += 1
        
        # Métricas
        row += 1
        ws_r.cell(row, 1, value='MÉTRICAS').font = header_font
        ws_r.cell(row, 1).fill = blue_fill
        ws_r.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        
        metricas = [
            ('Total documentos antes de filtrar', resultado.filas_originales),
            ('Total documentos después de filtrar', resultado.filas_filtradas),
            ('Valor total filtrado', resultado.total_valor),
        ]
        
        for label, valor in metricas:
            ws_r.cell(row, 1, value=label).font = Font(bold=True)
            c = ws_r.cell(row, 2, value=valor)
            if isinstance(valor, (int, float)) and valor > 1000:
                c.number_format = '"$"#,##0'
            row += 1
        
        # Advertencias
        if resultado.advertencias:
            row += 1
            ws_r.cell(row, 1, value='ADVERTENCIAS').font = header_font
            ws_r.cell(row, 1).fill = PatternFill('solid', start_color='ED7D31')
            ws_r.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            row += 1
            for adv in resultado.advertencias:
                ws_r.cell(row, 1, value=f'⚠ {adv}')
                ws_r.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                row += 1
        
        ws_r.column_dimensions['A'].width = 38
        ws_r.column_dimensions['B'].width = 35
        ws_r.column_dimensions['C'].width = 20
    
    # =========================================================================
    # HOJA 2: DETALLE
    # =========================================================================
    ws_d = wb.create_sheet('Detalle')
    df_export = resultado.df.copy()
    
    if df_export.empty:
        ws_d['A1'] = 'No hay documentos que coincidan con los filtros aplicados.'
    else:
        # Escribir filas con dataframe_to_rows
        for r_idx, row_data in enumerate(dataframe_to_rows(df_export, index=False, header=True), 1):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws_d.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = blue_fill
                    cell.alignment = center
                    cell.border = border
        
        # Formato de números
        col_total = list(df_export.columns).index('Total') + 1 if 'Total' in df_export.columns else None
        if col_total:
            for r_idx in range(2, len(df_export) + 2):
                ws_d.cell(r_idx, col_total).number_format = '"$"#,##0'
        
        # Ancho de columnas
        for i, col in enumerate(df_export.columns, 1):
            letra = ws_d.cell(1, i).column_letter
            if col in ['CUFE/CUDE']:
                ws_d.column_dimensions[letra].width = 25
            elif col in ['Nombre Emisor', 'Nombre Receptor']:
                ws_d.column_dimensions[letra].width = 35
            elif col in ['Tipo de documento', 'Estado']:
                ws_d.column_dimensions[letra].width = 22
            else:
                ws_d.column_dimensions[letra].width = 14
        
        ws_d.freeze_panes = 'A2'
    
    # =========================================================================
    # HOJA 3: POR PROVEEDOR (si grupo == Recibido)
    # =========================================================================
    grupo_aplicado = resultado.filtros_aplicados.get('Grupo')
    if grupo_aplicado == 'Recibido' and not df_export.empty:
        df_prov = resumir_por_proveedor(df_export)
        if not df_prov.empty:
            ws_p = wb.create_sheet('Por Proveedor')
            for r_idx, row_data in enumerate(dataframe_to_rows(df_prov, index=False, header=True), 1):
                for c_idx, value in enumerate(row_data, 1):
                    cell = ws_p.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = blue_fill
                        cell.alignment = center
            
            # Formato de moneda en columnas de totales
            cols_moneda = [c for c in df_prov.columns if 'total' in c.lower()]
            for col_name in cols_moneda:
                col_idx = list(df_prov.columns).index(col_name) + 1
                for r_idx in range(2, len(df_prov) + 2):
                    ws_p.cell(r_idx, col_idx).number_format = '"$"#,##0'
            
            ws_p.column_dimensions['A'].width = 15
            ws_p.column_dimensions['B'].width = 40
            ws_p.freeze_panes = 'A2'
    
    # =========================================================================
    # HOJA 4: POR CLIENTE (si grupo == Emitido)
    # =========================================================================
    if grupo_aplicado == 'Emitido' and not df_export.empty:
        df_cli = resumir_por_cliente(df_export)
        if not df_cli.empty:
            ws_c = wb.create_sheet('Por Cliente')
            for r_idx, row_data in enumerate(dataframe_to_rows(df_cli, index=False, header=True), 1):
                for c_idx, value in enumerate(row_data, 1):
                    cell = ws_c.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = blue_fill
                        cell.alignment = center
            
            cols_moneda = [c for c in df_cli.columns if 'total' in c.lower()]
            for col_name in cols_moneda:
                col_idx = list(df_cli.columns).index(col_name) + 1
                for r_idx in range(2, len(df_cli) + 2):
                    ws_c.cell(r_idx, col_idx).number_format = '"$"#,##0'
            
            ws_c.column_dimensions['A'].width = 15
            ws_c.column_dimensions['B'].width = 40
            ws_c.freeze_panes = 'A2'
    
    # Guardar en memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================================================
# TESTS
# ============================================================================

def test_cargar_excel():
    df = cargar_excel_radian('/home/claude/radian_muestra.xlsx')
    assert len(df) == 83569, f"Filas esperadas 83569, got {len(df)}"
    assert 'Forma de Pago' in df.columns
    assert 'Grupo' in df.columns
    print('✅ test_cargar_excel: 83.569 filas, 32 columnas')


def test_resumen():
    df = cargar_excel_radian('/home/claude/radian_muestra.xlsx')
    r = resumir_acuses(df)
    assert r.total_documentos == 83569
    assert r.total_application_response == 11942
    assert r.total_credito == 4771
    assert r.total_recibidos > 0
    print(f'✅ test_resumen: {r.total_credito} crédito, {r.total_recibidos} recibidos')


def test_filtro_credito_recibido():
    df = cargar_excel_radian('/home/claude/radian_muestra.xlsx')
    res = filtrar(df, forma_pago=2, grupo='Recibido')
    # Esperado: 4,127 facturas (excluyendo application response)
    assert res.filas_filtradas == 4127, f"Esperado 4127, got {res.filas_filtradas}"
    assert all(res.df['Forma de Pago'] == 2)
    assert all(res.df['Grupo'] == 'Recibido')
    assert all(res.df['Tipo de documento'] != TIPO_APPLICATION_RESPONSE)
    print(f'✅ test_filtro_credito_recibido: {res.filas_filtradas} facturas, ${res.total_valor:,.0f}')


def test_filtro_sin_filtros():
    df = cargar_excel_radian('/home/claude/radian_muestra.xlsx')
    res = filtrar(df, forma_pago=None, grupo=None, excluir_application_response=False, solo_aprobados=False)
    assert res.filas_filtradas == 83569
    print(f'✅ test_filtro_sin_filtros: {res.filas_filtradas} filas (todas)')


def test_nit_limpio():
    assert _limpiar_nit(900533491) == '900533491'
    assert _limpiar_nit('900533491.0') == '900533491'
    assert _limpiar_nit('900533491-5') == '900533491'
    assert _limpiar_nit(None) == ''
    assert _limpiar_nit(pd.NA) == ''
    print('✅ test_nit_limpio')


def test_resumen_por_proveedor():
    df = cargar_excel_radian('/home/claude/radian_muestra.xlsx')
    res = filtrar(df, forma_pago=2, grupo='Recibido')
    df_prov = resumir_por_proveedor(res.df)
    assert not df_prov.empty
    # El top proveedor por valor debe ser Costa del Campo SAS
    top = df_prov.iloc[0]
    assert top['NIT Emisor'] == '900555892'
    assert top['total_general'] > 900_000_000
    print(f'✅ test_resumen_por_proveedor: top = {top["Nombre Emisor"][:40]} ${top["total_general"]:,.0f}')


def test_exportar_excel():
    df = cargar_excel_radian('/home/claude/radian_muestra.xlsx')
    res = filtrar(df, forma_pago=2, grupo='Recibido')
    bytes_excel = exportar_a_excel(res, nombre_empresa='JIPER S.A.S')
    assert len(bytes_excel) > 1000
    # Guardar para inspección visual
    with open('/tmp/test_radian_export.xlsx', 'wb') as f:
        f.write(bytes_excel)
    print(f'✅ test_exportar_excel: {len(bytes_excel):,} bytes generados')


def test_filtro_por_fecha():
    df = cargar_excel_radian('/home/claude/radian_muestra.xlsx')
    res = filtrar(df, forma_pago=2, grupo='Recibido',
                  fecha_desde='2026-04-01', fecha_hasta='2026-04-30')
    assert res.filas_filtradas < 4127  # Debe ser un subset
    print(f'✅ test_filtro_por_fecha: {res.filas_filtradas} facturas en abril 2026')


def test_filtro_por_nit():
    df = cargar_excel_radian('/home/claude/radian_muestra.xlsx')
    # Costa del Campo es el top proveedor
    res = filtrar(df, forma_pago=2, grupo='Recibido', nit_contraparte='900555892')
    assert res.filas_filtradas == 6  # 6 facturas según análisis
    assert all(res.df['NIT Emisor'] == '900555892')
    print(f'✅ test_filtro_por_nit: 900555892 → {res.filas_filtradas} facturas')


if __name__ == '__main__':
    print('Ejecutando tests del procesador RADIAN...\n')
    test_cargar_excel()
    test_resumen()
    test_filtro_credito_recibido()
    test_filtro_sin_filtros()
    test_nit_limpio()
    test_resumen_por_proveedor()
    test_filtro_por_fecha()
    test_filtro_por_nit()
    test_exportar_excel()
    print('\n🎉 Todos los tests pasaron')
