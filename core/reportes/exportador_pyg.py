"""
Exportador Excel del P&G detallado por centro de costos.

Genera un workbook con:
    - Hoja "Índice" con enlaces a cada centro.
    - Hoja "Consolidado" (todos los CC; aplica inventario del formulario).
    - Una hoja por cada centro de costo con actividad (inventario en blanco).

Cada hoja replica la plantilla: Concepto | <meses> | Acumulado | A.V., con
secciones, subtotales en negrita y celdas de inventario editables (azul).
"""
from __future__ import annotations

import io
import re
from typing import List, Optional, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .reporte_centros_costos import BalanceCC, fusionar_prr
from .pyg_detallado import construir_pyg, _mes_de_periodo, _plantilla

FUENTE = "Calibri"
FMT_NUM = '#,##0;(#,##0);"-"'
FMT_PCT = '0.0%;(0.0%);"-"'
AZUL = "1F4E78"
AZUL_CLARO = "D9E1F2"
AZUL_INPUT = "0000FF"     # texto azul = celda editable
GRIS = "808080"
NARANJA_FILL = "FCE4D6"


def _sheet_name(nombre: str, usados: set) -> str:
    s = re.sub(r'[\[\]\:\*\?\/\\]', " ", nombre or "CC").strip()[:31] or "CC"
    base, i = s, 1
    while s.lower() in usados:
        suf = f" ({i})"
        s = base[:31 - len(suf)] + suf
        i += 1
    usados.add(s.lower())
    return s


def _escribir_hoja(ws, df: pd.DataFrame, titulo1: str, titulo2: str, titulo3: str,
                   plantilla=None, con_formulas: bool = True):
    meses = [c for c in df.columns if c not in ("Concepto", "Acumulado", "A.V.", "_tipo")]
    columnas = ["Concepto"] + meses + ["Acumulado", "A.V."]
    nmes = len(meses)

    ws.cell(row=1, column=1, value=titulo1).font = Font(name=FUENTE, bold=True, size=12)
    ws.cell(row=2, column=1, value=titulo2).font = Font(name=FUENTE, italic=True, size=9, color=GRIS)
    ws.cell(row=3, column=1, value=titulo3).font = Font(name=FUENTE, italic=True, size=9, color=GRIS)

    fila_enc = 5
    for j, col in enumerate(columnas, start=1):
        c = ws.cell(row=fila_enc, column=j, value=col)
        c.font = Font(name=FUENTE, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fila0 = fila_enc + 1                       # primera fila de datos (Excel)
    col_mes = [get_column_letter(2 + j) for j in range(nmes)]   # letras de meses
    col_acum = get_column_letter(2 + nmes)
    col_av = get_column_letter(3 + nmes)

    usar_formulas = con_formulas and plantilla is not None and len(plantilla) == len(df)

    # mapa label -> fila Excel (para referencias de fórmulas)
    label2row = {}
    if usar_formulas:
        for i, (kind, label, opts) in enumerate(plantilla):
            label2row[label] = fila0 + i
    # fila base del A.V. (Total ingresos operacionales netos)
    base_row = label2row.get("Total ingresos operacionales (netos)")

    def _expr(refs_mas, refs_menos, col):
        partes = []
        for lb in refs_mas:
            r = label2row.get(lb)
            if r:
                partes.append(f"+{col}{r}")
        for lb in refs_menos:
            r = label2row.get(lb)
            if r:
                partes.append(f"-{col}{r}")
        if not partes:
            return None
        s = "".join(partes)
        if s.startswith("+"):
            s = s[1:]
        return "=" + s

    fila = fila0
    for idx, (_, r) in enumerate(df.iterrows()):
        tipo = r["_tipo"]
        spec = plantilla[idx] if usar_formulas else (tipo, r["Concepto"], {})
        opts = spec[2] if usar_formulas else {}
        ws.cell(row=fila, column=1, value=r["Concepto"])

        if tipo not in ("header", "subhdr", "blank"):
            es_calc = tipo in ("sub", "ebitda")
            # --- meses ---
            for j, col in enumerate(meses):
                L = col_mes[j]
                if usar_formulas and es_calc:
                    f = _expr(opts.get("mas", []), opts.get("menos", []), L)
                    cell = ws.cell(row=fila, column=2 + j, value=f if f else 0)
                else:
                    v = r[col]
                    cell = ws.cell(row=fila, column=2 + j,
                                   value=(float(v) if v is not None else 0.0))
                cell.number_format = FMT_NUM
                cell.alignment = Alignment(horizontal="right")
            # --- Acumulado ---
            jac = 2 + nmes
            if usar_formulas and es_calc:
                f = _expr(opts.get("mas", []), opts.get("menos", []), col_acum)
                cell = ws.cell(row=fila, column=jac, value=f if f else 0)
            elif usar_formulas:
                # suma de los meses de esta misma fila
                rng = f"{col_mes[0]}{fila}:{col_mes[-1]}{fila}" if nmes > 1 else f"{col_mes[0]}{fila}"
                cell = ws.cell(row=fila, column=jac, value=f"=SUM({rng})")
            else:
                v = r["Acumulado"]
                cell = ws.cell(row=fila, column=jac, value=(float(v) if v is not None else 0.0))
            cell.number_format = FMT_NUM
            cell.alignment = Alignment(horizontal="right")
            # --- A.V. ---
            if r["A.V."] is not None or (usar_formulas and base_row):
                if usar_formulas and base_row:
                    f = f'=IFERROR({col_acum}{fila}/{col_acum}${base_row},"")'
                    cell = ws.cell(row=fila, column=jac + 1, value=f)
                else:
                    cell = ws.cell(row=fila, column=jac + 1, value=float(r["A.V."]))
                cell.number_format = FMT_PCT
                cell.alignment = Alignment(horizontal="right")

        # estilos por tipo
        es_sub = tipo in ("sub", "ebitda")
        for j in range(1, len(columnas) + 1):
            c = ws.cell(row=fila, column=j)
            if tipo == "header":
                c.font = Font(name=FUENTE, bold=True, size=10, color=AZUL)
            elif tipo == "subhdr":
                c.font = Font(name=FUENTE, bold=True, italic=True, size=9)
            elif es_sub:
                c.font = Font(name=FUENTE, bold=True, size=10)
                c.fill = PatternFill("solid", fgColor=AZUL_CLARO)
            elif tipo == "inv":
                c.font = Font(name=FUENTE, size=9, color=AZUL_INPUT)   # editable
            elif tipo == "trasl":
                c.font = Font(name=FUENTE, size=9, color=AZUL_INPUT)   # editable (manual)
            elif tipo == "mil":
                c.font = Font(name=FUENTE, size=9, italic=True)
                c.fill = PatternFill("solid", fgColor=NARANJA_FILL)    # Milagros
            else:
                c.font = Font(name=FUENTE, size=9)
        fila += 1

    ws.freeze_panes = ws.cell(row=fila_enc + 1, column=2)
    ws.column_dimensions["A"].width = 48
    for j in range(2, len(columnas) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 15


def exportar_pyg_excel(balances: List[BalanceCC],
                       balances_milagros: Optional[List[BalanceCC]] = None,
                       inventarios: Optional[Dict] = None,
                       traslados: Optional[Dict] = None,
                       mil_label: str = "Milagros",
                       fusionar_prr_cc: bool = True,
                       con_formulas: bool = True,
                       solo_con_actividad: bool = True) -> bytes:
    """Genera el workbook completo (consolidado + una hoja por CC).

    `con_formulas=True`: los subtotales, utilidades, EBITDA, Acumulado y A.V.
    salen como fórmulas de Excel que referencian las celdas; las celdas de
    inventario y traslados son las entradas editables. Así, al cambiar un dato
    a mano, Excel recalcula todo automáticamente.
    """
    if not balances:
        raise ValueError("Se requiere al menos un balance mensual.")

    if fusionar_prr_cc:
        balances = [fusionar_prr(b)[0] for b in balances]
        if balances_milagros:
            balances_milagros = [fusionar_prr(b)[0] for b in balances_milagros]

    bal_ref = sorted(balances, key=lambda b: _mes_de_periodo(b.periodo)[0])[-1]
    empresa, nit = bal_ref.empresa, bal_ref.nit
    t1 = f"{empresa}  ·  {nit}"
    t3 = "Cifras en COP · Mensual = movimiento del mes · Acumulado = suma de meses · A.V. = % sobre ingresos netos"
    # plantilla alineada con las filas del DataFrame (para construir fórmulas)
    plantilla = _plantilla(con_mil=bool(balances_milagros), mil=mil_label)

    wb = Workbook()
    usados = set()

    # --- Consolidado ---
    ws_cons = wb.active
    ws_cons.title = "Consolidado"
    usados.add("consolidado")
    df_cons = construir_pyg(balances, balances_milagros=balances_milagros,
                            cc=None, inventarios=inventarios, traslados=traslados,
                            mil_label=mil_label)
    sub2 = "Estado de Resultados — CONSOLIDADO (todos los centros)"
    if balances_milagros:
        sub2 += f" · incluye {mil_label}"
    _escribir_hoja(ws_cons, df_cons, t1, sub2, t3,
                   plantilla=plantilla, con_formulas=con_formulas)

    # --- CC con actividad (unión de centros de TODOS los meses y ambas empresas) ---
    etq = dict(bal_ref.etiqueta_cc())
    ccs = set()
    for b in balances:
        ccs |= set(b.detalle["cc"].unique())
    if balances_milagros:
        for b in balances_milagros:
            ccs |= set(b.detalle["cc"].unique())
        mref = sorted(balances_milagros, key=lambda b: _mes_de_periodo(b.periodo)[0])[-1]
        for cc_, lbl in mref.etiqueta_cc().items():
            etq.setdefault(cc_, lbl)
    ccs = sorted(ccs)
    info_index = []
    for cc in ccs:
        df = construir_pyg(balances, balances_milagros=balances_milagros,
                           cc=cc, inventarios=inventarios, traslados=traslados,
                           mil_label=mil_label)
        ing = df.loc[df["Concepto"] == "Total ingresos operacionales (netos)", "Acumulado"]
        cost = df.loc[df["Concepto"] == "TOTAL COSTO DE VENTAS Y DE PRODUCCIÓN", "Acumulado"]
        ingv = float(ing.iloc[0]) if not ing.empty else 0.0
        costv = float(cost.iloc[0]) if not cost.empty else 0.0
        if solo_con_actividad and abs(ingv) < 1 and abs(costv) < 1:
            continue
        nombre = etq.get(cc, cc)
        hoja = _sheet_name(nombre, usados)
        ws = wb.create_sheet(hoja)
        _escribir_hoja(ws, df, t1, f"Estado de Resultados — {nombre}  [{cc}]", t3,
                       plantilla=plantilla, con_formulas=con_formulas)
        un = df.loc[df["Concepto"] == "UTILIDAD (PÉRDIDA) NETA DEL EJERCICIO", "Acumulado"]
        info_index.append((cc, nombre, hoja, ingv,
                           float(un.iloc[0]) if not un.empty else 0.0))

    # --- Índice (al inicio) ---
    ws_idx = wb.create_sheet("Índice", 0)
    ws_idx.cell(row=1, column=1, value=t1).font = Font(name=FUENTE, bold=True, size=12)
    ws_idx.cell(row=2, column=1, value="Índice de centros de costo").font = Font(
        name=FUENTE, italic=True, size=10, color=GRIS)
    hdr = ["Código", "Centro de Costo", "Ingresos (acum)", "Utilidad neta (acum)", "Hoja"]
    for j, h in enumerate(hdr, start=1):
        c = ws_idx.cell(row=4, column=j, value=h)
        c.font = Font(name=FUENTE, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
    ws_idx.cell(row=5, column=2, value="CONSOLIDADO")
    ws_idx.cell(row=5, column=5, value="Consolidado").hyperlink = "#'Consolidado'!A1"
    ws_idx.cell(row=5, column=5).font = Font(name=FUENTE, color="0563C1", underline="single")
    fila = 6
    for cc, nombre, hoja, ingv, unv in info_index:
        ws_idx.cell(row=fila, column=1, value=cc)
        ws_idx.cell(row=fila, column=2, value=nombre)
        ci = ws_idx.cell(row=fila, column=3, value=ingv); ci.number_format = FMT_NUM
        cu = ws_idx.cell(row=fila, column=4, value=unv); cu.number_format = FMT_NUM
        link = ws_idx.cell(row=fila, column=5, value=hoja)
        link.hyperlink = f"#'{hoja}'!A1"
        link.font = Font(name=FUENTE, color="0563C1", underline="single")
        fila += 1
    ws_idx.column_dimensions["A"].width = 12
    ws_idx.column_dimensions["B"].width = 32
    ws_idx.column_dimensions["C"].width = 18
    ws_idx.column_dimensions["D"].width = 20
    ws_idx.column_dimensions["E"].width = 24
    ws_idx.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
