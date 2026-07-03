"""
Lectura de ventas del INFORME (EEFF "P&G por punto") por centro de costo y mes.

El P&G del informe trae, en cada hoja (Grupo1..4, CDP, Admon, ML...), un bloque
de 10 columnas por punto. La fila "% PARTICIPACION" contiene los códigos de CC
(en el desplazamiento +2 / +4 del bloque); dos filas más abajo está la fila de
meses (ENERO en el inicio del bloque, FEBRERO en +2, MARZO en +4, luego
DIFERENCIA y ACUMULADO que se ignoran). El renglón "Ventas" trae el valor de
ventas y "Devoluciones en ventas" las devoluciones (en negativo).

Devuelve un dict {(cc4, mes_num): {"ventas": x, "devol": y}} donde cc4 es el
código de CC a 4 dígitos (p. ej. "1107") y mes_num ∈ 1..12. El ingreso neto del
mes = ventas + devol (las devoluciones vienen en negativo).
"""
from __future__ import annotations

from typing import Dict, Tuple, Union, IO

from openpyxl import load_workbook

_MES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}

_PASO = 10  # ancho del bloque de cada punto


def _num(x) -> float:
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def _mes_de(v) -> Union[int, None]:
    """Devuelve el número de mes si la celda nombra un mes; None si es
    DIFERENCIA, ACUMULADO, vacío u otra cosa."""
    if not v:
        return None
    u = str(v).strip().upper()
    for nombre, num in _MES.items():
        if nombre in u:
            return num
    return None


def cargar_ventas_eeff(archivo: Union[str, IO[bytes]]) -> Dict[Tuple[str, int], Dict[str, float]]:
    """Lee el EEFF y devuelve ventas/devoluciones por (cc4, mes)."""
    wb = load_workbook(archivo, read_only=True, data_only=True)
    return _ventas_de_wb_eeff(wb)


def _ventas_de_wb_eeff(wb) -> Dict[Tuple[str, int], Dict[str, float]]:
    ventas: Dict[Tuple[str, int], Dict[str, float]] = {}

    for hoja in wb.sheetnames:
        ws = wb[hoja]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        rp = next((i for i, x in enumerate(rows)
                   if x and x[0] and str(x[0]).strip().upper().startswith("% PARTICIPACION")), None)
        if rp is None or rp + 2 >= len(rows):
            continue
        rv = next((i for i, x in enumerate(rows)
                   if x and x[0] and str(x[0]).strip().lower() == "ventas"), None)
        if rv is None:
            continue
        rd = next((i for i, x in enumerate(rows)
                   if x and x[0] and str(x[0]).strip().lower().startswith("devoluciones")), None)

        fcc = rows[rp]
        fnom = rows[rp + 1] if rp + 1 < len(rows) else ()
        fmes = rows[rp + 2]
        fv = rows[rv]
        fd = rows[rd] if rd is not None else None
        ncol = len(fmes)

        # primera columna que nombra un mes -> inicio del primer bloque
        b0 = next((c for c in range(1, ncol) if _mes_de(fmes[c]) is not None), None)
        if b0 is None:
            continue

        # mapa offset->mes del layout de la hoja (para celdas de mes sin
        # etiqueta, como bloques donde borraron el texto del mes): se toma el
        # mes más frecuente en cada offset entre todos los bloques.
        off_mes: Dict[int, Dict[int, int]] = {}
        b = b0
        while b < ncol:
            for off in (0, 2, 4, 6, 8):
                if b + off >= ncol:
                    break
                m = _mes_de(fmes[b + off])
                if m is not None:
                    off_mes.setdefault(off, {})
                    off_mes[off][m] = off_mes[off].get(m, 0) + 1
            b += _PASO
        off2mes = {off: max(c, key=c.get) for off, c in off_mes.items()}

        b = b0
        while b < ncol:
            # el código de CC del bloque está en +2 (o +4)
            cc = None
            for off in (2, 4):
                if b + off < len(fcc) and fcc[b + off] not in (None, ""):
                    s = str(fcc[b + off]).strip()
                    if s and s[-4:].isdigit():
                        cc = s[-4:]
                        break
            if cc:
                # validar que el bloque sea un PUNTO real y no un agregado
                # (la hoja "Total" trae bloques GRUPO 1..4 / TOTAL con códigos sueltos)
                nom_blk = ""
                for off in (0, 2, 4):
                    if b + off < len(fnom) and fnom[b + off]:
                        nom_blk = str(fnom[b + off]).strip().upper()
                        break
                if nom_blk.startswith("GRUPO") or nom_blk.startswith("TOTAL"):
                    b += _PASO
                    continue
                for off in (0, 2, 4, 6, 8):
                    if b + off >= ncol:
                        break
                    mes = _mes_de(fmes[b + off])
                    if mes is None:
                        celda = fmes[b + off]
                        if celda is None or not str(celda).strip():
                            # celda sin etiqueta: usar el mes del layout
                            mes = off2mes.get(off)
                    if mes is None:
                        continue
                    v = _num(fv[b + off]) if b + off < len(fv) else 0.0
                    d = _num(fd[b + off]) if (fd is not None and b + off < len(fd)) else 0.0
                    e = ventas.setdefault((cc, mes), {"ventas": 0.0, "devol": 0.0})
                    e["ventas"] += v
                    e["devol"] += d
            b += _PASO

    return ventas

# ------------------------------------------------------------------
# Informe administrativo MENSUAL ("RESULTADOS <MES> <AÑO>")
# ------------------------------------------------------------------
# Estructura: una hoja con los conceptos en la columna A; el renglón
# "VENTA EJECUTADA" trae la venta del mes; la fila inmediatamente anterior
# tiene el NOMBRE de cada punto (una columna por punto). Las columnas
# "TOTAL ..." y de control se ignoran. Como no trae códigos de CC, se mapea
# el nombre del punto a su código con el catálogo de centros.

_CATALOGO_CC = {
    # CDP / producción / administración
    "1001": "ADMINISTRACION", "1002": "CDP LA REGIONAL", "1003": "CDP MEDELLIN",
    "1004": "ZONA FRANCA",
    # Puntos JIPER (SL)
    "1101": "INDIANA", "1102": "OVIEDO", "1103": "EL TESORO", "1104": "SAN LUCAS",
    "1105": "DEL ESTE", "1106": "VIVA ENVIGADO", "1107": "LAURELES",
    "1108": "LLANOGRANDE", "1109": "BURBUJA POBLADO", "1110": "MIXY LOS COLORES",
    "1111": "CIUDAD DEL RIO", "1112": "FABRICATO", "1113": "LOS MOLINOS",
    "1114": "UNICENTRO", "1115": "LAS VEGAS", "1116": "MEDICAL TOWER",
    "1117": "MILLA DE ORO", "1118": "HOTEL NOCK",
    # Puntos Milagros (ML)
    "1201": "EL TESORO", "1202": "REMEDIOS", "1203": "VIVA ENVIGADO",
    "1204": "LAURELES", "1205": "FABRICATO", "1206": "MANILA",
    "1207": "MILLA DE ORO",
}

_TOKENS_RUIDO = {"CDP", "SL", "ML", "PARQUE", "LOS", "DE", "Y", "EL", "LA",
                 "DEL", "HOTEL", "PUNTO", "PV"}


def _norm_tokens(nombre: str) -> set:
    u = str(nombre or "").upper().split("(")[0]
    u = u.replace("MIXI", "MIXY").replace("–", " ").replace("-", " ")
    toks = {t for t in u.split() if t and t not in _TOKENS_RUIDO}
    return toks


def cc_de_nombre(nombre: str) -> Union[str, None]:
    """Mapea el nombre de un punto del informe a su código de CC (4 dígitos)."""
    u = str(nombre or "").upper()
    if not u.strip():
        return None
    # casos especiales primero
    if "REMEDIOS" in u:
        return "1202"
    if "ZONA FRANCA" in u or "INSTITUCIONAL" in u:
        return "1004"
    if "CDP" in u and "REGIONAL" in u:
        return "1002"
    if "CDP" in u and ("MEDELLIN" in u or "MEDELLÍN" in u or "AMERICA" in u or "AMÉRICA" in u):
        return "1003"
    if "ADMINISTRACION" in u or "ADMINISTRACIÓN" in u or u.strip() == "ADMON":
        return "1001"

    es_ml = bool(re_ml.search(u))
    toks = _norm_tokens(nombre)
    if not toks:
        return None
    familia = "12" if es_ml else "11"
    candidatos = []
    for cod, nom_cat in _CATALOGO_CC.items():
        if not cod.startswith(familia):
            continue
        toks_cat = _norm_tokens(nom_cat)
        if toks == toks_cat or toks <= toks_cat or toks_cat <= toks:
            candidatos.append(cod)
    if len(candidatos) == 1:
        return candidatos[0]
    # desempate: mayor intersección de tokens
    if candidatos:
        return max(candidatos, key=lambda c: len(toks & _norm_tokens(_CATALOGO_CC[c])))
    return None


import re as _re
re_ml = _re.compile(r"(^|\s)ML(\s|$)")


def cargar_ventas_resultados(archivo: Union[str, IO[bytes]]) -> Dict[Tuple[str, int], Dict[str, float]]:
    """Lee un informe administrativo mensual (RESULTADOS) y devuelve las
    ventas por (cc4, mes) a partir del renglón VENTA EJECUTADA."""
    wb = load_workbook(archivo, read_only=True, data_only=True)
    return _ventas_de_wb_resultados(wb)


def _ventas_de_wb_resultados(wb) -> Dict[Tuple[str, int], Dict[str, float]]:
    ventas: Dict[Tuple[str, int], Dict[str, float]] = {}
    for hoja in wb.sheetnames:
        ws = wb[hoja]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        rv = next((i for i, x in enumerate(rows)
                   if x and x[0] and "VENTA EJECUTADA" in str(x[0]).upper()), None)
        if rv is None or rv == 0:
            continue
        # mes: del nombre de la hoja, o de la fila de encabezado (col A)
        mes = _mes_de(hoja)
        if mes is None:
            mes = _mes_de(rows[rv - 1][0] if rows[rv - 1] else None)
        if mes is None:
            for r in rows[:6]:
                mes = _mes_de(r[0] if r else None)
                if mes:
                    break
        if mes is None:
            continue
        enc = rows[rv - 1]
        fv = rows[rv]
        for j in range(1, len(enc)):
            nom = enc[j]
            if nom is None or not str(nom).strip():
                continue
            u = str(nom).upper()
            if u.startswith("TOTAL") or "REVISADO" in u:
                continue
            cc = cc_de_nombre(nom)
            if not cc:
                continue
            v = _num(fv[j]) if j < len(fv) else 0.0
            e = ventas.setdefault((cc, mes), {"ventas": 0.0, "devol": 0.0})
            e["ventas"] += v
    return ventas


def cargar_ventas(archivo: Union[str, IO[bytes]]) -> Dict[Tuple[str, int], Dict[str, float]]:
    """Autodetecta el formato del informe y devuelve ventas por (cc4, mes).

    - Si alguna hoja tiene la fila "% PARTICIPACION" -> formato EEFF (por punto).
    - Si alguna hoja tiene "VENTA EJECUTADA" en la columna A -> informe
      administrativo mensual (RESULTADOS).
    Se pueden combinar varios archivos sumando los dicts resultantes.
    """
    wb = load_workbook(archivo, read_only=True, data_only=True)
    for hoja in wb.sheetnames:
        for row in wb[hoja].iter_rows(min_row=1, max_row=15, values_only=True):
            if row and row[0] and str(row[0]).strip().upper().startswith("% PARTICIPACION"):
                return _ventas_de_wb_eeff(wb)
    return _ventas_de_wb_resultados(wb)


def combinar_ventas(*dicts) -> Dict[Tuple[str, int], Dict[str, float]]:
    """Une varios dicts de ventas. Si un (cc, mes) aparece en más de un
    archivo, el último gana (permite corregir un mes re-subiendo el informe)."""
    out: Dict[Tuple[str, int], Dict[str, float]] = {}
    for d in dicts:
        if d:
            out.update(d)
    return out
