"""
Informe de detalle por cuenta (4 → 6 dígitos) por centro de costos.

Estructura del estado de resultados en cada hoja:
  Ingresos (informe administrativo)
  (−) Costo de materia prima con juego de inventario, separado en
      Alimentos y bebidas / Aseo y cafetería: inventario inicial + compras
      (incluida la compañía secundaria, p. ej. Milagros) + traslados CDP −
      inventario final, más el resto de la clase 6
  (−) Mano de obra (72) y CIF (73)         = UTILIDAD BRUTA (margen bruto)
  (−) Gastos operacionales (51 + 52 + 55)  = UTILIDAD OPERACIONAL
  (+) Ingresos financieros (42–43)
  (−) Gastos financieros (53)              = UTILIDAD NETA ANTES DE IMPUESTOS
  (−) Impuesto de renta (54, visible aun en cero)
                                           = UTILIDAD NETA
  EBITDA (último)

Incluye: Índice con resumen por centro (EBITDA al final) e hipervínculos
internos, hoja de Conciliación Global contra la cuenta 41 de libros
(dinámica), detalle de 6 dígitos contraíble bajo grupos de 4 dígitos,
análisis vertical mes a mes (base: ingresos del mes, o el resultado final
en centros sin ingresos) y vínculo "← Volver al Índice" en cada hoja.
"""
from __future__ import annotations

import io
import re
import collections
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.properties import Outline

from .reporte_centros_costos import BalanceCC, fusionar_prr, SIN_CC
from .pyg_detallado import _mes_de_periodo

F = "Calibri"
AZUL = "1F4E78"; AZUL_C = "D9E1F2"; AZUL_M = "BDD7EE"
GRIS = "808080"; NAR = "C55A11"
FMT = '#,##0;(#,##0);"-"'
FPCT = '0.0%;(0.0%);"-"'

SEC_COSTOS_MO = [
    ("COSTO MANO DE OBRA (72)", ["72"], 1, "Total mano de obra (72)"),
    ("COSTO CIF (73)", ["73"], 1, "Total CIF (73)"),
]
SEC_GASTOS = [
    ("GASTOS DE ADMINISTRACIÓN (51)", ["51"], 1, "Total gastos de administración"),
    ("GASTOS DE VENTAS (52)", ["52"], 1, "Total gastos de ventas"),
    ("GASTOS DIVERSOS Y EXTRAORDINARIOS (55)", ["55"], 1, "Total gastos diversos (55)"),
]
SEC_FINANCIEROS = [
    ("INGRESOS FINANCIEROS (4205 + 43)", ["4205", "43"], -1, "Total ingresos financieros (4205+43)"),
    ("GASTOS FINANCIEROS (53)", ["53"], 1, "Total gastos financieros (53)"),
]
SEC_IMPUESTO = ("IMPUESTO DE RENTA (54)", ["54"], 1, "Total impuesto de renta (54)")
_DEP = ["5160", "5165", "7360", "7365"]
_DEP_4 = ["5160", "5165", "7360", "7365"]
_C6_NOMBRADAS = ["614005", "614010", "614095"]  # costo alim, costo aseo, traslados CDP


def _sane(n: str) -> str:
    n = re.sub(r"[\[\]\*\?/\\:]", "", str(n)).strip()
    return n[:31] or "CC"


def _preparar(balances: List[BalanceCC], balances_milagros: Optional[List[BalanceCC]],
              fusionar: bool):
    """Prepara estructuras: movimientos 6 díg por empresa, inventarios 14 y
    resultados por empresa para la conciliación."""
    bal_ord = sorted(balances, key=lambda b: _mes_de_periodo(b.periodo)[0])
    mil_ord = sorted(balances_milagros or [], key=lambda b: _mes_de_periodo(b.periodo)[0])
    if fusionar:
        bal_ord = [fusionar_prr(b)[0] for b in bal_ord]
        mil_ord = [fusionar_prr(b)[0] for b in mil_ord]
    nmes = len(bal_ord)
    nombres_mes = [_mes_de_periodo(b.periodo)[1] for b in bal_ord]
    nums_mes = [_mes_de_periodo(b.periodo)[0] for b in bal_ord]
    mil_por_mes = {_mes_de_periodo(b.periodo)[0]: b for b in mil_ord}

    def vacio():
        return collections.defaultdict(lambda: collections.defaultdict(lambda: [0.0] * nmes))
    data_p, data_s = vacio(), vacio()
    inv = collections.defaultdict(lambda: {"ini": [0.0] * nmes, "fin": [0.0] * nmes})
    tras = collections.defaultdict(lambda: [0.0] * nmes)   # (grupo, cc) -> mov por mes
    full = {}   # (cc, cuenta) -> {mov[], sa, ns, nom, ncc} para el balance de prueba
    nombres: Dict[str, str] = {}
    etq: Dict[str, str] = {}
    res_emp = {"principal": [0.0] * nmes, "secundaria": [0.0] * nmes}

    def acumular(b: BalanceCC, i: int, data, clave: str):
        d = b.detalle.copy()
        d["c6"] = d["cuenta"].str[:6]
        for (cc, c6), g in d.groupby(["cc", "c6"]):
            data[cc][c6][i] += float(g["mov_periodo"].sum())
        for grupo, pref in (("alim", "14050501"), ("aseo", "14050502")):
            sub = d[d["cuenta"].str.startswith(pref)]
            for ccx, g in sub.groupby("cc"):
                inv[(grupo, ccx)]["ini"][i] += float(g["saldo_ant"].sum())
                inv[(grupo, ccx)]["fin"][i] += float(g["nuevo_saldo"].sum())
        # traslado CDP a nivel de cuenta completa (61409501 alim / 61409502 aseo),
        # que en el dato agregado a 6 dígitos ("614095") no se distinguen
        for grupo, pref in (("alim", "61409501"), ("aseo", "61409502")):
            sub = d[d["cuenta"].str.startswith(pref)]
            for ccx, g in sub.groupby("cc"):
                tras[(grupo, ccx)][i] += float(g["mov_periodo"].sum())
        # detalle completo (balance de prueba) por cuenta auxiliar y CC
        for rr in d.itertuples(index=False):
            k = (rr.cc, rr.cuenta)
            e = full.get(k)
            if e is None:
                e = {"mov": [0.0] * nmes, "sa": float(rr.saldo_ant),
                     "ns": 0.0, "nom": rr.nombre, "ncc": rr.nombre_cc}
                full[k] = e
            e["mov"][i] += float(rr.mov_periodo)
            e["ns"] = float(rr.nuevo_saldo)
        for c, n in b.nombres_cuenta.items():
            nombres.setdefault(c, n)
        for c, n in b.etiqueta_cc().items():
            etq.setdefault(c, n)
        m4 = -float(d.loc[d["cuenta"].str.startswith("4"), "mov_periodo"].sum())
        m567 = float(d.loc[d["cuenta"].str.startswith(("5", "6", "7")), "mov_periodo"].sum())
        res_emp[clave][i] += m4 - m567

    for i, b in enumerate(bal_ord):
        acumular(b, i, data_p, "principal")
        bm = mil_por_mes.get(nums_mes[i])
        if bm is not None:
            acumular(bm, i, data_s, "secundaria")
    return data_p, data_s, inv, tras, full, nombres, etq, nmes, nombres_mes, nums_mes, res_emp


def exportar_detalle_cuentas(balances: List[BalanceCC],
                             balances_milagros: Optional[List[BalanceCC]] = None,
                             ventas_eeff: Optional[Dict] = None,
                             mil_label: str = "Milagros",
                             empresa: Optional[str] = None,
                             nit: Optional[str] = None,
                             fusionar_prr_cc: bool = True,
                             estimar_fin_ultimo: bool = True) -> bytes:
    """Genera el libro de detalle por cuenta y lo devuelve como bytes.

    estimar_fin_ultimo: si el último mes trae en cero un ingreso/gasto no
    operacional o financiero (cuentas 42, 43, 53 y 55) que sí tuvo movimiento
    en los meses anteriores, lo completa con el promedio de esos meses (útil
    cuando aún no se han causado los financieros del mes en curso). Las celdas
    estimadas se marcan en naranja.
    """
    ventas_eeff = ventas_eeff or {}
    (data_p, data_s, INV, TRAS, FULL, NOMC, ETQ, nmes, MESN, NUMS, res_emp) = _preparar(
        balances, balances_milagros, fusionar_prr_cc)

    # --- estimación de financieros del último mes (promedio de los anteriores) ---
    _EST_PREF = ("42", "43", "53", "55")
    est_por_cc = collections.defaultdict(set)   # cc -> {c6 estimados}
    if estimar_fin_ultimo and nmes >= 2:
        u = nmes - 1
        for data in (data_p, data_s):
            for cc, dcc in data.items():
                for c6, vals in dcc.items():
                    if not c6.startswith(_EST_PREF):
                        continue
                    previos = vals[:u]
                    prom = round(sum(previos) / u) if u else 0
                    if abs(prom) >= 1 and abs(vals[u]) < 0.5:
                        vals[u] = float(prom)
                        est_por_cc[cc].add(c6)
    est_any = set().union(*est_por_cc.values()) if est_por_cc else set()
    empresa = empresa or (balances[0].empresa if balances else "")
    nit = nit or (balances[0].nit if balances else "")
    con_mil = bool(balances_milagros)
    titulo_emp = f"{empresa} · {nit}" + (f" (consolidado con {mil_label})" if con_mil else "")
    MILU = str(mil_label).upper()

    todos_cc = set(data_p) | set(data_s)

    def cuentas_de(data, cc):
        if cc is None:
            out = collections.defaultdict(lambda: [0.0] * nmes)
            for d in data.values():
                for c6, v in d.items():
                    for i in range(nmes):
                        out[c6][i] += v[i]
            return out
        return data.get(cc, {})

    def merged(cc):
        out = collections.defaultdict(lambda: [0.0] * nmes)
        for data in (data_p, data_s):
            for c6, v in cuentas_de(data, cc).items():
                for i in range(nmes):
                    out[c6][i] += v[i]
        return out

    def suma(d, prefs, sg=1, menos=None):
        out = [0.0] * nmes
        menos = menos or []
        for c6, v in d.items():
            if any(c6.startswith(p) for p in prefs) and not any(c6.startswith(m) for m in menos):
                for i in range(nmes):
                    out[i] += sg * v[i]
        return out

    def inv_cc(grupo, cc):
        ini = [0.0] * nmes; fin = [0.0] * nmes
        for (g, ccx), e in INV.items():
            if g == grupo and (cc is None or ccx == cc):
                for i in range(nmes):
                    ini[i] += e["ini"][i]; fin[i] += e["fin"][i]
        return ini, fin

    def tr_cc(grupo, cc):
        out = [0.0] * nmes
        for (g, ccx), v in TRAS.items():
            if g == grupo and (cc is None or ccx == cc):
                for i in range(nmes):
                    out[i] += v[i]
        return out

    def vinf_cc(cc):
        out = [0.0] * nmes
        for (c4, m), e in ventas_eeff.items():
            if m in NUMS and (cc is None or str(c4) == str(cc)[-4:]):
                out[NUMS.index(m)] += float(e.get("ventas", 0.0)) + float(e.get("devol", 0.0))
        return out

    def vinf_partes_cc(cc):
        """(ventas_brutas, devoluciones) por mes. El neto = brutas + devol
        respeta el total del informe (las devoluciones vienen en negativo; en
        el informe mensual RESULTADOS la venta ya es neta y devol queda en 0)."""
        vg = [0.0] * nmes
        dv = [0.0] * nmes
        for (c4, m), e in ventas_eeff.items():
            if m in NUMS and (cc is None or str(c4) == str(cc)[-4:]):
                i = NUMS.index(m)
                vg[i] += float(e.get("ventas", 0.0))
                dv[i] += float(e.get("devol", 0.0))
        return vg, dv

    wb = Workbook()
    wsi = wb.active
    wsi.title = "Índice"

    def hoja_detalle(ws, cc, titulo):
        dM = merged(cc)               # ambas empresas
        dP = cuentas_de(data_p, cc)   # principal
        dS = cuentas_de(data_s, cc)   # secundaria (Milagros)
        vi = vinf_cc(cc)
        # --- materia prima (juego de inventario) ---
        ini_al, fin_al = inv_cc("alim", cc)
        ini_as, fin_as = inv_cc("aseo", cc)
        cp_alim = suma(dP, ["710501"]); cp_beb = suma(dP, ["710502"])
        cp_cafe = suma(dP, ["710503"]); cp_aseo = suma(dP, ["710504"])
        cs_ab = suma(dS, ["710501", "710502"]); cs_ac = suma(dS, ["710503", "710504"])
        # traslados desde el centro de producción (CDP), por dependencia
        # (se leen por cuenta completa; en el dato de 6 díg "614095" no se separan)
        tr_al = tr_cc("alim", cc)   # 61409501 traslado CDP alimentos y bebidas
        tr_as = tr_cc("aseo", cc)   # 61409502 traslado CDP aseo y cafetería
        # averías, retiros y pérdida de inventario (clase 6 fuera de costo/traslado)
        averias = suma(dM, ["6"], menos=_C6_NOMBRADAS)
        # juego de inventario: inv inicial + compras + traslados + averías − inv final
        juego_al = [ini_al[i] + cp_alim[i] + cp_beb[i] + cs_ab[i] + tr_al[i]
                    + averias[i] - fin_al[i] for i in range(nmes)]
        juego_as = [ini_as[i] + cp_cafe[i] + cp_aseo[i] + cs_ac[i] + tr_as[i]
                    - fin_as[i] for i in range(nmes)]
        c6_71 = [suma(dM, ["6"])[i] + suma(dM, ["71"])[i] for i in range(nmes)]
        ajuste = [c6_71[i] - juego_al[i] - juego_as[i] for i in range(nmes)]
        mp = [juego_al[i] + juego_as[i] + ajuste[i] for i in range(nmes)]  # = c6_71
        g72 = suma(dM, ["72"]); g73 = suma(dM, ["73"])
        costo = [mp[i] + g72[i] + g73[i] for i in range(nmes)]
        otros_ing = suma(dM, ["42"], -1)
        ing_fin = suma(dM, ["43"], -1)
        g51 = suma(dM, ["51"]); g52 = suma(dM, ["52"]); g55 = suma(dM, ["55"])
        g53 = suma(dM, ["53"]); g54 = suma(dM, ["54"])
        gop = [g51[i] + g52[i] + g55[i] for i in range(nmes)]
        bruta = [vi[i] - costo[i] for i in range(nmes)]
        operacional = [bruta[i] + otros_ing[i] - gop[i] for i in range(nmes)]
        antes_imp = [operacional[i] + ing_fin[i] - g53[i] for i in range(nmes)]
        neta = [antes_imp[i] - g54[i] for i in range(nmes)]
        dep = suma(dM, _DEP, 1)
        ebitda = [vi[i] - costo[i] - g51[i] - g52[i] + dep[i] for i in range(nmes)]
        base = [vi[i] if abs(vi[i]) > 0.5 else abs(neta[i]) for i in range(nmes)]
        base.append(sum(vi) if abs(sum(vi)) > 0.5 else abs(sum(neta)))
        sin_ing = abs(sum(vi)) <= 0.5

        ws.cell(row=1, column=1, value=titulo_emp).font = Font(name=F, bold=True, size=11)
        sub2 = ("A.V. = % sobre el RESULTADO FINAL del mes (centro sin ingresos)"
                if sin_ing else "A.V. = % sobre ingresos del mes")
        ws.cell(row=2, column=1,
                value=f"{titulo} · ingresos del informe administrativo · detalle 4 → 6 dígitos · {sub2}"
                ).font = Font(name=F, italic=True, size=9, color=GRIS)
        cv = ws.cell(row=3, column=1, value="← Volver al Índice")
        cv.hyperlink = Hyperlink(ref=cv.coordinate, location="'Índice'!A1",
                                 display="← Volver al Índice")
        cv.font = Font(name=F, size=9, color="0563C1", underline="single")
        r = 4
        heads = ["Cuenta", "Nombre"]
        for m in MESN + ["Acumulado"]:
            heads += [m, "A.V."]
        for j, h in enumerate(heads, start=1):
            c = ws.cell(row=r, column=j, value=h)
            c.font = Font(name=F, bold=True, color="FFFFFF", size=9)
            c.fill = PatternFill("solid", fgColor=AZUL)
            c.alignment = Alignment(horizontal="center")
        r += 1

        filas_valor = []   # filas numéricas (para A.V. en pasada final)
        col_v = [3 + 2 * i for i in range(nmes)]          # columnas de valor
        col_ac = 3 + 2 * nmes                              # columna acumulado
        let = get_column_letter

        def _estilo(row, jc, bold, fill, color, italic, sz, pct=False):
            c = ws.cell(row=row, column=jc)
            c.number_format = FPCT if pct else FMT
            c.alignment = Alignment(horizontal="right")
            if pct:
                c.font = Font(name=F, size=8, bold=bold, color=color or GRIS, italic=True)
            else:
                c.font = Font(name=F, size=sz, bold=bold, color=color or "000000", italic=italic)
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)
            return c

        def _acum_formula(row):
            return "=" + "+".join(f"{let(j)}{row}" for j in col_v)

        def num(row, vals, bold=False, fill=None, color=None, italic=False, sz=9):
            """Valores por mes + acumulado como FÓRMULA (=suma de meses)."""
            for i, v in enumerate(vals):
                _estilo(row, col_v[i], bold, fill, color, italic, sz).value = round(v)
                _estilo(row, col_v[i] + 1, bold, fill, color, italic, sz, pct=True)
            _estilo(row, col_ac, bold, fill, color, italic, sz).value = _acum_formula(row)
            _estilo(row, col_ac + 1, bold, fill, color, italic, sz, pct=True)
            filas_valor.append(row)

        def numf(row, formulas, bold=False, fill=None, color=None, italic=False, sz=10):
            """Fórmulas por mes + acumulado (=suma de meses)."""
            for i, f_ in enumerate(formulas):
                _estilo(row, col_v[i], bold, fill, color, italic, sz).value = f_
                _estilo(row, col_v[i] + 1, bold, fill, color, italic, sz, pct=True)
            _estilo(row, col_ac, bold, fill, color, italic, sz).value = _acum_formula(row)
            _estilo(row, col_ac + 1, bold, fill, color, italic, sz, pct=True)
            filas_valor.append(row)

        def f_sum_rango(r0, r1):
            return [f"=SUM({let(j)}{r0}:{let(j)}{r1})" for j in col_v]

        def f_sum_filas(filas, restar=()):
            out = []
            for j in col_v:
                t = "+".join(f"{let(j)}{r}" for r in filas) or "0"
                for r in restar:
                    t += f"-{let(j)}{r}"
                out.append("=" + t)
            return out

        def linea(txt, vals, cuenta="", bold=False, fill=None, nivel=0):
            nonlocal r
            if cuenta:
                ws.cell(row=r, column=1, value=cuenta).font = Font(name=F, size=9)
            ws.cell(row=r, column=2, value=txt).font = Font(name=F, size=9, bold=bold)
            if fill:
                for j in (1, 2):
                    ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor=fill)
            num(r, vals, bold=bold, fill=fill)
            if nivel:
                ws.row_dimensions[r].outline_level = nivel
                ws.row_dimensions[r].hidden = True
            r += 1

        def subtotal(txt, vals, nivel=0, salto=2, formulas=None):
            nonlocal r
            c = ws.cell(row=r, column=2, value=txt)
            c.font = Font(name=F, bold=True, size=10)
            for j in (1, 2):
                ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor=AZUL_C)
            if formulas is not None:
                numf(r, formulas, bold=True, fill=AZUL_C)
            else:
                num(r, vals, bold=True, fill=AZUL_C)
            if nivel:
                ws.row_dimensions[r].outline_level = nivel
                ws.row_dimensions[r].hidden = True
            r += salto

        def encabezado(txt):
            nonlocal r
            c = ws.cell(row=r, column=1, value=txt)
            c.font = Font(name=F, bold=True, size=10, color=AZUL)
            r += 1

        def subencabezado(txt):
            nonlocal r
            c = ws.cell(row=r, column=2, value=txt)
            c.font = Font(name=F, bold=True, size=9, italic=True)
            r += 1

        def seccion(tit, prefs, sg, sub_t, forzar=False, menos=None):
            nonlocal r
            menos = menos or []
            c6s = sorted(x for x in dM if any(x.startswith(p) for p in prefs)
                         and not any(x.startswith(m) for m in menos)
                         and any(abs(v) > 0.5 for v in dM[x]))
            sub = [0.0] * nmes
            if not c6s and not forzar:
                return sub
            por4 = collections.OrderedDict()
            for c6 in c6s:
                por4.setdefault(c6[:4], []).append(c6)
            v4s = {}
            for c4, hijos in por4.items():
                v4 = [0.0] * nmes
                for h in hijos:
                    for i in range(nmes):
                        v4[i] += sg * dM[h][i]
                v4s[c4] = v4
                for i in range(nmes):
                    sub[i] += v4[i]
            encabezado(tit)
            fila_total = r
            r += 1                                # se escribe al final con fórmula
            filas_padre = []
            for c4, hijos in por4.items():
                ws.cell(row=r, column=1, value=c4).font = Font(name=F, size=9, bold=True)
                ws.cell(row=r, column=2, value=NOMC.get(c4, "")).font = Font(name=F, size=9, bold=True)
                fila_padre = r
                filas_padre.append(fila_padre)
                for j in (1, 2):
                    ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor="F2F2F2")
                ws.row_dimensions[r].outline_level = 1
                ws.row_dimensions[r].hidden = True
                r += 1
                h0 = r
                for h in hijos:
                    ws.cell(row=r, column=1, value="   " + h).font = Font(name=F, size=9)
                    ws.cell(row=r, column=2, value=NOMC.get(h, "")).font = Font(name=F, size=9)
                    num(r, [sg * x for x in dM[h]])
                    ws.row_dimensions[r].outline_level = 2
                    ws.row_dimensions[r].hidden = True
                    r += 1
                numf(fila_padre, f_sum_rango(h0, r - 1), bold=True, fill="F2F2F2", sz=9)
            fin_r = r
            r = fila_total
            subtotal(sub_t, sub, salto=1, formulas=f_sum_filas(filas_padre))
            filas_seccion[sub_t] = fila_total
            padres_seccion.update({c4: fp for c4, fp in zip(por4.keys(), filas_padre)})
            r = fin_r + 1
            return sub

        filas_total = {}

        def total(txt, vals, formulas=None):
            nonlocal r
            ws.cell(row=r, column=2, value=txt).font = Font(name=F, bold=True, size=10)
            for j in (1, 2):
                ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor=AZUL_M)
            if formulas is not None:
                numf(r, formulas, bold=True, fill=AZUL_M, sz=10)
            else:
                num(r, vals, bold=True, fill=AZUL_M, sz=10)
            filas_total[txt] = r
            r += 2

        filas_seccion = {}   # nombre de subtotal -> fila
        padres_seccion = {}  # código 4 díg -> fila
        # 1) Ingresos (informe) — total arriba (fórmula a las líneas de ventas)
        # Devoluciones tomadas de la cuenta 4175 por CC (libros). La venta bruta
        # se ajusta para que el neto (bruta − devoluciones) = total del informe.
        dv = suma(dM, ["4175"], -1)                       # devoluciones (negativo)
        vg = [vi[i] - dv[i] for i in range(nmes)]         # venta bruta = neto informe + devoluciones
        encabezado("INGRESOS OPERACIONALES (informe administrativo)")
        fila_ing = r
        r += 1
        fila_vtas = r
        linea("Ventas brutas (informe)", vg, "4135", nivel=1)
        fila_dev = r
        linea("(−) Devoluciones en ventas [4175]", dv, "4175", nivel=1)
        fin_r = r
        r = fila_ing
        subtotal("Total ingresos operacionales (neto informe)", vi, salto=1,
                 formulas=f_sum_filas([fila_vtas, fila_dev]))
        filas_seccion["Total ingresos operacionales (informe)"] = fila_ing
        r = fin_r + 1
        # 2) Costo de materia prima (juego de inventario) — total arriba,
        #    sin cuentas de traslado (solo comparativo si llegan a tener valor)
        encabezado("COSTO DE MATERIA PRIMA (juego de inventario)")
        fila_mp = r
        r += 1
        fila_jal = r
        r += 1
        h0 = r
        linea("   (+) Inventario inicial alimentos/bebidas", ini_al, "14050501", nivel=2)
        linea("   (+) Compras de alimentos", cp_alim, "710501", nivel=2)
        linea("   (+) Compras de bebidas, vinos y licores", cp_beb, "710502", nivel=2)
        if con_mil and any(abs(x) > 0.5 for x in cs_ab):
            linea(f"   (+) Compras alimentos y bebidas {MILU}", cs_ab, "7105", nivel=2)
        linea("   (+) Traslado desde CDP alimentos y bebidas", tr_al, "61409501", nivel=2)
        linea("   (+) Averías, retiros y pérdida de inventario", averias, "6305", nivel=2)
        linea("   (−) Inventario final alimentos/bebidas", [-x for x in fin_al], "14050501", nivel=2)
        h1 = r - 1
        fin_r = r
        r = fila_jal
        subtotal("Costo alimentos y bebidas", juego_al, nivel=1, salto=1,
                 formulas=f_sum_rango(h0, h1))
        r = fin_r
        fila_jas = r
        r += 1
        h0 = r
        linea("   (+) Inventario inicial aseo/cafetería", ini_as, "14050502", nivel=2)
        linea("   (+) Compras cafetería y empaque", cp_cafe, "710503", nivel=2)
        linea("   (+) Compras de aseo", cp_aseo, "710504", nivel=2)
        if con_mil and any(abs(x) > 0.5 for x in cs_ac):
            linea(f"   (+) Compras cafetería y aseo {MILU}", cs_ac, "7105", nivel=2)
        linea("   (+) Traslado desde CDP aseo y cafetería", tr_as, "61409502", nivel=2)
        linea("   (−) Inventario final aseo/cafetería", [-x for x in fin_as], "14050502", nivel=2)
        h1 = r - 1
        fin_r = r
        r = fila_jas
        subtotal("Costo aseo y cafetería", juego_as, nivel=1, salto=1,
                 formulas=f_sum_rango(h0, h1))
        r = fin_r
        filas_mp = [fila_jal, fila_jas]
        if any(abs(x) > 0.5 for x in ajuste):
            linea("(±) Ajuste juego de inventario vs libros (comparativo)", ajuste, "71", nivel=1)
            filas_mp.append(r - 1)
        fin_r = r
        r = fila_mp
        subtotal("Total costo de materia prima", mp, salto=1, formulas=f_sum_filas(filas_mp))
        filas_seccion["Total costo de materia prima"] = fila_mp
        r = fin_r + 1
        S = filas_seccion

        def ref(nombres, restar=()):
            f_mas = [S[n] for n in nombres if n in S]
            f_men = [S[n] for n in restar if n in S]
            out = []
            for j in col_v:
                t = "+".join(f"{let(j)}{x}" for x in f_mas) or "0"
                for x in f_men:
                    t += f"-{let(j)}{x}"
                out.append("=" + t)
            return out

        # 3) Mano de obra y CIF -> utilidad bruta
        for tit, prefs, sg, sub_t in SEC_COSTOS_MO:
            seccion(tit, prefs, sg, sub_t)
        f_bruta = []
        for j in col_v:
            t = f"={let(j)}{S['Total ingresos operacionales (informe)']}-{let(j)}{S['Total costo de materia prima']}"
            for n in ("Total mano de obra (72)", "Total CIF (73)"):
                if n in S:
                    t += f"-{let(j)}{S[n]}"
            f_bruta.append(t)
        total("UTILIDAD BRUTA (MARGEN BRUTO)", bruta, formulas=f_bruta)
        # 4) Otros ingresos (42) + gastos operacionales -> utilidad operacional
        seccion("OTROS INGRESOS (42)", ["42"], -1, "Total otros ingresos (42)", menos=["4205"])
        for tit, prefs, sg, sub_t in SEC_GASTOS:
            seccion(tit, prefs, sg, sub_t)
        f_oper = []
        for j in col_v:
            t = f"={let(j)}{filas_total['UTILIDAD BRUTA (MARGEN BRUTO)']}"
            if "Total otros ingresos (42)" in S:
                t += f"+{let(j)}{S['Total otros ingresos (42)']}"
            for n in ("Total gastos de administración", "Total gastos de ventas", "Total gastos diversos (55)"):
                if n in S:
                    t += f"-{let(j)}{S[n]}"
            f_oper.append(t)
        total("UTILIDAD OPERACIONAL", operacional, formulas=f_oper)
        # 5) Financieros -> utilidad antes de impuestos
        for tit, prefs, sg, sub_t in SEC_FINANCIEROS:
            seccion(tit, prefs, sg, sub_t)
        f_ant = []
        for j in col_v:
            t = f"={let(j)}{filas_total['UTILIDAD OPERACIONAL']}"
            if "Total ingresos financieros (4205+43)" in S:
                t += f"+{let(j)}{S['Total ingresos financieros (4205+43)']}"
            if "Total gastos financieros (53)" in S:
                t += f"-{let(j)}{S['Total gastos financieros (53)']}"
            f_ant.append(t)
        total("UTILIDAD NETA ANTES DE IMPUESTOS", antes_imp, formulas=f_ant)
        # 6) Impuesto de renta (visible aunque sea cero) -> utilidad neta
        tit, prefs, sg, sub_t = SEC_IMPUESTO
        seccion(tit, prefs, sg, sub_t, forzar=True)
        f_neta = []
        for j in col_v:
            t = f"={let(j)}{filas_total['UTILIDAD NETA ANTES DE IMPUESTOS']}"
            if "Total impuesto de renta (54)" in S:
                t += f"-{let(j)}{S['Total impuesto de renta (54)']}"
            f_neta.append(t)
        total("UTILIDAD NETA", neta, formulas=f_neta)
        # 7) EBITDA de último = bruta − 51 − 52 + depreciaciones/amortizaciones
        f_eb = []
        for j in col_v:
            t = f"={let(j)}{filas_total['UTILIDAD BRUTA (MARGEN BRUTO)']}"
            for n in ("Total gastos de administración", "Total gastos de ventas"):
                if n in S:
                    t += f"-{let(j)}{S[n]}"
            for c4 in _DEP_4:
                if c4 in padres_seccion:
                    t += f"+{let(j)}{padres_seccion[c4]}"
            f_eb.append(t)
        total("EBITDA", ebitda, formulas=f_eb)

        # ---- A.V. como fórmulas (pasada final) ----
        fila_base_ing = S["Total ingresos operacionales (informe)"]
        fila_neta = filas_total["UTILIDAD NETA"]
        for fr in filas_valor:
            for i, j in enumerate(col_v + [col_ac]):
                vcell = f"{let(j)}{fr}"
                if sin_ing:
                    bcell = f"{let(j)}${fila_neta}"
                    formula = f'=IF(ABS({bcell})=0,"",{vcell}/ABS({bcell}))'
                else:
                    bcell = f"{let(j)}${fila_base_ing}"
                    formula = f'=IF({bcell}=0,"",{vcell}/{bcell})'
                ws.cell(row=fr, column=j + 1, value=formula)

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 44
        for i in range(nmes + 1):
            ws.column_dimensions[get_column_letter(3 + 2 * i)].width = 13
            ws.column_dimensions[get_column_letter(4 + 2 * i)].width = 7
        ws.freeze_panes = "C5"
        return (vi, costo, bruta, otros_ing, gop, operacional, ing_fin, g53,
                antes_imp, g54, neta, ebitda), dict(S), dict(filas_total)

    wsc = wb.create_sheet("Consolidado")
    cons, consS, consT = hoja_detalle(wsc, None, "CONSOLIDADO")
    ccs = [c for c in sorted(todos_cc) if c != SIN_CC] + ([SIN_CC] if SIN_CC in todos_cc else [])
    ccs = [cc for cc in ccs
           if any(c6[0] in "4567" and any(abs(v) > 0.5 for v in vals)
                  for c6, vals in merged(cc).items())]
    usados = {"Índice", "Consolidado", "Conciliación Global"}
    filas = []
    for cc in ccs:
        nom = ETQ.get(cc, cc)
        sn = _sane(nom); base_sn = sn; k = 2
        while sn in usados:
            sn = f"{base_sn[:28]} {k}"; k += 1
        usados.add(sn)
        ws = wb.create_sheet(sn)
        vals, vS, vT = hoja_detalle(ws, cc, f"{cc} · {nom}")
        filas.append((cc, nom, sn, vals, vS, vT))

    # ----- Índice -----
    (viC, costoC, brutaC, otrosC, gopC, operC, ifinC, g53C, antesC, g54C, netaC, ebC) = cons
    wsi.cell(row=1, column=1, value=titulo_emp).font = Font(name=F, bold=True, size=12)
    wsi.cell(row=2, column=1,
             value=f"Resumen por centro de costos · acumulado {MESN[0]}–{MESN[-1]} · ingresos del informe administrativo"
             ).font = Font(name=F, italic=True, size=9, color=GRIS)
    heads = ["CC", "Centro de costos", "Ingresos (informe)", "Costo (CMV + cta 7)",
             "Utilidad bruta", "Otros ingresos (42)", "Gastos operac. (51+52+55)",
             "Utilidad operacional", "Ingresos financieros (4205+43)", "Gastos financieros (53)",
             "Impuesto renta (54)", "Utilidad neta", "EBITDA"]
    r = 4
    for j, h in enumerate(heads, start=1):
        c = wsi.cell(row=r, column=j, value=h)
        c.font = Font(name=F, bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r += 1

    def numrow(row, vals, bold=False, fill=None):
        for j, v in enumerate(vals, start=3):
            c = wsi.cell(row=row, column=j,
                         value=v if isinstance(v, str) else round(v))
            c.number_format = FMT
            c.alignment = Alignment(horizontal="right")
            c.font = Font(name=F, size=9, bold=bold)
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)

    _LAC = get_column_letter(3 + 2 * nmes)  # columna del acumulado en las hojas

    def vals_idx(hoja, S, T):
        """Fórmulas del índice referenciando los totales (acumulado) de la hoja."""
        h = str(hoja).replace("'", "''")

        def ref(fila):
            return f"='{h}'!{_LAC}{fila}"

        def suma_refs(nombres):
            fs = [S[n] for n in nombres if n in S]
            if not fs:
                return 0
            return "='" + h + "'!" + ("+'" + h + "'!").join(f"{_LAC}{x}" for x in fs)

        return [
            ref(S["Total ingresos operacionales (informe)"]),
            suma_refs(["Total costo de materia prima", "Total mano de obra (72)", "Total CIF (73)"]),
            ref(T["UTILIDAD BRUTA (MARGEN BRUTO)"]),
            suma_refs(["Total otros ingresos (42)"]),
            suma_refs(["Total gastos de administración", "Total gastos de ventas",
                       "Total gastos diversos (55)"]),
            ref(T["UTILIDAD OPERACIONAL"]),
            suma_refs(["Total ingresos financieros (4205+43)"]),
            suma_refs(["Total gastos financieros (53)"]),
            suma_refs(["Total impuesto de renta (54)"]),
            ref(T["UTILIDAD NETA"]),
            ref(T["EBITDA"]),
        ]

    c = wsi.cell(row=r, column=2, value="CONSOLIDADO")
    c.hyperlink = Hyperlink(ref=c.coordinate, location="'Consolidado'!A1", display="CONSOLIDADO")
    c.font = Font(name=F, bold=True, size=9, color="0563C1", underline="single")
    numrow(r, vals_idx("Consolidado", consS, consT), bold=True, fill=AZUL_C)
    for j in (1, 2):
        wsi.cell(row=r, column=j).fill = PatternFill("solid", fgColor=AZUL_C)
    r += 1
    for fila_cc in filas:
        cc, nom, sn = fila_cc[0], fila_cc[1], fila_cc[2]
        wsi.cell(row=r, column=1, value=cc[-4:] if cc != SIN_CC else "—").font = Font(name=F, size=9)
        c = wsi.cell(row=r, column=2, value=nom)
        c.hyperlink = Hyperlink(ref=c.coordinate, location=f"'{sn}'!A1", display=str(nom))
        c.font = Font(name=F, size=9, color="0563C1", underline="single")
        numrow(r, vals_idx(sn, fila_cc[4], fila_cc[5]))
        r += 1
    r += 1
    for t in ["Ingresos: únicamente ventas de los informes administrativos (la cuenta 41 de libros NO se incluye; ver hoja Conciliación Global).",
              "Costo de materia prima = inventario inicial + compras (ambas compañías) + traslados desde CDP (614095) − inventario final, separado en alimentos/bebidas y aseo/cafetería, más el costo por retiros, averías y cortesías. La línea de ajuste (comparativo) concilia el juego de inventario contra el costo de libros.",
              "Utilidad bruta = ingresos − costo (materia prima + 72 + 73). Utilidad operacional = bruta + otros ingresos (42) − gastos operacionales (51+52+55).",
              "Utilidad antes de impuestos = operacional + ingresos financieros (43) − gastos financieros (53). Utilidad neta = antes de impuestos − renta (54).",
              "EBITDA = ingresos − costo − gastos 51+52 + depreciaciones y amortizaciones (5160, 5165, 7360, 7365).",
              "A.V. mes a mes: % sobre los ingresos del informe del mes; en centros sin ingresos, % sobre el resultado final del mes.",
              "Totales arriba: en cada hoja los botones +/− del margen izquierdo abren el detalle (totales → grupos de 4 dígitos → cuentas de 6 dígitos)."]:
        wsi.cell(row=r, column=2, value="· " + t).font = Font(name=F, italic=True, size=8, color=GRIS)
        r += 1
    wsi.column_dimensions["A"].width = 7
    wsi.column_dimensions["B"].width = 28
    for j in range(3, 14):
        wsi.column_dimensions[get_column_letter(j)].width = 13
    wsi.freeze_panes = "A5"

    # ----- Conciliación Global -----
    ws = wb.create_sheet("Conciliación Global", 1)
    ws.cell(row=1, column=1, value=f"{empresa} · {nit}").font = Font(name=F, bold=True, size=12)
    ws.cell(row=2, column=1,
            value=f"Conciliación global: ventas del informe administrativo vs libros contables (41) · {MESN[0]}–{MESN[-1]}"
            ).font = Font(name=F, italic=True, size=9, color=GRIS)
    cv = ws.cell(row=3, column=1, value="← Volver al Índice")
    cv.hyperlink = Hyperlink(ref=cv.coordinate, location="'Índice'!A1", display="← Volver al Índice")
    cv.font = Font(name=F, size=9, color="0563C1", underline="single")
    r = 4
    for j, h in enumerate(["Concepto"] + MESN + ["Acumulado"], start=1):
        c = ws.cell(row=r, column=j, value=h)
        c.font = Font(name=F, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center")
    r += 1

    def fila(row, txt, vals=None, bold=False, fill=None, color=None, italic=False, sz=10):
        c = ws.cell(row=row, column=1, value=txt)
        c.font = Font(name=F, bold=bold, italic=italic, size=sz, color=color or "000000")
        if fill:
            for j in range(1, nmes + 3):
                ws.cell(row=row, column=j).fill = PatternFill("solid", fgColor=fill)
        if vals is not None:
            vv = list(vals) + [sum(vals)]
            for j, v in enumerate(vv, start=2):
                c2 = ws.cell(row=row, column=j, value=round(v))
                c2.number_format = FMT
                c2.alignment = Alignment(horizontal="right")
                c2.font = Font(name=F, bold=bold, size=sz, color=color or "000000", italic=italic)
        return row + 1

    d_all = merged(None)
    l41 = suma(d_all, ["41"], -1)
    dif = [viC[i] - l41[i] for i in range(nmes)]
    r = fila(r, "1. INGRESOS", bold=True, color=AZUL)
    r = fila(r, "   Ventas según informes administrativos", viC)
    r = fila(r, "   Ingresos según libros (cuenta 41)", l41)
    r = fila(r, "   DIFERENCIA (informe − libros)", dif, bold=True, fill=AZUL_C)
    focos = []
    for cc in todos_cc:
        dcc = merged(cc)
        l41cc = suma(dcc, ["41"], -1)
        vicc = vinf_cc(cc)
        dcc_dif = [vicc[i] - l41cc[i] for i in range(nmes)]
        if any(abs(x) > 0.5 for x in dcc_dif + [sum(dcc_dif)]):
            focos.append((cc, dcc_dif))
    focos.sort(key=lambda x: -max(abs(v) for v in x[1] + [sum(x[1])]))
    for cc, v in focos[:6]:
        r = fila(r, f"      · {ETQ.get(cc, cc)}", v, italic=True, sz=9)
    resto = [sum(f[1][i] for f in focos[6:]) for i in range(nmes)]
    if any(abs(x) > 0.5 for x in resto):
        r = fila(r, "      · Otros centros (menores)", resto, italic=True, sz=9)
    r += 1
    r = fila(r, "2. RESULTADO", bold=True, color=AZUL)
    r = fila(r, f"   Resultado según libros {empresa.split()[0] if empresa else 'principal'}",
             res_emp["principal"])
    if con_mil:
        r = fila(r, f"   Resultado según libros {mil_label}", res_emp["secundaria"])
    res_lib = [res_emp["principal"][i] + res_emp["secundaria"][i] for i in range(nmes)]
    r = fila(r, "   Resultado libros combinado (4−5−6−7)", res_lib, bold=True)
    r = fila(r, "   (+) Diferencia de ingresos informe vs libros", dif, italic=True, color=NAR)
    r = fila(r, "   = UTILIDAD NETA según este informe", netaC, bold=True, fill=AZUL_C)
    r = fila(r, "   EBITDA según este informe", ebC, bold=True, fill=AZUL_C)
    r += 1
    r = fila(r, "3. VERIFICACIÓN (debe ser 0)", bold=True, color=AZUL)
    chk = [res_lib[i] + dif[i] - netaC[i] for i in range(nmes)]
    r = fila(r, "   Puente: libros + dif. ingresos − utilidad neta informe", chk)
    r += 1
    for t in ["· La diferencia de utilidad frente a libros es atribuible únicamente a los INGRESOS",
              "  (el informe administrativo y la contabilidad pueden reconocer ventas en meses distintos).",
              "· Costos y gastos de este informe salen directamente de libros, peso a peso."]:
        r = fila(r, t, italic=True, sz=9)
    ws.column_dimensions["A"].width = 58
    for j in range(2, nmes + 3):
        ws.column_dimensions[get_column_letter(j)].width = 16
    ws.freeze_panes = "B5"

    # ===== Hoja "Movimientos" (balance de prueba por cuenta auxiliar y CC) =====
    wsm = wb.create_sheet("Movimientos")
    wsm.cell(row=1, column=1, value=f"{empresa} · {nit} — Balance de prueba por cuenta auxiliar y centro de costos").font = Font(name=F, bold=True, size=12)
    wsm.cell(row=2, column=1, value=f"Composición de los saldos del informe · {MESN[0]}–{MESN[-1]} · usa el filtro de la fila de títulos").font = Font(name=F, italic=True, size=9, color=GRIS)
    cvb = wsm.cell(row=3, column=1, value="← Volver al Índice")
    cvb.hyperlink = Hyperlink(ref=cvb.coordinate, location="'Índice'!A1", display="← Volver al Índice")
    cvb.font = Font(name=F, size=9, color="0563C1", underline="single")
    hm = ["CC", "Nombre CC", "Cuenta", "Nombre cuenta", "Saldo inicial"] + MESN + ["Mov. acumulado", "Saldo final"]
    r = 5
    for j, h in enumerate(hm, start=1):
        c = wsm.cell(row=r, column=j, value=h)
        c.font = Font(name=F, bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center")
    r += 1
    fila_titulos = 5
    # ordenar por CC y cuenta; guardar fila de aterrizaje por (cc, cuenta)
    filas_mov = sorted(FULL.items(), key=lambda kv: (kv[0][0], kv[0][1]))
    landing = {}      # (cc, cuenta) -> fila ; por_cc para búsquedas por prefijo
    por_cc = collections.defaultdict(list)
    for (cc, cta), e in filas_mov:
        if not (any(abs(x) > 0.5 for x in e["mov"]) or abs(e["sa"]) > 0.5 or abs(e["ns"]) > 0.5):
            continue
        wsm.cell(row=r, column=1, value=cc).font = Font(name=F, size=9)
        wsm.cell(row=r, column=2, value=e["ncc"]).font = Font(name=F, size=9)
        wsm.cell(row=r, column=3, value=cta).font = Font(name=F, size=9)
        wsm.cell(row=r, column=4, value=e["nom"]).font = Font(name=F, size=9)
        vals = [e["sa"]] + e["mov"] + [None, e["ns"]]
        for k, v in enumerate(vals):
            col = 5 + k
            cc2 = wsm.cell(row=r, column=col)
            if col == 5 + nmes + 1:   # mov acumulado = suma de meses
                lets = get_column_letter(6); lete = get_column_letter(5 + nmes)
                cc2.value = f"=SUM({lets}{r}:{lete}{r})"
            else:
                cc2.value = round(v) if isinstance(v, (int, float)) else v
            cc2.number_format = FMT
            cc2.alignment = Alignment(horizontal="right")
            cc2.font = Font(name=F, size=9)
        landing[(cc, cta)] = r
        por_cc[cc].append((cta, r))
        r += 1
    ult = r - 1
    if ult >= fila_titulos + 1:
        wsm.auto_filter.ref = f"A{fila_titulos}:{get_column_letter(len(hm))}{ult}"
    wsm.freeze_panes = "E6"
    anchos = [10, 26, 12, 40, 15] + [14] * nmes + [15, 15]
    for j, w in enumerate(anchos, start=1):
        wsm.column_dimensions[get_column_letter(j)].width = w

    def landing_row(cc, code):
        """Primera fila de Movimientos para (cc, code) o para code en cualquier CC."""
        code = str(code).strip()
        if not code:
            return None
        candidatos = por_cc.get(cc, []) if cc is not None else None
        if candidatos is not None:
            for cta, fila in candidatos:
                if cta.startswith(code):
                    return fila
            return None
        # consolidado: primera coincidencia en orden CC, cuenta
        for (ccx, cta), e in filas_mov:
            if cta.startswith(code) and (ccx, cta) in landing:
                return landing[(ccx, cta)]
        return None

    # ---- enlazar cada cuenta del informe con su composición ----
    sn_a_cc = {f[2]: f[0] for f in filas}
    for ws2 in wb.worksheets:
        if ws2.title in ("Índice", "Conciliación Global", "Movimientos"):
            continue
        cc_hoja = None if ws2.title == "Consolidado" else sn_a_cc.get(ws2.title)
        for rr in range(6, ws2.max_row + 1):
            cod = str(ws2.cell(row=rr, column=1).value or "").strip()
            if not cod or not cod[0].isdigit():
                continue
            fila_dest = landing_row(cc_hoja, cod)
            if not fila_dest:
                continue
            cel = ws2.cell(row=rr, column=1)
            cel.hyperlink = Hyperlink(ref=cel.coordinate,
                                      location=f"'Movimientos'!A{fila_dest}",
                                      display=cod)
            fz = cel.font
            cel.font = Font(name=F, size=fz.size or 9, bold=fz.bold,
                            italic=fz.italic, color="0563C1", underline="single")

    for ws2 in wb.worksheets:
        if ws2.title in ("Índice", "Conciliación Global"):
            continue
        ws2.sheet_properties.outlinePr = Outline(summaryBelow=False, summaryRight=False)
        maxr = ws2.max_row
        for rr in range(1, maxr + 1):
            dn = ws2.row_dimensions.get(rr + 1)
            dt = ws2.row_dimensions.get(rr)
            if (dn.outline_level if dn else 0) > (dt.outline_level if dt else 0):
                ws2.row_dimensions[rr].collapsed = True

    # ---- marcar en naranja las celdas estimadas del último mes ----
    if est_por_cc:
        col_um = 3 + 2 * (nmes - 1)          # columna de valor del último mes
        col_ac = 3 + 2 * nmes                # columna acumulado
        sn_por_cc = {f[2]: f[0] for f in filas}   # nombre de hoja -> cc

        def _estimados_hoja(titulo):
            if titulo == "Consolidado":
                return est_any
            cc = sn_por_cc.get(titulo)
            return est_por_cc.get(cc, set())

        for ws2 in wb.worksheets:
            if ws2.title in ("Índice", "Conciliación Global"):
                continue
            est = _estimados_hoja(ws2.title)
            if not est:
                continue
            c4s = {c[:4] for c in est}
            for rr in range(5, ws2.max_row + 1):
                cod = str(ws2.cell(row=rr, column=1).value or "").strip()
                if cod in est or cod in c4s:
                    for col in (col_um, col_um + 1, col_ac, col_ac + 1):
                        cel = ws2.cell(row=rr, column=col)
                        fz = cel.font
                        cel.font = Font(name=F, size=fz.size or 9, bold=fz.bold,
                                        italic=True, color=NAR)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
