"""
Plantilla de datos manuales por centro de costo.

Genera un Excel para que el usuario capture, por centro de costo y mes:
    - Inventario inicial y final de alimentos/bebidas y de aseo/cafetería.
    - Traslados desde producción de alimentos y de aseo.

Y lo vuelve a leer para alimentar el P&G detallado.

Estructura de columnas (una hoja "Datos"):
    Código | Centro de Costo |
    <Mes> Inv Ini Alim | <Mes> Inv Fin Alim | <Mes> Inv Ini Aseo |
    <Mes> Inv Fin Aseo | <Mes> Traslado Alim | <Mes> Traslado Aseo | ...

Cada mes aporta 6 columnas. Los valores en blanco se toman como 0.
"""
from __future__ import annotations

import io
from typing import List, Tuple, Dict

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .pyg_detallado import _MESES

FUENTE = "Calibri"
AZUL = "1F4E78"
GRIS = "808080"

# (sufijo de columna, grupo, clave). clave: "ini"/"fin" para inventario,
# "trasl" para traslado.
_CAMPOS = [
    ("Inv Ini Alim", "alim", "ini"),
    ("Inv Fin Alim", "alim", "fin"),
    ("Inv Ini Aseo", "aseo", "ini"),
    ("Inv Fin Aseo", "aseo", "fin"),
    ("Traslado Alim", "alim", "trasl"),
    ("Traslado Aseo", "aseo", "trasl"),
]

_NOMBRE_A_NUM = {nom: num for num, nom in _MESES.values()}


def generar_plantilla_datos(centros: List[Tuple[str, str]],
                            meses_nombres: List[str]) -> bytes:
    """Genera la plantilla en blanco.

    `centros`: lista de (codigo_cc, nombre_cc).
    `meses_nombres`: nombres de mes en orden (p. ej. ["Enero","Febrero",...]).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    ws.cell(row=1, column=1,
            value="Plantilla de datos manuales por centro de costo").font = Font(
        name=FUENTE, bold=True, size=12)
    ws.cell(row=2, column=1,
            value=("Inventario inicial/final y traslados desde producción. "
                   "Deja en 0 lo que no aplique. No cambies la fila de encabezados "
                   "(fila 4) ni la columna Código.")).font = Font(
        name=FUENTE, italic=True, size=9, color=GRIS)

    # Encabezados (fila 4)
    headers = ["Código", "Centro de Costo"]
    for mes in meses_nombres:
        for suf, _g, _k in _CAMPOS:
            headers.append(f"{mes} · {suf}")
    fila_enc = 4
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=fila_enc, column=j, value=h)
        c.font = Font(name=FUENTE, bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Filas de centros
    fila = fila_enc + 1
    for cod, nom in centros:
        ws.cell(row=fila, column=1, value=cod).font = Font(name=FUENTE, size=9)
        ws.cell(row=fila, column=2, value=nom).font = Font(name=FUENTE, size=9)
        for j in range(3, len(headers) + 1):
            cell = ws.cell(row=fila, column=j, value=0)
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
        fila += 1

    ws.freeze_panes = "C5"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    for j in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def leer_plantilla_datos(fuente) -> Tuple[Dict, Dict]:
    """Lee la plantilla diligenciada.

    Devuelve (inventarios, traslados):
        inventarios: {(grupo, cc, mes_num): {"ini": x, "fin": y}}
        traslados:   {(grupo, cc, mes_num): valor}
    """
    wb = load_workbook(fuente, read_only=True, data_only=True)
    ws = wb["Datos"] if "Datos" in wb.sheetnames else wb.active
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(filas) < 5:
        return {}, {}

    headers = [str(h).strip() if h is not None else "" for h in filas[3]]
    # mapear índice de columna -> (grupo, clave, mes_num)
    colmap = {}
    for j, h in enumerate(headers):
        if " · " not in h:
            continue
        mespart, suf = h.split(" · ", 1)
        mes_num = _NOMBRE_A_NUM.get(mespart.strip())
        if mes_num is None:
            continue
        for s, g, k in _CAMPOS:
            if suf.strip() == s:
                colmap[j] = (g, k, mes_num)
                break

    inventarios: Dict = {}
    traslados: Dict = {}
    for row in filas[4:]:
        if not row or row[0] is None:
            continue
        cc = str(row[0]).strip()
        if not cc:
            continue
        for j, (g, k, mnum) in colmap.items():
            if j >= len(row):
                continue
            val = row[j]
            try:
                val = float(val) if val is not None else 0.0
            except (TypeError, ValueError):
                val = 0.0
            if not val:
                continue
            if k == "trasl":
                traslados[(g, cc, mnum)] = val
            else:
                inventarios.setdefault((g, cc, mnum), {})[k] = val
    return inventarios, traslados
