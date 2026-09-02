# -*- coding: utf-8 -*-
"""
Lector del "resumen de ventas por punto" (el cuadro NOMBRE | VALOR).

Sirve como fuente de valores para los Certificados de Ventas cuando la
informacion llega:
  - en imagen / pantallazo  -> OCR (tesseract), o
  - en Excel (dos columnas: nombre del punto + valor).

El cuadro trae dos secciones separadas por las filas "TOTAL SANTA LENA" y
"TOTAL MILAGROS"; por eso el mapeo al centro de costo se hace por SECCION
(familia 11xx = Santa Lena / JIPER, familia 12xx = Milagros), no solo por el
nombre.

IMPORTANTE: el OCR puede confundir digitos (1<->4, 6<->8). El resultado SIEMPRE
debe revisarse en una tabla editable antes de generar certificados.
"""
from __future__ import annotations

import io
import re
from typing import Dict, List, Optional

from openpyxl import load_workbook

from core.reportes.ventas_informe import _CATALOGO_CC, _norm_tokens


# ------------------------------------------------------------------ OCR
def ocr_disponible() -> bool:
    try:
        import pytesseract  # noqa: F401
        from PIL import Image  # noqa: F401
        pytesseract.get_tesseract_version()
        return True
    except Exception:
        return False


def leer_imagen_texto(img_bytes: bytes) -> str:
    """OCR de la imagen del resumen. Escala la imagen para mejorar los digitos."""
    import pytesseract
    from PIL import Image
    img = Image.open(io.BytesIO(img_bytes)).convert("L")
    w, h = img.size
    factor = 3 if max(w, h) < 1600 else 2
    img = img.resize((w * factor, h * factor))
    return pytesseract.image_to_string(img, lang="spa+eng", config="--psm 6")


# ------------------------------------------------------------------ mapeo por familia
def cc_de_nombre_en_familia(nombre: str, familia: str) -> Optional[str]:
    """Mapea el nombre a un CC dentro de la familia dada ('11' o '12')."""
    toks = _norm_tokens(nombre)
    if not toks:
        return None
    candidatos = []
    for cod, nom_cat in _CATALOGO_CC.items():
        if not cod.startswith(familia):
            continue
        toks_cat = _norm_tokens(nom_cat)
        if toks == toks_cat or toks <= toks_cat or toks_cat <= toks:
            candidatos.append(cod)
    if len(candidatos) == 1:
        return candidatos[0]
    if candidatos:
        return max(candidatos, key=lambda c: len(toks & _norm_tokens(_CATALOGO_CC[c])))
    return None


# ------------------------------------------------------------------ helpers de parseo
_SALTAR = ("TOTAL", "CRECIMIENTO", "VENTAS ACUMULA", "VENTAS ACUMULADAS")


def _solo_numero(s: str) -> Optional[int]:
    digs = re.sub(r"[^\d]", "", s or "")
    return int(digs) if digs else None


def _fila_valida(nombre: str) -> bool:
    u = (nombre or "").strip().upper()
    if not u:
        return False
    for k in _SALTAR:
        if u.startswith(k):
            return False
    if u in ("VENTAS", "DAS", "$"):
        return False
    return True


SECCIONES = ["Santa Leña", "Milagros"]


def familia_de_seccion(seccion: str) -> str:
    return "12" if str(seccion).strip().lower().startswith("mila") else "11"


def seccion_de_familia(familia: str) -> str:
    return "Milagros" if str(familia) == "12" else "Santa Leña"


def cc_por_nombre_seccion(nombre: str, seccion: str) -> str:
    """CC a partir del nombre y la seccion elegida por el usuario."""
    return cc_de_nombre_en_familia(nombre, familia_de_seccion(seccion)) or ""


def _cc_familia_por_seccion(nombre_upper: str, familia_actual: str) -> str:
    """Cambia a la familia Milagros al cruzar el separador de Santa Lena.

    Se aceptan varias formas por si el OCR lee distinto: 'TOTAL SANTA...',
    'CRECIMIENTO SANTA...', o cualquier linea con SANTA + (TOTAL o LENA/LEÑA).
    """
    u = nombre_upper
    if ("SANTA" in u and ("TOTAL" in u or "CRECIMIENTO" in u)) \
       or ("TOTAL SANTA" in u) or ("SANTA LE" in u and "TOTAL" in u):
        return "12"   # de aqui en adelante: Milagros
    return familia_actual


def _prefijo_marca(nombre: str):
    """Detecta el prefijo SL (Santa Lena) / ML (Milagros) al inicio del nombre.

    Devuelve (familia | None, nombre_sin_prefijo). Acepta 'SL OVIEDO', 'ML TESORO'
    y hasta 'MLCANAL' (sin espacio).
    """
    s = str(nombre or "").strip()
    u = s.upper()
    if u.startswith("SL"):
        return "11", s[2:].lstrip(" .-\t")
    if u.startswith("ML"):
        return "12", s[2:].lstrip(" .-\t")
    return None, s


def _col_valor_acumuladas(ws, defecto: int = 19) -> int:
    """Indice (1-based) de la columna de VENTAS ACUMULADAS; por defecto S (19)."""
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and "ACUMULAD" in c.value.upper():
                return c.column
    return defecto


def _arma_fila(nombre_bruto: str, valor, familia_actual: str):
    """Construye la fila con la familia dada por el prefijo SL/ML (o la seccion)."""
    pref, nombre = _prefijo_marca(nombre_bruto)
    familia = pref if pref else familia_actual
    if not _fila_valida(nombre) or not valor:
        return None
    cc4 = cc_de_nombre_en_familia(nombre, familia) or ""
    return {"nombre": nombre, "familia": familia,
            "seccion": seccion_de_familia(familia), "cc4": cc4, "valor": int(valor)}


# ------------------------------------------------------------------ parseo de texto (OCR)
def parsear_texto(texto: str) -> List[Dict]:
    """Filas [{nombre,familia,seccion,cc4,valor}] desde el texto OCR.

    La marca la da el prefijo SL/ML del nombre; si no hay prefijo, se usa la
    seccion (arriba/abajo de TOTAL SANTA LENA) como respaldo.
    """
    filas: List[Dict] = []
    familia = "11"
    for linea in (texto or "").splitlines():
        up = linea.upper()
        if "TOTAL MILAGROS" in up:      # fin de los puntos; lo de abajo es P&G
            break
        familia = _cc_familia_por_seccion(up, familia)
        if "TOTAL" in up or "CRECIMIENTO" in up:
            continue
        m = re.match(r"^\s*(.*?)\$\s*([\d.,\s]+)", linea)
        if not m:
            continue
        nombre = re.sub(r"\s+", " ", m.group(1)).strip(" .:-|")
        valor = _solo_numero(m.group(2))
        fila = _arma_fila(nombre, valor, familia)
        if fila:
            filas.append(fila)
    return filas


# ------------------------------------------------------------------ parseo de Excel
def parsear_excel(xlsx_bytes: bytes) -> List[Dict]:
    """Lee el resumen desde el Excel de ventas.

    - La marca la da el prefijo SL/ML del nombre (col A).
    - El valor se toma de la columna 'VENTAS ACUMULADAS' (col S por defecto).
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    filas: List[Dict] = []
    for ws in wb.worksheets:
        col_val = _col_valor_acumuladas(ws)     # 1-based
        familia = "11"
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            nombre = next((c for c in row if isinstance(c, str) and c.strip()), "")
            valor_celda = row[col_val - 1] if len(row) >= col_val else None
            up = (nombre or "").upper()
            if "TOTAL MILAGROS" in up:      # fin de los puntos; lo de abajo es P&G
                break
            familia = _cc_familia_por_seccion(up, familia)
            if "TOTAL" in up or "CRECIMIENTO" in up:
                continue
            valor = None
            if isinstance(valor_celda, (int, float)):
                valor = int(round(valor_celda))
            if not valor or valor <= 0:
                continue
            fila = _arma_fila(nombre, valor, familia)
            if fila:
                filas.append(fila)
    return filas


# ------------------------------------------------------------------ API principal
def leer_resumen(archivo_bytes: bytes, nombre_archivo: str = "") -> List[Dict]:
    """Detecta si es imagen o Excel y devuelve las filas [{nombre,familia,cc4,valor}]."""
    ext = (nombre_archivo or "").lower().rsplit(".", 1)[-1]
    if ext in ("xlsx", "xlsm", "xls"):
        return parsear_excel(archivo_bytes)
    if ext in ("png", "jpg", "jpeg", "tif", "tiff", "webp", "bmp"):
        if not ocr_disponible():
            raise RuntimeError(
                "OCR no disponible: el servidor no tiene tesseract instalado.")
        return parsear_texto(leer_imagen_texto(archivo_bytes))
    # sin extension: intentar imagen
    if not ocr_disponible():
        raise RuntimeError("Formato no reconocido y OCR no disponible.")
    return parsear_texto(leer_imagen_texto(archivo_bytes))


def filas_a_ventas(filas: List[Dict], mes_num: int) -> Dict:
    """Convierte filas revisadas en el dict de ventas {(cc4, mes): {'ventas': v}}."""
    out: Dict = {}
    for f in filas:
        cc4 = str(f.get("cc4") or "").strip()
        val = f.get("valor")
        if cc4 and val:
            out[(cc4, int(mes_num))] = {"ventas": float(val), "devol": 0.0}
    return out
