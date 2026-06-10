"""
Motor de reportes por Centro de Costos.

Procesa un "Balance de Prueba por Cuenta con CC" (exportado del software
contable) y produce:

    1) Balance de prueba por centro de costos  -> balance_por_cc()
    2) Estado de resultados por centro de costos -> estado_resultados_por_cc()

El archivo de entrada tiene esta estructura (hoja única "Datos"):

    Fila 1:  EMPRESA - NIT
    Fila 2:  Balance de Prueba por Cuenta con CC (Normal)  <Mes>-<dd>-<aaaa>
    Fila 3:  Cuenta | Equivalencia | Nombre | Centro de Costos | Nombre CC |
             Saldo Anterior | Débitos | Créditos | Nuevo Saldo
    Fila 4+: datos

Reglas contables aplicadas (PUC Colombia):
    - Clases de resultado: 4 Ingresos, 5 Gastos, 6 Costos de ventas,
      7 Costos de producción/operación.
    - Las cuentas de ingreso (clase 4) tienen naturaleza crédito y vienen
      con saldo NEGATIVO; para presentación se invierte el signo.
    - Nuevo Saldo = Saldo Anterior + Débitos - Créditos  (saldo ACUMULADO).
    - El movimiento DEL PERIODO de una cuenta = Débitos - Créditos.
    - Solo se suman las cuentas HOJA (las que no son prefijo de otra cuenta
      más larga) para no duplicar con los subtotales jerárquicos.
    - Para una cuenta hoja con detalle por CC se usan SUS líneas por CC;
      si una cuenta hoja solo trae línea sin CC, se asigna al bucket
      "(SIN CC)" para no perder saldos.

Este módulo no depende de Streamlit ni de Supabase: recibe una ruta o un
objeto tipo archivo (BytesIO) y devuelve DataFrames de pandas.
"""
from __future__ import annotations

from dataclasses import dataclass
from typing import Optional, Union, BinaryIO
import re

import pandas as pd
from openpyxl import load_workbook


# ------------------------------------------------------------------
# Constantes contables
# ------------------------------------------------------------------

CLASES_RESULTADO = {"4", "5", "6", "7"}

NOMBRE_CLASE = {
    "4": "INGRESOS",
    "5": "GASTOS",
    "6": "COSTOS DE VENTAS",
    "7": "COSTOS DE PRODUCCIÓN Y OPERACIÓN",
}

# Grupos (2 dígitos) y su rol en el estado de resultados.
SIN_CC = "(SIN CC)"
ETIQUETA_TOTAL = "TOTAL"


# ------------------------------------------------------------------
# Estructura de datos
# ------------------------------------------------------------------

@dataclass
class BalanceCC:
    """Balance de prueba por centro de costos ya normalizado."""
    empresa: str
    nit: str
    periodo: str            # texto tal cual del archivo, p. ej. "Abr-30-2026"
    titulo: str             # título completo de la fila 2
    detalle: pd.DataFrame   # cuentas HOJA (una fila por cuenta+CC)
    completo: pd.DataFrame  # todas las filas tal cual (incluye subtotales)
    nombres_cuenta: dict     # código de cuenta (cualquier nivel) -> nombre
    nombres_cc: dict         # código de CC -> nombre CC

    # ------- helpers de centros de costo -------
    def centros_de_costo(self) -> pd.DataFrame:
        """DataFrame único cc/nombre_cc presente en el detalle."""
        d = (
            self.detalle[["cc", "nombre_cc"]]
            .drop_duplicates()
            .sort_values("cc")
            .reset_index(drop=True)
        )
        return d

    def etiqueta_cc(self) -> dict:
        """Mapa código_CC -> etiqueta a mostrar (nombre del centro de costo).

        Si dos centros distintos comparten nombre, se desambigua con el código.
        SIN_CC se muestra tal cual.
        """
        nombres = (
            self.detalle[["cc", "nombre_cc"]].drop_duplicates()
        )
        # detectar nombres repetidos en códigos distintos
        repetidos = (
            nombres[nombres["cc"] != SIN_CC]
            .groupby("nombre_cc")["cc"].nunique()
        )
        repetidos = set(repetidos[repetidos > 1].index)
        mapa = {}
        for _, r in nombres.iterrows():
            cc, nom = r["cc"], (r["nombre_cc"] or "").strip()
            if cc == SIN_CC:
                mapa[cc] = SIN_CC
            elif not nom:
                mapa[cc] = cc
            elif nom in repetidos:
                mapa[cc] = f"{nom} ({cc})"
            else:
                mapa[cc] = nom
        return mapa


# ------------------------------------------------------------------
# Carga / parsing
# ------------------------------------------------------------------

def _to_float(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("\u00a0", "")
    if s in ("", "#N/A", "-"):
        return 0.0
    # quitar separadores de miles y normalizar coma decimal si aplica
    s = s.replace(",", "") if s.count(",") and s.count(".") else s
    try:
        return float(s)
    except ValueError:
        return 0.0


def cargar_balance(fuente: Union[str, BinaryIO]) -> BalanceCC:
    """Lee el .xlsx y devuelve un BalanceCC normalizado.

    `fuente` puede ser una ruta o un objeto tipo archivo (p. ej. el
    `UploadedFile` de Streamlit o un BytesIO).
    """
    wb = load_workbook(fuente, read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(filas) < 4:
        raise ValueError("El archivo no tiene la estructura esperada (muy pocas filas).")

    # --- Encabezado ---
    cab_empresa = str(filas[0][0] or "").strip()
    if " - " in cab_empresa:
        empresa, nit = cab_empresa.rsplit(" - ", 1)
        empresa, nit = empresa.strip(), nit.strip()
    else:
        empresa, nit = cab_empresa, ""

    titulo = str(filas[1][0] or "").strip()
    m = re.search(r"([A-Za-z]{3}-\d{1,2}-\d{4})", titulo)
    periodo = m.group(1) if m else titulo

    # --- Validar columnas (fila 3 = índice 2) ---
    encabezados = [str(c or "").strip() for c in filas[2]]
    esperado = ["Cuenta", "Equivalencia", "Nombre", "Centro de Costos",
                "Nombre CC", "Saldo Anterior", "Débitos", "Créditos", "Nuevo Saldo"]
    if encabezados[:len(esperado)] != esperado:
        raise ValueError(
            "Las columnas no coinciden con el formato esperado.\n"
            f"Esperado: {esperado}\nEncontrado: {encabezados[:len(esperado)]}"
        )

    # --- Datos ---
    registros = []
    for r in filas[3:]:
        if r[0] is None:
            continue
        cuenta = str(r[0]).strip()
        if not cuenta or cuenta == "#N/A":
            continue
        registros.append({
            "cuenta": cuenta,
            "nombre": str(r[2] or "").strip(),
            "cc": str(r[3] or "").strip(),
            "nombre_cc": str(r[4] or "").strip(),
            "saldo_ant": _to_float(r[5]),
            "debitos": _to_float(r[6]),
            "creditos": _to_float(r[7]),
            "nuevo_saldo": _to_float(r[8]),
        })

    completo = pd.DataFrame(registros)
    if completo.empty:
        raise ValueError("No se encontraron filas de datos válidas.")

    # marcar cuentas numéricas y nivel
    completo["es_numerica"] = completo["cuenta"].str.fullmatch(r"\d+")
    completo["clase"] = completo["cuenta"].str[0].where(completo["es_numerica"], "")
    completo["nivel"] = completo["cuenta"].str.len().where(completo["es_numerica"], 0)

    detalle = _extraer_hojas(completo)

    # Mapas de nombres (de TODAS las filas, cualquier nivel)
    nombres_cuenta = {}
    for _, r in completo.iterrows():
        c = r["cuenta"]
        if c not in nombres_cuenta and r["nombre"]:
            nombres_cuenta[c] = r["nombre"]
    nombres_cc = {}
    for _, r in completo.iterrows():
        cc = r["cc"]
        if cc and cc not in nombres_cc and r["nombre_cc"]:
            nombres_cc[cc] = r["nombre_cc"]

    return BalanceCC(
        empresa=empresa, nit=nit, periodo=periodo, titulo=titulo,
        detalle=detalle, completo=completo,
        nombres_cuenta=nombres_cuenta, nombres_cc=nombres_cc,
    )


def _extraer_hojas(completo: pd.DataFrame) -> pd.DataFrame:
    """Devuelve solo las cuentas HOJA, una fila por cuenta+CC.

    - Una cuenta es HOJA si ningún otro código numérico la tiene como prefijo.
    - Para cada cuenta hoja: si tiene líneas con CC se conservan esas;
      si solo tiene línea sin CC se etiqueta con SIN_CC.
    """
    num = completo[completo["es_numerica"]].copy()
    codigos = sorted(num["cuenta"].unique())

    # detectar padres: en orden lexicográfico, un código es padre si el
    # siguiente código empieza por él.
    padres = set()
    for i in range(len(codigos) - 1):
        if codigos[i + 1].startswith(codigos[i]):
            padres.add(codigos[i])

    hojas = num[~num["cuenta"].isin(padres)].copy()

    out = []
    for cuenta, grupo in hojas.groupby("cuenta", sort=False):
        con_cc = grupo[grupo["cc"] != ""]
        usar = con_cc if not con_cc.empty else grupo
        for _, row in usar.iterrows():
            d = row.to_dict()
            if d["cc"] == "":
                d["cc"] = SIN_CC
                d["nombre_cc"] = SIN_CC
            out.append(d)

    res = pd.DataFrame(out)
    # movimiento del periodo (naturaleza débito positiva)
    res["mov_periodo"] = res["debitos"] - res["creditos"]
    res = res.sort_values(["cuenta", "cc"]).reset_index(drop=True)
    return res


def fusionar_prr(bal: BalanceCC):
    """Devuelve (nuevo_balance, no_emparejados) con los centros 'PRR ...' y
    'PRR MMTO ...' fusionados dentro de su punto principal del mismo nombre.

    - Un centro es PRR si su nombre empieza por 'PRR '.
    - Se quita el prefijo 'PRR MMTO ' o 'PRR ' para obtener el nombre base y se
      busca el centro NO-PRR con ese nombre (exacto; si no, por tokens cuando
      es inequívoco). Su detalle se reasigna a ese código (gasto en su misma
      cuenta).
    - Los que no emparejan se conservan tal cual y se listan en no_emparejados.
    """
    import re as _re
    from dataclasses import replace as _replace

    det = bal.detalle.copy()
    principales = {}
    tokens_princ = {}
    for cc, nom in det[["cc", "nombre_cc"]].drop_duplicates().itertuples(index=False):
        if cc == SIN_CC:
            continue
        if not str(nom).upper().startswith("PRR "):
            principales.setdefault(str(nom).strip().upper(), cc)
            tokens_princ[cc] = set(str(nom).strip().upper().split())

    def _base(nom):
        n = str(nom).strip().upper()
        n = _re.sub(r"^PRR\s+MMTO\s+", "", n)
        n = _re.sub(r"^PRR\s+", "", n)
        return n.strip()

    mapa = {}
    no_emparejados = []
    for cc, nom in det[["cc", "nombre_cc"]].drop_duplicates().itertuples(index=False):
        if not str(nom).upper().startswith("PRR "):
            continue
        base = _base(nom)
        if base in principales:
            mapa[cc] = principales[base]
            continue
        tprr = set(base.split())
        cand = [c for c, tk in tokens_princ.items() if tk and tk <= tprr]
        if len(cand) == 1:
            mapa[cc] = cand[0]
        else:
            no_emparejados.append((cc, nom))

    if mapa:
        det["cc"] = det["cc"].map(lambda c: mapa.get(c, c))
        nombre_de = {cc: bal.nombres_cc.get(cc, cc) for cc in set(mapa.values())}
        det["nombre_cc"] = det.apply(
            lambda r: nombre_de.get(r["cc"], r["nombre_cc"]), axis=1)
        num = ["saldo_ant", "debitos", "creditos", "nuevo_saldo", "mov_periodo"]
        meta = [m for m in ["nombre", "nombre_cc", "clase", "es_numerica", "nivel"]
                if m in det.columns]
        agg = {c: "sum" for c in num if c in det.columns}
        agg.update({m: "first" for m in meta})
        det = det.groupby(["cuenta", "cc"], as_index=False).agg(agg)

    nuevo = _replace(bal, detalle=det)
    return nuevo, no_emparejados


# ------------------------------------------------------------------
# Reporte 1: Balance de prueba por centro de costos
# ------------------------------------------------------------------

def _rollup_detalle(bal: "BalanceCC", nivel: int) -> pd.DataFrame:
    """Agrega el detalle (cuentas hoja) al nivel de cuenta indicado.

    nivel: número de dígitos del código de cuenta (p. ej. 6 = subcuenta).
    Suma por (cuenta truncada a `nivel`, CC) y toma el nombre de ese código.
    """
    d = bal.detalle.copy()
    d["cuenta_n"] = d["cuenta"].str[:nivel]
    agg = (
        d.groupby(["cuenta_n", "cc"], as_index=False)
        .agg(saldo_ant=("saldo_ant", "sum"),
             debitos=("debitos", "sum"),
             creditos=("creditos", "sum"),
             nuevo_saldo=("nuevo_saldo", "sum"),
             mov_periodo=("mov_periodo", "sum"))
    )
    agg = agg.rename(columns={"cuenta_n": "cuenta"})
    agg["nombre"] = agg["cuenta"].map(
        lambda c: bal.nombres_cuenta.get(c, bal.nombres_cuenta.get(c, "")))
    agg["nombre_cc"] = agg["cc"].map(lambda cc: SIN_CC if cc == SIN_CC
                                     else bal.nombres_cc.get(cc, cc))
    agg["clase"] = agg["cuenta"].str[0]
    return agg


def balance_por_cc(
    bal: BalanceCC,
    cc: Optional[str] = None,
    solo_resultado: bool = False,
    nivel_detalle: int = 6,
    usar_nombre_cc: bool = True,
) -> pd.DataFrame:
    """Balance de prueba (saldo ant / débitos / créditos / nuevo saldo).

    - `cc`: filtra a un centro de costo (código). None = todos.
    - `solo_resultado=True`: solo clases 4..7.
    - `nivel_detalle`: dígitos del código de cuenta a mostrar (6 = subcuenta,
      8 = auxiliar). Se agregan los movimientos a ese nivel.
    - `usar_nombre_cc`: muestra el nombre del CC en lugar del código.
    """
    df = _rollup_detalle(bal, nivel_detalle)
    if solo_resultado:
        df = df[df["clase"].isin(CLASES_RESULTADO)]
    if cc is not None:
        df = df[df["cc"] == cc]

    df = df.sort_values(["cuenta", "cc"]).reset_index(drop=True)
    etiqueta = bal.etiqueta_cc() if usar_nombre_cc else {c: c for c in df["cc"].unique()}
    df["Centro de Costo"] = df["cc"].map(lambda c: etiqueta.get(c, c))

    cols = ["cuenta", "nombre", "Centro de Costo",
            "saldo_ant", "debitos", "creditos", "nuevo_saldo"]
    return df[cols]


def balance_matriz_por_cc(bal: BalanceCC, valor: str = "nuevo_saldo",
                          solo_resultado: bool = False,
                          nivel_detalle: int = 6,
                          usar_nombre_cc: bool = True) -> pd.DataFrame:
    """Matriz cuenta (filas) x centro de costo (columnas).

    `valor` ∈ {"nuevo_saldo", "saldo_ant", "debitos", "creditos", "mov_periodo"}.
    """
    df = _rollup_detalle(bal, nivel_detalle)
    if solo_resultado:
        df = df[df["clase"].isin(CLASES_RESULTADO)]
    if usar_nombre_cc:
        etiqueta = bal.etiqueta_cc()
        df["cc_lbl"] = df["cc"].map(lambda c: etiqueta.get(c, c))
    else:
        df["cc_lbl"] = df["cc"]
    piv = pd.pivot_table(
        df, index=["cuenta", "nombre"], columns="cc_lbl",
        values=valor, aggfunc="sum", fill_value=0.0,
    )
    piv[ETIQUETA_TOTAL] = piv.sum(axis=1)
    piv = piv.reset_index()
    return piv


# ------------------------------------------------------------------
# Reporte 2: Estado de resultados por centro de costos
# ------------------------------------------------------------------

# Estructura del Estado de Resultados (14 renglones + EBITDA).
#   tipo:
#     "ingreso"  -> línea que SUMA a la utilidad (se muestra en positivo)
#     "egreso"   -> línea que RESTA a la utilidad (se muestra en positivo)
#     "subtotal" -> combinación lineal de renglones anteriores (+/-)
#     "memo"     -> línea informativa (no entra en la cadena de utilidad)
#     "ebitda"   -> subtotal informativo (utilidad op + dep. y amort.)
#
# Los predicados son EXHAUSTIVOS y SIN SOLAPAMIENTO sobre clases 4 a 7:
#   - clase 4 = 41 + 42 + 43
#   - clase 5 = 51 + 52 + 53 + 54 + resto (otros gastos)
#   - clase 6 + 7 = costos  (clase 6 incluida por defecto; ver clase6_en_costos)
# así la suma reconstruye las clases y la utilidad neta cuadra al peso.
#
# `clase6_en_costos`: si True, la clase 6 (costo de mercancía) entra en
# "Costos"; si False, entra en "Otros gastos".
def _renglones_pyg(clase6_en_costos: bool = True):
    pref = lambda *ps: (lambda c: any(c.startswith(x) for x in ps))
    resto = lambda clase, *tomados: (
        lambda c: c.startswith(clase) and not any(c.startswith(t) for t in tomados))
    # Depreciaciones y amortizaciones: cuenta de 4 dígitos termina en 60 ó 65
    es_dep_amort = lambda c: len(c) >= 4 and c[2:4] in ("60", "65")

    if clase6_en_costos:
        pred_costos = pref("6", "7")
        pred_otros_gastos = resto("5", "51", "52", "53", "54")
    else:
        pred_costos = pref("7")
        # la clase 6 cae en otros gastos junto al resto de la clase 5
        pred_otros_gastos = lambda c: (c.startswith("6")
                                       or resto("5", "51", "52", "53", "54")(c))

    return [
        ("ingreso",  "Ingresos",                     pref("41")),
        ("egreso",   "Costos",                        pred_costos),
        ("subtotal", "UTILIDAD BRUTA",                {"+": ["Ingresos"],
                                                        "-": ["Costos"]}),
        ("ingreso",  "Otros ingresos",                pref("42")),
        ("egreso",   "Gastos de ventas",              pref("52")),
        ("egreso",   "Gastos de administración",      pref("51")),
        ("egreso",   "Otros gastos",                  pred_otros_gastos),
        ("subtotal", "UTILIDAD OPERACIONAL",          {"+": ["UTILIDAD BRUTA", "Otros ingresos"],
                                                        "-": ["Gastos de ventas",
                                                              "Gastos de administración",
                                                              "Otros gastos"]}),
        ("ingreso",  "Ingresos financieros",          pref("43")),
        ("egreso",   "Gastos financieros",            pref("53")),
        ("subtotal", "UTILIDAD ANTES DE IMPUESTOS",   {"+": ["UTILIDAD OPERACIONAL",
                                                              "Ingresos financieros"],
                                                        "-": ["Gastos financieros"]}),
        ("egreso",   "Impuesto de renta",             pref("54")),
        ("subtotal", "UTILIDAD NETA",                 {"+": ["UTILIDAD ANTES DE IMPUESTOS"],
                                                        "-": ["Impuesto de renta"]}),
        ("memo",     "(+) Depreciaciones y amortizaciones", es_dep_amort),
        ("ebitda",   "EBITDA",                        {"+": ["UTILIDAD OPERACIONAL",
                                                              "(+) Depreciaciones y amortizaciones"]}),
    ]


def estado_resultados_por_cc(bal: BalanceCC, base: str = "acumulado",
                             clase6_en_costos: bool = True,
                             usar_nombre_cc: bool = True) -> pd.DataFrame:
    """Estado de resultados con una columna por centro de costo + TOTAL.

    `base` ∈ {"acumulado", "periodo"}:
        - "acumulado": usa Nuevo Saldo (resultado del año hasta la fecha).
        - "periodo":   usa el movimiento del periodo (Débitos - Créditos).
    `clase6_en_costos`: incluir la clase 6 dentro de "Costos" (True) o en
        "Otros gastos" (False).
    `usar_nombre_cc`: encabezados de columna con el nombre del CC en vez del código.

    Devuelve un DataFrame con columna 'Concepto', una columna por cada CC,
    'TOTAL' y '_tipo'. Ingresos y egresos se presentan en positivo; los
    subtotales/EBITDA ya traen el signo correcto.
    """
    if base not in ("acumulado", "periodo"):
        raise ValueError("base debe ser 'acumulado' o 'periodo'")
    valor_col = "nuevo_saldo" if base == "acumulado" else "mov_periodo"

    df = bal.detalle[bal.detalle["clase"].isin(CLASES_RESULTADO)].copy()

    centros = sorted([c for c in df["cc"].unique() if c != SIN_CC])
    if SIN_CC in df["cc"].unique():
        centros = centros + [SIN_CC]

    def suma_por_cc(mask, signo):
        """signo: +1 ingreso (se invierte saldo), -1 egreso (saldo tal cual)."""
        por_cc = {cc: 0.0 for cc in centros}
        sub = df[mask]
        if not sub.empty:
            # display positivo: ingreso = -saldo ; egreso = +saldo
            factor = -1.0 if signo > 0 else 1.0
            agg = (sub[valor_col] * factor).groupby(sub["cc"]).sum()
            for cc, v in agg.items():
                por_cc[cc] = float(v)
        return por_cc

    renglones = _renglones_pyg(clase6_en_costos=clase6_en_costos)
    valores = {}   # etiqueta -> {cc: valor_display}
    filas = []

    for tipo, etiqueta, regla in renglones:
        if tipo in ("ingreso", "egreso"):
            por_cc = suma_por_cc(df["cuenta"].map(regla), +1 if tipo == "ingreso" else -1)
        elif tipo == "memo":
            # depreciaciones/amortizaciones: magnitud positiva (son egresos)
            por_cc = suma_por_cc(df["cuenta"].map(regla), -1)
        else:  # subtotal / ebitda
            por_cc = {cc: 0.0 for cc in centros}
            for k in regla.get("+", []):
                for cc in centros:
                    por_cc[cc] += valores[k][cc]
            for k in regla.get("-", []):
                for cc in centros:
                    por_cc[cc] -= valores[k][cc]

        valores[etiqueta] = por_cc
        fila = {"Concepto": etiqueta}
        fila.update(por_cc)
        fila[ETIQUETA_TOTAL] = sum(por_cc.values())
        fila["_tipo"] = tipo
        filas.append(fila)

    out = pd.DataFrame(filas)
    cols = ["Concepto"] + centros + [ETIQUETA_TOTAL, "_tipo"]
    out = out[cols]

    if usar_nombre_cc:
        etiqueta = bal.etiqueta_cc()
        out = out.rename(columns={c: etiqueta.get(c, c) for c in centros})
    return out


def resumen_por_cc(bal: BalanceCC, base: str = "acumulado",
                   clase6_en_costos: bool = True,
                   usar_nombre_cc: bool = True) -> pd.DataFrame:
    """Tabla compacta: una fila por CC con los principales subtotales.

    Se deriva del estado de resultados para que los números coincidan
    exactamente con el reporte detallado.
    """
    # internamente usamos códigos para indexar
    er = estado_resultados_por_cc(bal, base=base, clase6_en_costos=clase6_en_costos,
                                  usar_nombre_cc=False)
    centros = [c for c in er.columns if c not in ("Concepto", ETIQUETA_TOTAL, "_tipo")]
    etiqueta = bal.etiqueta_cc() if usar_nombre_cc else {c: c for c in centros}

    def fila(concepto):
        r = er[er["Concepto"] == concepto].iloc[0]
        return {cc: float(r[cc]) for cc in centros}

    ingresos = fila("Ingresos")
    ub = fila("UTILIDAD BRUTA")
    uop = fila("UTILIDAD OPERACIONAL")
    un = fila("UTILIDAD NETA")
    ebitda = fila("EBITDA")

    filas = []
    for cc in centros:
        filas.append({
            "Centro de Costo": etiqueta.get(cc, cc),
            "Ingresos": ingresos[cc],
            "Utilidad bruta": ub[cc],
            "Utilidad operacional": uop[cc],
            "Utilidad neta": un[cc],
            "EBITDA": ebitda[cc],
            "Margen neto %": (un[cc] / ingresos[cc] * 100.0) if ingresos[cc] else 0.0,
            "_cc": cc,
        })
    res = pd.DataFrame(filas).sort_values("_cc").reset_index(drop=True)
    res = res.drop(columns="_cc")

    total = {
        "Centro de Costo": ETIQUETA_TOTAL,
        "Ingresos": res["Ingresos"].sum(),
        "Utilidad bruta": res["Utilidad bruta"].sum(),
        "Utilidad operacional": res["Utilidad operacional"].sum(),
        "Utilidad neta": res["Utilidad neta"].sum(),
        "EBITDA": res["EBITDA"].sum(),
        "Margen neto %": (res["Utilidad neta"].sum() / res["Ingresos"].sum() * 100.0)
                          if res["Ingresos"].sum() else 0.0,
    }
    res = pd.concat([res, pd.DataFrame([total])], ignore_index=True)
    return res
