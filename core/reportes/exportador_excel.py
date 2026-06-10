"""
Exportador a Excel de los reportes por centro de costos.

Genera un workbook con:
    - Hoja "Estado de Resultados": P&G con una columna por CC + TOTAL.
    - Hoja "Resumen por CC": una fila por CC (ingresos/costos/gastos/utilidad).
    - Hoja "Balance por CC": detalle de cuentas hoja con sus saldos.

Devuelve bytes (BytesIO) listos para st.download_button.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .reporte_centros_costos import (
    BalanceCC, estado_resultados_por_cc, resumen_por_cc, balance_por_cc,
    ETIQUETA_TOTAL,
)

FUENTE = "Calibri"
FMT_NUM = '#,##0;(#,##0);"-"'

AZUL = "1F4E78"
AZUL_CLARO = "D9E1F2"
GRIS = "808080"


def _estilar_encabezado(ws, fila, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=fila, column=col)
        c.font = Font(name=FUENTE, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _hoja_estado_resultados(wb, bal: BalanceCC, base: str, clase6_en_costos: bool):
    er = estado_resultados_por_cc(bal, base=base, clase6_en_costos=clase6_en_costos)
    ws = wb.active
    ws.title = "Estado de Resultados"

    centros = [c for c in er.columns if c not in ("Concepto", ETIQUETA_TOTAL, "_tipo")]
    columnas = ["Concepto"] + centros + [ETIQUETA_TOTAL]

    # Título
    base_txt = "Acumulado del año" if base == "acumulado" else "Movimiento del periodo"
    ws.cell(row=1, column=1, value=f"{bal.empresa}  ({bal.nit})").font = Font(
        name=FUENTE, bold=True, size=12)
    ws.cell(row=2, column=1,
            value=f"Estado de Resultados por Centro de Costos · {bal.periodo} · {base_txt}"
            ).font = Font(name=FUENTE, italic=True, size=10, color=GRIS)

    fila_enc = 4
    for j, col in enumerate(columnas, start=1):
        ws.cell(row=fila_enc, column=j, value=col)
    _estilar_encabezado(ws, fila_enc, len(columnas))

    fila = fila_enc + 1
    for _, r in er.iterrows():
        tipo = r["_tipo"]
        es_fuerte = tipo in ("subtotal", "ebitda")
        es_memo = tipo == "memo"
        ws.cell(row=fila, column=1, value=r["Concepto"])
        for j, col in enumerate(columnas[1:], start=2):
            cell = ws.cell(row=fila, column=j, value=float(r[col]))
            cell.number_format = FMT_NUM
            cell.alignment = Alignment(horizontal="right")
        for j in range(1, len(columnas) + 1):
            c = ws.cell(row=fila, column=j)
            c.font = Font(name=FUENTE, bold=es_fuerte, italic=es_memo,
                          size=10, color=(GRIS if es_memo else "000000"))
            if es_fuerte:
                c.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        fila += 1

    ws.freeze_panes = ws.cell(row=fila_enc + 1, column=2)
    ws.column_dimensions["A"].width = 36
    for j in range(2, len(columnas) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 18
    return ws


def _hoja_resumen(wb, bal: BalanceCC, base: str, clase6_en_costos: bool):
    res = resumen_por_cc(bal, base=base, clase6_en_costos=clase6_en_costos)
    ws = wb.create_sheet("Resumen por CC")
    cols = list(res.columns)
    for j, col in enumerate(cols, start=1):
        ws.cell(row=1, column=j, value=col)
    _estilar_encabezado(ws, 1, len(cols))
    for i, (_, r) in enumerate(res.iterrows(), start=2):
        es_total = r["Centro de Costo"] == ETIQUETA_TOTAL
        for j, col in enumerate(cols, start=1):
            v = r[col]
            cell = ws.cell(row=i, column=j, value=v)
            if isinstance(v, (int, float)):
                cell.number_format = ('0.0"%"' if col == "Margen neto %" else FMT_NUM)
                cell.alignment = Alignment(horizontal="right")
            cell.font = Font(name=FUENTE, bold=es_total, size=10)
            if es_total:
                cell.fill = PatternFill("solid", fgColor=AZUL_CLARO)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 26
    for j in range(3, len(cols) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 18
    return ws


def _hoja_balance(wb, bal: BalanceCC, nivel_detalle: int = 6):
    df = balance_por_cc(bal, nivel_detalle=nivel_detalle, usar_nombre_cc=True)
    ws = wb.create_sheet("Balance por CC")
    headers = ["Cuenta", "Nombre", "Centro de Costo",
               "Saldo Anterior", "Débitos", "Créditos", "Nuevo Saldo"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    _estilar_encabezado(ws, 1, len(headers))
    for i, (_, r) in enumerate(df.iterrows(), start=2):
        ws.cell(row=i, column=1, value=r["cuenta"]).font = Font(name=FUENTE, size=9)
        ws.cell(row=i, column=2, value=r["nombre"]).font = Font(name=FUENTE, size=9)
        ws.cell(row=i, column=3, value=r["Centro de Costo"]).font = Font(name=FUENTE, size=9)
        for j, key in zip(range(4, 8),
                          ["saldo_ant", "debitos", "creditos", "nuevo_saldo"]):
            cell = ws.cell(row=i, column=j, value=float(r[key]))
            cell.number_format = FMT_NUM
            cell.alignment = Alignment(horizontal="right")
            cell.font = Font(name=FUENTE, size=9)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 28
    for j in range(4, 8):
        ws.column_dimensions[get_column_letter(j)].width = 16
    return ws


def exportar_excel(bal: BalanceCC, base: str = "acumulado",
                   clase6_en_costos: bool = True,
                   nivel_detalle: int = 6) -> bytes:
    """Genera el workbook completo y devuelve los bytes."""
    wb = Workbook()
    _hoja_estado_resultados(wb, bal, base, clase6_en_costos)
    _hoja_resumen(wb, bal, base, clase6_en_costos)
    _hoja_balance(wb, bal, nivel_detalle)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
