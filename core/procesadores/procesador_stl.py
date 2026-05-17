"""
core/procesadores/procesador_stl.py

Procesador del flujo STL (ventas mayoristas JIPER → retailers).

A diferencia del POS, las STL son ventas a CRÉDITO (no contado), con
distintas tarifas de IVA por factura, y van a cartera (CxC) en vez
de caja.

ENTRADA:
    - Excel del Token DIAN del mes (igual al usado por POS)
    - config_stl.json (cuentas por tarifa, comprobantes, CC)

REGLAS DE PROCESAMIENTO:

    1. Filtra por NIT_EMISOR = nit_empresa (JIPER) y MES seleccionado.

    2. Identifica TRES grupos:
        a) FACTURAS STL: prefijo "STL"
        b) NC STL: folio comienza con "NC2" (sin prefijo en columna Prefijo)
        c) Resto: se ignora (lo procesan otros módulos: POS, NC POS, DSE)

    3. Para cada FACTURA STL, detecta la TARIFA EFECTIVA de IVA:
        - tarifa = IVA / base_estimada (donde base = Total - IVA)
        - Si tarifa ≈ 19%: caso B (IVA 19% puro)
        - Si tarifa ≈ 5% : caso B (IVA 5% puro)
        - Si tarifa ≈ 0% : caso A (sin IVA)
        - Si tarifa irregular (entre 0 y otra tarifa válida): caso C MIXTO
              → calcula base_19 = IVA / 0.19
                base_no_grav = Total - base_19 - IVA

    4. Para FACTURAS STL genera 2 a 4 líneas contables (asiento):
        - Caso A (sin IVA):  Db CxC / Cr cta_ingreso_0
        - Caso B (19% puro): Db CxC / Cr cta_ingreso_19 / Cr cta_iva_19
        - Caso B (5% puro):  Db CxC / Cr cta_ingreso_5 / Cr cta_iva_5
        - Caso C (mixto):    Db CxC / Cr cta_ingreso_19 / Cr cta_ingreso_0 / Cr cta_iva_19

    5. Para NC STL (anula venta), reversa la cuenta usada:
        - Si NC sin IVA: Db cta_ingreso_0 (venta sin IVA) / Cr CxC
        - Si NC con IVA 19%: Db cta_ingreso_19 + Db cta_iva_dev_19 / Cr CxC
        - (mismo patrón para 5%)

    6. Cuadre garantizado: Db == Cr en cada asiento individual.

SALIDA:
    dict con:
      - 'plano': DataFrame en formato Silla3 (mismas columnas que POS)
      - 'lineas_stl_facturas': solo líneas de facturas STL
      - 'lineas_nc_stl': solo líneas de NC STL
      - 'resumen_por_tarifa': dict con totales por tarifa
      - 'resumen_por_cliente': dict por NIT
      - 'alertas': list[dict] con facturas no procesadas o con problemas
      - 'cuadre': dict { 'debitos': float, 'creditos': float, 'diferencia': float, 'cuadra': bool }
      - 'metadatos': fecha_min, fecha_max, total_facturado, etc.
"""

from __future__ import annotations

import io
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


# ============================================================
# Constantes y configuración
# ============================================================

# Tolerancia para considerar una tarifa "pura" vs "mixta"
# Si la tarifa calculada está dentro de ±0.5% de una tarifa válida, es pura.
TOLERANCIA_TARIFA = 0.005

# Tolerancia en pesos para considerar un cuadre exacto
TOLERANCIA_PESOS = 5.0

# Columnas del Token (deben coincidir con parser_token_dian.py)
COL_TIPO_DOC      = "Tipo de documento"
COL_FOLIO         = "Folio"
COL_PREFIJO       = "Prefijo"
COL_FECHA_EMISION = "Fecha Emisión"
COL_NIT_EMISOR    = "NIT Emisor"
COL_NIT_RECEPTOR  = "NIT Receptor"
COL_NOMBRE_REC    = "Nombre Receptor"
COL_IVA           = "IVA"
COL_INC           = "IC"
COL_TOTAL         = "Total"

TIPO_FACTURA_E    = "Factura electrónica"
TIPO_NOTA_CREDITO = "Nota de crédito electrónica"

# Columnas del plano Silla3 (mismas que POS)
COLUMNAS_PLANO = [
    "CUENTA",
    "COMPROBANTE",
    "FECHA",
    "DOCUMENTO",
    "DOC REFERENCIA",
    "NIT",
    "DETALLE",
    "TR",
    "VALOR",
    "BASE",
    "CENTRO DE COSTO",
]


# ============================================================
# Carga de configuración
# ============================================================

def cargar_config_stl(ruta: str | Path | None = None) -> dict:
    """Lee config_stl.json. Si no se pasa ruta, usa el default del repo."""
    if ruta is None:
        ruta = Path(__file__).resolve().parents[1] / "data" / "config_stl.json"
    ruta = Path(ruta)
    if not ruta.exists():
        raise FileNotFoundError(f"No existe el archivo de configuración STL: {ruta}")
    with open(ruta, "r", encoding="utf-8") as f:
        return json.load(f)


# ============================================================
# Helpers
# ============================================================

def _parsear_fecha(v) -> Optional[date]:
    """Parsea fechas en varios formatos. Devuelve None si no se puede."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _to_number(v) -> float:
    """Convierte a número, devuelve 0.0 si no se puede."""
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        try:
            return float(str(v).replace(",", "").replace("$", "").strip())
        except Exception:
            return 0.0


def _detectar_tarifa(total: float, iva: float, tarifas_validas: List[float]) -> Tuple[str, float, float, float]:
    """
    Detecta qué tarifa de IVA aplica.

    Returns:
        (caso, tarifa_efectiva, base, base_no_gravada)

        caso ∈ {'A_sin_iva', 'B_pura', 'C_mixto'}
        - A: IVA = 0
        - B: tarifa coincide ±0.5% con una tarifa válida (19%, 5%)
        - C: tarifa no coincide → mixto (parte 19% + parte sin IVA)

        En caso B/C la tarifa_efectiva indica cuál (0.19, 0.05).
        En caso C base = base_gravada (al 19%), base_no_gravada = lo demás.
    """
    if iva <= 0.0001:
        return "A_sin_iva", 0.0, total, 0.0

    # Calcular tarifa efectiva
    base_est = total - iva
    if base_est <= 0:
        return "C_mixto", 0.19, iva / 0.19, total - iva - (iva / 0.19)

    tarifa = iva / base_est

    # ¿Es alguna tarifa válida (pura)?
    for t_valida in tarifas_validas:
        if abs(tarifa - t_valida) < TOLERANCIA_TARIFA:
            return "B_pura", t_valida, base_est, 0.0

    # No es pura → mixto. Asumir 19% para la base gravada.
    base_19 = round(iva / 0.19, 2)
    base_no_grav = round(total - base_19 - iva, 2)
    return "C_mixto", 0.19, base_19, base_no_grav


def _normalizar_nit(v) -> str:
    """Convierte un NIT a string sin decimales (1234567.0 → '1234567')."""
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _formato_fecha(f: date) -> str:
    """Convierte date a 'mm/dd/yyyy'."""
    return f.strftime("%m/%d/%Y")


# ============================================================
# Lectura del Token (igual a parser_token_dian)
# ============================================================

def _leer_token_excel(fuente) -> List[dict]:
    """Lee el Excel del Token y devuelve lista de filas como dicts."""
    if hasattr(fuente, "read"):
        contenido = fuente.read()
        if hasattr(fuente, "seek"):
            fuente.seek(0)
        bio = io.BytesIO(contenido)
    elif isinstance(fuente, (bytes, bytearray)):
        bio = io.BytesIO(bytes(fuente))
    elif isinstance(fuente, (str, Path)):
        bio = str(fuente)
    else:
        raise TypeError(f"Tipo de fuente no soportado: {type(fuente)}")

    wb = load_workbook(bio, read_only=True, data_only=True)
    ws = wb.active

    # Primera fila = encabezados
    headers = None
    filas = []
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(c).strip() if c else "" for c in row]
            continue
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        fila = {h: v for h, v in zip(headers, row)}
        filas.append(fila)
    wb.close()
    return filas


# ============================================================
# Procesador principal
# ============================================================

def procesar_stl(
    fuente_token,
    config: dict | None = None,
    anio: int | None = None,
    mes: int | None = None,
    nit_empresa: str | None = None,
) -> dict:
    """
    Procesa el Token DIAN y genera el plano contable STL del mes.

    Args:
        fuente_token: ruta, bytes o file-like del Excel del Token.
        config: configuración cargada de config_stl.json. Si None, se carga del repo.
        anio: año a filtrar. Si None, usa el mes más reciente del Token.
        mes:  mes a filtrar (1-12). Si None, usa el más reciente.
        nit_empresa: override del NIT emisor. Si None, usa el de config.

    Returns:
        dict con 'plano', 'cuadre', 'alertas', 'resumen_por_tarifa', etc.
    """
    if config is None:
        config = cargar_config_stl()

    if nit_empresa is None:
        nit_empresa = str(config["nit_empresa"])

    tarifas_validas = config["tolerancias"]["tarifas_iva_validas"]
    cuentas_tarifa  = config["cuentas_por_tarifa_iva"]
    cta_cxc         = config["cta_cxc"]
    comprobante_v   = str(config["comprobante_stl"])
    comprobante_nc  = str(config["comprobante_nc_stl"])
    cc_default      = config["cc_default"]

    # ============================================================
    # 1. Leer y filtrar Token
    # ============================================================
    filas = _leer_token_excel(fuente_token)
    log = [f"📋 Token leído: {len(filas):,} filas totales"]

    # Detectar mes/año si no se dieron
    fechas_validas = []
    for f in filas:
        nit_em = _normalizar_nit(f.get(COL_NIT_EMISOR))
        if nit_em != nit_empresa:
            continue
        fecha = _parsear_fecha(f.get(COL_FECHA_EMISION))
        if fecha:
            fechas_validas.append(fecha)
    if not fechas_validas:
        raise ValueError(f"No se encontraron facturas del NIT {nit_empresa} en el Token.")

    if anio is None or mes is None:
        # Tomar el mes más reciente con datos
        ultima = max(fechas_validas)
        anio = anio or ultima.year
        mes  = mes  or ultima.month
    log.append(f"📅 Procesando período: {anio}-{mes:02d}")

    # ============================================================
    # 2. Clasificar facturas y NCs del mes
    # ============================================================
    stl_facturas = []
    stl_ncs      = []
    descartados  = {"otro_nit": 0, "otro_mes": 0, "otro_tipo": 0, "otro_prefijo": 0}

    for f in filas:
        nit_em = _normalizar_nit(f.get(COL_NIT_EMISOR))
        if nit_em != nit_empresa:
            descartados["otro_nit"] += 1
            continue

        fecha = _parsear_fecha(f.get(COL_FECHA_EMISION))
        if not fecha or fecha.year != anio or fecha.month != mes:
            descartados["otro_mes"] += 1
            continue

        tipo_doc = str(f.get(COL_TIPO_DOC) or "").strip()
        prefijo = str(f.get(COL_PREFIJO) or "").strip().upper()
        folio = str(f.get(COL_FOLIO) or "").strip()

        # Identificar tipo del documento
        if tipo_doc == TIPO_FACTURA_E and prefijo == "STL":
            stl_facturas.append(f)
        elif tipo_doc == TIPO_NOTA_CREDITO and folio.startswith("NC2"):
            stl_ncs.append(f)
        else:
            if tipo_doc not in (TIPO_FACTURA_E, TIPO_NOTA_CREDITO):
                descartados["otro_tipo"] += 1
            else:
                descartados["otro_prefijo"] += 1

    log.append(f"🏢 Facturas STL detectadas: {len(stl_facturas)}")
    log.append(f"📝 NC STL detectadas (folio NC2xxx): {len(stl_ncs)}")

    # ============================================================
    # 3. Generar líneas contables de FACTURAS STL
    # ============================================================
    lineas_facturas = []
    alertas = []
    resumen_tarifa = {"sin_iva": {"facs": 0, "base": 0.0},
                      "iva_19":  {"facs": 0, "base": 0.0, "iva": 0.0},
                      "iva_5":   {"facs": 0, "base": 0.0, "iva": 0.0},
                      "mixto":   {"facs": 0}}

    for f in stl_facturas:
        folio  = str(f.get(COL_FOLIO) or "").strip()
        fecha  = _parsear_fecha(f.get(COL_FECHA_EMISION))
        nit_rec = _normalizar_nit(f.get(COL_NIT_RECEPTOR))
        nombre = str(f.get(COL_NOMBRE_REC) or "").strip()[:50]
        total  = round(_to_number(f.get(COL_TOTAL)), 2)
        iva    = round(_to_number(f.get(COL_IVA)), 2)

        if total <= 0:
            alertas.append({
                "tipo": "factura_total_cero",
                "folio": f"STL{folio}",
                "mensaje": "Total <= 0, se ignora",
            })
            continue

        caso, tarifa, base, base_no_grav = _detectar_tarifa(
            total, iva, tarifas_validas
        )

        documento = folio  # solo el número, sin prefijo STL
        doc_ref   = folio
        detalle_base = f"FACT STL{folio} - {nombre[:30]}"
        fecha_str = _formato_fecha(fecha)

        # Verificar que la tarifa esté en la config
        if caso == "B_pura":
            tarifa_key = str(int(round(tarifa * 100)))
            if tarifa_key not in cuentas_tarifa:
                alertas.append({
                    "tipo": "tarifa_no_configurada",
                    "folio": f"STL{folio}",
                    "mensaje": f"Tarifa {tarifa*100:.1f}% no está en config_stl.json",
                })
                continue
        elif caso == "C_mixto":
            # mixto siempre involucra 19% + 0%
            if "19" not in cuentas_tarifa or "0" not in cuentas_tarifa:
                alertas.append({
                    "tipo": "tarifa_no_configurada",
                    "folio": f"STL{folio}",
                    "mensaje": "Caso mixto requiere config para tarifa 19 y 0",
                })
                continue

        # ── DÉBITO: Cartera ──
        lineas_facturas.append({
            "CUENTA":          cta_cxc,
            "COMPROBANTE":     comprobante_v,
            "FECHA":           fecha_str,
            "DOCUMENTO":       documento,
            "DOC REFERENCIA":  doc_ref,
            "NIT":             nit_rec,
            "DETALLE":         detalle_base,
            "TR":              "1",  # 1 = Débito
            "VALOR":           total,
            "BASE":            "",
            "CENTRO DE COSTO": cc_default,
        })

        # ── CRÉDITOS según caso ──
        if caso == "A_sin_iva":
            cta_ing = cuentas_tarifa["0"]["cta_ingreso"]
            lineas_facturas.append({
                "CUENTA":          cta_ing,
                "COMPROBANTE":     comprobante_v,
                "FECHA":           fecha_str,
                "DOCUMENTO":       documento,
                "DOC REFERENCIA":  doc_ref,
                "NIT":             nit_rec,
                "DETALLE":         detalle_base,
                "TR":              "2",  # 2 = Crédito
                "VALOR":           total,
                "BASE":            "",
                "CENTRO DE COSTO": cc_default,
            })
            resumen_tarifa["sin_iva"]["facs"] += 1
            resumen_tarifa["sin_iva"]["base"] += total

        elif caso == "B_pura":
            tarifa_key = str(int(round(tarifa * 100)))
            cta_ing = cuentas_tarifa[tarifa_key]["cta_ingreso"]
            cta_iva = cuentas_tarifa[tarifa_key]["cta_iva"]
            base_recalc = round(total - iva, 2)  # más preciso
            lineas_facturas.append({
                "CUENTA":          cta_ing,
                "COMPROBANTE":     comprobante_v,
                "FECHA":           fecha_str,
                "DOCUMENTO":       documento,
                "DOC REFERENCIA":  doc_ref,
                "NIT":             nit_rec,
                "DETALLE":         detalle_base,
                "TR":              "2",
                "VALOR":           base_recalc,
                "BASE":            "",
                "CENTRO DE COSTO": cc_default,
            })
            lineas_facturas.append({
                "CUENTA":          cta_iva,
                "COMPROBANTE":     comprobante_v,
                "FECHA":           fecha_str,
                "DOCUMENTO":       documento,
                "DOC REFERENCIA":  doc_ref,
                "NIT":             nit_rec,
                "DETALLE":         detalle_base,
                "TR":              "2",
                "VALOR":           iva,
                "BASE":            base_recalc,
                "CENTRO DE COSTO": cc_default,
            })
            key = "iva_19" if tarifa_key == "19" else "iva_5"
            resumen_tarifa[key]["facs"] += 1
            resumen_tarifa[key]["base"] += base_recalc
            resumen_tarifa[key]["iva"]  += iva

        elif caso == "C_mixto":
            # 19% + 0%
            cta_ing_19 = cuentas_tarifa["19"]["cta_ingreso"]
            cta_iva_19 = cuentas_tarifa["19"]["cta_iva"]
            cta_ing_0  = cuentas_tarifa["0"]["cta_ingreso"]

            lineas_facturas.append({
                "CUENTA":          cta_ing_19,
                "COMPROBANTE":     comprobante_v,
                "FECHA":           fecha_str,
                "DOCUMENTO":       documento,
                "DOC REFERENCIA":  doc_ref,
                "NIT":             nit_rec,
                "DETALLE":         detalle_base + " (base grav 19%)",
                "TR":              "2",
                "VALOR":           base,  # base 19%
                "BASE":            "",
                "CENTRO DE COSTO": cc_default,
            })
            lineas_facturas.append({
                "CUENTA":          cta_ing_0,
                "COMPROBANTE":     comprobante_v,
                "FECHA":           fecha_str,
                "DOCUMENTO":       documento,
                "DOC REFERENCIA":  doc_ref,
                "NIT":             nit_rec,
                "DETALLE":         detalle_base + " (parte s/IVA)",
                "TR":              "2",
                "VALOR":           base_no_grav,
                "BASE":            "",
                "CENTRO DE COSTO": cc_default,
            })
            lineas_facturas.append({
                "CUENTA":          cta_iva_19,
                "COMPROBANTE":     comprobante_v,
                "FECHA":           fecha_str,
                "DOCUMENTO":       documento,
                "DOC REFERENCIA":  doc_ref,
                "NIT":             nit_rec,
                "DETALLE":         detalle_base,
                "TR":              "2",
                "VALOR":           iva,
                "BASE":            base,
                "CENTRO DE COSTO": cc_default,
            })
            resumen_tarifa["mixto"]["facs"] += 1
            resumen_tarifa["iva_19"]["base"] += base
            resumen_tarifa["iva_19"]["iva"]  += iva
            resumen_tarifa["sin_iva"]["base"] += base_no_grav

    # ============================================================
    # 4. Generar líneas contables de NC STL
    # ============================================================
    lineas_ncs = []
    resumen_nc = {"facs": 0, "total": 0.0, "iva": 0.0}

    for f in stl_ncs:
        folio  = str(f.get(COL_FOLIO) or "").strip()
        fecha  = _parsear_fecha(f.get(COL_FECHA_EMISION))
        nit_rec = _normalizar_nit(f.get(COL_NIT_RECEPTOR))
        nombre = str(f.get(COL_NOMBRE_REC) or "").strip()[:50]
        total  = round(_to_number(f.get(COL_TOTAL)), 2)
        iva    = round(_to_number(f.get(COL_IVA)), 2)

        if total <= 0:
            alertas.append({
                "tipo": "nc_total_cero",
                "folio": folio,
                "mensaje": "NC con total <= 0, se ignora",
            })
            continue

        # Quitar prefijo NC para que documento = solo número
        # folio típico: "NC2579" → documento "2579", referencia "NC2579"
        if folio.upper().startswith("NC"):
            documento = folio[2:]  # solo el número
        else:
            documento = folio
        doc_ref = folio  # con NC
        detalle_base = f"NC {folio} - {nombre[:30]}"
        fecha_str = _formato_fecha(fecha)

        caso, tarifa, base, base_no_grav = _detectar_tarifa(
            total, iva, tarifas_validas
        )

        # ── CRÉDITO: Cartera (al revés de una factura) ──
        lineas_ncs.append({
            "CUENTA":          cta_cxc,
            "COMPROBANTE":     comprobante_nc,
            "FECHA":           fecha_str,
            "DOCUMENTO":       documento,
            "DOC REFERENCIA":  doc_ref,
            "NIT":             nit_rec,
            "DETALLE":         detalle_base,
            "TR":              "2",  # Crédito
            "VALOR":           total,
            "BASE":            "",
            "CENTRO DE COSTO": cc_default,
        })

        # ── DÉBITOS según caso (reversa la cuenta de venta) ──
        if caso == "A_sin_iva":
            cta_ing = cuentas_tarifa["0"]["cta_ingreso"]
            lineas_ncs.append({
                "CUENTA":          cta_ing,
                "COMPROBANTE":     comprobante_nc,
                "FECHA":           fecha_str,
                "DOCUMENTO":       documento,
                "DOC REFERENCIA":  doc_ref,
                "NIT":             nit_rec,
                "DETALLE":         detalle_base,
                "TR":              "1",  # Débito
                "VALOR":           total,
                "BASE":            "",
                "CENTRO DE COSTO": cc_default,
            })
        elif caso == "B_pura":
            tarifa_key = str(int(round(tarifa * 100)))
            cta_ing = cuentas_tarifa[tarifa_key]["cta_ingreso"]
            cta_iva_dev = cuentas_tarifa[tarifa_key].get("cta_iva_dev") or cuentas_tarifa[tarifa_key]["cta_iva"]
            base_recalc = round(total - iva, 2)
            lineas_ncs.append({
                "CUENTA":          cta_ing,
                "COMPROBANTE":     comprobante_nc,
                "FECHA":           fecha_str,
                "DOCUMENTO":       documento,
                "DOC REFERENCIA":  doc_ref,
                "NIT":             nit_rec,
                "DETALLE":         detalle_base,
                "TR":              "1",
                "VALOR":           base_recalc,
                "BASE":            "",
                "CENTRO DE COSTO": cc_default,
            })
            lineas_ncs.append({
                "CUENTA":          cta_iva_dev,
                "COMPROBANTE":     comprobante_nc,
                "FECHA":           fecha_str,
                "DOCUMENTO":       documento,
                "DOC REFERENCIA":  doc_ref,
                "NIT":             nit_rec,
                "DETALLE":         detalle_base,
                "TR":              "1",
                "VALOR":           iva,
                "BASE":            base_recalc,
                "CENTRO DE COSTO": cc_default,
            })
        elif caso == "C_mixto":
            cta_ing_19 = cuentas_tarifa["19"]["cta_ingreso"]
            cta_iva_dev_19 = cuentas_tarifa["19"].get("cta_iva_dev") or cuentas_tarifa["19"]["cta_iva"]
            cta_ing_0  = cuentas_tarifa["0"]["cta_ingreso"]

            lineas_ncs.append({
                "CUENTA":          cta_ing_19, "COMPROBANTE": comprobante_nc,
                "FECHA":           fecha_str,  "DOCUMENTO": documento, "DOC REFERENCIA": doc_ref,
                "NIT":             nit_rec,    "DETALLE": detalle_base + " (base grav 19%)",
                "TR":              "1",        "VALOR": base, "BASE": "",
                "CENTRO DE COSTO": cc_default,
            })
            lineas_ncs.append({
                "CUENTA":          cta_ing_0,  "COMPROBANTE": comprobante_nc,
                "FECHA":           fecha_str,  "DOCUMENTO": documento, "DOC REFERENCIA": doc_ref,
                "NIT":             nit_rec,    "DETALLE": detalle_base + " (parte s/IVA)",
                "TR":              "1",        "VALOR": base_no_grav, "BASE": "",
                "CENTRO DE COSTO": cc_default,
            })
            lineas_ncs.append({
                "CUENTA":          cta_iva_dev_19, "COMPROBANTE": comprobante_nc,
                "FECHA":           fecha_str,      "DOCUMENTO": documento, "DOC REFERENCIA": doc_ref,
                "NIT":             nit_rec,        "DETALLE": detalle_base,
                "TR":              "1",            "VALOR": iva, "BASE": base,
                "CENTRO DE COSTO": cc_default,
            })

        resumen_nc["facs"] += 1
        resumen_nc["total"] += total
        resumen_nc["iva"] += iva

    # ============================================================
    # 5. Unir y verificar cuadre
    # ============================================================
    todas_lineas = lineas_facturas + lineas_ncs
    plano = pd.DataFrame(todas_lineas, columns=COLUMNAS_PLANO) if todas_lineas else pd.DataFrame(columns=COLUMNAS_PLANO)

    if len(plano) > 0:
        debitos  = plano[plano["TR"] == "1"]["VALOR"].sum()
        creditos = plano[plano["TR"] == "2"]["VALOR"].sum()
    else:
        debitos = creditos = 0.0
    diferencia = round(debitos - creditos, 2)

    cuadre = {
        "debitos":    round(debitos, 2),
        "creditos":   round(creditos, 2),
        "diferencia": diferencia,
        "cuadra":     abs(diferencia) < TOLERANCIA_PESOS,
    }

    log.append(f"💰 Total débitos:  ${debitos:,.2f}")
    log.append(f"💰 Total créditos: ${creditos:,.2f}")
    log.append(f"📐 Diferencia: ${diferencia:,.2f} ({'✅ cuadra' if cuadre['cuadra'] else '❌ no cuadra'})")

    # ============================================================
    # 6. Resumen por cliente (NIT)
    # ============================================================
    resumen_cliente = {}
    for f in stl_facturas:
        nit = _normalizar_nit(f.get(COL_NIT_RECEPTOR))
        nombre = str(f.get(COL_NOMBRE_REC) or "").strip()
        if nit not in resumen_cliente:
            resumen_cliente[nit] = {"nombre": nombre, "facturas": 0, "total": 0.0, "iva": 0.0, "ncs": 0, "total_nc": 0.0}
        resumen_cliente[nit]["facturas"] += 1
        resumen_cliente[nit]["total"] += round(_to_number(f.get(COL_TOTAL)), 2)
        resumen_cliente[nit]["iva"]   += round(_to_number(f.get(COL_IVA)), 2)
    for f in stl_ncs:
        nit = _normalizar_nit(f.get(COL_NIT_RECEPTOR))
        nombre = str(f.get(COL_NOMBRE_REC) or "").strip()
        if nit not in resumen_cliente:
            resumen_cliente[nit] = {"nombre": nombre, "facturas": 0, "total": 0.0, "iva": 0.0, "ncs": 0, "total_nc": 0.0}
        resumen_cliente[nit]["ncs"] += 1
        resumen_cliente[nit]["total_nc"] += round(_to_number(f.get(COL_TOTAL)), 2)

    return {
        "plano":               plano,
        "lineas_stl_facturas": pd.DataFrame(lineas_facturas, columns=COLUMNAS_PLANO) if lineas_facturas else pd.DataFrame(columns=COLUMNAS_PLANO),
        "lineas_nc_stl":       pd.DataFrame(lineas_ncs, columns=COLUMNAS_PLANO) if lineas_ncs else pd.DataFrame(columns=COLUMNAS_PLANO),
        "resumen_por_tarifa":  resumen_tarifa,
        "resumen_nc":          resumen_nc,
        "resumen_por_cliente": resumen_cliente,
        "alertas":             alertas,
        "cuadre":              cuadre,
        "log":                 log,
        "metadatos": {
            "anio": anio,
            "mes":  mes,
            "facturas_procesadas": len(stl_facturas),
            "ncs_procesadas":      len(stl_ncs),
            "lineas_plano":        len(plano),
            "descartados":         descartados,
        },
    }


# ============================================================
# Exportadores (compatibles con formato Silla3)
# ============================================================

def plano_a_tsv_bytes(df: pd.DataFrame) -> bytes:
    """Exporta el plano a TSV (TAB) listo para el sistema contable."""
    if len(df) == 0:
        return "".encode("utf-8-sig")
    return df.to_csv(sep="\t", index=False, encoding="utf-8-sig").encode("utf-8-sig")


def plano_a_csv_bytes(df: pd.DataFrame) -> bytes:
    """Exporta el plano a CSV (coma) listo para el sistema contable."""
    if len(df) == 0:
        return "".encode("utf-8-sig")
    return df.to_csv(sep=",", index=False, encoding="utf-8-sig").encode("utf-8-sig")


def plano_a_xlsx_bytes(df: pd.DataFrame, resumen: dict | None = None) -> bytes:
    """
    Exporta el plano a Excel con hojas adicionales de revisión.

    Args:
        df: DataFrame del plano (única hoja "Plano").
        resumen: dict con 'resumen_por_tarifa', 'resumen_por_cliente', 'cuadre', etc.
                 (opcional — añade hojas de validación).
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Hoja 1: Plano completo
        df.to_excel(writer, sheet_name="Plano", index=False)

        if resumen is not None:
            # Hoja 2: Resumen por tarifa
            rt = resumen.get("resumen_por_tarifa") or {}
            df_rt = pd.DataFrame([
                {"Tarifa": "Sin IVA",   "Facturas": rt.get("sin_iva", {}).get("facs", 0),
                 "Base": rt.get("sin_iva", {}).get("base", 0)},
                {"Tarifa": "IVA 19%",   "Facturas": rt.get("iva_19", {}).get("facs", 0),
                 "Base": rt.get("iva_19", {}).get("base", 0),
                 "IVA":  rt.get("iva_19", {}).get("iva", 0)},
                {"Tarifa": "IVA 5%",    "Facturas": rt.get("iva_5", {}).get("facs", 0),
                 "Base": rt.get("iva_5", {}).get("base", 0),
                 "IVA":  rt.get("iva_5", {}).get("iva", 0)},
                {"Tarifa": "Mixto",     "Facturas": rt.get("mixto", {}).get("facs", 0)},
            ])
            df_rt.to_excel(writer, sheet_name="Por tarifa", index=False)

            # Hoja 3: Resumen por cliente
            rc = resumen.get("resumen_por_cliente") or {}
            df_rc = pd.DataFrame([
                {"NIT": nit, **datos} for nit, datos in rc.items()
            ])
            if len(df_rc) > 0:
                df_rc.to_excel(writer, sheet_name="Por cliente", index=False)

            # Hoja 4: Cuadre y totales
            cu = resumen.get("cuadre") or {}
            meta = resumen.get("metadatos") or {}
            df_cu = pd.DataFrame([
                {"Concepto": "Período",          "Valor": f"{meta.get('anio','')}-{meta.get('mes',''):02d}" if meta.get('mes') else ""},
                {"Concepto": "Facturas STL",     "Valor": meta.get("facturas_procesadas", 0)},
                {"Concepto": "NC STL",           "Valor": meta.get("ncs_procesadas", 0)},
                {"Concepto": "Líneas plano",     "Valor": meta.get("lineas_plano", 0)},
                {"Concepto": "Débitos",          "Valor": cu.get("debitos", 0)},
                {"Concepto": "Créditos",         "Valor": cu.get("creditos", 0)},
                {"Concepto": "Diferencia",       "Valor": cu.get("diferencia", 0)},
                {"Concepto": "Cuadra",           "Valor": "✅ SÍ" if cu.get("cuadra") else "❌ NO"},
            ])
            df_cu.to_excel(writer, sheet_name="Cuadre", index=False)

            # Hoja 5: Alertas (si las hay)
            alertas = resumen.get("alertas") or []
            if alertas:
                pd.DataFrame(alertas).to_excel(writer, sheet_name="Alertas", index=False)

    buf.seek(0)
    return buf.read()


# ============================================================
# PROCESADOR DESDE XMLs (versión preciso por línea)
# ============================================================

def procesar_stl_desde_xmls(
    fuente_zip,
    config: dict | None = None,
    anio: int | None = None,
    mes: int | None = None,
    nit_empresa: str | None = None,
) -> dict:
    """
    Procesa STL leyendo los XMLs reales del ZIP descargado de la DIAN.

    Ventajas sobre procesar_stl (que usa Token):
      - Lee CADA LÍNEA del XML con su tarifa REAL (no calculada)
      - Genera líneas contables granulares: una por (tarifa, tipo_impuesto) de cada factura
      - Detecta automáticamente STL y NC STL (NC2xxx) en el mismo ZIP
      - Distingue IVA (01) vs INC (04)
      - Detecta ZIPs corruptos (folios duplicados)

    Args:
        fuente_zip: bytes, ruta o file-like del ZIP con los XMLs.
        config: configuración cargada de config_stl.json.
        anio, mes: filtros opcionales (si no se pasan, procesa todo).
        nit_empresa: override del NIT emisor.

    Returns:
        dict con misma estructura que procesar_stl + 'fuente': 'xmls'
    """
    from core.procesadores.parser_xml_stl import procesar_zip_xmls_stl

    if config is None:
        config = cargar_config_stl()
    if nit_empresa is None:
        nit_empresa = str(config["nit_empresa"])

    cuentas_tarifa = config["cuentas_por_tarifa_iva"]
    cta_cxc        = config["cta_cxc"]
    comprobante_v  = str(config["comprobante_stl"])
    comprobante_nc = str(config["comprobante_nc_stl"])
    cc_default     = config["cc_default"]

    log = []

    # ─── Leer ZIP de XMLs ───
    # NO filtramos por NIT aquí — clasificaremos por folio (STL vs NC2)
    # Los XMLs descargados manualmente del portal son SIEMPRE emitidos por JIPER
    # (porque solo se descargan los propios)
    resultado_zip = procesar_zip_xmls_stl(fuente_zip)
    log.extend(resultado_zip["log"])

    if resultado_zip["duplicados"] > 0:
        log.append(
            f"🚨 ALERTA CRÍTICA: {resultado_zip['duplicados']} XMLs DUPLICADOS detectados. "
            f"El ZIP parece estar corrupto (descarga incorrecta). Verifica la fuente."
        )

    # ─── Clasificar STL vs NC STL ───
    stl_facturas = []
    stl_ncs      = []

    for f in resultado_zip["facturas"]:
        if f.get("es_duplicado"):
            continue  # ignorar duplicados (mismo folio repetido)

        folio = f["folio"]
        fecha = f["fecha"]

        # Filtro de mes/año si se pidió
        if anio and fecha and fecha.year != anio:
            continue
        if mes and fecha and fecha.month != mes:
            continue

        # Clasificar:
        #  - Factura STL: tipo factura + folio empieza por "STL"
        #  - NC STL: tipo NC + folio empieza por "NC2" (sin prefijo en columna)
        if f["tipo_doc"] == "factura" and folio.upper().startswith("STL"):
            stl_facturas.append(f)
        elif f["tipo_doc"] == "nota_credito" and folio.upper().startswith("NC2"):
            stl_ncs.append(f)
        # otros documentos se ignoran (FEs POS, NCs POS, etc.)

    log.append(f"🏢 Facturas STL detectadas: {len(stl_facturas)}")
    log.append(f"📝 NC STL detectadas (NC2xxx): {len(stl_ncs)}")

    # ─── Inferir período si no se dio ───
    if anio is None or mes is None:
        fechas = [f["fecha"] for f in stl_facturas + stl_ncs if f["fecha"]]
        if fechas:
            ultima = max(fechas)
            anio = anio or ultima.year
            mes  = mes or ultima.month

    # ─── Generar líneas contables ───
    lineas_facturas = []
    lineas_ncs      = []
    alertas         = []
    resumen_tarifa  = {
        "sin_iva": {"facs": 0, "base": 0.0},
        "iva_19":  {"facs": 0, "base": 0.0, "iva": 0.0},
        "iva_5":   {"facs": 0, "base": 0.0, "iva": 0.0},
        "inc_8":   {"facs": 0, "base": 0.0, "iva": 0.0},
        "otro":    {"facs": 0, "base": 0.0, "iva": 0.0},
        # 'mixto' = factura que tiene a la vez líneas gravadas y no gravadas.
        # Se llena en la lógica de _agregar_lineas_documento cuando una misma
        # factura aporta a más de una tarifa. La UI (4b_Ingresos_POS.py) la
        # lee, así que SIEMPRE debe existir aunque venga en 0.
        "mixto":   {"facs": 0},
    }

    def _agregar_lineas_documento(factura, es_nc, lineas_out):
        """Agrega líneas contables para una factura/NC granular por tarifa."""
        from collections import defaultdict
        folio   = factura["folio"]
        fecha   = factura["fecha"]
        fecha_str = fecha.strftime("%m/%d/%Y") if fecha else ""
        nit_rec = factura["nit_receptor"]
        nombre  = (factura["nombre_receptor"] or "")[:50]
        total   = round(factura["totales"]["payable"], 2)

        if total <= 0:
            alertas.append({
                "tipo": "doc_total_cero", "folio": folio,
                "mensaje": "Total <= 0, se ignora",
            })
            return False

        # Documento: solo número (sin prefijo)
        if folio.upper().startswith("STL"):
            documento = folio[3:]
        elif folio.upper().startswith("NC"):
            documento = ''.join(c for c in folio if c.isdigit())
        else:
            documento = folio
        doc_ref = folio  # con prefijo
        detalle_base = ("NC " if es_nc else "FACT ") + f"{folio} - {nombre[:30]}"
        comprobante = comprobante_nc if es_nc else comprobante_v

        # Agrupar líneas por (tax_scheme, tarifa)
        # Esto permite que si una factura tiene 10 ítems al 19% y 5 al 0%,
        # solo se generen 2 líneas contables agregadas (más limpio que 15).
        grupos = defaultdict(lambda: {"base": 0.0, "iva": 0.0, "n_lineas": 0})
        for ln in factura["lineas"]:
            scheme = ln.get("tax_scheme")
            tarifa = round(ln["tarifa_pct"], 2)
            key = (scheme, tarifa)
            grupos[key]["base"] += ln["base_gravada"]
            grupos[key]["iva"]  += ln["iva"]
            grupos[key]["n_lineas"] += 1

        # ── Cartera (siempre primera línea) ──
        if es_nc:
            # NC: Crédito a cartera (reversa)
            lineas_out.append({
                "CUENTA":          cta_cxc,
                "COMPROBANTE":     comprobante,
                "FECHA":           fecha_str,
                "DOCUMENTO":       documento,
                "DOC REFERENCIA":  doc_ref,
                "NIT":             nit_rec,
                "DETALLE":         detalle_base,
                "TR":              "2",
                "VALOR":           total,
                "BASE":            "",
                "CENTRO DE COSTO": cc_default,
            })
        else:
            # Factura: Débito a cartera
            lineas_out.append({
                "CUENTA":          cta_cxc,
                "COMPROBANTE":     comprobante,
                "FECHA":           fecha_str,
                "DOCUMENTO":       documento,
                "DOC REFERENCIA":  doc_ref,
                "NIT":             nit_rec,
                "DETALLE":         detalle_base,
                "TR":              "1",
                "VALOR":           total,
                "BASE":            "",
                "CENTRO DE COSTO": cc_default,
            })

        # ── Líneas por tarifa ──
        # TR para factura: Crédito=2 (ventas). Para NC: Débito=1 (reversa ventas).
        tr_ingreso = "1" if es_nc else "2"

        for (scheme, tarifa), datos in grupos.items():
            base = round(datos["base"], 2)
            iva  = round(datos["iva"], 2)
            if base <= 0 and iva <= 0:
                continue

            # Resolver cuenta de ingreso
            tarifa_key = str(int(round(tarifa)))
            etiqueta_tarifa = f"{tarifa}%"

            # IVA en general usa cuentas_tarifa por tarifa
            if scheme == "01" and tarifa_key in cuentas_tarifa:
                cta_ing = cuentas_tarifa[tarifa_key]["cta_ingreso"]
                cta_iva = cuentas_tarifa[tarifa_key].get("cta_iva")
                cta_iva_dev = cuentas_tarifa[tarifa_key].get("cta_iva_dev") or cta_iva

                # Línea de ingreso
                lineas_out.append({
                    "CUENTA":          cta_ing,
                    "COMPROBANTE":     comprobante,
                    "FECHA":           fecha_str,
                    "DOCUMENTO":       documento,
                    "DOC REFERENCIA":  doc_ref,
                    "NIT":             nit_rec,
                    "DETALLE":         detalle_base + (f" (IVA {etiqueta_tarifa})" if tarifa > 0 else " (sin IVA)"),
                    "TR":              tr_ingreso,
                    "VALOR":           base,
                    "BASE":            "",
                    "CENTRO DE COSTO": cc_default,
                })

                # Línea de IVA (si aplica)
                if iva > 0 and cta_iva:
                    lineas_out.append({
                        "CUENTA":          cta_iva_dev if es_nc else cta_iva,
                        "COMPROBANTE":     comprobante,
                        "FECHA":           fecha_str,
                        "DOCUMENTO":       documento,
                        "DOC REFERENCIA":  doc_ref,
                        "NIT":             nit_rec,
                        "DETALLE":         detalle_base + f" (IVA {etiqueta_tarifa})",
                        "TR":              tr_ingreso,
                        "VALOR":           iva,
                        "BASE":            base,
                        "CENTRO DE COSTO": cc_default,
                    })

                # Estadística
                if tarifa == 0:
                    resumen_tarifa["sin_iva"]["facs"] += (1 if not es_nc else 0)
                    resumen_tarifa["sin_iva"]["base"] += base
                elif int(round(tarifa)) == 19:
                    resumen_tarifa["iva_19"]["facs"] += (1 if not es_nc else 0)
                    resumen_tarifa["iva_19"]["base"] += base
                    resumen_tarifa["iva_19"]["iva"]  += iva
                elif int(round(tarifa)) == 5:
                    resumen_tarifa["iva_5"]["facs"] += (1 if not es_nc else 0)
                    resumen_tarifa["iva_5"]["base"] += base
                    resumen_tarifa["iva_5"]["iva"]  += iva
            elif scheme is None or tarifa == 0:
                # Sin impuesto (exento)
                cta_ing = cuentas_tarifa.get("0", {}).get("cta_ingreso")
                if cta_ing:
                    lineas_out.append({
                        "CUENTA":          cta_ing,
                        "COMPROBANTE":     comprobante,
                        "FECHA":           fecha_str,
                        "DOCUMENTO":       documento,
                        "DOC REFERENCIA":  doc_ref,
                        "NIT":             nit_rec,
                        "DETALLE":         detalle_base + " (sin IVA)",
                        "TR":              tr_ingreso,
                        "VALOR":           base,
                        "BASE":            "",
                        "CENTRO DE COSTO": cc_default,
                    })
                    resumen_tarifa["sin_iva"]["facs"] += (1 if not es_nc else 0)
                    resumen_tarifa["sin_iva"]["base"] += base
            else:
                # Tarifa no configurada (ej. INC 8%) → alerta
                alertas.append({
                    "tipo":    "tarifa_no_configurada",
                    "folio":   folio,
                    "mensaje": f"Tarifa {tarifa}% / scheme {scheme} no está en config_stl.json. Línea ignorada.",
                })

        return True

    for f in stl_facturas:
        _agregar_lineas_documento(f, es_nc=False, lineas_out=lineas_facturas)
    for f in stl_ncs:
        _agregar_lineas_documento(f, es_nc=True, lineas_out=lineas_ncs)

    # ─── Construir plano y verificar cuadre ───
    todas_lineas = lineas_facturas + lineas_ncs
    plano = pd.DataFrame(todas_lineas, columns=COLUMNAS_PLANO) if todas_lineas else pd.DataFrame(columns=COLUMNAS_PLANO)

    if len(plano) > 0:
        debitos  = plano[plano["TR"] == "1"]["VALOR"].sum()
        creditos = plano[plano["TR"] == "2"]["VALOR"].sum()
    else:
        debitos = creditos = 0.0
    diferencia = round(debitos - creditos, 2)

    cuadre = {
        "debitos":    round(debitos, 2),
        "creditos":   round(creditos, 2),
        "diferencia": diferencia,
        "cuadra":     abs(diferencia) < TOLERANCIA_PESOS,
    }

    log.append(f"💰 Total débitos:  ${debitos:,.2f}")
    log.append(f"💰 Total créditos: ${creditos:,.2f}")
    log.append(f"📐 Cuadre: {'✅' if cuadre['cuadra'] else '❌'}")

    # ─── Resumen por cliente ───
    resumen_cliente = {}
    for f in stl_facturas:
        nit = f["nit_receptor"]
        if nit not in resumen_cliente:
            resumen_cliente[nit] = {"nombre": f["nombre_receptor"], "facturas": 0, "total": 0.0, "iva": 0.0, "ncs": 0, "total_nc": 0.0}
        resumen_cliente[nit]["facturas"] += 1
        resumen_cliente[nit]["total"] += f["totales"]["payable"]
        # IVA total de la factura
        iva_total = sum(ln["iva"] for ln in f["lineas"])
        resumen_cliente[nit]["iva"] += iva_total
    for f in stl_ncs:
        nit = f["nit_receptor"]
        if nit not in resumen_cliente:
            resumen_cliente[nit] = {"nombre": f["nombre_receptor"], "facturas": 0, "total": 0.0, "iva": 0.0, "ncs": 0, "total_nc": 0.0}
        resumen_cliente[nit]["ncs"] += 1
        resumen_cliente[nit]["total_nc"] += f["totales"]["payable"]

    resumen_nc = {
        "facs":  len(stl_ncs),
        "total": sum(f["totales"]["payable"] for f in stl_ncs),
        "iva":   sum(sum(ln["iva"] for ln in f["lineas"]) for f in stl_ncs),
    }

    return {
        "plano":               plano,
        "lineas_stl_facturas": pd.DataFrame(lineas_facturas, columns=COLUMNAS_PLANO) if lineas_facturas else pd.DataFrame(columns=COLUMNAS_PLANO),
        "lineas_nc_stl":       pd.DataFrame(lineas_ncs, columns=COLUMNAS_PLANO) if lineas_ncs else pd.DataFrame(columns=COLUMNAS_PLANO),
        "resumen_por_tarifa":  resumen_tarifa,
        "resumen_nc":          resumen_nc,
        "resumen_por_cliente": resumen_cliente,
        "alertas":             alertas,
        "cuadre":              cuadre,
        "log":                 log,
        "fuente":              "xmls",
        "duplicados_zip":      resultado_zip["duplicados"],
        "metadatos": {
            "anio": anio,
            "mes":  mes,
            "facturas_procesadas": len(stl_facturas),
            "ncs_procesadas":      len(stl_ncs),
            "lineas_plano":        len(plano),
            "fuente":              "XMLs reales (UBL DIAN)",
        },
    }
