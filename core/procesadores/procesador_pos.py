"""
Procesador de ingresos diarios por punto de venta (POS).

Toma un Excel multi-hoja con reportes de varios sistemas POS
y genera el plano contable de ingresos por día por sucursal.

Hojas esperadas:
    - DATOS PUNTO        : maestro de sucursales (obligatoria)
    - REPORTE CHILI      : reporte estilo Chili (Net + Final)
    - L3AF MILLA DE ORO  : reporte estilo L3AF (Total Neto + Impuestos + Total Final)
    - HENKO MILAGROS     : reporte estilo Henko multi-sucursal
    - HENKO REMEDIOS     : reporte estilo Henko de una sucursal
    - PLANO              : hoja de salida (se ignora a la entrada)

DATOS PUNTO debe tener estas columnas (obligatorias):
    COMPROBANTE | NOMBRE EN REPORTE | SEDE | CC | CUENTA DE CAJA | CTA BASE V | CTA ICO | CLASE DE SEDE

Asiento generado (3 líneas por día/sucursal):
    Db CUENTA DE CAJA    = Total Final
    Cr CTA BASE V        = Neto (subtotal sin IC)
    Cr CTA ICO           = IC (Final - Neto)

El cuadre Db = Cr es exacto por construcción.
"""
from __future__ import annotations
import io
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime, date
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


# ============================================================
# Constantes
# ============================================================

NIT_GENERICO_POS = "222222222"
DETALLE_PREFIJO = "VENTAS POS "

# Columnas finales del plano (orden exacto de salida)
COLUMNAS_PLANO = [
    "CUENTA",
    "COMPROBANTE",
    "FECHA",
    "DOCUMENTO",
    "DOC REFERENCIA",
    "NIT",
    "DETALLE",
    "TR",
    "VALOR",
    "BASE",
    "CENTRO DE COSTO",
]

# Columnas que requiere DATOS PUNTO
COLUMNAS_MAESTRO = (
    "COMPROBANTE",
    "NOMBRE EN REPORTE",
    "SEDE",
    "CC",
    "CUENTA DE CAJA",
    "CTA BASE V",
    "CTA ICO",
)

# Hojas a ignorar (no son reportes)
HOJAS_NO_REPORTE = {"DATOS PUNTO", "PLANO"}


# Mapeo: si el nombre de la hoja contiene esta palabra clave (normalizada),
# las sucursales que se procesen deben tener esa CLASE DE SEDE.
# Si una sucursal no coincide, se ignora y se reporta al usuario.
HOJA_CLASE_ESPERADA = (
    # (palabra_clave_en_nombre_hoja, clase_de_sede_esperada)
    ("CHILI",     "SANTA LEÑA"),
    ("MILAGROS",  "RESTAURANTE MILAGROS"),
    ("REMEDIOS",  "RESTAURANTE MILAGROS"),
    ("L3AF",      "RESTAURANTE MILAGROS"),
)


def _clase_esperada_para_hoja(nombre_hoja: str) -> Optional[str]:
    """Determina qué CLASE DE SEDE deberían tener las sucursales de esta hoja.

    Retorna None si el nombre de la hoja no coincide con ningún patrón
    conocido (en ese caso no se aplica validación de clase).
    """
    nombre_norm = _normalizar_texto(nombre_hoja)
    for palabra, clase in HOJA_CLASE_ESPERADA:
        if palabra in nombre_norm:
            return _normalizar_texto(clase)
    return None


# ============================================================
# Helpers de normalización
# ============================================================

def _normalizar_texto(s) -> str:
    """Normaliza texto para comparación: sin acentos, mayúsculas, sin espacios extra."""
    if s is None:
        return ""
    t = str(s).strip()
    if not t:
        return ""
    # Quitar acentos
    t = unicodedata.normalize("NFKD", t)
    t = "".join(c for c in t if not unicodedata.combining(c))
    # Mayúsculas y espacios colapsados
    t = re.sub(r"\s+", " ", t).upper().strip()
    return t


def _to_int(v) -> int:
    """Convierte cualquier valor numérico a entero. Retorna 0 si falla."""
    if v is None or v == "":
        return 0
    try:
        return int(round(float(v)))
    except (ValueError, TypeError):
        return 0


def _parsear_fecha(v) -> Optional[date]:
    """Convierte a date desde varios formatos posibles:
        - datetime / date
        - 'DD/MM/YYYY' (CHILI, HENKO)
        - 'YYYY-MM-DD' (L3AF)
        - 'D/M/YYYY' (variantes)
    """
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    # Intentar varios formatos
    formatos = [
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%m/%d/%Y",
        "%m/%d/%y",
    ]
    for fmt in formatos:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _formato_fecha_plano(d: date) -> str:
    """Formato MM/DD/AAAA (lo pidió el usuario)."""
    return d.strftime("%m/%d/%Y")


def _formato_cc(cc) -> str:
    """Centro de costo siempre como string conservando ceros a la izquierda."""
    if cc is None:
        return ""
    s = str(cc).strip()
    # Si vino como float (12345.0), quitar la parte decimal
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _formato_cuenta(c) -> str:
    """Cuenta como string sin decimales."""
    if c is None:
        return ""
    s = str(c).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


# ============================================================
# Lectura del maestro DATOS PUNTO
# ============================================================

@dataclass
class Sucursal:
    """Información de una sucursal extraída de DATOS PUNTO."""
    comprobante: str
    nombre_reporte: str       # cómo aparece en los reportes (puede tener variaciones)
    sede: str                 # nombre oficial de la sede
    cc: str                   # centro de costo (con ceros a la izq)
    cuenta_caja: str
    cta_base_v: str
    cta_ico: str
    clase: str = ""           # CLASE DE SEDE: SANTA LEÑA, RESTAURANTE MILAGROS, etc.

    @property
    def clave_busqueda(self) -> str:
        """Clave normalizada para buscar la sucursal en los reportes."""
        return _normalizar_texto(self.nombre_reporte)

    @property
    def clase_norm(self) -> str:
        """Clase normalizada para comparar."""
        return _normalizar_texto(self.clase)


def _leer_datos_punto(wb) -> List[Sucursal]:
    """Lee la hoja DATOS PUNTO y retorna lista de sucursales."""
    if "DATOS PUNTO" not in wb.sheetnames:
        raise ValueError(
            "No se encontró la hoja 'DATOS PUNTO' en el Excel. "
            "Esta hoja es obligatoria y contiene el maestro de sucursales."
        )
    ws = wb["DATOS PUNTO"]
    filas = list(ws.iter_rows(values_only=True))
    if len(filas) < 2:
        raise ValueError("La hoja 'DATOS PUNTO' está vacía o sin datos.")

    # Mapear índices de columnas
    encabezado = [str(c).strip().upper() if c else "" for c in filas[0]]
    idx = {}
    for col_req in COLUMNAS_MAESTRO:
        try:
            idx[col_req] = encabezado.index(col_req)
        except ValueError:
            raise ValueError(
                f"DATOS PUNTO: falta la columna obligatoria '{col_req}'. "
                f"Columnas encontradas: {encabezado}"
            )

    # CLASE DE SEDE es opcional pero MUY RECOMENDADA para validación cruzada
    idx_clase = None
    try:
        idx_clase = encabezado.index("CLASE DE SEDE")
    except ValueError:
        pass

    sucursales: List[Sucursal] = []
    for fila in filas[1:]:
        if not fila or not fila[idx["NOMBRE EN REPORTE"]]:
            continue
        try:
            clase = ""
            if idx_clase is not None and idx_clase < len(fila) and fila[idx_clase]:
                clase = str(fila[idx_clase]).strip()
            s = Sucursal(
                comprobante=str(fila[idx["COMPROBANTE"]]).strip(),
                nombre_reporte=str(fila[idx["NOMBRE EN REPORTE"]]).strip(),
                sede=str(fila[idx["SEDE"]]).strip() if fila[idx["SEDE"]] else "",
                cc=_formato_cc(fila[idx["CC"]]),
                cuenta_caja=_formato_cuenta(fila[idx["CUENTA DE CAJA"]]),
                cta_base_v=_formato_cuenta(fila[idx["CTA BASE V"]]),
                cta_ico=_formato_cuenta(fila[idx["CTA ICO"]]),
                clase=clase,
            )
            sucursales.append(s)
        except (ValueError, IndexError):
            continue

    if not sucursales:
        raise ValueError("La hoja 'DATOS PUNTO' no produjo ninguna sucursal válida.")

    return sucursales


def _buscar_sucursal(
    nombre_reporte: str,
    sucursales: List[Sucursal],
) -> Optional[Sucursal]:
    """Busca una sucursal por su nombre tal como aparece en el reporte.

    Estrategia:
      1. Match exacto normalizado.
      2. Si el reporte trae solo parte del nombre (ej. 'Tesoro' en HENKO REMEDIOS),
         intentar match parcial: la cadena del reporte está contenida en el
         nombre_reporte de DATOS PUNTO o viceversa.
    """
    if not nombre_reporte:
        return None
    clave = _normalizar_texto(nombre_reporte)

    # 1) Match exacto
    for s in sucursales:
        if s.clave_busqueda == clave:
            return s

    return None


# ============================================================
# Detector de tipo de hoja
# ============================================================

def _detectar_columnas_reporte(filas: list) -> dict:
    """Examina las primeras filas de un reporte y deduce dónde están las
    columnas clave: fecha, sucursal (opcional), neto, impuestos (opcional), final.

    Retorna dict con:
        {
            "fila_encabezado": int,
            "col_fecha": int,
            "col_sucursal": Optional[int],
            "col_neto": int,
            "col_impuestos": Optional[int],
            "col_final": int,
        }
    Lanza ValueError si no se puede determinar.
    """
    NOMBRES_FECHA = ("FECHA", "DATE", "FECHA DE APERTURA", "OPEN DATE")
    NOMBRES_NETO = ("TOTAL NETO", "NET TOTAL", "SUBTOTAL")
    NOMBRES_FINAL = ("TOTAL FINAL", "FINAL TOTAL", "TOTAL")
    NOMBRES_SUC = ("SUCURSAL", "ALMACEN", "ALMACÉN")
    NOMBRES_IMP = ("IMPUESTOS",)

    # Buscar fila de encabezado en las primeras 10 filas:
    # debe tener AL MENOS columna de fecha + columna de neto o final juntas.
    fila_enc_idx = None
    encabezados = None
    for i, fila in enumerate(filas[:10]):
        if not fila:
            continue
        celdas_norm = [_normalizar_texto(c) for c in fila]
        tiene_fecha = any(c in NOMBRES_FECHA for c in celdas_norm)
        tiene_neto = any(c in NOMBRES_NETO for c in celdas_norm)
        tiene_final = any(c in NOMBRES_FINAL for c in celdas_norm)
        if tiene_fecha and (tiene_neto or tiene_final):
            fila_enc_idx = i
            encabezados = celdas_norm
            break

    if fila_enc_idx is None or not encabezados:
        raise ValueError(
            "No se pudo encontrar la fila de encabezado. Buscaba "
            "una fila con 'Fecha/Date' + 'Total Neto/Net Total/Subtotal' "
            "o 'Total/Total Final/Final Total'."
        )

    def _find(*nombres) -> Optional[int]:
        for j, h in enumerate(encabezados):
            if h in nombres:
                return j
        return None

    col_fecha = _find(*NOMBRES_FECHA)
    col_sucursal = _find(*NOMBRES_SUC)
    col_neto = _find(*NOMBRES_NETO)
    col_impuestos = _find(*NOMBRES_IMP)
    col_final = _find(*NOMBRES_FINAL)

    if col_fecha is None or col_neto is None or col_final is None:
        raise ValueError(
            f"Falta alguna columna obligatoria. "
            f"col_fecha={col_fecha}, col_neto={col_neto}, col_final={col_final}. "
            f"Encabezados detectados: {encabezados}"
        )

    return {
        "fila_encabezado": fila_enc_idx,
        "col_fecha": col_fecha,
        "col_sucursal": col_sucursal,
        "col_neto": col_neto,
        "col_impuestos": col_impuestos,
        "col_final": col_final,
    }


# ============================================================
# Procesar una hoja
# ============================================================

def _procesar_hoja(
    nombre_hoja: str,
    ws,
    sucursales: List[Sucursal],
) -> Tuple[List[dict], List[str], List[dict]]:
    """Procesa una hoja de reporte y produce las filas de plano.

    Soporta dos formatos:
      A) Hoja con múltiples reportes (estilo CHILI): cada sección tiene su
         propio título de sucursal seguido de un encabezado y filas de fechas.
      B) Hoja con un único reporte (HENKO con columna 'Almacén' o L3AF mono-sucursal).

    Aplica validación de CLASE DE SEDE: cada hoja tiene una clase esperada
    (CHILI → SANTA LEÑA, HENKO/L3AF → RESTAURANTE MILAGROS). Las sucursales
    cuya CLASE no coincida se ignoran y se reportan.

    Returns:
        - filas_plano: lista de dicts (cada uno una línea del plano)
        - log_mensajes: lista de strings informativos
        - sucursales_no_encontradas: lista de dicts {nombre, hoja, total, motivo}
    """
    log: List[str] = []
    filas_plano: List[dict] = []
    sucs_no_enc_dict = {}  # {nombre_no_encontrado: {"total": int, "motivo": str}}

    # ==================== Validación de clase ====================
    clase_esperada = _clase_esperada_para_hoja(nombre_hoja)
    if clase_esperada:
        log.append(f"  📌 '{nombre_hoja}' → solo procesa sucursales de clase '{clase_esperada}'.")

    def _validar_clase(suc: Sucursal) -> bool:
        """True si la clase de la sucursal coincide con la esperada de la hoja.
        Si la hoja no tiene clase esperada, siempre retorna True."""
        if not clase_esperada:
            return True
        return suc.clase_norm == clase_esperada

    def _registrar_descarte(nombre: str, total: int, motivo: str):
        nom_norm = _normalizar_texto(nombre)
        if nom_norm not in sucs_no_enc_dict:
            sucs_no_enc_dict[nom_norm] = {"total": 0, "motivo": motivo}
        sucs_no_enc_dict[nom_norm]["total"] += total

    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        log.append(f"  Hoja '{nombre_hoja}' está vacía, se ignora.")
        return filas_plano, log, []

    # Detectar todas las "secciones de encabezado" en la hoja
    NOMBRES_FECHA = ("FECHA", "DATE", "FECHA DE APERTURA", "OPEN DATE")
    NOMBRES_NETO = ("TOTAL NETO", "NET TOTAL", "SUBTOTAL")
    NOMBRES_FINAL = ("TOTAL FINAL", "FINAL TOTAL", "TOTAL")

    secciones = []
    for i, fila in enumerate(filas):
        if not fila:
            continue
        celdas_norm = [_normalizar_texto(c) for c in fila]
        tiene_fecha = any(c in NOMBRES_FECHA for c in celdas_norm)
        tiene_neto = any(c in NOMBRES_NETO for c in celdas_norm)
        tiene_final = any(c in NOMBRES_FINAL for c in celdas_norm)
        if tiene_fecha and (tiene_neto or tiene_final):
            secciones.append({"idx_encabezado": i, "encabezados": celdas_norm})

    if not secciones:
        log.append(f"  ⚠️ Hoja '{nombre_hoja}': no se encontró encabezado, se ignora.")
        return filas_plano, log, []

    # Pool de sucursales a considerar para esta hoja: si hay clase esperada,
    # solo las que coinciden con esa clase. Si no, todas.
    if clase_esperada:
        sucursales_pool = [s for s in sucursales if s.clase_norm == clase_esperada]
        if not sucursales_pool:
            log.append(
                f"  ⚠️ '{nombre_hoja}': no hay sucursales en DATOS PUNTO con "
                f"CLASE DE SEDE '{clase_esperada}'. Hoja ignorada."
            )
            return filas_plano, log, []
    else:
        sucursales_pool = sucursales

    def _buscar_en_pool(nombre: str) -> Optional[Sucursal]:
        return _buscar_sucursal(nombre, sucursales_pool)

    def _localizar_sucursal_default(idx_enc: int) -> Optional[Sucursal]:
        # Buscar hacia arriba (hasta 5 filas) un texto que coincida con alguna
        # sucursal del pool válido (filtrado por clase si corresponde)
        for j in range(idx_enc - 1, max(idx_enc - 5, -1), -1):
            if j < 0 or j >= len(filas) or not filas[j]:
                continue
            for celda in filas[j]:
                if celda is None:
                    continue
                celda_norm = _normalizar_texto(celda)
                if not celda_norm:
                    continue
                # Match exacto contra el pool válido
                for s in sucursales_pool:
                    if s.clave_busqueda == celda_norm:
                        return s
        # Fallback: nombre de la hoja
        nombre_hoja_norm = _normalizar_texto(nombre_hoja)
        for s in sucursales_pool:
            if s.clave_busqueda and s.clave_busqueda in nombre_hoja_norm:
                return s
        return None

    def _find_in(encabezados: list, *nombres) -> Optional[int]:
        for j, h in enumerate(encabezados):
            if h in nombres:
                return j
        return None

    NOMBRES_SUC = ("SUCURSAL", "ALMACEN", "ALMACÉN")
    NOMBRES_IMP = ("IMPUESTOS",)

    for s_idx, seccion in enumerate(secciones):
        idx_enc = seccion["idx_encabezado"]
        encabezados = seccion["encabezados"]

        col_fecha = _find_in(encabezados, *NOMBRES_FECHA)
        col_sucursal = _find_in(encabezados, *NOMBRES_SUC)
        col_neto = _find_in(encabezados, *NOMBRES_NETO)
        col_impuestos = _find_in(encabezados, *NOMBRES_IMP)
        col_final = _find_in(encabezados, *NOMBRES_FINAL)

        if col_fecha is None or col_neto is None or col_final is None:
            continue

        # Sucursal default si NO hay columna de sucursal
        sucursal_default = None
        if col_sucursal is None:
            sucursal_default = _localizar_sucursal_default(idx_enc)
            if sucursal_default is None:
                # ¿La sucursal aparece arriba pero NO está en el pool válido?
                # → buscarla en TODAS las sucursales (sin filtro) para reportar
                #   por qué se descartó
                for j in range(idx_enc - 1, max(idx_enc - 5, -1), -1):
                    if j < 0 or j >= len(filas) or not filas[j]:
                        continue
                    for celda in filas[j]:
                        if celda is None:
                            continue
                        nombre_visto = str(celda).strip()
                        if not nombre_visto:
                            continue
                        suc_global = _buscar_sucursal(nombre_visto, sucursales)
                        # Sumar total aproximado de la sección para reportarlo
                        total_seccion = 0
                        idx_fin_tmp = (
                            secciones[s_idx + 1]["idx_encabezado"]
                            if s_idx + 1 < len(secciones)
                            else len(filas)
                        )
                        for k in range(idx_enc + 1, idx_fin_tmp):
                            f_tmp = filas[k]
                            if not f_tmp:
                                continue
                            primera = str(f_tmp[0]).strip().upper() if f_tmp[0] else ""
                            if primera.startswith(("GRAND", "END", "PRINTED")):
                                continue
                            if col_final < len(f_tmp):
                                total_seccion += _to_int(f_tmp[col_final])
                        if suc_global and not _validar_clase(suc_global):
                            _registrar_descarte(
                                nombre_visto,
                                total_seccion,
                                f"clase '{suc_global.clase}' no coincide con "
                                f"esperada '{clase_esperada}' para hoja '{nombre_hoja}'",
                            )
                        elif not suc_global and clase_esperada:
                            _registrar_descarte(
                                nombre_visto,
                                total_seccion,
                                f"sucursal no encontrada en DATOS PUNTO con "
                                f"clase '{clase_esperada}'",
                            )
                        break  # solo primera celda no vacía
                    if any(c for c in filas[j] if c):
                        break
                continue

        # Determinar dónde termina esta sección
        idx_fin = (
            secciones[s_idx + 1]["idx_encabezado"]
            if s_idx + 1 < len(secciones)
            else len(filas)
        )

        agregados = {}

        for i in range(idx_enc + 1, idx_fin):
            fila = filas[i]
            if not fila:
                continue
            primera = str(fila[0]).strip().upper() if fila[0] else ""
            if primera.startswith(("GRAND", "END", "PRINTED")):
                continue

            f_raw = fila[col_fecha] if col_fecha < len(fila) else None
            f = _parsear_fecha(f_raw)
            if not f:
                continue

            # Sucursal
            suc = None
            if col_sucursal is not None and col_sucursal < len(fila):
                nombre_suc_raw = fila[col_sucursal]
                if nombre_suc_raw:
                    nombre_str = str(nombre_suc_raw)
                    suc = _buscar_en_pool(nombre_str)
                    if suc is None:
                        # Verificar si existe pero con clase incorrecta
                        suc_global = _buscar_sucursal(nombre_str, sucursales)
                        valor_final = _to_int(
                            fila[col_final] if col_final < len(fila) else 0
                        )
                        if suc_global and not _validar_clase(suc_global):
                            _registrar_descarte(
                                nombre_str, valor_final,
                                f"clase '{suc_global.clase}' ≠ esperada '{clase_esperada}'",
                            )
                        else:
                            _registrar_descarte(
                                nombre_str, valor_final,
                                "no encontrada en DATOS PUNTO" + (
                                    f" (clase esperada: {clase_esperada})"
                                    if clase_esperada else ""
                                ),
                            )
                        continue
                else:
                    suc = sucursal_default
            else:
                suc = sucursal_default

            if suc is None:
                continue

            # Validación final de seguridad: la sucursal pool ya está filtrada,
            # pero si por alguna razón llegó otra, validamos.
            if not _validar_clase(suc):
                _registrar_descarte(
                    suc.nombre_reporte,
                    _to_int(fila[col_final] if col_final < len(fila) else 0),
                    f"clase '{suc.clase}' ≠ esperada '{clase_esperada}'",
                )
                continue

            # Montos
            neto = _to_int(fila[col_neto] if col_neto < len(fila) else 0)
            final = _to_int(fila[col_final] if col_final < len(fila) else 0)
            if col_impuestos is not None and col_impuestos < len(fila):
                impuestos = _to_int(fila[col_impuestos])
            else:
                impuestos = final - neto

            if final <= 0 and neto <= 0:
                continue

            clave = (f, suc.cc)
            if clave not in agregados:
                agregados[clave] = {
                    "neto": 0, "impuestos": 0, "final": 0, "sucursal": suc,
                }
            agregados[clave]["neto"] += neto
            agregados[clave]["impuestos"] += impuestos
            agregados[clave]["final"] += final

        # Generar líneas del plano
        for (f, _cc), info in sorted(agregados.items(), key=lambda x: (x[0][0], x[0][1])):
            suc: Sucursal = info["sucursal"]
            neto = info["neto"]
            impuestos = info["impuestos"]
            final = info["final"]

            if neto + impuestos != final:
                impuestos = final - neto

            fecha_str = _formato_fecha_plano(f)
            documento = str(f.day)
            detalle = f"{DETALLE_PREFIJO}{suc.nombre_reporte}".upper()
            comp = suc.comprobante

            filas_plano.append({
                "CUENTA": suc.cuenta_caja, "COMPROBANTE": comp,
                "FECHA": fecha_str, "DOCUMENTO": documento,
                "DOC REFERENCIA": "", "NIT": NIT_GENERICO_POS,
                "DETALLE": detalle, "TR": "1",
                "VALOR": final, "BASE": 0,
                "CENTRO DE COSTO": suc.cc,
            })
            if neto > 0:
                filas_plano.append({
                    "CUENTA": suc.cta_base_v, "COMPROBANTE": comp,
                    "FECHA": fecha_str, "DOCUMENTO": documento,
                    "DOC REFERENCIA": "", "NIT": NIT_GENERICO_POS,
                    "DETALLE": detalle, "TR": "2",
                    "VALOR": neto, "BASE": 0,
                    "CENTRO DE COSTO": suc.cc,
                })
            if impuestos > 0:
                filas_plano.append({
                    "CUENTA": suc.cta_ico, "COMPROBANTE": comp,
                    "FECHA": fecha_str, "DOCUMENTO": documento,
                    "DOC REFERENCIA": "", "NIT": NIT_GENERICO_POS,
                    "DETALLE": detalle, "TR": "2",
                    "VALOR": impuestos, "BASE": neto,
                    "CENTRO DE COSTO": suc.cc,
                })

    n_lineas = len(filas_plano)
    n_asientos = n_lineas // 3 if n_lineas > 0 else 0
    if n_lineas > 0:
        df_tmp = pd.DataFrame(filas_plano, columns=COLUMNAS_PLANO)
        total_db = int(df_tmp[df_tmp["TR"] == "1"]["VALOR"].astype(int).sum())
        total_cr = int(df_tmp[df_tmp["TR"] == "2"]["VALOR"].astype(int).sum())
        log.append(
            f"  ✓ '{nombre_hoja}': {len(secciones)} secciones, ~{n_asientos} asientos. "
            f"Db=${total_db:,} Cr=${total_cr:,}".replace(",", ".")
        )
    else:
        log.append(f"  ⚠️ '{nombre_hoja}': no se generaron asientos.")

    sucs_no_enc = [
        {
            "nombre": n,
            "hoja": nombre_hoja,
            "total_aproximado": data["total"],
            "motivo": data["motivo"],
        }
        for n, data in sucs_no_enc_dict.items()
    ]
    return filas_plano, log, sucs_no_enc


# ============================================================
# Función principal
# ============================================================

def procesar_pos(
    archivo,
) -> Tuple[pd.DataFrame, List[str], List[dict]]:
    """Procesa el Excel completo y produce el plano contable.

    Args:
        archivo: file-like (BytesIO, UploadedFile) o bytes del Excel.

    Returns:
        (df_plano, log_mensajes, sucursales_no_encontradas)

        df_plano: DataFrame con las 11 columnas del plano.
        log_mensajes: lista de strings informativos.
        sucursales_no_encontradas: lista de dicts con sucursales que
            aparecieron en los reportes pero no están en DATOS PUNTO.
    """
    log: List[str] = []

    # Normalizar entrada a BytesIO
    if hasattr(archivo, "read"):
        contenido = archivo.read()
        if hasattr(archivo, "seek"):
            archivo.seek(0)
        bio = io.BytesIO(contenido)
    elif isinstance(archivo, (bytes, bytearray)):
        bio = io.BytesIO(bytes(archivo))
    else:
        raise TypeError(f"Tipo de archivo no soportado: {type(archivo)}")

    wb = load_workbook(bio, data_only=True, read_only=True)

    # Leer maestro
    sucursales = _leer_datos_punto(wb)
    log.append(f"📋 DATOS PUNTO: {len(sucursales)} sucursales registradas.")

    # Procesar cada hoja de reporte
    todas_filas: List[dict] = []
    todas_sucs_no_enc: List[dict] = []

    hojas_reporte = [h for h in wb.sheetnames if h.upper() not in HOJAS_NO_REPORTE]
    log.append(f"📂 Hojas de reporte detectadas: {len(hojas_reporte)}")

    for nombre_hoja in hojas_reporte:
        ws = wb[nombre_hoja]
        filas, sub_log, sucs_no = _procesar_hoja(nombre_hoja, ws, sucursales)
        todas_filas.extend(filas)
        log.extend(sub_log)
        todas_sucs_no_enc.extend(sucs_no)

    # Resumen general
    df_plano = pd.DataFrame(todas_filas, columns=COLUMNAS_PLANO)
    if len(df_plano) > 0:
        total_db = int(df_plano[df_plano["TR"] == "1"]["VALOR"].astype(int).sum())
        total_cr = int(df_plano[df_plano["TR"] == "2"]["VALOR"].astype(int).sum())
        log.append("")
        log.append(f"📊 TOTAL plano: {len(df_plano)} líneas")
        log.append(f"   Total Db: $ {total_db:,}".replace(",", "."))
        log.append(f"   Total Cr: $ {total_cr:,}".replace(",", "."))
        if total_db == total_cr:
            log.append(f"   ✅ Cuadre perfecto Db = Cr")
        else:
            log.append(f"   ❌ DESCUADRE: diferencia $ {total_db - total_cr:,}".replace(",", "."))

    return df_plano, log, todas_sucs_no_enc


# ============================================================
# Exportar a TSV (mismo formato que Caja Menor / DIAN)
# ============================================================

def dataframe_a_plano_tsv(
    df: pd.DataFrame,
    incluir_encabezado_excel: bool = True,
) -> bytes:
    """Convierte el DataFrame a bytes TSV compatible con el sistema contable.

    Si incluir_encabezado_excel=True, antepone 'sep=\\t' para que Excel abra el
    archivo con las columnas alineadas.
    """
    # Asegurar todas las columnas y orden correcto
    df_out = df[COLUMNAS_PLANO].copy()
    # Convertir a string para evitar problemas de tipos
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str).str.replace("\t", " ", regex=False)

    tsv = df_out.to_csv(sep="\t", index=False, header=True, lineterminator="\r\n")

    if incluir_encabezado_excel:
        tsv = "sep=\t\r\n" + tsv

    return tsv.encode("utf-8")


# ============================================================
# Combinar archivos individuales (modo "subir por separado")
# ============================================================

def combinar_archivos_pos(
    archivo_datos_punto,
    reportes: dict,
) -> bytes:
    """Combina archivos individuales en un único Excel listo para procesar.

    Esta función permite que el usuario suba los reportes POS por separado
    (cada empresa POS exporta su reporte en un archivo distinto) en lugar
    de tener que copiar-pegar todo en un solo Excel multi-hoja.

    Args:
        archivo_datos_punto: file-like o bytes del Excel con la hoja DATOS PUNTO.
            Debe contener una hoja llamada 'DATOS PUNTO'.
        reportes: dict con los reportes individuales:
            {
                'REPORTE CHILI': file-like|bytes (opcional),
                'L3AF MILLA DE ORO': file-like|bytes (opcional),
                'HENKO MILAGROS': file-like|bytes (opcional),
                'HENKO REMEDIOS': file-like|bytes (opcional),
            }
            Para cada clave presente, se toma la PRIMERA hoja del archivo
            y se renombra al nombre de la clave.

    Returns:
        bytes del Excel combinado (xlsx). Listo para pasar a procesar_pos().
    """
    # Cargar DATOS PUNTO
    if hasattr(archivo_datos_punto, "read"):
        contenido_dp = archivo_datos_punto.read()
        if hasattr(archivo_datos_punto, "seek"):
            archivo_datos_punto.seek(0)
    elif isinstance(archivo_datos_punto, (bytes, bytearray)):
        contenido_dp = bytes(archivo_datos_punto)
    else:
        raise TypeError(
            f"archivo_datos_punto: tipo no soportado {type(archivo_datos_punto)}"
        )

    wb_dp = load_workbook(io.BytesIO(contenido_dp), data_only=True)

    if "DATOS PUNTO" not in wb_dp.sheetnames:
        raise ValueError(
            "El archivo de DATOS PUNTO debe contener una hoja llamada "
            "exactamente 'DATOS PUNTO'."
        )

    # Construir el workbook combinado: copio las hojas que necesito
    from openpyxl import Workbook
    from copy import copy

    wb_out = Workbook()
    # Eliminar la hoja por defecto
    default = wb_out.active
    wb_out.remove(default)

    def _copiar_hoja(ws_orig, nombre_destino):
        """Copia los valores de una hoja a otra dándole un nombre nuevo."""
        ws_new = wb_out.create_sheet(title=nombre_destino[:31])
        for row in ws_orig.iter_rows(values_only=True):
            ws_new.append(row)

    # Copiar DATOS PUNTO primero (será requisito de procesar_pos)
    _copiar_hoja(wb_dp["DATOS PUNTO"], "DATOS PUNTO")

    # Copiar cada reporte
    for nombre_destino, archivo in reportes.items():
        if archivo is None:
            continue
        if hasattr(archivo, "read"):
            contenido = archivo.read()
            if hasattr(archivo, "seek"):
                archivo.seek(0)
        elif isinstance(archivo, (bytes, bytearray)):
            contenido = bytes(archivo)
        else:
            continue

        try:
            wb_rep = load_workbook(io.BytesIO(contenido), data_only=True)
        except Exception as e:
            raise ValueError(f"No se pudo leer el archivo de '{nombre_destino}': {e}")

        # Tomar la primera hoja con datos
        ws_rep = wb_rep.active
        if ws_rep is None or ws_rep.max_row <= 1:
            # Buscar otra hoja con datos
            for nh in wb_rep.sheetnames:
                if wb_rep[nh].max_row > 1:
                    ws_rep = wb_rep[nh]
                    break

        if ws_rep is None or ws_rep.max_row <= 1:
            raise ValueError(
                f"El archivo de '{nombre_destino}' no contiene datos."
            )

        _copiar_hoja(ws_rep, nombre_destino)

    # Guardar a bytes
    bio_out = io.BytesIO()
    wb_out.save(bio_out)
    bio_out.seek(0)
    return bio_out.getvalue()


# ============================================================
# Persistencia de DATOS PUNTO en Supabase Storage
# ============================================================

NOMBRE_DATOS_PUNTO = "datos_punto.xlsx"
BUCKET_CONFIG = "empresas-config"


def cargar_datos_punto_desde_storage(empresa_id: str) -> Optional[bytes]:
    """Descarga el archivo de DATOS PUNTO guardado para la empresa.

    Retorna None si la empresa todavía no ha guardado un archivo.
    """
    from db.supabase_client import get_supabase

    sb = get_supabase()
    ruta = f"{empresa_id}/{NOMBRE_DATOS_PUNTO}"
    try:
        resp = sb.storage.from_(BUCKET_CONFIG).download(ruta)
    except Exception:
        return None
    if not resp:
        return None
    if isinstance(resp, (bytes, bytearray)):
        return bytes(resp)
    contenido = getattr(resp, "content", None)
    return contenido if isinstance(contenido, (bytes, bytearray)) else None


def guardar_datos_punto_en_storage(empresa_id: str, contenido: bytes) -> bool:
    """Sube el archivo de DATOS PUNTO al bucket de la empresa."""
    from db.supabase_client import get_supabase_admin, get_supabase

    sb = get_supabase_admin() or get_supabase()
    ruta = f"{empresa_id}/{NOMBRE_DATOS_PUNTO}"
    try:
        sb.storage.from_(BUCKET_CONFIG).upload(
            path=ruta,
            file=contenido,
            file_options={
                "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "upsert": "true",
            },
        )
        return True
    except Exception:
        try:
            sb.storage.from_(BUCKET_CONFIG).update(
                path=ruta,
                file=contenido,
                file_options={
                    "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
            )
            return True
        except Exception:
            raise


def eliminar_datos_punto_de_storage(empresa_id: str) -> bool:
    """Elimina el archivo de DATOS PUNTO del Storage."""
    from db.supabase_client import get_supabase_admin, get_supabase

    sb = get_supabase_admin() or get_supabase()
    ruta = f"{empresa_id}/{NOMBRE_DATOS_PUNTO}"
    try:
        sb.storage.from_(BUCKET_CONFIG).remove([ruta])
        return True
    except Exception:
        return False


def info_datos_punto(contenido_xlsx: bytes) -> dict:
    """Lee el archivo y devuelve estadísticas básicas."""
    bio = io.BytesIO(contenido_xlsx)
    wb = load_workbook(bio, data_only=True, read_only=True)
    if "DATOS PUNTO" not in wb.sheetnames:
        raise ValueError("El archivo no contiene la hoja 'DATOS PUNTO'.")
    sucursales = _leer_datos_punto(wb)

    # Contar por clase
    por_clase = {}
    for s in sucursales:
        por_clase[s.clase or "(sin clase)"] = por_clase.get(s.clase or "(sin clase)", 0) + 1

    return {
        "total": len(sucursales),
        "por_clase": por_clase,
        "sucursales": [
            {
                "nombre": s.nombre_reporte,
                "sede": s.sede,
                "cc": s.cc,
                "clase": s.clase,
                "cuenta_caja": s.cuenta_caja,
            }
            for s in sucursales
        ],
    }
