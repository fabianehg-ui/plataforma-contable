"""
Procesador de Compras y Egresos.

Procesa 3 archivos relacionados y genera 4 comprobantes contables:

1. TOKEN DIAN (.xlsx) → Comprobante 3 (Facturas) + Comprobante 7 (NC)
   - Filtra Tipo='Factura electrónica' → Comp 3 (compras)
   - Filtra Tipo='Nota de crédito electrónica' → Comp 7 (devoluciones)
   - Numeración global: YYYYMMNNN ordenando por fecha+folio

2. DOCUMENTO SOPORTE (.xlsx) → Comprobante 18
   - Compras a personas naturales que no facturan electrónicamente
   - Mapeo Producto → Cuenta vía catálogo
   - Retefuente 3.5% si supera base UVT en servicios

3. EGRESOS CAJA MENOR (HTML/.xls) → Comprobante 13
   - Pagos a proveedores desde caja menor
   - CEG cancela DS → toma cuenta del proveedor del DS
   - CEG a Banco Caja Social → cuenta Db 11100501
   - Resto → Db 23359501

REGLAS CRÍTICAS:
- TODOS los valores en VALOR son POSITIVOS. El signo lo da el TR (1=Db, 2=Cr).
- Las cuentas 2365xxxx solo se usan para CALCULAR retefuente sobre bases legales.
  NO se usan como Db en pagos o causaciones.
- En NC se usa 622505 (devolución en compra) — la cuenta correcta según normatividad.
"""
from __future__ import annotations
import io
import json
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


# ============================================================
# Constantes
# ============================================================

# Comprobantes
COMPROBANTE_COMPRAS = "3"
COMPROBANTE_NC_PROVEEDORES = "7"
COMPROBANTE_CEG = "13"
COMPROBANTE_DS = "18"

# Tipos de transacción
TR_DEBITO = "1"
TR_CREDITO = "2"

# Cuentas
CUENTA_CAJA = "11050500"
CUENTA_BANCO_DEFAULT = "11100501"
CUENTA_PROVEEDORES = "22050501"
CUENTA_ACREEDORES = "23359501"
CUENTA_RETEFUENTE_SERVICIOS = "23654003"
CUENTA_RETEFUENTE_JURIDICA = "23654005"
CUENTA_RETEFUENTE_NATURAL = "23654025"
CUENTA_RETEFUENTE_HONORARIOS = "23651501"
CUENTA_RETEFUENTE_COMISIONES = "23652001"
CUENTA_RETEFUENTE_ARRIENDO = "23653000"
CUENTA_RETEIVA = "23653525"
CUENTA_IVA_COMPRAS = "24080201"
CUENTA_IVA_SERVICIOS = "24080308"
CUENTA_IVA_NC = "24080204"
CUENTA_IMPUESTO_CONSUMO = "511575"
CUENTA_COMPRAS_GRAVADAS = "620505"
CUENTA_COMPRAS_NO_GRAVADAS = "620510"
CUENTA_DEVOLUCION_COMPRAS = "622505"

# Defaults DS
CUENTA_DB_DEFAULT_DS = "519530"
CUENTA_CR_DEFAULT_DS = "23359501"

# Otras
TASA_RETEFUENTE_SERVICIOS = 0.035
UVT_DEFAULT = 49799
CENTRO_COSTOS_DEFAULT = "PRINCIPAL"
NIT_GENERICO = "222222222"

# Columnas del plano
COLUMNAS_PLANO = [
    "CUENTA",
    "COMPROBANTE",
    "FECHA",
    "DOCUMENTO",
    "DOC REFERENCIA",
    "NIT",
    "DETALLE",
    "TR",
    "VALOR",
    "BASE",
    "CENTRO DE COSTO",
]


# ============================================================
# Estructuras de datos
# ============================================================

@dataclass
class ProductoDS:
    codigo: str
    nombre: str
    cuenta_db: str
    cuenta_cr_proveedor: str
    retiene_servicios: bool


@dataclass
class BancoConfig:
    nit: str
    nombre: str
    cuenta: str


@dataclass
class ProveedorConsumo:
    nit: str
    nombre: str
    tasa: float


@dataclass
class TipoRetencion:
    """Configuración tributaria para un tipo de proveedor.

    base_min_uvt = 0 significa "sin mínimo legal" (ej: honorarios).
    cuenta_iva_descontable: 24080201 para compras de bienes,
                           24080308 para servicios/honorarios/comisiones/arrendamientos.
    """
    codigo: str               # 'juridica', 'natural', 'honorarios', etc.
    descripcion: str
    cuenta_renta: str         # vacía si no_retiene
    tasa_renta: float
    base_min_uvt: int
    aplica_reteiva: bool
    tasa_reteiva: float
    cuenta_iva_descontable: str  # 24080201 o 24080308 según tipo de operación


@dataclass
class ProveedorRetencion:
    """Mapeo NIT → tipo de retención."""
    nit: str
    nombre: str
    tipo: str


@dataclass
class DocumentoToken:
    """Una factura o NC del archivo TOKEN DIAN."""
    tipo: str             # 'Factura electrónica' o 'Nota de crédito electrónica'
    folio: str            # solo el número
    prefijo: str          # 'FEGB', 'FE', etc.
    fecha: date
    nit_emisor: str
    nombre_emisor: str
    iva: float
    rete_iva: float
    rete_renta: float
    rete_ica: float
    inc: float
    total: float

    @property
    def folio_completo(self) -> str:
        """Folio con prefijo concatenado, como aparece en el plano."""
        if self.prefijo:
            return f"{self.prefijo}{self.folio}"
        return self.folio

    @property
    def es_factura(self) -> bool:
        return "FACTURA" in self.tipo.upper()

    @property
    def es_nota_credito(self) -> bool:
        return "NOTA DE CRÉDITO" in self.tipo.upper() or "NOTA DE CREDITO" in self.tipo.upper()


@dataclass
class DocumentoSoporte:
    consecutivo: str
    fecha: date
    proveedor: str
    identificacion: str
    producto_codigo: str
    producto_nombre: str
    producto_completo: str
    total: float
    anulado: bool


@dataclass
class EgresoCajaMenor:
    consecutivo: str
    cancela: str
    identificacion: str
    nombres: str
    fecha: date
    sucursal: str
    concepto: str
    valor: float


# ============================================================
# Helpers
# ============================================================

def _formato_cuenta(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and not pd.isna(valor):
        valor = int(valor)
    return str(valor).strip()


def _formato_cc(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and not pd.isna(valor):
        valor = int(valor)
    s = str(valor).strip()
    return re.sub(r"\D", "", s)


def _to_float(valor) -> float:
    """Convierte string/numérico a float manejando formatos latino y anglo."""
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        if pd.isna(valor):
            return 0.0
        return float(valor)
    s = str(valor).strip().replace("$", "").replace(" ", "").replace("\xa0", "")
    if not s or s == "-":
        return 0.0
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        partes = s.split(",")
        if len(partes) == 2 and len(partes[1]) <= 2 and len(partes[1]) > 0:
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "." in s:
        partes = s.split(".")
        if len(partes) == 2 and len(partes[1]) <= 2:
            pass
        else:
            s = s.replace(".", "")
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _redondear(valor: float) -> int:
    """Redondea a entero. SIEMPRE devuelve absoluto positivo."""
    return abs(int(round(valor)))


def _formato_fecha_plano(d: date) -> str:
    return d.strftime("%m/%d/%Y")


def _crear_linea_plano(
    cuenta: str, comprobante: str, fecha: date, documento: str,
    doc_referencia: str, nit: str, detalle: str, tr: str,
    valor: int, base: int = 0, centro_costo: str = CENTRO_COSTOS_DEFAULT,
) -> dict:
    """Crea una fila del plano con valores SIEMPRE positivos."""
    return {
        "CUENTA": cuenta,
        "COMPROBANTE": comprobante,
        "FECHA": _formato_fecha_plano(fecha),
        "DOCUMENTO": documento,
        "DOC REFERENCIA": doc_referencia,
        "NIT": nit,
        "DETALLE": detalle,
        "TR": tr,
        "VALOR": abs(int(valor)),  # ← garantiza positivo
        "BASE": abs(int(base)),
        "CENTRO DE COSTO": centro_costo,
    }


# ============================================================
# Catálogo
# ============================================================

_RUTA_CATALOGO = (
    Path(__file__).resolve().parent.parent / "data" / "catalogo_compras.json"
)


def cargar_catalogo() -> dict:
    """Carga el catálogo embebido."""
    if not _RUTA_CATALOGO.exists():
        raise FileNotFoundError(f"No se encontró el catálogo: {_RUTA_CATALOGO}")
    with open(_RUTA_CATALOGO, "r", encoding="utf-8") as f:
        data = json.load(f)

    productos: Dict[str, ProductoDS] = {}
    for item in data.get("productos_ds", []):
        p = ProductoDS(
            codigo=str(item.get("codigo", "")).strip().upper(),
            nombre=str(item.get("nombre", "")).strip(),
            cuenta_db=str(item.get("cuenta_db", CUENTA_DB_DEFAULT_DS)).strip(),
            cuenta_cr_proveedor=str(item.get("cuenta_cr_proveedor", CUENTA_CR_DEFAULT_DS)).strip(),
            retiene_servicios=bool(item.get("retiene_servicios", False)),
        )
        if p.codigo:
            productos[p.codigo] = p

    bancos: List[BancoConfig] = []
    for item in data.get("bancos", []):
        b = BancoConfig(
            nit=_formato_cc(item.get("nit", "")),
            nombre=str(item.get("nombre", "")).strip(),
            cuenta=str(item.get("cuenta", CUENTA_BANCO_DEFAULT)).strip(),
        )
        if b.nit:
            bancos.append(b)

    proveedores_consumo: Dict[str, ProveedorConsumo] = {}
    for item in data.get("proveedores_impuesto_consumo", []):
        nit = _formato_cc(item.get("nit", ""))
        if nit:
            proveedores_consumo[nit] = ProveedorConsumo(
                nit=nit,
                nombre=str(item.get("nombre", "")).strip(),
                tasa=float(item.get("tasa", 0.08)),
            )

    # Tipos de retención
    tipos_retencion: Dict[str, TipoRetencion] = {}
    for codigo, conf in data.get("tipos_retencion", {}).items():
        tipos_retencion[codigo] = TipoRetencion(
            codigo=codigo,
            descripcion=str(conf.get("descripcion", "")).strip(),
            cuenta_renta=str(conf.get("cuenta_renta", "")).strip(),
            tasa_renta=float(conf.get("tasa_renta", 0.0)),
            base_min_uvt=int(conf.get("base_min_uvt", 0)),
            aplica_reteiva=bool(conf.get("aplica_reteiva", False)),
            tasa_reteiva=float(conf.get("tasa_reteiva", 0.0)),
            cuenta_iva_descontable=str(conf.get("cuenta_iva_descontable", CUENTA_IVA_COMPRAS)).strip(),
        )

    # Proveedores con tipo de retención
    proveedores_retencion: Dict[str, ProveedorRetencion] = {}
    for item in data.get("proveedores_retencion", []):
        nit = _formato_cc(item.get("nit", ""))
        tipo = str(item.get("tipo", "")).strip().lower()
        if nit and tipo:
            proveedores_retencion[nit] = ProveedorRetencion(
                nit=nit,
                nombre=str(item.get("nombre", "")).strip(),
                tipo=tipo,
            )

    tipo_retencion_default = str(data.get("tipo_retencion_default", "juridica")).strip().lower()
    aplicar_calculo_retenciones = bool(data.get("aplicar_calculo_retenciones", False))

    doc = data.get("_doc", {})
    uvt = int(doc.get("uvt_2026", UVT_DEFAULT))
    base_uvt = int(doc.get("base_servicios_uvt", 4))

    return {
        "productos": productos,
        "bancos": bancos,
        "proveedores_consumo": proveedores_consumo,
        "tipos_retencion": tipos_retencion,
        "proveedores_retencion": proveedores_retencion,
        "tipo_retencion_default": tipo_retencion_default,
        "aplicar_calculo_retenciones": aplicar_calculo_retenciones,
        "uvt": uvt,
        "base_servicios_uvt": base_uvt,
    }


def info_catalogo() -> dict:
    cat = cargar_catalogo()
    return {
        "total_productos": len(cat["productos"]),
        "total_bancos": len(cat["bancos"]),
        "total_proveedores_consumo": len(cat["proveedores_consumo"]),
        "total_proveedores_retencion": len(cat["proveedores_retencion"]),
        "total_tipos_retencion": len(cat["tipos_retencion"]),
        "tipo_retencion_default": cat["tipo_retencion_default"],
        "aplicar_calculo_retenciones": cat["aplicar_calculo_retenciones"],
        "uvt_2026": cat["uvt"],
        "base_minima_retefuente": cat["uvt"] * cat["base_servicios_uvt"],
        "productos": [
            {
                "codigo": p.codigo, "nombre": p.nombre,
                "cuenta_db": p.cuenta_db, "cuenta_cr_proveedor": p.cuenta_cr_proveedor,
                "retiene": "Sí" if p.retiene_servicios else "No",
            }
            for p in cat["productos"].values()
        ],
        "bancos": [
            {"nit": b.nit, "nombre": b.nombre, "cuenta": b.cuenta}
            for b in cat["bancos"]
        ],
        "tipos_retencion": [
            {
                "codigo": t.codigo,
                "descripcion": t.descripcion,
                "cuenta_renta": t.cuenta_renta,
                "tasa_renta": f"{t.tasa_renta*100:.1f}%",
                "base_min_uvt": t.base_min_uvt,
                "base_min_pesos": t.base_min_uvt * cat["uvt"],
                "reteiva": f"{t.tasa_reteiva*100:.0f}%" if t.aplica_reteiva else "No",
                "cuenta_iva": t.cuenta_iva_descontable,
            }
            for t in cat["tipos_retencion"].values()
        ],
        "proveedores_retencion": [
            {"nit": p.nit, "nombre": p.nombre, "tipo": p.tipo}
            for p in cat["proveedores_retencion"].values()
        ],
    }


# ============================================================
# Lectura: TOKEN DIAN
# ============================================================

def _leer_token(archivo) -> List[DocumentoToken]:
    """Lee el archivo TOKEN y retorna facturas + NC (NO notas débito)."""
    if hasattr(archivo, "read"):
        contenido = archivo.read()
        if hasattr(archivo, "seek"):
            archivo.seek(0)
        bio = io.BytesIO(contenido)
    elif isinstance(archivo, (bytes, bytearray)):
        bio = io.BytesIO(bytes(archivo))
    else:
        raise TypeError(f"Tipo de archivo no soportado: {type(archivo)}")

    wb = load_workbook(bio, data_only=True, read_only=True)
    ws = wb.active
    iter_rows = ws.iter_rows(values_only=True)
    headers = next(iter_rows, None)
    if not headers:
        return []
    headers = [str(h).strip() if h is not None else "" for h in headers]
    h_upper = [h.upper() for h in headers]

    def find_idx_exact(clave):
        for i, h in enumerate(h_upper):
            if h == clave.upper():
                return i
        return None

    def find_idx_contains(*claves):
        for i, h in enumerate(h_upper):
            for c in claves:
                if c.upper() in h:
                    return i
        return None

    idx_tipo = find_idx_contains("TIPO DE DOCUMENTO")
    idx_folio = find_idx_exact("FOLIO")
    idx_prefijo = find_idx_exact("PREFIJO")
    idx_fecha = find_idx_contains("FECHA EMI")
    idx_nit_em = find_idx_contains("NIT EMISOR")
    idx_nom_em = find_idx_contains("NOMBRE EMISOR")
    idx_iva = find_idx_exact("IVA")
    idx_rete_iva = find_idx_contains("RETE IVA")
    idx_rete_renta = find_idx_contains("RETE RENTA")
    idx_rete_ica = find_idx_contains("RETE ICA")
    idx_inc = find_idx_exact("INC")
    idx_total = find_idx_exact("TOTAL")

    if idx_tipo is None or idx_folio is None or idx_total is None:
        return []

    docs: List[DocumentoToken] = []
    for fila in iter_rows:
        if not fila or all(c is None for c in fila):
            continue
        tipo = str(fila[idx_tipo] or "").strip()
        # Solo facturas y notas crédito
        if "FACTURA" not in tipo.upper() and "NOTA DE CRÉDITO" not in tipo.upper() and "NOTA DE CREDITO" not in tipo.upper():
            continue

        folio = str(fila[idx_folio] or "").strip()
        if not folio:
            continue

        prefijo = ""
        if idx_prefijo is not None and idx_prefijo < len(fila):
            p = fila[idx_prefijo]
            if p is not None and not pd.isna(p):
                prefijo = str(p).strip()

        # Fecha
        fecha_raw = fila[idx_fecha] if idx_fecha is not None and idx_fecha < len(fila) else None
        fecha_obj = None
        if isinstance(fecha_raw, datetime):
            fecha_obj = fecha_raw.date()
        elif isinstance(fecha_raw, date):
            fecha_obj = fecha_raw
        elif isinstance(fecha_raw, str):
            for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
                try:
                    fecha_obj = datetime.strptime(fecha_raw.strip(), fmt).date()
                    break
                except ValueError:
                    continue
        if fecha_obj is None:
            try:
                fecha_obj = pd.to_datetime(fecha_raw).date()
            except (ValueError, TypeError):
                continue

        nit_em = _formato_cc(fila[idx_nit_em] if idx_nit_em is not None and idx_nit_em < len(fila) else None)
        nom_em = ""
        if idx_nom_em is not None and idx_nom_em < len(fila):
            n = fila[idx_nom_em]
            if n is not None and not pd.isna(n):
                nom_em = str(n).strip()

        iva = _to_float(fila[idx_iva]) if idx_iva is not None and idx_iva < len(fila) else 0.0
        rete_iva = _to_float(fila[idx_rete_iva]) if idx_rete_iva is not None and idx_rete_iva < len(fila) else 0.0
        rete_renta = _to_float(fila[idx_rete_renta]) if idx_rete_renta is not None and idx_rete_renta < len(fila) else 0.0
        rete_ica = _to_float(fila[idx_rete_ica]) if idx_rete_ica is not None and idx_rete_ica < len(fila) else 0.0
        inc = _to_float(fila[idx_inc]) if idx_inc is not None and idx_inc < len(fila) else 0.0
        total = _to_float(fila[idx_total]) if idx_total is not None and idx_total < len(fila) else 0.0

        if total <= 0:
            continue

        docs.append(DocumentoToken(
            tipo=tipo,
            folio=folio,
            prefijo=prefijo,
            fecha=fecha_obj,
            nit_emisor=nit_em,
            nombre_emisor=nom_em,
            iva=iva,
            rete_iva=rete_iva,
            rete_renta=rete_renta,
            rete_ica=rete_ica,
            inc=inc,
            total=total,
        ))

    return docs


# ============================================================
# Lectura: Documento Soporte
# ============================================================

def _leer_documentos_soporte(archivo) -> List[DocumentoSoporte]:
    """Lee el archivo DOCUMENTO_SOPORTE.xlsx."""
    if hasattr(archivo, "read"):
        contenido = archivo.read()
        if hasattr(archivo, "seek"):
            archivo.seek(0)
        bio = io.BytesIO(contenido)
    elif isinstance(archivo, (bytes, bytearray)):
        bio = io.BytesIO(bytes(archivo))
    else:
        raise TypeError(f"Tipo de archivo no soportado: {type(archivo)}")

    wb = load_workbook(bio, data_only=True, read_only=True)
    hoja = None
    for nombre in wb.sheetnames:
        u = nombre.upper()
        if "COMPRA" in u or "SOPORTE" in u or "DOCUMENTO" in u:
            hoja = nombre
            break
    if hoja is None:
        hoja = wb.sheetnames[0]

    ws = wb[hoja]
    iter_rows = ws.iter_rows(values_only=True)
    headers = next(iter_rows, None)
    if not headers:
        return []
    headers = [str(h).strip() if h is not None else "" for h in headers]
    h_upper = [h.upper() for h in headers]

    def find_idx(*claves):
        for i, h in enumerate(h_upper):
            for c in claves:
                if c in h:
                    return i
        return None

    idx_anul = find_idx("ANULADO")
    idx_doc = find_idx("DOCUMENTO")
    idx_fecha = find_idx("FECHA")
    idx_prov = find_idx("PROVEEDOR")
    idx_id = find_idx("IDENTIF")
    idx_prod = find_idx("PRODUCTO")
    idx_total = find_idx("TOTAL")

    if idx_doc is None:
        return []

    docs: List[DocumentoSoporte] = []
    for fila in iter_rows:
        if not fila or all(c is None for c in fila):
            continue
        doc_raw = fila[idx_doc] if idx_doc < len(fila) else None
        if doc_raw is None or pd.isna(doc_raw):
            continue
        s_doc = str(doc_raw).strip()
        if not s_doc or s_doc.upper() == "NAN" or "SUM" in s_doc.upper():
            continue

        if s_doc.replace(".0", "").isdigit():
            consecutivo = f"DS{int(float(s_doc))}"
        elif s_doc.upper().startswith("DS"):
            consecutivo = s_doc.upper()
        else:
            consecutivo = f"DS{s_doc}"

        anulado = False
        if idx_anul is not None and idx_anul < len(fila):
            v = fila[idx_anul]
            if v is not None and not pd.isna(v):
                anulado = str(v).strip().upper() in ("SI", "TRUE", "YES", "1")

        fecha_raw = fila[idx_fecha] if idx_fecha is not None and idx_fecha < len(fila) else None
        if isinstance(fecha_raw, datetime):
            fecha_obj = fecha_raw.date()
        elif isinstance(fecha_raw, date):
            fecha_obj = fecha_raw
        else:
            try:
                fecha_obj = pd.to_datetime(fecha_raw).date()
            except (ValueError, TypeError):
                continue

        proveedor = ""
        if idx_prov is not None and idx_prov < len(fila):
            p = fila[idx_prov]
            if p is not None and not pd.isna(p):
                proveedor = str(p).strip()

        ident = ""
        if idx_id is not None and idx_id < len(fila):
            ident = _formato_cc(fila[idx_id])

        prod_completo = ""
        if idx_prod is not None and idx_prod < len(fila):
            p = fila[idx_prod]
            if p is not None and not pd.isna(p):
                prod_completo = str(p).strip()

        prod_codigo = ""
        prod_nombre = prod_completo
        if prod_completo:
            partes = prod_completo.split(maxsplit=1)
            prod_codigo = partes[0].upper() if partes else ""
            prod_nombre = partes[1] if len(partes) > 1 else ""

        total = _to_float(fila[idx_total]) if idx_total is not None and idx_total < len(fila) else 0.0
        if total <= 0 or anulado:
            continue

        docs.append(DocumentoSoporte(
            consecutivo=consecutivo,
            fecha=fecha_obj,
            proveedor=proveedor,
            identificacion=ident,
            producto_codigo=prod_codigo,
            producto_nombre=prod_nombre,
            producto_completo=prod_completo,
            total=total,
            anulado=anulado,
        ))

    return docs


# ============================================================
# Lectura: Egresos Caja Menor
# ============================================================

def _leer_egresos_caja_menor(archivo) -> List[EgresoCajaMenor]:
    """Lee el archivo de egresos. Soporta HTML y xlsx real."""
    if hasattr(archivo, "read"):
        contenido = archivo.read()
        if hasattr(archivo, "seek"):
            archivo.seek(0)
    elif isinstance(archivo, (bytes, bytearray)):
        contenido = bytes(archivo)
    else:
        raise TypeError(f"Tipo de archivo no soportado: {type(archivo)}")

    if contenido[:9].decode("utf-8", errors="ignore").upper().startswith("<!DOCTYPE"):
        df = pd.read_html(io.BytesIO(contenido))[0]
    else:
        df = pd.read_excel(io.BytesIO(contenido))

    egresos: List[EgresoCajaMenor] = []
    for _, row in df.iterrows():
        cons = row.get("Consecutivo")
        if cons is None or pd.isna(cons):
            continue
        cons_str = str(cons).strip()
        if not cons_str or cons_str.upper() == "NAN" or "SUM" in cons_str.upper():
            continue

        cancela = row.get("Cancela", cons_str)
        if cancela is None or pd.isna(cancela):
            cancela = cons_str
        cancela = str(cancela).strip()

        ident = _formato_cc(row.get("Identificación"))
        nombres = row.get("Nombres", "")
        if pd.isna(nombres):
            nombres = ""
        nombres = str(nombres).strip()

        fecha_raw = row.get("Fecha")
        if isinstance(fecha_raw, datetime):
            fecha_obj = fecha_raw.date()
        elif isinstance(fecha_raw, date):
            fecha_obj = fecha_raw
        elif isinstance(fecha_raw, str):
            try:
                fecha_obj = datetime.strptime(fecha_raw.strip(), "%d/%m/%Y").date()
            except ValueError:
                try:
                    fecha_obj = pd.to_datetime(fecha_raw).date()
                except (ValueError, TypeError):
                    continue
        else:
            try:
                fecha_obj = pd.to_datetime(fecha_raw).date()
            except (ValueError, TypeError):
                continue

        sucursal = row.get("Sucursal", "")
        if pd.isna(sucursal):
            sucursal = ""
        sucursal = str(sucursal).strip()

        concepto = row.get("Concepto", "")
        if pd.isna(concepto):
            concepto = ""
        concepto = str(concepto).strip()

        valor = _to_float(row.get("Valor"))
        if valor <= 0:
            continue

        egresos.append(EgresoCajaMenor(
            consecutivo=cons_str,
            cancela=cancela,
            identificacion=ident,
            nombres=nombres,
            fecha=fecha_obj,
            sucursal=sucursal or CENTRO_COSTOS_DEFAULT,
            concepto=concepto,
            valor=valor,
        ))

    return egresos


# ============================================================
# Cálculo de retenciones (FASE 2)
# ============================================================

@dataclass
class RetencionesCalculadas:
    """Resultado del cálculo de retenciones para un documento."""
    rete_renta: float
    cuenta_renta: str
    base_renta: float
    rete_iva: float
    base_reteiva: float
    rete_ica: float
    cuenta_iva_descontable: str  # cuenta IVA Db a usar en factura (24080201 o 24080308)
    fuente: str           # 'token', 'catalogo', 'token+catalogo', 'sin_retencion'
    tipo_aplicado: str    # código del tipo de retención usado
    advertencia: Optional[str] = None


def _calcular_retenciones(
    doc: DocumentoToken,
    valor_subtotal: float,
    valor_iva: float,
    tipos_retencion: Dict[str, TipoRetencion],
    proveedores_retencion: Dict[str, ProveedorRetencion],
    tipo_default: str,
    uvt: int,
    aplicar_calculo: bool = True,
) -> RetencionesCalculadas:
    """Calcula retenciones aplicando regla: TOKEN manda, catálogo es fallback.

    Lógica:
      1. Si TOKEN trae rete_renta > 0 → se respeta (fuente de verdad).
         Si no Y aplicar_calculo=True → se calcula desde catálogo (si supera base mínima).
         Si no Y aplicar_calculo=False → no retiene (modo legacy v2.0).
      2. Si TOKEN trae rete_iva > 0 → se respeta.
         Si no Y aplicar_calculo=True → se calcula 15% del IVA si el tipo lo aplica.
      3. ICA SIEMPRE viene del TOKEN (depende del municipio, no del tipo).
      4. Para cuenta de renta: se determina por NIT en catálogo.
         Si no está → default con advertencia.

    Importante: cuando TOKEN trae el valor, NO se valida base mínima
    (asumimos que quien generó el TOKEN ya validó eso).
    """
    nit = doc.nit_emisor or NIT_GENERICO
    advertencia = None
    fuente_partes: List[str] = []

    # ----- Determinar tipo de retención -----
    proveedor = proveedores_retencion.get(nit)
    if proveedor is not None:
        tipo_codigo = proveedor.tipo
        if tipo_codigo not in tipos_retencion:
            advertencia = (
                f"⚠️ Proveedor {nit} ({doc.nombre_emisor}) tiene tipo '{tipo_codigo}' "
                f"que no existe en tipos_retencion. Usando default '{tipo_default}'."
            )
            tipo_codigo = tipo_default
    else:
        tipo_codigo = tipo_default
        # Solo advertir si el TOKEN no trae retenciones Y el cálculo está activo
        if aplicar_calculo and doc.rete_renta == 0 and doc.rete_iva == 0:
            tipo_default_obj = tipos_retencion.get(tipo_default)
            tasa_str = f"{tipo_default_obj.tasa_renta*100:.1f}%" if tipo_default_obj else "?"
            advertencia = (
                f"⚠️ Proveedor {nit} ({doc.nombre_emisor}) no está en catálogo. "
                f"Aplicando default '{tipo_default}' ({tasa_str})."
            )

    tipo = tipos_retencion.get(tipo_codigo) or tipos_retencion.get(tipo_default)
    if tipo is None:
        # Caso extremo: ni el default existe — devolver sin retención
        return RetencionesCalculadas(
            rete_renta=0.0, cuenta_renta="", base_renta=0.0,
            rete_iva=0.0, base_reteiva=0.0, rete_ica=_redondear(doc.rete_ica),
            cuenta_iva_descontable=CUENTA_IVA_COMPRAS,
            fuente="sin_retencion", tipo_aplicado="", advertencia=advertencia,
        )

    # ----- RETEFUENTE -----
    rete_renta = 0.0
    base_renta = 0.0
    cuenta_renta = ""
    if doc.rete_renta > 0:
        # TOKEN manda
        rete_renta = _redondear(doc.rete_renta)
        base_renta = _redondear(valor_subtotal)
        cuenta_renta = tipo.cuenta_renta or CUENTA_RETEFUENTE_JURIDICA
        fuente_partes.append("token_renta")
    elif aplicar_calculo and tipo.tasa_renta > 0 and tipo.cuenta_renta:
        # Calcular desde catálogo si supera la base mínima
        base_min = tipo.base_min_uvt * uvt
        if valor_subtotal > base_min:
            rete_renta = _redondear(valor_subtotal * tipo.tasa_renta)
            base_renta = _redondear(valor_subtotal)
            cuenta_renta = tipo.cuenta_renta
            fuente_partes.append("calc_renta")

    # ----- RETEIVA -----
    rete_iva = 0.0
    base_reteiva = 0.0
    if doc.rete_iva > 0:
        # TOKEN manda
        rete_iva = _redondear(doc.rete_iva)
        base_reteiva = _redondear(valor_iva)
        fuente_partes.append("token_iva")
    elif aplicar_calculo and tipo.aplica_reteiva and tipo.tasa_reteiva > 0 and valor_iva > 0:
        # Calcular desde catálogo
        rete_iva = _redondear(valor_iva * tipo.tasa_reteiva)
        base_reteiva = _redondear(valor_iva)
        fuente_partes.append("calc_iva")

    # ----- RETEICA -----
    # Siempre del TOKEN — depende del municipio, no del tipo de proveedor
    rete_ica = _redondear(doc.rete_ica)
    if rete_ica > 0:
        fuente_partes.append("token_ica")

    fuente = "+".join(fuente_partes) if fuente_partes else "sin_retencion"

    return RetencionesCalculadas(
        rete_renta=rete_renta,
        cuenta_renta=cuenta_renta,
        base_renta=base_renta,
        rete_iva=rete_iva,
        base_reteiva=base_reteiva,
        rete_ica=rete_ica,
        cuenta_iva_descontable=tipo.cuenta_iva_descontable or CUENTA_IVA_COMPRAS,
        fuente=fuente,
        tipo_aplicado=tipo.codigo,
        advertencia=advertencia,
    )


# ============================================================
# Generación: Comprobante 3 (Compras DIAN) y 7 (NC)
# ============================================================

def _generar_asientos_token(
    docs: List[DocumentoToken],
    anio: int,
    mes: int,
    proveedores_consumo: Dict[str, ProveedorConsumo],
    tipos_retencion: Dict[str, TipoRetencion],
    proveedores_retencion: Dict[str, ProveedorRetencion],
    tipo_retencion_default: str,
    uvt: int,
    aplicar_calculo_retenciones: bool = True,
    consecutivo_inicial_c3: int = 1,
    consecutivo_inicial_c7: int = 1,
) -> Tuple[List[dict], List[str], List[dict]]:
    """Genera asientos para Compras (Comp 3) y NC Proveedores (Comp 7).

    Numeración INDEPENDIENTE por comprobante (v3.2):
      - Comp 3 (facturas): YYYYMMNNN empezando en consecutivo_inicial_c3
      - Comp 7 (NC):       YYYYMMNNN empezando en consecutivo_inicial_c7
      - Cada uno ordenado por (fecha, folio).

    Aplica retenciones automáticas (FASE 2):
      - Si TOKEN trae rete_renta/rete_iva/rete_ica > 0 → se respeta.
      - Si no → se calcula desde catálogo según tipo del proveedor.
      - Default 'juridica' (2.5%) si NIT no está catalogado.

    Returns:
        (filas, advertencias, log_retenciones)
        log_retenciones: lista de dicts con detalle por documento para auditoría.
    """
    filas: List[dict] = []
    advertencias: List[str] = []
    log_retenciones: List[dict] = []

    # Separar facturas y NC y ordenarlas por separado
    facturas = sorted(
        [d for d in docs if d.es_factura],
        key=lambda d: (d.fecha, d.folio),
    )
    notas_credito = sorted(
        [d for d in docs if d.es_nota_credito],
        key=lambda d: (d.fecha, d.folio),
    )

    # Concatenar con sus respectivos consecutivos iniciales
    docs_con_idx: List[Tuple[DocumentoToken, int]] = []
    for i, doc in enumerate(facturas):
        docs_con_idx.append((doc, consecutivo_inicial_c3 + i))
    for i, doc in enumerate(notas_credito):
        docs_con_idx.append((doc, consecutivo_inicial_c7 + i))

    for doc, consecutivo_n in docs_con_idx:
        documento = f"{anio:04d}{mes:02d}{consecutivo_n:03d}"
        folio_completo = doc.folio_completo
        nit = doc.nit_emisor or NIT_GENERICO
        nombre = doc.nombre_emisor.upper() if doc.nombre_emisor else ""

        valor_total = _redondear(doc.total)
        valor_iva = _redondear(doc.iva)
        valor_inc = _redondear(doc.inc)
        valor_subtotal = valor_total - valor_iva - valor_inc

        # Inferir impuesto al consumo si proveedor está en lista
        if doc.es_factura and nit in proveedores_consumo and valor_inc == 0 and valor_iva == 0:
            tasa = proveedores_consumo[nit].tasa
            valor_subtotal = _redondear(doc.total / (1 + tasa))
            valor_inc = valor_total - valor_subtotal

        # Calcular retenciones (TOKEN manda, catálogo es fallback si flag activo)
        retenciones = _calcular_retenciones(
            doc=doc,
            valor_subtotal=valor_subtotal,
            valor_iva=valor_iva,
            tipos_retencion=tipos_retencion,
            proveedores_retencion=proveedores_retencion,
            tipo_default=tipo_retencion_default,
            uvt=uvt,
            aplicar_calculo=aplicar_calculo_retenciones,
        )
        if retenciones.advertencia:
            advertencias.append(retenciones.advertencia)

        total_retenciones = (
            retenciones.rete_renta + retenciones.rete_iva + retenciones.rete_ica
        )

        # Log de auditoría por documento
        if doc.es_factura and (total_retenciones > 0 or retenciones.fuente != "sin_retencion"):
            log_retenciones.append({
                "documento": documento,
                "folio": folio_completo,
                "nit": nit,
                "nombre": nombre,
                "subtotal": int(valor_subtotal),
                "iva": int(valor_iva),
                "rete_renta": int(retenciones.rete_renta),
                "cuenta_renta": retenciones.cuenta_renta,
                "rete_iva": int(retenciones.rete_iva),
                "rete_ica": int(retenciones.rete_ica),
                "tipo": retenciones.tipo_aplicado,
                "fuente": retenciones.fuente,
            })

        if doc.es_factura:
            # === FACTURA — Comp 3 ===
            # Db cuenta de gasto/costo (subtotal)
            cuenta_compra = CUENTA_COMPRAS_GRAVADAS if valor_iva > 0 else CUENTA_COMPRAS_NO_GRAVADAS

            filas.append(_crear_linea_plano(
                cuenta=cuenta_compra,
                comprobante=COMPROBANTE_COMPRAS,
                fecha=doc.fecha,
                documento=documento,
                doc_referencia=folio_completo,
                nit=nit,
                detalle=f"COMPRA FE {folio_completo} - {nombre}",
                tr=TR_DEBITO,
                valor=valor_subtotal,
            ))

            # Db IVA si aplica — cuenta según tipo de operación:
            # 24080201 para bienes (juridica/natural)
            # 24080308 para servicios (honorarios/comisiones/arrendamiento/servicios_natural)
            if valor_iva > 0:
                filas.append(_crear_linea_plano(
                    cuenta=retenciones.cuenta_iva_descontable,
                    comprobante=COMPROBANTE_COMPRAS,
                    fecha=doc.fecha,
                    documento=documento,
                    doc_referencia=folio_completo,
                    nit=nit,
                    detalle=f"IVA FE {folio_completo} - {nombre}",
                    tr=TR_DEBITO,
                    valor=valor_iva,
                    base=valor_subtotal,
                ))

            # Db Impuesto al Consumo si aplica
            if valor_inc > 0:
                filas.append(_crear_linea_plano(
                    cuenta=CUENTA_IMPUESTO_CONSUMO,
                    comprobante=COMPROBANTE_COMPRAS,
                    fecha=doc.fecha,
                    documento=documento,
                    doc_referencia=folio_completo,
                    nit=nit,
                    detalle=f"IMP CONSUMO FE {folio_completo} - {nombre}",
                    tr=TR_DEBITO,
                    valor=valor_inc,
                ))

            # Cr Retefuente (si aplica)
            if retenciones.rete_renta > 0 and retenciones.cuenta_renta:
                filas.append(_crear_linea_plano(
                    cuenta=retenciones.cuenta_renta,
                    comprobante=COMPROBANTE_COMPRAS,
                    fecha=doc.fecha,
                    documento=documento,
                    doc_referencia=folio_completo,
                    nit=nit,
                    detalle=f"RETEFUENTE {retenciones.tipo_aplicado.upper()} FE {folio_completo}",
                    tr=TR_CREDITO,
                    valor=retenciones.rete_renta,
                    base=retenciones.base_renta,
                ))

            # Cr ReteIVA (si aplica)
            if retenciones.rete_iva > 0:
                filas.append(_crear_linea_plano(
                    cuenta=CUENTA_RETEIVA,
                    comprobante=COMPROBANTE_COMPRAS,
                    fecha=doc.fecha,
                    documento=documento,
                    doc_referencia=folio_completo,
                    nit=nit,
                    detalle=f"RETEIVA FE {folio_completo} - {nombre}",
                    tr=TR_CREDITO,
                    valor=retenciones.rete_iva,
                    base=retenciones.base_reteiva,
                ))

            # Cr ICA (si TOKEN lo trae)
            # Nota: la cuenta de ICA depende del municipio y no la inferimos.
            # Si llegase ICA del TOKEN sin cuenta definida, NO se asienta y NO se descuenta del proveedor.
            ica_asentado = 0.0
            if retenciones.rete_ica > 0:
                advertencias.append(
                    f"⚠️ Doc {folio_completo} trae ReteICA ${int(retenciones.rete_ica):,}. "
                    f"Cuenta de ICA no está definida en catálogo — NO se asienta. "
                    f"Revisar manualmente."
                    .replace(",", ".")
                )

            # Cr proveedor por el total NETO (total - retenciones efectivamente asentadas)
            retenciones_asentadas = retenciones.rete_renta + retenciones.rete_iva + ica_asentado
            valor_proveedor = valor_total - retenciones_asentadas
            filas.append(_crear_linea_plano(
                cuenta=CUENTA_PROVEEDORES,
                comprobante=COMPROBANTE_COMPRAS,
                fecha=doc.fecha,
                documento=documento,
                doc_referencia=folio_completo,
                nit=nit,
                detalle=f"PROVEEDOR FE {folio_completo} - {nombre}",
                tr=TR_CREDITO,
                valor=valor_proveedor,
            ))

        elif doc.es_nota_credito:
            # === NOTA CRÉDITO — Comp 7 ===
            # Si la NC tiene retenciones, hay que reversarlas también.
            # Importante: solo reversamos las que efectivamente se asentaron como Cr en su día.
            # El ICA no se asienta (no hay cuenta), así que tampoco se reversa.
            retenciones_asentadas_nc = retenciones.rete_renta + retenciones.rete_iva
            valor_proveedor_nc = valor_total - retenciones_asentadas_nc
            filas.append(_crear_linea_plano(
                cuenta=CUENTA_PROVEEDORES,
                comprobante=COMPROBANTE_NC_PROVEEDORES,
                fecha=doc.fecha,
                documento=documento,
                doc_referencia=folio_completo,
                nit=nit,
                detalle=f"NC {folio_completo} PROVEEDOR - {nombre}",
                tr=TR_DEBITO,
                valor=valor_proveedor_nc,
            ))

            # Db reverso de retefuente (si aplica)
            if retenciones.rete_renta > 0 and retenciones.cuenta_renta:
                filas.append(_crear_linea_plano(
                    cuenta=retenciones.cuenta_renta,
                    comprobante=COMPROBANTE_NC_PROVEEDORES,
                    fecha=doc.fecha,
                    documento=documento,
                    doc_referencia=folio_completo,
                    nit=nit,
                    detalle=f"NC {folio_completo} REVERSO RETEFUENTE",
                    tr=TR_DEBITO,
                    valor=retenciones.rete_renta,
                    base=retenciones.base_renta,
                ))

            # Db reverso de ReteIVA (si aplica)
            if retenciones.rete_iva > 0:
                filas.append(_crear_linea_plano(
                    cuenta=CUENTA_RETEIVA,
                    comprobante=COMPROBANTE_NC_PROVEEDORES,
                    fecha=doc.fecha,
                    documento=documento,
                    doc_referencia=folio_completo,
                    nit=nit,
                    detalle=f"NC {folio_completo} REVERSO RETEIVA",
                    tr=TR_DEBITO,
                    valor=retenciones.rete_iva,
                    base=retenciones.base_reteiva,
                ))

            # ICA en NC: igual que en factura, no se reversa si no hay cuenta
            if retenciones.rete_ica > 0:
                advertencias.append(
                    f"⚠️ NC {folio_completo} trae ReteICA ${int(retenciones.rete_ica):,}. "
                    f"Cuenta de ICA no está definida — NO se reversa. Revisar manualmente."
                    .replace(",", ".")
                )

            # Cr 622505 (devolución en compras) por subtotal
            if valor_subtotal > 0:
                filas.append(_crear_linea_plano(
                    cuenta=CUENTA_DEVOLUCION_COMPRAS,
                    comprobante=COMPROBANTE_NC_PROVEEDORES,
                    fecha=doc.fecha,
                    documento=documento,
                    doc_referencia=folio_completo,
                    nit=nit,
                    detalle=f"NC {folio_completo} REVERSA - {nombre}",
                    tr=TR_CREDITO,
                    valor=valor_subtotal,
                ))

            # Cr IVA reverso
            if valor_iva > 0:
                filas.append(_crear_linea_plano(
                    cuenta=CUENTA_IVA_NC,
                    comprobante=COMPROBANTE_NC_PROVEEDORES,
                    fecha=doc.fecha,
                    documento=documento,
                    doc_referencia=folio_completo,
                    nit=nit,
                    detalle=f"NC {folio_completo} IVA - {nombre}",
                    tr=TR_CREDITO,
                    valor=valor_iva,
                    base=valor_subtotal,
                ))

    return filas, advertencias, log_retenciones


# ============================================================
# Generación: Comprobante 18 (DS)
# ============================================================

def _generar_asientos_ds(
    docs: List[DocumentoSoporte],
    productos_map: Dict[str, ProductoDS],
    base_minima_retefuente: int,
) -> Tuple[List[dict], Dict[str, str], List[str]]:
    filas: List[dict] = []
    mapa: Dict[str, str] = {}
    advertencias: List[str] = []

    for ds in docs:
        prod = productos_map.get(ds.producto_codigo)
        if prod is None:
            advertencias.append(
                f"⚠️ {ds.consecutivo}: producto '{ds.producto_codigo}' no está en "
                f"catálogo. Usando defaults (519530 → 23359501)."
            )
            cta_db = CUENTA_DB_DEFAULT_DS
            cta_cr_prov = CUENTA_CR_DEFAULT_DS
            retiene = False
        else:
            cta_db = prod.cuenta_db
            cta_cr_prov = prod.cuenta_cr_proveedor
            retiene = prod.retiene_servicios

        valor_total = _redondear(ds.total)
        nit = ds.identificacion or NIT_GENERICO
        documento = ds.consecutivo
        doc_ref = ds.consecutivo
        detalle_db = ds.producto_completo

        # Db gasto/costo
        filas.append(_crear_linea_plano(
            cuenta=cta_db,
            comprobante=COMPROBANTE_DS,
            fecha=ds.fecha,
            documento=documento,
            doc_referencia=doc_ref,
            nit=nit,
            detalle=detalle_db,
            tr=TR_DEBITO,
            valor=valor_total,
        ))

        # ¿Aplica retefuente? Solo cálculo Cr
        retefuente = 0
        if retiene and valor_total > base_minima_retefuente:
            retefuente = _redondear(valor_total * TASA_RETEFUENTE_SERVICIOS)
            if retefuente > 0:
                filas.append(_crear_linea_plano(
                    cuenta=CUENTA_RETEFUENTE_SERVICIOS,
                    comprobante=COMPROBANTE_DS,
                    fecha=ds.fecha,
                    documento=documento,
                    doc_referencia=doc_ref,
                    nit=nit,
                    detalle=f"RETEFUENTE COMPRA_NATURAL - {ds.proveedor.upper()}",
                    tr=TR_CREDITO,
                    valor=retefuente,
                    base=valor_total,
                ))

        # Cr proveedor
        valor_proveedor = valor_total - retefuente
        if valor_proveedor > 0:
            filas.append(_crear_linea_plano(
                cuenta=cta_cr_prov,
                comprobante=COMPROBANTE_DS,
                fecha=ds.fecha,
                documento=documento,
                doc_referencia=doc_ref,
                nit=nit,
                detalle="PROVEEDORES",
                tr=TR_CREDITO,
                valor=valor_proveedor,
            ))

        # Mapeo CEG↔DS: cuenta del proveedor (NO retefuente)
        mapa[ds.consecutivo] = cta_cr_prov

    return filas, mapa, advertencias


# ============================================================
# Generación: Comprobante 13 (CEG)
# ============================================================

def _generar_asientos_ceg(
    egresos: List[EgresoCajaMenor],
    mapa_ds_cuenta_cr: Dict[str, str],
    bancos: List[BancoConfig],
) -> Tuple[List[dict], List[str]]:
    filas: List[dict] = []
    advertencias: List[str] = []
    bancos_map = {b.nit: b for b in bancos}

    for eg in egresos:
        valor = _redondear(eg.valor)
        nit = eg.identificacion or NIT_GENERICO
        documento = eg.consecutivo
        cancela = eg.cancela.strip().upper()
        es_cruce_ds = cancela.startswith("DS")
        doc_referencia = eg.cancela if es_cruce_ds else eg.consecutivo

        if es_cruce_ds:
            cta_db = mapa_ds_cuenta_cr.get(eg.cancela)
            if cta_db is None:
                advertencias.append(
                    f"⚠️ {eg.consecutivo} cancela {eg.cancela}, pero ese DS no se encontró. "
                    f"Usando 23359501 por defecto."
                )
                cta_db = CUENTA_ACREEDORES
        elif nit in bancos_map:
            cta_db = bancos_map[nit].cuenta
        else:
            cta_db = CUENTA_ACREEDORES

        nombre_limpio = eg.nombres.strip() if eg.nombres else ""
        detalle = nombre_limpio.upper() if nombre_limpio else f"EGRESO {eg.consecutivo}"

        # Db
        filas.append(_crear_linea_plano(
            cuenta=cta_db,
            comprobante=COMPROBANTE_CEG,
            fecha=eg.fecha,
            documento=documento,
            doc_referencia=doc_referencia,
            nit=nit,
            detalle=detalle,
            tr=TR_DEBITO,
            valor=valor,
            centro_costo=eg.sucursal or CENTRO_COSTOS_DEFAULT,
        ))
        # Cr Caja
        filas.append(_crear_linea_plano(
            cuenta=CUENTA_CAJA,
            comprobante=COMPROBANTE_CEG,
            fecha=eg.fecha,
            documento=documento,
            doc_referencia=doc_referencia,
            nit=nit,
            detalle=detalle,
            tr=TR_CREDITO,
            valor=valor,
            centro_costo=eg.sucursal or CENTRO_COSTOS_DEFAULT,
        ))

    return filas, advertencias


# ============================================================
# Función principal
# ============================================================

def procesar_compras_y_egresos(
    archivo_token=None,
    archivo_ds=None,
    archivo_egresos=None,
    anio: int = 2026,
    mes: int = 3,
    consecutivo_inicial_c3: int = 1,
    consecutivo_inicial_c7: int = 1,
    consecutivo_token_inicial: Optional[int] = None,  # alias retrocompat
) -> Tuple[pd.DataFrame, List[str], dict]:
    """Procesa los 3 archivos y genera el plano contable unificado.

    Args:
        archivo_token: TOKEN DIAN (.xlsx) — genera Comp 3 y Comp 7
        archivo_ds: DOCUMENTO_SOPORTE (.xlsx) — genera Comp 18
        archivo_egresos: EGRESOS_CAJA_MENOR (.xls/.html) — genera Comp 13
        anio, mes: para construir el consecutivo YYYYMMNNN del Comp 3 y Comp 7
        consecutivo_inicial_c3: número desde el cual empezar Comp 3 (default 1)
        consecutivo_inicial_c7: número desde el cual empezar Comp 7 (default 1)
        consecutivo_token_inicial: ALIAS retrocompatible — si se pasa, se usa
            como inicial para AMBOS comp 3 y 7.

    Returns:
        (df_plano, log, resumen)
    """
    # Retrocompat: si pasaron el parámetro viejo, lo respetamos
    if consecutivo_token_inicial is not None:
        consecutivo_inicial_c3 = consecutivo_token_inicial
        consecutivo_inicial_c7 = consecutivo_token_inicial

    log: List[str] = []
    todas_filas: List[dict] = []
    log_retenciones: List[dict] = []

    # Cargar catálogo
    cat = cargar_catalogo()
    base_minima_retefuente = cat["uvt"] * cat["base_servicios_uvt"]
    log.append(f"📋 Catálogo: {len(cat['productos'])} productos DS, {len(cat['bancos'])} bancos, "
               f"{len(cat['proveedores_retencion'])} proveedores con tipo retención.")
    log.append(f"   UVT 2026: ${cat['uvt']:,} | Base mínima retefuente servicios: ${base_minima_retefuente:,}".replace(",","."))
    log.append(f"   Tipo de retención default: '{cat['tipo_retencion_default']}'")
    modo_calc = "ACTIVADO (calcula desde catálogo)" if cat["aplicar_calculo_retenciones"] else "DESACTIVADO (solo respeta TOKEN)"
    log.append(f"   Cálculo automático de retenciones: {modo_calc}")

    mapa_ds_a_cr: Dict[str, str] = {}

    # 1) Documentos Soporte (Comp 18) PRIMERO para generar mapa
    n_ds = 0
    if archivo_ds is not None:
        log.append("")
        log.append("📂 Leyendo DOCUMENTO SOPORTE...")
        ds_list = _leer_documentos_soporte(archivo_ds)
        n_ds = len(ds_list)
        log.append(f"   Documentos DS leídos: {n_ds}")
        if ds_list:
            filas_ds, mapa_ds_a_cr, adv_ds = _generar_asientos_ds(
                ds_list, cat["productos"], base_minima_retefuente
            )
            todas_filas.extend(filas_ds)
            log.extend(adv_ds)
            log.append(f"   Líneas plano C18: {len(filas_ds)}")

    # 2) TOKEN DIAN (Comp 3 + Comp 7)
    n_facturas = n_nc = 0
    if archivo_token is not None:
        log.append("")
        log.append("📂 Leyendo TOKEN DIAN...")
        token_docs = _leer_token(archivo_token)
        n_facturas = sum(1 for d in token_docs if d.es_factura)
        n_nc = sum(1 for d in token_docs if d.es_nota_credito)
        log.append(f"   Facturas leídas: {n_facturas}")
        log.append(f"   Notas crédito leídas: {n_nc}")

        if token_docs:
            filas_token, adv_token, log_retenciones = _generar_asientos_token(
                token_docs,
                anio,
                mes,
                cat["proveedores_consumo"],
                cat["tipos_retencion"],
                cat["proveedores_retencion"],
                cat["tipo_retencion_default"],
                cat["uvt"],
                cat["aplicar_calculo_retenciones"],
                consecutivo_inicial_c3=consecutivo_inicial_c3,
                consecutivo_inicial_c7=consecutivo_inicial_c7,
            )
            todas_filas.extend(filas_token)
            log.extend(adv_token)
            n_c3 = sum(1 for f in filas_token if f["COMPROBANTE"] == COMPROBANTE_COMPRAS)
            n_c7 = sum(1 for f in filas_token if f["COMPROBANTE"] == COMPROBANTE_NC_PROVEEDORES)
            log.append(f"   Líneas plano C3: {n_c3} (desde consecutivo {consecutivo_inicial_c3:03d})")
            log.append(f"   Líneas plano C7: {n_c7} (desde consecutivo {consecutivo_inicial_c7:03d})")

            # Resumen de retenciones aplicadas
            if log_retenciones:
                total_renta = sum(r["rete_renta"] for r in log_retenciones)
                total_riva = sum(r["rete_iva"] for r in log_retenciones)
                total_rica = sum(r["rete_ica"] for r in log_retenciones)
                docs_con_rete = sum(
                    1 for r in log_retenciones
                    if r["rete_renta"] > 0 or r["rete_iva"] > 0 or r["rete_ica"] > 0
                )
                docs_token = sum(1 for r in log_retenciones if "token" in r["fuente"])
                docs_calc = sum(1 for r in log_retenciones if "calc" in r["fuente"])
                log.append(
                    f"   💰 Retenciones: {docs_con_rete} docs | "
                    f"Rte=${total_renta:,} | RIVA=${total_riva:,} | RICA=${total_rica:,}".replace(",", ".")
                )
                log.append(
                    f"      Fuente: {docs_token} desde TOKEN, {docs_calc} calculadas desde catálogo"
                )

    # 3) Egresos Caja Menor (Comp 13)
    n_ceg = 0
    if archivo_egresos is not None:
        log.append("")
        log.append("📂 Leyendo EGRESOS CAJA MENOR...")
        egresos = _leer_egresos_caja_menor(archivo_egresos)
        n_ceg = len(egresos)
        log.append(f"   Egresos CEG leídos: {n_ceg}")
        if egresos:
            filas_ceg, adv_ceg = _generar_asientos_ceg(
                egresos, mapa_ds_a_cr, cat["bancos"]
            )
            todas_filas.extend(filas_ceg)
            log.extend(adv_ceg)
            log.append(f"   Líneas plano C13: {len(filas_ceg)}")

    df_plano = pd.DataFrame(todas_filas, columns=COLUMNAS_PLANO)

    # Verificar que todos los valores sean positivos
    if len(df_plano) > 0:
        n_negativos = (df_plano["VALOR"] < 0).sum()
        if n_negativos > 0:
            log.append(f"⚠️ {n_negativos} líneas con valor negativo (debería ser 0)!")

    # Resumen
    log.append("")
    log.append("=" * 60)
    log.append("📊 RESUMEN GENERAL")
    log.append("=" * 60)
    if len(df_plano) > 0:
        total_db = int(df_plano[df_plano["TR"] == TR_DEBITO]["VALOR"].sum())
        total_cr = int(df_plano[df_plano["TR"] == TR_CREDITO]["VALOR"].sum())
        log.append(f"   Total líneas: {len(df_plano)}")
        log.append(f"   Total Db: $ {total_db:,}".replace(",","."))
        log.append(f"   Total Cr: $ {total_cr:,}".replace(",","."))
        if total_db == total_cr:
            log.append("   ✅ Cuadre perfecto Db = Cr")
        else:
            log.append(f"   ❌ DESCUADRE: dif $ {total_db - total_cr:,}".replace(",","."))

        log.append("")
        log.append("   Asientos por comprobante:")
        for comp, sub in df_plano.groupby("COMPROBANTE"):
            db = int(sub[sub["TR"] == TR_DEBITO]["VALOR"].sum())
            cr = int(sub[sub["TR"] == TR_CREDITO]["VALOR"].sum())
            n_docs = sub["DOCUMENTO"].nunique()
            estado = "✅" if db == cr else "❌"
            log.append(f"      {estado} Comp {comp}: {n_docs} docs, {len(sub)} líneas, "
                       f"Db=Cr=${db:,}".replace(",","."))

    resumen = {
        "total_lineas": len(df_plano),
        "n_facturas": n_facturas,
        "n_nc_proveedores": n_nc,
        "n_ds": n_ds,
        "n_egresos": n_ceg,
        "comprobantes": sorted(df_plano["COMPROBANTE"].unique().tolist()) if len(df_plano) > 0 else [],
        "retenciones_detalle": log_retenciones,
        "total_retefuente": sum(r["rete_renta"] for r in log_retenciones),
        "total_reteiva": sum(r["rete_iva"] for r in log_retenciones),
        "total_reteica": sum(r["rete_ica"] for r in log_retenciones),
    }
    return df_plano, log, resumen


# ============================================================
# Exportación
# ============================================================

def dataframe_a_plano_tsv(
    df: pd.DataFrame,
    incluir_encabezado_excel: bool = True,
) -> bytes:
    df_out = df[COLUMNAS_PLANO].copy()
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str).str.replace("\t", " ", regex=False)
    tsv = df_out.to_csv(sep="\t", index=False, header=True, lineterminator="\r\n")
    if incluir_encabezado_excel:
        tsv = "sep=\t\r\n" + tsv
    return tsv.encode("utf-8")
