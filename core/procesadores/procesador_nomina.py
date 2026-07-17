"""
Procesador de Nómina mensual.

Toma un Excel con UNA HOJA POR EMPLEADA, donde cada hoja contiene
DOS desprendibles apilados verticalmente (quincena 1 y quincena 2).
Genera el plano contable con:

    - 2 asientos de CAUSACIÓN DE NÓMINA (uno por quincena, comprobante 11)
    - 1 asiento de PROVISIÓN MENSUAL (aportes patronales + prestaciones, comprobante 9)

Estructura esperada del Excel (basada en el formato manual de Casa UnoTres SAS):
    - 1 hoja por empleada
    - Cada hoja tiene 2 desprendibles (Q1 y Q2)
    - Cada desprendible identifica:
        * NOMBRE: nombre completo del empleado
        * C.C.: cédula
        * PERIODO: rango de fechas
        * NUMERO: 1 (Q1) o 2 (Q2)
        * SALARIO BASICO: valor mensual
    - Sección DEVENGADO con conceptos: Salario Básico, Subsidio de transporte,
      Horas extras, Horas ordinarias dominicales, Horas nocturnas,
      Incapacidad, Otros, Prima servicios, Intereses cesantías
    - Sección DEDUCCIONES con: Salud, Pensión, Otros
    - Total Devengado, Total Deducciones, Neto a Pagar

Reglas contables (Casa UnoTres SAS):
    - IBC seguridad social = Devengado − Aux. Transporte − Incapacidad
    - Base prestaciones    = Total Devengado (todo)
    - Base vacaciones      = Devengado − Aux. Transporte
    - Empresa exonerada de Salud patronal 8.5%, SENA 2%, ICBF 3%
    - Solo aplica: Pensión 12%, ARL (variable), Caja Compensación 4%
    - Las DEDUCCIONES del empleado se LEEN del Excel (no se recalculan)

Cuentas contables (por tipo de empleado):
    - ADMIN     → serie 510xxx
    - OPERACION → serie 520xxx
"""
from __future__ import annotations
import io
import json
import re
import calendar
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Iterable

import pandas as pd
from openpyxl import load_workbook


# ============================================================
# Constantes y configuración contable
# ============================================================

# Comprobantes
COMPROBANTE_NOMINA = "11"      # causación quincenal de nómina
COMPROBANTE_PROVISION = "9"    # provisión mensual de aportes y prestaciones

# Detalles del plano
DETALLE_NOMINA_Q1 = "NOMINA QUINCENA 1"
DETALLE_NOMINA_Q2 = "NOMINA QUINCENA 2"
DETALLE_PROVISION = "PROVISION NOMINA"

# Tasas legales (Colombia 2026)
TASA_PENSION_PATRONAL = 0.12       # 12%
TASA_CAJA_COMPENSACION = 0.04      # 4%
TASA_ARL_DEFAULT = 0.00522         # Clase I (fallback si no se encuentra empleada)
TASA_CESANTIAS = 0.0833            # 8.33%
TASA_INTERESES_CESANTIAS = 0.12    # 12% anual sobre cesantías
TASA_PRIMA_SERVICIOS = 0.0833      # 8.33%
TASA_VACACIONES = 0.0417           # 4.17%

# Columnas finales del plano (orden exacto de salida — mismo formato que POS/DIAN)
COLUMNAS_PLANO = [
    "CUENTA",
    "COMPROBANTE",
    "FECHA",
    "DOCUMENTO",
    "DOC REFERENCIA",
    "NIT",
    "DETALLE",
    "TR",
    "VALOR",
    "BASE",
    "CENTRO DE COSTO",
]

# ============================================================
# Plan de cuentas (parametrizado por tipo: ADMIN o OPERACION)
# ============================================================
#
# La estructura permite agregar fácilmente otros tipos en el futuro.
# Las cuentas de PASIVO son únicas (no dependen del tipo).
#
# ADMIN     → serie 510xxx (Gastos de administración)
# OPERACION → serie 520xxx (Gastos de operación / producción)

CUENTAS_GASTO: Dict[str, Dict[str, str]] = {
    "ADMIN": {
        "sueldo":          "510506",
        "aux_transporte":  "510527",
        "horas_extras":    "510515",
        "otros":           "510518",
        "incapacidad":     "510524",
        # Aportes patronales
        "pension":         "510470",
        "arl":             "510468",
        "caja":            "510472",
        # Prestaciones
        "cesantias":       "510530",
        "int_cesantias":   "510533",
        "prima":           "510536",
        "vacaciones":      "510539",
    },
    "OPERACION": {
        "sueldo":          "520506",
        "aux_transporte":  "520527",
        "horas_extras":    "520515",
        "otros":           "520518",
        "incapacidad":     "520524",
        "pension":         "520470",
        "arl":             "520468",
        "caja":            "520472",
        "cesantias":       "520530",
        "int_cesantias":   "520533",
        "prima":           "520536",
        "vacaciones":      "520539",
    },
}

# Cuentas de pasivo (mismas para todos los empleados)
CUENTAS_PASIVO: Dict[str, str] = {
    # Causación nómina
    "ded_salud":           "25500502",   # Deducción Salud trabajador
    "ded_pension":         "25503002",   # Aporte Pensión trabajador
    "ded_otros":           "25500502",   # 'Otros' del Excel (cae en deducción salud por defecto)
    "salarios_por_pagar":  "25050501",   # NETO a pagar al empleado
    # Provisión aportes patronales
    "prov_pension":        "25503001",
    "prov_arl":            "25500601",
    "prov_caja":           "25501001",
    # Provisión prestaciones
    "prov_cesantias":      "25300500",
    "prov_int_cesantias":  "25301000",
    "prov_prima":          "25302000",
    "prov_vacaciones":     "25301500",
}


# ============================================================
# Estructuras de datos
# ============================================================

@dataclass
class Empleado:
    """Maestro embebido de empleados (cargado desde empleados.json)."""
    cc: str
    nombre: str
    tipo: str               # 'ADMIN' o 'OPERACION'
    tasa_arl: float
    fondo_pension: str = ""
    eps: str = ""
    arl: str = ""
    caja: str = ""

    def cuentas_gasto(self) -> Dict[str, str]:
        """Retorna el diccionario de cuentas de gasto según el tipo."""
        if self.tipo not in CUENTAS_GASTO:
            raise ValueError(
                f"Tipo de empleado '{self.tipo}' no reconocido para "
                f"empleado {self.cc} ({self.nombre}). "
                f"Usa: {', '.join(CUENTAS_GASTO.keys())}."
            )
        return CUENTAS_GASTO[self.tipo]


@dataclass
class Quincena:
    """Datos extraídos de UN desprendible (UNA quincena de UNA empleada)."""
    cc: str
    nombre_excel: str         # nombre como viene en el Excel
    numero: int               # 1 o 2
    fecha_inicio: Optional[date] = None
    fecha_fin: Optional[date] = None
    salario_basico_mes: float = 0.0
    # Devengado
    salario_basico: float = 0.0
    aux_transporte: float = 0.0
    horas_extras_recargos: float = 0.0
    incapacidad: float = 0.0
    otros: float = 0.0
    total_devengado: float = 0.0
    # Deducciones (LEÍDAS del Excel, no calculadas)
    ded_salud: float = 0.0
    ded_pension: float = 0.0
    ded_otros: float = 0.0
    total_deducciones: float = 0.0
    # Neto
    neto: float = 0.0


@dataclass
class ResumenEmpleadoMes:
    """Resumen mensual de UN empleado (sumando Q1 + Q2) para cálculo de provisiones."""
    cc: str
    nombre: str
    devengado_total: float = 0.0
    aux_transporte_total: float = 0.0
    incapacidad_total: float = 0.0

    @property
    def ibc(self) -> float:
        """IBC seguridad social = Devengado − Aux Transporte − Incapacidad."""
        return self.devengado_total - self.aux_transporte_total - self.incapacidad_total

    @property
    def base_prestaciones(self) -> float:
        """Base de cesantías, intereses y prima = Total Devengado."""
        return self.devengado_total

    @property
    def base_vacaciones(self) -> float:
        """Base de vacaciones = Devengado − Aux Transporte."""
        return self.devengado_total - self.aux_transporte_total


# ============================================================
# Carga del maestro de empleados
# ============================================================

_RUTA_EMPLEADOS_JSON = (
    Path(__file__).resolve().parent.parent / "data" / "empleados.json"
)


def cargar_empleados_embebido() -> Dict[str, Empleado]:
    """Carga el maestro de empleados desde el JSON embebido.

    Returns:
        Diccionario {cc: Empleado}. La CC se normaliza (sin puntos, sin guiones).
    """
    if not _RUTA_EMPLEADOS_JSON.exists():
        raise FileNotFoundError(
            f"No se encontró el maestro de empleados: {_RUTA_EMPLEADOS_JSON}. "
            "Verifica que el archivo core/data/empleados.json esté en el repo."
        )

    try:
        with open(_RUTA_EMPLEADOS_JSON, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError) as e:
        raise ValueError(f"No se pudo leer empleados.json: {e}")

    items = data.get("empleados", [])
    if not items:
        raise ValueError("empleados.json no contiene empleados válidos.")

    empleados: Dict[str, Empleado] = {}
    for item in items:
        cc = _normalizar_cc(item.get("cc", ""))
        if not cc:
            continue
        empleados[cc] = Empleado(
            cc=cc,
            nombre=str(item.get("nombre", "")).strip().upper(),
            tipo=str(item.get("tipo", "OPERACION")).strip().upper(),
            tasa_arl=float(item.get("tasa_arl", TASA_ARL_DEFAULT)),
            fondo_pension=str(item.get("fondo_pension", "")).strip(),
            eps=str(item.get("eps", "")).strip(),
            arl=str(item.get("arl", "")).strip(),
            caja=str(item.get("caja", "")).strip(),
        )

    return empleados


def info_empleados_embebido() -> dict:
    """Retorna estadísticas del maestro embebido (para mostrar en UI)."""
    empleados = cargar_empleados_embebido()
    por_tipo: Dict[str, int] = {}
    for emp in empleados.values():
        por_tipo[emp.tipo] = por_tipo.get(emp.tipo, 0) + 1
    return {
        "total": len(empleados),
        "por_tipo": por_tipo,
        "empleados": [
            {
                "cc": e.cc,
                "nombre": e.nombre,
                "tipo": e.tipo,
                "tasa_arl": e.tasa_arl,
                "tasa_arl_pct": f"{e.tasa_arl * 100:.3f}%",
                "fondo_pension": e.fondo_pension,
                "eps": e.eps,
                "arl": e.arl,
                "caja": e.caja,
            }
            for e in empleados.values()
        ],
    }


# ============================================================
# Helpers de normalización
# ============================================================

def _normalizar_cc(valor) -> str:
    """Normaliza una cédula: solo dígitos, sin puntos ni guiones."""
    if valor is None:
        return ""
    if isinstance(valor, float) and not pd.isna(valor):
        valor = int(valor)
    s = str(valor).strip()
    return re.sub(r"\D", "", s)


def _normalizar_texto(s) -> str:
    """Quita acentos, espacios extra y pasa a mayúsculas para comparar.

    Maneja correctamente NaN/None devolviendo cadena vacía (NO 'NAN').
    """
    if s is None:
        return ""
    # Detectar NaN de pandas/numpy
    try:
        if pd.isna(s):
            return ""
    except (TypeError, ValueError):
        pass
    s = str(s).strip().upper()
    # str(nan) genera 'NAN' literal — descartar
    if s in ("NAN", "NONE", "NAT"):
        return ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"\s+", " ", s)
    return s


def _to_float(valor) -> float:
    """Convierte cualquier valor a float, robusto contra '$' y '.000'."""
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        if pd.isna(valor):
            return 0.0
        return float(valor)
    s = str(valor).strip().replace("$", "").replace(",", "").replace(" ", "")
    if not s or s == "-":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _redondear(valor: float) -> int:
    """Redondea a entero (los planos contables no usan decimales)."""
    return int(round(valor))


def _formato_fecha_plano(d: date) -> str:
    """Formato MM/DD/AAAA exigido por el plano."""
    return d.strftime("%m/%d/%Y")


def _formato_cuenta(valor) -> str:
    """Cuenta como string sin '.0'."""
    if valor is None:
        return ""
    if isinstance(valor, float) and not pd.isna(valor):
        valor = int(valor)
    return str(valor).strip()


# ============================================================
# Lectura del Excel — extracción de quincenas
# ============================================================

# Mapeo: si el texto en la columna izquierda contiene esta palabra clave,
# el valor numérico de la fila pertenece a este concepto del DEVENGADO.
# El orden importa: se busca el primer match.
PATRONES_DEVENGADO = [
    # (palabras_clave, campo)
    (["SALARIO BASICO"],                 "salario_basico"),
    (["SUBSIDIO", "TRANSPORTE"],         "aux_transporte"),
    (["AUXILIO", "TRANSPORTE"],          "aux_transporte"),
    (["HORAS EXTRA"],                    "horas_extras_recargos"),
    (["HORA EXTRA"],                     "horas_extras_recargos"),
    (["DOMINICAL"],                      "horas_extras_recargos"),
    (["NOCTURN"],                        "horas_extras_recargos"),
    (["RECARGO"],                        "horas_extras_recargos"),
    (["INCAPACIDAD"],                    "incapacidad"),
    (["RETROACTIVO"],                    "otros"),
    (["AJUSTE"],                         "otros"),
    (["BONIFICACION"],                   "otros"),
    (["BONIFICACIÓN"],                   "otros"),
    # 'Otros' debe ir al final porque es genérico
    (["OTROS"],                          "otros"),
    # Conceptos que vienen pero típicamente en 0 (los ignoramos sumándolos a otros si hay valor)
    (["PRIMA"],                          "otros"),
    (["INTERES"],                        "otros"),
]

PATRONES_DEDUCCION = [
    (["SALUD"],                          "ded_salud"),
    (["PENSION"],                        "ded_pension"),
    (["LICENCIA NO REMU"],               "ded_otros"),
    (["OTROS"],                          "ded_otros"),
]


def _coincide_patron(texto_norm: str, patron: list) -> bool:
    """True si TODAS las palabras del patrón están en el texto normalizado."""
    return all(p in texto_norm for p in patron)


def _detectar_columna_devengado(ws_data: List[List]) -> int:
    """Detecta en qué columna está el VALOR del devengado. Busca celda 'DEVENGADO'."""
    for fila in ws_data:
        for j, celda in enumerate(fila):
            if _normalizar_texto(celda) == "DEVENGADO":
                return j
    return 5  # default basado en el formato observado


def _detectar_columna_deduccion(ws_data: List[List]) -> Tuple[int, int]:
    """Detecta columnas para etiqueta y valor de DEDUCCIONES.

    Retorna (col_etiqueta, col_valor).
    """
    for i, fila in enumerate(ws_data):
        for j, celda in enumerate(fila):
            if _normalizar_texto(celda) == "DEDUCCIONES":
                # El valor está unas columnas a la derecha
                # (en el formato observado, la etiqueta está en col 7 y el valor en col 9)
                return j, j + 2
    return 7, 9


def _extraer_periodo(texto: str) -> Tuple[Optional[date], Optional[date]]:
    """Extrae fechas de un periodo tipo 'Marzo 1 de 2026 a Marzo 15 de 2026'."""
    if not texto:
        return None, None

    texto_norm = _normalizar_texto(texto)

    meses = {
        "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
        "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
        "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
    }

    # Patrón: "MES DIA DE AAAA A MES DIA DE AAAA"
    pat = r"(\w+)\s+(\d+)\s+DE\s+(\d{4})\s+A\s+(\w+)\s+(\d+)\s+DE\s+(\d{4})"
    m = re.search(pat, texto_norm)
    if m:
        try:
            mes1 = meses.get(m.group(1))
            dia1 = int(m.group(2))
            anio1 = int(m.group(3))
            mes2 = meses.get(m.group(4))
            dia2 = int(m.group(5))
            anio2 = int(m.group(6))
            if mes1 and mes2:
                return (
                    date(anio1, mes1, dia1),
                    date(anio2, mes2, dia2),
                )
        except (ValueError, TypeError):
            pass

    return None, None


def _localizar_secciones_quincena(ws_data: List[List]) -> List[int]:
    """Encuentra las filas donde EMPIEZA cada desprendible (encabezado).

    Busca filas que digan 'CASA UNOTRES' o similar como marcador de inicio.
    Returns:
        Lista de índices de fila donde empieza cada desprendible.
    """
    inicios = []
    for i, fila in enumerate(ws_data):
        for celda in fila:
            t = _normalizar_texto(celda)
            if "CASA UNOTRES" in t or "CASA TRECE" in t:
                # Solo marcar si es la primera celda con texto significativo
                inicios.append(i)
                break

    # Deduplicar (a veces el nombre de la empresa aparece en varias celdas seguidas)
    inicios_filtrados = []
    for idx in inicios:
        if not inicios_filtrados or idx - inicios_filtrados[-1] > 5:
            inicios_filtrados.append(idx)

    return inicios_filtrados


def _extraer_quincena_desde_seccion(
    ws_data: List[List],
    fila_inicio: int,
    fila_fin: int,
    nombre_hoja: str,
) -> Optional[Quincena]:
    """Extrae UNA quincena leyendo desde fila_inicio hasta fila_fin del worksheet."""

    # Recortar el bloque
    bloque = ws_data[fila_inicio:fila_fin]
    if not bloque:
        return None

    quincena = Quincena(cc="", nombre_excel="", numero=0)

    # 1) Localizar campos clave: NOMBRE, C.C., PERIODO, NÚMERO, SALARIO BÁSICO
    for fila in bloque:
        for j, celda in enumerate(fila):
            t = _normalizar_texto(celda)
            if t == "NOMBRE":
                # El valor está en alguna celda más a la derecha
                for k in range(j + 1, len(fila)):
                    val = fila[k]
                    if val is not None and str(val).strip() and _normalizar_texto(val) != "NOMBRE":
                        quincena.nombre_excel = str(val).strip().upper()
                        break
            elif t in ("C.C.", "CC", "CEDULA", "CÉDULA"):
                for k in range(j + 1, len(fila)):
                    val = fila[k]
                    if val is not None and str(val).strip():
                        cc_norm = _normalizar_cc(val)
                        if cc_norm:
                            quincena.cc = cc_norm
                            break
            elif t == "PERIODO":
                for k in range(j + 1, len(fila)):
                    val = fila[k]
                    if val is not None and str(val).strip():
                        fi, ff = _extraer_periodo(str(val))
                        if fi:
                            quincena.fecha_inicio = fi
                            quincena.fecha_fin = ff
                            break
            elif t == "NUMERO":
                # El número está en la fila SIGUIENTE en la misma columna
                # (en el formato: una fila tiene 'MES NUMERO' y la siguiente 'MAR  1')
                idx_fila = bloque.index(fila)
                if idx_fila + 1 < len(bloque):
                    fila_sig = bloque[idx_fila + 1]
                    if j < len(fila_sig):
                        val = fila_sig[j]
                        if val is not None:
                            try:
                                quincena.numero = int(_to_float(val))
                            except (ValueError, TypeError):
                                pass
            elif t == "SALARIO BASICO":
                # El valor está más a la derecha
                for k in range(j + 1, len(fila)):
                    val = fila[k]
                    v = _to_float(val)
                    if v > 0:
                        quincena.salario_basico_mes = v
                        break

    # Si no encontramos CC pero sí nombre, dejamos vacío y lo resolveremos por nombre de hoja
    if not quincena.cc and not quincena.nombre_excel:
        return None

    # 2) Detectar columnas
    col_devengado = _detectar_columna_devengado(bloque)
    _, col_deduccion_val = _detectar_columna_deduccion(bloque)

    # 3) Iterar filas y mapear conceptos
    for fila in bloque:
        # Etiqueta de devengado: típicamente en col 1, pero puede estar en col 0
        # (la hoja varía según el formato del Excel manual de cada empleada).
        # Buscamos la PRIMERA columna con texto descriptivo (que no sea solo números).
        # IMPORTANTE: las etiquetas pueden tener dígitos ('Incapacidad 66,66%').
        etiqueta_dev = ""
        for k in (1, 0, 2):
            if k < len(fila):
                txt = _normalizar_texto(fila[k])
                if not txt or len(txt) < 3:
                    continue
                # Aceptar etiqueta si tiene al menos UNA letra
                # (las puras numéricas no son etiquetas)
                if any(c.isalpha() for c in txt):
                    etiqueta_dev = txt
                    break

        etiqueta_ded = ""
        # La etiqueta de deducción puede estar en col 7 (formato estándar)
        # o en col 6 (variante)
        for k in (7, 6, 8):
            if k < len(fila):
                txt = _normalizar_texto(fila[k])
                if not txt or len(txt) < 3:
                    continue
                if any(c.isalpha() for c in txt):
                    etiqueta_ded = txt
                    break

        # Devengado
        if etiqueta_dev:
            # Saltar totales para no doblar
            if "TOTAL DEVENGADO" in etiqueta_dev or "NETO" in etiqueta_dev:
                # Capturar Total Devengado y Neto
                if "TOTAL DEVENGADO" in etiqueta_dev:
                    if col_devengado < len(fila):
                        v = _to_float(fila[col_devengado])
                        if v > 0:
                            quincena.total_devengado = v
                if "NETO" in etiqueta_dev:
                    # NETO está en otra columna (típicamente justo después de '$')
                    for k in range(2, len(fila)):
                        v = _to_float(fila[k])
                        if v > 1000:  # los netos son > 1000
                            quincena.neto = v
                            break
            else:
                for patrones, campo in PATRONES_DEVENGADO:
                    if _coincide_patron(etiqueta_dev, patrones):
                        if col_devengado < len(fila):
                            v = _to_float(fila[col_devengado])
                            if v > 0:
                                actual = getattr(quincena, campo)
                                setattr(quincena, campo, actual + v)
                        break

        # Deducción
        if etiqueta_ded:
            if "TOTAL DEDUCCIONES" in etiqueta_ded:
                if col_deduccion_val < len(fila):
                    v = _to_float(fila[col_deduccion_val])
                    if v > 0:
                        quincena.total_deducciones = v
            else:
                for patrones, campo in PATRONES_DEDUCCION:
                    if _coincide_patron(etiqueta_ded, patrones):
                        if col_deduccion_val < len(fila):
                            v = _to_float(fila[col_deduccion_val])
                            if v > 0:
                                actual = getattr(quincena, campo)
                                setattr(quincena, campo, actual + v)
                        break

    # Validación mínima: tiene que tener algún devengado
    if quincena.total_devengado <= 0 and quincena.salario_basico <= 0:
        return None

    # Si no se capturó total_devengado pero sí los componentes, sumar
    if quincena.total_devengado <= 0:
        quincena.total_devengado = (
            quincena.salario_basico + quincena.aux_transporte +
            quincena.horas_extras_recargos + quincena.incapacidad + quincena.otros
        )

    # Si no se capturó total_deducciones pero sí los componentes, sumar
    if quincena.total_deducciones <= 0:
        quincena.total_deducciones = (
            quincena.ded_salud + quincena.ded_pension + quincena.ded_otros
        )

    # Si no se capturó neto, calcular
    if quincena.neto <= 0:
        quincena.neto = quincena.total_devengado - quincena.total_deducciones

    return quincena


def _leer_quincenas_de_hoja(
    ws_data: List[List], nombre_hoja: str
) -> List[Quincena]:
    """Lee TODAS las quincenas (típicamente 2) de UNA hoja del Excel."""
    inicios = _localizar_secciones_quincena(ws_data)
    if not inicios:
        return []

    quincenas: List[Quincena] = []
    for k, fila_inicio in enumerate(inicios):
        fila_fin = inicios[k + 1] if k + 1 < len(inicios) else len(ws_data)
        q = _extraer_quincena_desde_seccion(ws_data, fila_inicio, fila_fin, nombre_hoja)
        if q:
            quincenas.append(q)

    return quincenas


def _hoja_a_matriz(ws) -> List[List]:
    """Convierte una hoja de openpyxl a lista de listas (más fácil de iterar)."""
    matriz = []
    for fila in ws.iter_rows(values_only=True):
        matriz.append(list(fila))
    return matriz


# ============================================================
# Construcción del plano contable
# ============================================================

def _crear_linea_plano(
    cuenta: str,
    fecha: date,
    documento: str,
    nit: str,
    detalle: str,
    tr: str,                # '1' = Db, '2' = Cr
    valor: float,
    comprobante: str,
    base: float = 0.0,
    centro_costo: str = "",
    doc_referencia: str = "",
) -> dict:
    """Crea una fila del plano con el formato estándar."""
    return {
        "CUENTA": cuenta,
        "COMPROBANTE": comprobante,
        "FECHA": _formato_fecha_plano(fecha),
        "DOCUMENTO": documento,
        # Por defecto la referencia es igual al documento (salvo que se indique otra)
        "DOC REFERENCIA": doc_referencia or documento,
        "NIT": nit,
        "DETALLE": detalle,
        "TR": tr,
        "VALOR": _redondear(valor),
        "BASE": _redondear(base),
        "CENTRO DE COSTO": centro_costo,
    }


def _generar_asiento_quincena(
    quincenas_q: List[Quincena],
    empleados: Dict[str, Empleado],
    numero_quincena: int,
    fecha_asiento: date,
    doc_base: int = 1,
) -> Tuple[List[dict], List[str], int]:
    """Genera las líneas del plano de UNA quincena (todas las empleadas).

    Numeración: 1 documento por EMPLEADO y por quincena. El consecutivo
    arranca en `doc_base` y avanza de a 1 por cada empleado procesado.

    Returns:
        (filas_plano, advertencias, siguiente_documento_disponible)
    """
    filas: List[dict] = []
    advertencias: List[str] = []
    detalle = DETALLE_NOMINA_Q1 if numero_quincena == 1 else DETALLE_NOMINA_Q2
    doc_actual = int(doc_base)

    for q in quincenas_q:
        emp = empleados.get(q.cc)
        if not emp:
            advertencias.append(
                f"⚠️ Empleada CC {q.cc} ({q.nombre_excel}) no está en el maestro "
                f"de empleados. Se descarta."
            )
            continue

        # Documento único para este empleado en esta quincena
        documento = str(doc_actual)
        doc_actual += 1

        cuentas = emp.cuentas_gasto()
        nit = emp.cc

        # Pre-redondear cada componente (los redondeos individuales pueden no sumar igual
        # al total redondeado, pero el neto se ajusta al final para garantizar Db=Cr)
        v_sal = _redondear(q.salario_basico)
        v_aux = _redondear(q.aux_transporte)
        v_he  = _redondear(q.horas_extras_recargos)
        v_otr = _redondear(q.otros)
        v_inc = _redondear(q.incapacidad)
        v_dsa = _redondear(q.ded_salud)
        v_dpe = _redondear(q.ded_pension)
        v_dot = _redondear(q.ded_otros)

        # Total Db redondeado real = suma de componentes redondeados
        total_db_emp = v_sal + v_aux + v_he + v_otr + v_inc
        total_cr_deducciones = v_dsa + v_dpe + v_dot
        # Neto = Db − deducciones (CUADRE GARANTIZADO por construcción)
        neto_calc = total_db_emp - total_cr_deducciones

        # ---- DEVENGADOS (Db) ----
        if v_sal > 0:
            filas.append(_crear_linea_plano(
                cuenta=cuentas["sueldo"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle} {emp.nombre}",
                tr="1", valor=v_sal, base=0,
                comprobante=COMPROBANTE_NOMINA,
            ))
        if v_aux > 0:
            filas.append(_crear_linea_plano(
                cuenta=cuentas["aux_transporte"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle} {emp.nombre}",
                tr="1", valor=v_aux, base=0,
                comprobante=COMPROBANTE_NOMINA,
            ))
        if v_he > 0:
            filas.append(_crear_linea_plano(
                cuenta=cuentas["horas_extras"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle} {emp.nombre}",
                tr="1", valor=v_he, base=0,
                comprobante=COMPROBANTE_NOMINA,
            ))
        if v_otr > 0:
            filas.append(_crear_linea_plano(
                cuenta=cuentas["otros"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle} {emp.nombre}",
                tr="1", valor=v_otr, base=0,
                comprobante=COMPROBANTE_NOMINA,
            ))
        if v_inc > 0:
            filas.append(_crear_linea_plano(
                cuenta=cuentas["incapacidad"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle} {emp.nombre}",
                tr="1", valor=v_inc, base=0,
                comprobante=COMPROBANTE_NOMINA,
            ))

        # ---- DEDUCCIONES (Cr) ----
        if v_dsa > 0:
            filas.append(_crear_linea_plano(
                cuenta=CUENTAS_PASIVO["ded_salud"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle} {emp.nombre}",
                tr="2", valor=v_dsa, base=0,
                comprobante=COMPROBANTE_NOMINA,
            ))
        if v_dpe > 0:
            filas.append(_crear_linea_plano(
                cuenta=CUENTAS_PASIVO["ded_pension"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle} {emp.nombre}",
                tr="2", valor=v_dpe, base=0,
                comprobante=COMPROBANTE_NOMINA,
            ))
        if v_dot > 0:
            filas.append(_crear_linea_plano(
                cuenta=CUENTAS_PASIVO["ded_otros"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle} {emp.nombre} OTRAS DEDUCCIONES",
                tr="2", valor=v_dot, base=0,
                comprobante=COMPROBANTE_NOMINA,
            ))

        # ---- NETO A PAGAR (Cr) — calculado para garantizar cuadre exacto ----
        if neto_calc > 0:
            filas.append(_crear_linea_plano(
                cuenta=CUENTAS_PASIVO["salarios_por_pagar"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle} {emp.nombre}",
                tr="2", valor=neto_calc, base=0,
                comprobante=COMPROBANTE_NOMINA,
            ))

        # Verificar coherencia con el neto del Excel (informativo; no bloquea)
        neto_excel = _redondear(q.neto)
        if abs(neto_calc - neto_excel) > 1:
            advertencias.append(
                f"ℹ️ {emp.nombre} Q{q.numero}: neto contable ${neto_calc:,} "
                f"vs neto Excel ${neto_excel:,} "
                f"(diferencia ${neto_calc - neto_excel:,} por redondeos)".replace(",", ".")
            )

    return filas, advertencias, doc_actual


def _generar_asiento_provision(
    resumenes: Dict[str, ResumenEmpleadoMes],
    empleados: Dict[str, Empleado],
    fecha_asiento: date,
    mes: int,
) -> Tuple[List[dict], List[str]]:
    """Genera el asiento mensual de provisiones (aportes patronales + prestaciones).

    Returns:
        (filas_plano, advertencias)
    """
    filas: List[dict] = []
    advertencias: List[str] = []
    documento = str(mes)
    detalle_base = DETALLE_PROVISION

    for cc, res in resumenes.items():
        emp = empleados.get(cc)
        if not emp:
            continue

        cuentas = emp.cuentas_gasto()
        nit = emp.cc
        ibc = res.ibc
        base_prest = res.base_prestaciones
        base_vac = res.base_vacaciones

        # -------- APORTES PATRONALES (Db gasto, Cr pasivo) --------
        # Pensión 12% IBC
        v_pension = ibc * TASA_PENSION_PATRONAL
        if v_pension > 0:
            filas.append(_crear_linea_plano(
                cuenta=cuentas["pension"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle_base} PENSION {emp.nombre}",
                tr="1", valor=v_pension, base=ibc,
                comprobante=COMPROBANTE_PROVISION,
            ))
            filas.append(_crear_linea_plano(
                cuenta=CUENTAS_PASIVO["prov_pension"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle_base} PENSION {emp.nombre}",
                tr="2", valor=v_pension, base=ibc,
                comprobante=COMPROBANTE_PROVISION,
            ))

        # ARL (tasa específica por empleada)
        v_arl = ibc * emp.tasa_arl
        if v_arl > 0:
            filas.append(_crear_linea_plano(
                cuenta=cuentas["arl"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle_base} ARL {emp.nombre}",
                tr="1", valor=v_arl, base=ibc,
                comprobante=COMPROBANTE_PROVISION,
            ))
            filas.append(_crear_linea_plano(
                cuenta=CUENTAS_PASIVO["prov_arl"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle_base} ARL {emp.nombre}",
                tr="2", valor=v_arl, base=ibc,
                comprobante=COMPROBANTE_PROVISION,
            ))

        # Caja Compensación 4% IBC
        v_caja = ibc * TASA_CAJA_COMPENSACION
        if v_caja > 0:
            filas.append(_crear_linea_plano(
                cuenta=cuentas["caja"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle_base} CAJA {emp.nombre}",
                tr="1", valor=v_caja, base=ibc,
                comprobante=COMPROBANTE_PROVISION,
            ))
            filas.append(_crear_linea_plano(
                cuenta=CUENTAS_PASIVO["prov_caja"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle_base} CAJA {emp.nombre}",
                tr="2", valor=v_caja, base=ibc,
                comprobante=COMPROBANTE_PROVISION,
            ))

        # -------- PRESTACIONES SOCIALES --------
        # Cesantías 8.33% sobre Devengado total
        v_ces = base_prest * TASA_CESANTIAS
        if v_ces > 0:
            filas.append(_crear_linea_plano(
                cuenta=cuentas["cesantias"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle_base} CESANTIAS {emp.nombre}",
                tr="1", valor=v_ces, base=base_prest,
                comprobante=COMPROBANTE_PROVISION,
            ))
            filas.append(_crear_linea_plano(
                cuenta=CUENTAS_PASIVO["prov_cesantias"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle_base} CESANTIAS {emp.nombre}",
                tr="2", valor=v_ces, base=base_prest,
                comprobante=COMPROBANTE_PROVISION,
            ))

        # Intereses cesantías = cesantías * 12% / 12 (mensual)
        # Equivalente a base_prest * 0.0833 * 0.12 / 12 = base_prest * 0.000833
        # Pero la fórmula correcta mensual es: cesantías_acumuladas * 12% / 12
        # Como provisionamos cesantías de 1 mes, intereses = cesantías * 0.01 (1% mensual aprox)
        v_int = v_ces * (TASA_INTERESES_CESANTIAS / 12)
        if v_int > 0:
            filas.append(_crear_linea_plano(
                cuenta=cuentas["int_cesantias"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle_base} INT CESANTIAS {emp.nombre}",
                tr="1", valor=v_int, base=v_ces,
                comprobante=COMPROBANTE_PROVISION,
            ))
            filas.append(_crear_linea_plano(
                cuenta=CUENTAS_PASIVO["prov_int_cesantias"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle_base} INT CESANTIAS {emp.nombre}",
                tr="2", valor=v_int, base=v_ces,
                comprobante=COMPROBANTE_PROVISION,
            ))

        # Prima 8.33% sobre Devengado total
        v_pri = base_prest * TASA_PRIMA_SERVICIOS
        if v_pri > 0:
            filas.append(_crear_linea_plano(
                cuenta=cuentas["prima"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle_base} PRIMA {emp.nombre}",
                tr="1", valor=v_pri, base=base_prest,
                comprobante=COMPROBANTE_PROVISION,
            ))
            filas.append(_crear_linea_plano(
                cuenta=CUENTAS_PASIVO["prov_prima"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle_base} PRIMA {emp.nombre}",
                tr="2", valor=v_pri, base=base_prest,
                comprobante=COMPROBANTE_PROVISION,
            ))

        # Vacaciones 4.17% sobre Devengado − Aux Transporte
        v_vac = base_vac * TASA_VACACIONES
        if v_vac > 0:
            filas.append(_crear_linea_plano(
                cuenta=cuentas["vacaciones"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle_base} VACACIONES {emp.nombre}",
                tr="1", valor=v_vac, base=base_vac,
                comprobante=COMPROBANTE_PROVISION,
            ))
            filas.append(_crear_linea_plano(
                cuenta=CUENTAS_PASIVO["prov_vacaciones"],
                fecha=fecha_asiento, documento=documento, nit=nit,
                detalle=f"{detalle_base} VACACIONES {emp.nombre}",
                tr="2", valor=v_vac, base=base_vac,
                comprobante=COMPROBANTE_PROVISION,
            ))

    return filas, advertencias


# ============================================================
# Función principal
# ============================================================

def procesar_nomina(
    archivo,
    anio: int,
    mes: int,
) -> Tuple[pd.DataFrame, List[str], dict]:
    """Procesa el Excel de nómina y produce el plano contable completo.

    Args:
        archivo: file-like, bytes, o path del Excel.
        anio: año del periodo a procesar (ej. 2026).
        mes: mes del periodo (1-12).

    Returns:
        (df_plano, log, resumen_dict)
        - df_plano: DataFrame con las 11 columnas estándar.
        - log: lista de mensajes informativos.
        - resumen_dict: estadísticas y desglose para mostrar en UI.
    """
    log: List[str] = []

    # Normalizar entrada a BytesIO
    if hasattr(archivo, "read"):
        contenido = archivo.read()
        if hasattr(archivo, "seek"):
            archivo.seek(0)
        bio = io.BytesIO(contenido)
    elif isinstance(archivo, (bytes, bytearray)):
        bio = io.BytesIO(bytes(archivo))
    else:
        raise TypeError(f"Tipo de archivo no soportado: {type(archivo)}")

    # Cargar maestro de empleados
    empleados = cargar_empleados_embebido()
    log.append(f"📋 Maestro de empleados: {len(empleados)} registros.")

    # Detectar formato (.xls vs .xlsx)
    contenido_bytes = bio.getvalue()
    bio.seek(0)
    es_xls_antiguo = contenido_bytes[:4] == b"\xd0\xcf\x11\xe0"

    if es_xls_antiguo:
        # Usar pandas+xlrd para .xls antiguo
        log.append("📂 Formato detectado: Excel 97-2003 (.xls)")
        try:
            xls = pd.ExcelFile(bio, engine="xlrd")
        except ImportError:
            raise ImportError(
                "Para leer archivos .xls se requiere el paquete 'xlrd>=2.0.1'. "
                "Agrégalo a requirements.txt."
            )
        hojas = xls.sheet_names
    else:
        log.append("📂 Formato detectado: Excel moderno (.xlsx)")
        wb = load_workbook(bio, data_only=True, read_only=True)
        hojas = wb.sheetnames

    log.append(f"📑 Hojas detectadas: {len(hojas)}")

    # Procesar cada hoja → quincenas
    todas_quincenas: List[Quincena] = []
    for nombre_hoja in hojas:
        if es_xls_antiguo:
            df_hoja = pd.read_excel(xls, sheet_name=nombre_hoja, header=None)
            ws_data = df_hoja.values.tolist()
        else:
            ws = wb[nombre_hoja]
            ws_data = _hoja_a_matriz(ws)

        quincenas = _leer_quincenas_de_hoja(ws_data, nombre_hoja)

        if not quincenas:
            log.append(f"  ⚠️ '{nombre_hoja}': no se detectaron desprendibles válidos.")
            continue

        # Si la quincena no tiene CC, intentar resolverla por el nombre de la hoja
        for q in quincenas:
            if not q.cc and q.nombre_excel:
                # Buscar empleado con nombre similar
                for cc_master, emp in empleados.items():
                    if _normalizar_texto(emp.nombre) in _normalizar_texto(q.nombre_excel) \
                            or _normalizar_texto(q.nombre_excel) in _normalizar_texto(emp.nombre):
                        q.cc = cc_master
                        break

        log.append(f"  ✓ '{nombre_hoja}': {len(quincenas)} quincenas detectadas.")
        todas_quincenas.extend(quincenas)

    if not todas_quincenas:
        raise ValueError(
            "No se detectó ninguna quincena en el archivo. "
            "Verifica que el formato sea el esperado (1 hoja por empleada, "
            "2 desprendibles por hoja con encabezado 'CASA UNOTRES SAS')."
        )

    # Separar por número de quincena
    q1 = [q for q in todas_quincenas if q.numero == 1]
    q2 = [q for q in todas_quincenas if q.numero == 2]
    log.append("")
    log.append(f"📊 Total quincenas: {len(todas_quincenas)} "
               f"(Q1: {len(q1)}, Q2: {len(q2)})")

    if not q1 and not q2:
        raise ValueError(
            "No se pudo identificar las quincenas (campo NUMERO no detectado). "
            "Verifica que cada desprendible tenga el campo 'NUMERO' con valor 1 o 2."
        )

    # Construir resúmenes mensuales para provisiones
    resumenes: Dict[str, ResumenEmpleadoMes] = {}
    for q in todas_quincenas:
        if not q.cc:
            continue
        emp = empleados.get(q.cc)
        if not emp:
            continue
        if q.cc not in resumenes:
            resumenes[q.cc] = ResumenEmpleadoMes(cc=q.cc, nombre=emp.nombre)
        r = resumenes[q.cc]
        r.devengado_total += q.total_devengado
        r.aux_transporte_total += q.aux_transporte
        r.incapacidad_total += q.incapacidad

    # Fechas de los asientos
    fecha_q1 = date(anio, mes, 15)
    # Último día real del mes (28, 29, 30 o 31 según el mes y si es bisiesto)
    ultimo_dia = calendar.monthrange(anio, mes)[1]
    fecha_q2 = date(anio, mes, ultimo_dia)
    fecha_provision = fecha_q2  # provisión al cierre del mes

    # Generar líneas
    todas_filas: List[dict] = []
    # Consecutivo corrido: 1 documento por empleado y por quincena.
    # Q1 usa 1..N, Q2 continúa en N+1..2N (documentos únicos dentro del comp 11).
    siguiente_doc = 1

    if q1:
        log.append("")
        log.append(f"📝 Generando asiento Q1 (comprobante {COMPROBANTE_NOMINA}, "
                   f"1 documento por empleado desde {siguiente_doc}, "
                   f"fecha {_formato_fecha_plano(fecha_q1)})...")
        filas_q1, adv_q1, siguiente_doc = _generar_asiento_quincena(
            q1, empleados, 1, fecha_q1, doc_base=siguiente_doc
        )
        todas_filas.extend(filas_q1)
        log.extend(adv_q1)
        log.append(f"   {len(filas_q1)} líneas generadas.")

    if q2:
        log.append("")
        log.append(f"📝 Generando asiento Q2 (comprobante {COMPROBANTE_NOMINA}, "
                   f"1 documento por empleado desde {siguiente_doc}, "
                   f"fecha {_formato_fecha_plano(fecha_q2)})...")
        filas_q2, adv_q2, siguiente_doc = _generar_asiento_quincena(
            q2, empleados, 2, fecha_q2, doc_base=siguiente_doc
        )
        todas_filas.extend(filas_q2)
        log.extend(adv_q2)
        log.append(f"   {len(filas_q2)} líneas generadas.")

    if resumenes:
        log.append("")
        log.append(f"📝 Generando asiento de PROVISIÓN mensual "
                   f"(comprobante {COMPROBANTE_PROVISION}, documento {mes}, "
                   f"fecha {_formato_fecha_plano(fecha_provision)})...")
        filas_prov, adv_prov = _generar_asiento_provision(
            resumenes, empleados, fecha_provision, mes
        )
        todas_filas.extend(filas_prov)
        log.extend(adv_prov)
        log.append(f"   {len(filas_prov)} líneas generadas.")

    # Construir DataFrame final
    df_plano = pd.DataFrame(todas_filas, columns=COLUMNAS_PLANO)

    # Resumen y validación
    log.append("")
    log.append("=" * 60)
    log.append("📊 RESUMEN GENERAL")
    log.append("=" * 60)

    if len(df_plano) > 0:
        total_db = int(df_plano[df_plano["TR"] == "1"]["VALOR"].astype(int).sum())
        total_cr = int(df_plano[df_plano["TR"] == "2"]["VALOR"].astype(int).sum())
        log.append(f"   Total líneas: {len(df_plano)}")
        log.append(f"   Total Db:    $ {total_db:,}".replace(",", "."))
        log.append(f"   Total Cr:    $ {total_cr:,}".replace(",", "."))
        if total_db == total_cr:
            log.append(f"   ✅ Cuadre perfecto Db = Cr")
        else:
            log.append(f"   ❌ DESCUADRE: diferencia $ "
                       f"{total_db - total_cr:,}".replace(",", "."))
    else:
        log.append("   ⚠️ No se generaron líneas en el plano.")

    # Resumen por empleado para mostrar en UI
    resumen_dict = {
        "total_lineas": len(df_plano),
        "empleados_procesados": len(resumenes),
        "quincenas_q1": len(q1),
        "quincenas_q2": len(q2),
        "periodo": f"{anio}-{mes:02d}",
        "fecha_q1": _formato_fecha_plano(fecha_q1),
        "fecha_q2": _formato_fecha_plano(fecha_q2),
        "detalle_empleados": [
            {
                "cc": cc,
                "nombre": r.nombre,
                "devengado_total": _redondear(r.devengado_total),
                "aux_transporte": _redondear(r.aux_transporte_total),
                "incapacidad": _redondear(r.incapacidad_total),
                "ibc": _redondear(r.ibc),
                "base_prestaciones": _redondear(r.base_prestaciones),
                "base_vacaciones": _redondear(r.base_vacaciones),
                "tasa_arl_pct": f"{empleados[cc].tasa_arl * 100:.3f}%"
                                if cc in empleados else "N/A",
                "tipo": empleados[cc].tipo if cc in empleados else "N/A",
            }
            for cc, r in resumenes.items()
        ],
    }

    return df_plano, log, resumen_dict


# ============================================================
# Exportar a TSV (mismo formato que los demás módulos)
# ============================================================

def dataframe_a_plano_tsv(
    df: pd.DataFrame,
    incluir_encabezado_excel: bool = True,
) -> bytes:
    """Convierte el DataFrame a bytes TSV compatible con el sistema contable.

    Si incluir_encabezado_excel=True, antepone 'sep=\\t' para que Excel abra el
    archivo con las columnas alineadas.
    """
    df_out = df[COLUMNAS_PLANO].copy()
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str).str.replace("\t", " ", regex=False)

    tsv = df_out.to_csv(sep="\t", index=False, header=True, lineterminator="\r\n")

    if incluir_encabezado_excel:
        tsv = "sep=\t\r\n" + tsv

    return tsv.encode("utf-8")
