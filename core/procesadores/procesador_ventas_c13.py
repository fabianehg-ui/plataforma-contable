"""
Procesador de Ventas Casa 13 — Facturación electrónica (Siigo).

Toma el archivo RadGridExport.xlsx exportado del sistema de facturación
y genera el plano contable con MÚLTIPLES comprobantes según el tipo de
documento.

Mapeo de prefijos a comprobantes:
    PVFE...   → COMP 19  (Venta POS, contado, NIT genérico)
    GBF...    → COMP 5   (Venta GBF, contado, NIT genérico)
    FE...     → COMP 4   (Venta a empresa, crédito, NIT real)
    DV4xxxx   → COMP 6   (NC de PVFE)
    DV3xxxx   → COMP 15  (NC de GBF)
    DW...     → COMP 16  (NC con IVA en 24080101)
    NCA...    → COMP 21  (NC misceláneas)

Cuentas de contrapartida:
    - Ventas POS/GBF (PVFE, GBF):  Db 11050500 (CAJA)
    - Ventas FE (a empresas):      Db 130506   (CLIENTES)
    - NC PVFE/GBF/DW (devolución): Cr 11050500 (CAJA)
    - NC NCA: Cr 11050500 si DocRef PVFE/GBF, Cr 130506 si DocRef FE

IVA en notas crédito:
    - DV3, DV4, NCA → cuenta 24080102 (IVA NC)
    - DW           → cuenta 24080101 (IVA generado normal)

Retenciones automáticas (solo C4 con conceptos múltiples):
    - 19551503 = 3.5%  × Arrendamientos (415505)   "RETENCION ARRENDAMIENTOS 3.5%"
    - 19551510 = 11%   × Comisiones    (415030)    "RETENCION COMISIONES 11%"
"""
from __future__ import annotations
import io
import re
from dataclasses import dataclass, field
from datetime import datetime, date
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


# ============================================================
# Constantes y configuración
# ============================================================

# Mapeo prefijo → comprobante
# El orden importa: prefijos más largos van primero
MAPEO_PREFIJO_COMPROBANTE = [
    # (prefijo_match, comprobante, tipo, descripción)
    # NOTAS CRÉDITO primero (los DV son más específicos que los DW)
    ("DV4",  "6",  "NC_PVFE",  "Nota crédito de factura PVFE"),
    ("DV3",  "15", "NC_GBF",   "Nota crédito de factura GBF"),
    ("DW",   "16", "NC_DW",    "Nota crédito DW"),
    ("NCA",  "21", "NC_NCA",   "Nota crédito misceláneas"),
    # VENTAS
    ("PVFE", "19", "VTA_PVFE", "Venta POS Siigo"),
    ("GBF",  "5",  "VTA_GBF",  "Venta GBF"),
    ("FE",   "4",  "VTA_FE",   "Venta a empresa"),
]

# Tipos de transacción
TR_DEBITO = "1"
TR_CREDITO = "2"

# Cuentas de contrapartida
CUENTA_CAJA = "11050500"
CUENTA_CLIENTES = "130506"

# IVA: dos cuentas según el caso
CUENTA_IVA_GENERADO = "24080101"  # Ventas normales y NC tipo DW
CUENTA_IVA_NC = "24080102"        # NC tipo DV/NCA

# Cuentas de retención (automáticas en C4)
CUENTA_RET_ARRENDAMIENTOS = "19551503"
CUENTA_RET_COMISIONES = "19551510"
TASA_RET_ARRENDAMIENTOS = 0.035  # 3.5%
TASA_RET_COMISIONES = 0.11       # 11%

# Cuentas de movimiento que vienen del archivo (Siigo)
CUENTA_VENTAS_GRAVADAS = "41359501"
CUENTA_VENTAS_EXENTAS = "41359602"
CUENTA_FLETES = "41359601"
CUENTA_EMPAQUE = "41559501"
CUENTA_ARRENDAMIENTOS = "415505"
CUENTA_COMISIONES = "415030"
CUENTA_PUBLICIDAD = "415555"
CUENTA_DEVOLUCION = "41750201"

# Mapeo nombre PUC → cuenta + nombre estándar de detalle
# (el archivo de Siigo trae "Cuenta #" y "Nombre Cuenta PUC", confiamos en eso)
NOMBRES_DETALLE: Dict[str, str] = {
    CUENTA_VENTAS_GRAVADAS:   "VENTAS GRAVADAS",
    CUENTA_VENTAS_EXENTAS:    "EXENTAS",
    CUENTA_FLETES:            "FLETES",
    CUENTA_EMPAQUE:           "SERVICIO DE EMPAQUE",
    CUENTA_ARRENDAMIENTOS:    "ARRENDAMIENTOS",
    CUENTA_COMISIONES:        "COMISIONES",
    CUENTA_PUBLICIDAD:        "PUBLICIDAD",
    CUENTA_IVA_GENERADO:      "IVA GENERADO 19%",
    CUENTA_IVA_NC:            "IVA GENERADO 19%",
    CUENTA_DEVOLUCION:        "DEVOLUCION EN VENTAS",
    CUENTA_CAJA:              "CAJA",
    CUENTA_CLIENTES:          "CLIENTES",
    CUENTA_RET_ARRENDAMIENTOS: "RETENCION ARRENDAMIENTOS 3.5%",
    CUENTA_RET_COMISIONES:    "RETENCION COMISIONES 11%",
}

# NIT genérico para ventas a consumidor final (PVFE, GBF, NC asociadas)
NIT_GENERICO = "222222222"

# Centro de costos (PRINCIPAL en este módulo, igual que en POS)
CENTRO_COSTOS_DEFAULT = "PRINCIPAL"

# Columnas finales del plano (mismo orden que los demás módulos)
COLUMNAS_PLANO = [
    "CUENTA",
    "COMPROBANTE",
    "FECHA",
    "DOCUMENTO",
    "DOC REFERENCIA",
    "NIT",
    "DETALLE",
    "TR",
    "VALOR",
    "BASE",
    "CENTRO DE COSTO",
]

# Columnas esperadas en el archivo de Siigo
COLUMNAS_ESPERADAS = {
    "A", "D", "Sucursal", "Fecha", "Documento", "Identificación",
    "Cliente", "Nombre Cuenta PUC", "Cuenta #", "Total", "Base Impuestos",
}


# ============================================================
# Estructuras de datos
# ============================================================

@dataclass
class LineaSiigo:
    """Una línea cruda del archivo RadGridExport."""
    anulado: bool
    es_devolucion: bool
    sucursal: str
    fecha: date
    documento: str
    identificacion: str
    cliente: str
    nombre_cuenta: str
    cuenta: str
    total: float
    base_impuestos: float


@dataclass
class Documento:
    """Un documento agrupado (factura o nota crédito) con sus líneas."""
    documento: str
    fecha: date
    cliente: str
    identificacion: str
    sucursal: str
    es_devolucion: bool
    lineas: List[LineaSiigo] = field(default_factory=list)

    @property
    def prefijo(self) -> str:
        """Determina el prefijo del documento (PVFE, GBF, FE, DV3, DV4, DW, NCA)."""
        return _detectar_prefijo(self.documento)

    @property
    def comprobante(self) -> str:
        """Retorna el comprobante asignado según el prefijo."""
        prefijo = self.prefijo
        for p, comp, _, _ in MAPEO_PREFIJO_COMPROBANTE:
            if prefijo == p or self.documento.startswith(p):
                # Manejar DV4 vs DV3 con cuidado
                if p in ("DV3", "DV4"):
                    if self.documento.startswith(p):
                        return comp
                    continue
                if self.documento.startswith(p):
                    return comp
        return ""

    @property
    def tipo(self) -> str:
        """Retorna el tipo (VTA_PVFE, NC_PVFE, etc.)."""
        for p, _, tipo, _ in MAPEO_PREFIJO_COMPROBANTE:
            if self.documento.startswith(p):
                # DV3 vs DV4 specific
                if p in ("DV3", "DV4"):
                    if self.documento.startswith(p):
                        return tipo
                    continue
                if self.documento.startswith(p):
                    return tipo
        return "DESCONOCIDO"


# ============================================================
# Helpers
# ============================================================

def _detectar_prefijo(documento: str) -> str:
    """Devuelve el prefijo de un documento, considerando DV3 vs DV4."""
    if not documento:
        return ""
    doc_upper = documento.upper().strip()
    # DV3 y DV4 son diferentes (4to caracter define rango)
    if doc_upper.startswith("DV"):
        if len(doc_upper) >= 3:
            tercer = doc_upper[2]
            if tercer == "3":
                return "DV3"
            elif tercer == "4":
                return "DV4"
        return "DV"
    # Otros prefijos: revisar por orden de longitud descendente
    for p, _, _, _ in MAPEO_PREFIJO_COMPROBANTE:
        if p in ("DV3", "DV4"):
            continue
        if doc_upper.startswith(p):
            return p
    return ""


def _detectar_comprobante(documento: str) -> Optional[Tuple[str, str]]:
    """Retorna (comprobante, tipo) según el prefijo, o None si no se reconoce."""
    prefijo = _detectar_prefijo(documento)
    for p, comp, tipo, _ in MAPEO_PREFIJO_COMPROBANTE:
        if p == prefijo:
            return (comp, tipo)
    return None


def _formato_cuenta(valor) -> str:
    """Cuenta como string sin '.0'."""
    if valor is None:
        return ""
    if isinstance(valor, float) and not pd.isna(valor):
        valor = int(valor)
    return str(valor).strip()


def _formato_cc(valor) -> str:
    """Identificación normalizada: solo dígitos."""
    if valor is None:
        return ""
    if isinstance(valor, float) and not pd.isna(valor):
        valor = int(valor)
    s = str(valor).strip()
    return re.sub(r"\D", "", s)


def _to_float(valor) -> float:
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        if pd.isna(valor):
            return 0.0
        return float(valor)
    s = str(valor).strip().replace("$", "").replace(",", "")
    if not s or s == "-":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _redondear(valor: float) -> int:
    """Redondea a entero."""
    return int(round(valor))


def _formato_fecha_plano(d: date) -> str:
    return d.strftime("%m/%d/%Y")


def _detalle_para(cuenta: str, nombre_archivo: Optional[str] = None) -> str:
    """Retorna el nombre estándar para la columna DETALLE."""
    if cuenta in NOMBRES_DETALLE:
        return NOMBRES_DETALLE[cuenta]
    # Fallback: usar el nombre del archivo si está disponible
    if nombre_archivo:
        return str(nombre_archivo).strip().upper()
    return ""


# ============================================================
# Lectura del archivo Siigo
# ============================================================

def _leer_radgrid(archivo) -> List[LineaSiigo]:
    """Lee el archivo RadGridExport.xlsx y retorna lista de líneas."""

    # Normalizar entrada a BytesIO
    if hasattr(archivo, "read"):
        contenido = archivo.read()
        if hasattr(archivo, "seek"):
            archivo.seek(0)
        bio = io.BytesIO(contenido)
    elif isinstance(archivo, (bytes, bytearray)):
        bio = io.BytesIO(bytes(archivo))
    else:
        raise TypeError(f"Tipo de archivo no soportado: {type(archivo)}")

    wb = load_workbook(bio, data_only=True, read_only=True)

    # Buscar la hoja correcta — preferimos 'RadGridExport' si existe, sino la primera
    hoja_objetivo = None
    for nombre in wb.sheetnames:
        if "RADGRID" in nombre.upper() or nombre.upper().startswith("VENTAS"):
            hoja_objetivo = nombre
            break
    if hoja_objetivo is None:
        hoja_objetivo = wb.sheetnames[0]

    ws = wb[hoja_objetivo]

    # Leer encabezados
    filas_iter = ws.iter_rows(values_only=True)
    encabezados = next(filas_iter, None)
    if not encabezados:
        raise ValueError(f"La hoja '{hoja_objetivo}' está vacía.")

    encabezados = [str(c).strip() if c is not None else "" for c in encabezados]
    indices = {nombre: i for i, nombre in enumerate(encabezados)}

    # Validar columnas mínimas
    faltantes = COLUMNAS_ESPERADAS - set(encabezados)
    if faltantes:
        raise ValueError(
            f"Faltan columnas en el archivo: {faltantes}. "
            f"Encontradas: {encabezados}"
        )

    def get(fila, nombre):
        idx = indices.get(nombre)
        if idx is None or idx >= len(fila):
            return None
        return fila[idx]

    lineas: List[LineaSiigo] = []
    for fila in filas_iter:
        if not fila or all(c is None for c in fila):
            continue

        documento = get(fila, "Documento")
        if not documento:
            continue
        documento = str(documento).strip()

        fecha_raw = get(fila, "Fecha")
        if isinstance(fecha_raw, datetime):
            fecha_obj = fecha_raw.date()
        elif isinstance(fecha_raw, date):
            fecha_obj = fecha_raw
        else:
            try:
                fecha_obj = pd.to_datetime(fecha_raw).date()
            except (ValueError, TypeError):
                continue  # sin fecha válida, descartar

        anulado = bool(get(fila, "A")) if get(fila, "A") is not None else False
        es_devol = bool(get(fila, "D")) if get(fila, "D") is not None else False

        cta = _formato_cuenta(get(fila, "Cuenta #"))
        if not cta or cta == "0":
            # Líneas "Sin Cuenta" — se descartan
            continue

        lineas.append(LineaSiigo(
            anulado=anulado,
            es_devolucion=es_devol,
            sucursal=str(get(fila, "Sucursal") or "").strip(),
            fecha=fecha_obj,
            documento=documento,
            identificacion=_formato_cc(get(fila, "Identificación")),
            cliente=str(get(fila, "Cliente") or "").strip(),
            nombre_cuenta=str(get(fila, "Nombre Cuenta PUC") or "").strip(),
            cuenta=cta,
            total=_to_float(get(fila, "Total")),
            base_impuestos=_to_float(get(fila, "Base Impuestos")),
        ))

    return lineas


def _agrupar_por_documento(lineas: List[LineaSiigo]) -> Dict[str, Documento]:
    """Agrupa líneas por documento."""
    docs: Dict[str, Documento] = {}
    for ln in lineas:
        if ln.documento not in docs:
            docs[ln.documento] = Documento(
                documento=ln.documento,
                fecha=ln.fecha,
                cliente=ln.cliente,
                identificacion=ln.identificacion,
                sucursal=ln.sucursal,
                es_devolucion=ln.es_devolucion,
            )
        docs[ln.documento].lineas.append(ln)
    return docs


# ============================================================
# Generación del plano por tipo de documento
# ============================================================

def _crear_linea_plano(
    cuenta: str,
    comprobante: str,
    fecha: date,
    documento: str,
    doc_referencia: str,
    nit: str,
    detalle: str,
    tr: str,
    valor: float,
    base: float = 0,
    centro_costo: str = CENTRO_COSTOS_DEFAULT,
) -> dict:
    """Crea una fila del plano con el formato estándar."""
    return {
        "CUENTA": cuenta,
        "COMPROBANTE": comprobante,
        "FECHA": _formato_fecha_plano(fecha),
        "DOCUMENTO": documento,
        "DOC REFERENCIA": doc_referencia,
        "NIT": nit,
        "DETALLE": detalle,
        "TR": tr,
        "VALOR": _redondear(valor),
        "BASE": _redondear(base),
        "CENTRO DE COSTO": centro_costo,
    }


def _generar_asiento_venta_contado(doc: Documento, comprobante: str) -> List[dict]:
    """Asiento para PVFE y GBF: ventas contado.
    Db 11050500 (Caja) ↔ Cr cuentas del archivo (positivas).
    NIT: 222222222 (consumidor final).
    DOCUMENTO = DOC REFERENCIA = consecutivo del archivo.

    IMPORTANTE: cada componente Cr se REDONDEA individualmente, y la CAJA Db
    se calcula como suma de los redondeos para garantizar cuadre exacto.
    """
    filas: List[dict] = []
    nit = NIT_GENERICO
    doc_ref = doc.documento  # En ventas, doc y doc_ref son lo mismo
    documento = doc.documento

    total_cr = 0  # int — suma de redondeos

    # Líneas de Cr (ventas, IVA, otros conceptos)
    for ln in doc.lineas:
        if ln.total <= 0:
            continue
        valor_redondeado = _redondear(abs(ln.total))
        if valor_redondeado <= 0:
            continue
        # IVA: la base sí va; otras cuentas: BASE = 0
        base = _redondear(abs(ln.base_impuestos)) if ln.cuenta in (CUENTA_IVA_GENERADO, CUENTA_IVA_NC) else 0
        filas.append(_crear_linea_plano(
            cuenta=ln.cuenta,
            comprobante=comprobante,
            fecha=ln.fecha,
            documento=documento,
            doc_referencia=doc_ref,
            nit=nit,
            detalle=_detalle_para(ln.cuenta, ln.nombre_cuenta),
            tr=TR_CREDITO,
            valor=valor_redondeado,
            base=base,
        ))
        total_cr += valor_redondeado

    # Línea de Db: CAJA por la suma EXACTA de redondeos (cuadre garantizado)
    if total_cr > 0:
        filas.append(_crear_linea_plano(
            cuenta=CUENTA_CAJA,
            comprobante=comprobante,
            fecha=doc.fecha,
            documento=documento,
            doc_referencia=doc_ref,
            nit=nit,
            detalle=_detalle_para(CUENTA_CAJA),
            tr=TR_DEBITO,
            valor=total_cr,
            base=0,
        ))

    return filas


def _generar_asiento_venta_fe(doc: Documento, comprobante: str) -> List[dict]:
    """Asiento para FE: venta a empresa con NIT real.
    Db 130506 (Clientes) + retenciones automáticas ↔ Cr cuentas del archivo.

    Cada Cr se redondea individualmente. CLIENTES Db se calcula como
    diferencia exacta para cuadre.
    """
    filas: List[dict] = []
    nit = doc.identificacion or NIT_GENERICO
    doc_ref = doc.documento
    documento = doc.documento

    total_cr = 0  # int — suma de redondeos

    # Líneas de Cr (cuentas del archivo, todas positivas)
    for ln in doc.lineas:
        if ln.total <= 0:
            continue
        valor_redondeado = _redondear(abs(ln.total))
        if valor_redondeado <= 0:
            continue
        base = _redondear(abs(ln.base_impuestos)) if ln.cuenta in (CUENTA_IVA_GENERADO, CUENTA_IVA_NC) else 0
        filas.append(_crear_linea_plano(
            cuenta=ln.cuenta,
            comprobante=comprobante,
            fecha=ln.fecha,
            documento=documento,
            doc_referencia=doc_ref,
            nit=nit,
            detalle=_detalle_para(ln.cuenta, ln.nombre_cuenta),
            tr=TR_CREDITO,
            valor=valor_redondeado,
            base=base,
        ))
        total_cr += valor_redondeado

    # Calcular retenciones automáticas si aplican
    total_db_retenciones = 0
    base_arrendamientos = 0.0
    base_comisiones = 0.0
    for ln in doc.lineas:
        if ln.cuenta == CUENTA_ARRENDAMIENTOS and ln.total > 0:
            base_arrendamientos += ln.total
        elif ln.cuenta == CUENTA_COMISIONES and ln.total > 0:
            base_comisiones += ln.total

    # Retención de arrendamientos 3.5%
    if base_arrendamientos > 0:
        v_ret_arr = _redondear(base_arrendamientos * TASA_RET_ARRENDAMIENTOS)
        if v_ret_arr > 0:
            filas.append(_crear_linea_plano(
                cuenta=CUENTA_RET_ARRENDAMIENTOS,
                comprobante=comprobante,
                fecha=doc.fecha,
                documento=documento,
                doc_referencia=doc_ref,
                nit=nit,
                detalle=_detalle_para(CUENTA_RET_ARRENDAMIENTOS),
                tr=TR_DEBITO,
                valor=v_ret_arr,
                base=_redondear(base_arrendamientos),
            ))
            total_db_retenciones += v_ret_arr

    # Retención de comisiones 11%
    if base_comisiones > 0:
        v_ret_com = _redondear(base_comisiones * TASA_RET_COMISIONES)
        if v_ret_com > 0:
            filas.append(_crear_linea_plano(
                cuenta=CUENTA_RET_COMISIONES,
                comprobante=comprobante,
                fecha=doc.fecha,
                documento=documento,
                doc_referencia=doc_ref,
                nit=nit,
                detalle=_detalle_para(CUENTA_RET_COMISIONES),
                tr=TR_DEBITO,
                valor=v_ret_com,
                base=_redondear(base_comisiones),
            ))
            total_db_retenciones += v_ret_com

    # Línea Db CLIENTES por el residual (cuadre garantizado)
    valor_clientes = total_cr - total_db_retenciones
    if valor_clientes > 0:
        filas.append(_crear_linea_plano(
            cuenta=CUENTA_CLIENTES,
            comprobante=comprobante,
            fecha=doc.fecha,
            documento=documento,
            doc_referencia=doc_ref,
            nit=nit,
            detalle=_detalle_para(CUENTA_CLIENTES),
            tr=TR_DEBITO,
            valor=valor_clientes,
            base=0,
        ))

    return filas


def _generar_asiento_nota_credito(doc: Documento, comprobante: str, tipo: str) -> List[dict]:
    """Asiento para notas crédito (DV4, DV3, DW, NCA).

    Db cuentas (todas positivas: DEVOLUCION, IVA) ↔ Cr CAJA o CLIENTES.
    DOCUMENTO = consecutivo NC; DOC REFERENCIA = factura asociada (de Documentos__3_).

    El IVA de NC va en cuenta:
        - 24080101 si tipo == NC_DW
        - 24080102 en los demás casos
    """
    filas: List[dict] = []
    documento = doc.documento

    # Determinar contrapartida y NIT
    es_factura_empresa = bool(
        doc.identificacion and len(doc.identificacion) >= 9
        and doc.identificacion.startswith(("8", "9"))
    )

    if tipo == "NC_NCA" and es_factura_empresa:
        cuenta_contra = CUENTA_CLIENTES
        nit = doc.identificacion
    else:
        cuenta_contra = CUENTA_CAJA
        nit = NIT_GENERICO

    cuenta_iva_objetivo = CUENTA_IVA_GENERADO if tipo == "NC_DW" else CUENTA_IVA_NC

    total_db = 0  # int — suma de redondeos
    doc_ref = doc.documento  # se sobrescribe afuera con notas_credito_map si está disponible

    for ln in doc.lineas:
        if ln.total == 0:
            continue
        valor_redondeado = _redondear(abs(ln.total))
        if valor_redondeado <= 0:
            continue
        cuenta_destino = ln.cuenta
        if ln.cuenta in (CUENTA_IVA_GENERADO, CUENTA_IVA_NC):
            cuenta_destino = cuenta_iva_objetivo
        base = _redondear(abs(ln.base_impuestos)) if cuenta_destino in (CUENTA_IVA_GENERADO, CUENTA_IVA_NC) else 0
        filas.append(_crear_linea_plano(
            cuenta=cuenta_destino,
            comprobante=comprobante,
            fecha=ln.fecha,
            documento=documento,
            doc_referencia=doc_ref,
            nit=nit,
            detalle=_detalle_para(cuenta_destino, ln.nombre_cuenta),
            tr=TR_DEBITO,
            valor=valor_redondeado,
            base=base,
        ))
        total_db += valor_redondeado

    if total_db > 0:
        filas.append(_crear_linea_plano(
            cuenta=cuenta_contra,
            comprobante=comprobante,
            fecha=doc.fecha,
            documento=documento,
            doc_referencia=doc_ref,
            nit=nit,
            detalle=_detalle_para(cuenta_contra),
            tr=TR_CREDITO,
            valor=total_db,
            base=0,
        ))

    return filas


# ============================================================
# Lectura de Documentos.xlsx (NOTAS CREDITO) — necesario para:
#   1. Enriquecer DOC REFERENCIA en NC con la factura asociada.
#   2. CONSTRUIR los asientos NCA (no vienen en RadGridExport).
# ============================================================

@dataclass
class NotaCreditoCabecera:
    """Cabecera de NC desde Documentos.xlsx para construir asientos NCA."""
    consecutivo: str
    factura_asociada: str
    fecha: date
    identificacion: str
    cliente: str
    sucursal: str
    impuesto: float       # IVA
    valor: float          # Total con IVA (lo que se devuelve al cliente)
    anulado: bool


def _leer_notas_credito(archivo_notas) -> Tuple[Dict[str, str], List[NotaCreditoCabecera]]:
    """Lee Documentos__3_.xlsx y retorna:
        - mapping {consecutivo_nc: consecutivo_factura_asociada}
        - lista completa de cabeceras de NC (para construir asientos NCA)

    Si no se puede leer, retorna ({}, []).
    """
    if archivo_notas is None:
        return {}, []

    if hasattr(archivo_notas, "read"):
        contenido = archivo_notas.read()
        if hasattr(archivo_notas, "seek"):
            archivo_notas.seek(0)
        bio = io.BytesIO(contenido)
    elif isinstance(archivo_notas, (bytes, bytearray)):
        bio = io.BytesIO(bytes(archivo_notas))
    else:
        return {}, []

    try:
        wb = load_workbook(bio, data_only=True, read_only=True)
    except Exception:
        return {}, []

    # Buscar la hoja apropiada
    hoja_obj = None
    for nombre in wb.sheetnames:
        upper = nombre.upper()
        if "NOTA" in upper or "DOCUMENTO" in upper:
            hoja_obj = nombre
            break
    if hoja_obj is None:
        hoja_obj = wb.sheetnames[0]

    ws = wb[hoja_obj]
    filas_iter = ws.iter_rows(values_only=True)
    encabezados = next(filas_iter, None)
    if not encabezados:
        return {}, []

    encabezados = [str(c).strip() if c is not None else "" for c in encabezados]
    h_upper = [h.upper() for h in encabezados]

    def find_idx(*claves):
        for i, h in enumerate(h_upper):
            for clave in claves:
                if all(p in h for p in clave.split("|")):
                    return i
        return None

    idx_tipo = find_idx("TIPO|DOCUMENTO")
    idx_cons = find_idx("CONSECUTIVO") if "ASOCIADA" not in h_upper[find_idx("CONSECUTIVO") or 0] else None
    # Mejor búsqueda: el primer "CONSECUTIVO" que NO contenga "ASOCIADA"
    idx_cons = next((i for i, h in enumerate(h_upper) if "CONSECUTIVO" in h and "ASOCIADA" not in h), None)
    idx_asoc = next((i for i, h in enumerate(h_upper) if "ASOCIADA" in h), None)
    idx_anul = next((i for i, h in enumerate(h_upper) if "ANULADO" in h), None)
    idx_fech = next((i for i, h in enumerate(h_upper) if "FECHA" in h), None)
    idx_id = next((i for i, h in enumerate(h_upper) if "IDENTIFICACI" in h), None)
    idx_cli = next((i for i, h in enumerate(h_upper) if "CLIENTE" in h), None)
    idx_suc = next((i for i, h in enumerate(h_upper) if "SUCURSAL" in h), None)
    idx_imp = next((i for i, h in enumerate(h_upper) if h.strip() == "IMPUESTO"), None)
    idx_val = next((i for i, h in enumerate(h_upper) if h.strip() == "VALOR"), None)

    if idx_cons is None:
        return {}, []

    mapping: Dict[str, str] = {}
    cabeceras: List[NotaCreditoCabecera] = []

    for fila in filas_iter:
        if not fila or len(fila) <= idx_cons:
            continue
        nc = fila[idx_cons]
        if not nc:
            continue
        nc = str(nc).strip()

        asoc = ""
        if idx_asoc is not None and idx_asoc < len(fila) and fila[idx_asoc]:
            asoc = str(fila[idx_asoc]).strip()
            mapping[nc] = asoc

        anulado = False
        if idx_anul is not None and idx_anul < len(fila):
            v = fila[idx_anul]
            if v is not None:
                v_str = str(v).strip().upper()
                anulado = v_str in ("SI", "TRUE", "1", "YES")

        fecha_obj = None
        if idx_fech is not None and idx_fech < len(fila):
            fr = fila[idx_fech]
            if isinstance(fr, datetime):
                fecha_obj = fr.date()
            elif isinstance(fr, date):
                fecha_obj = fr
            elif fr is not None:
                try:
                    fecha_obj = pd.to_datetime(fr).date()
                except (ValueError, TypeError):
                    fecha_obj = None

        ident = _formato_cc(fila[idx_id]) if idx_id is not None and idx_id < len(fila) else ""
        cliente = str(fila[idx_cli]).strip() if idx_cli is not None and idx_cli < len(fila) and fila[idx_cli] else ""
        sucursal = str(fila[idx_suc]).strip() if idx_suc is not None and idx_suc < len(fila) and fila[idx_suc] else ""
        impuesto = _to_float(fila[idx_imp]) if idx_imp is not None and idx_imp < len(fila) else 0.0
        valor = _to_float(fila[idx_val]) if idx_val is not None and idx_val < len(fila) else 0.0

        cabeceras.append(NotaCreditoCabecera(
            consecutivo=nc,
            factura_asociada=asoc,
            fecha=fecha_obj or date.today(),
            identificacion=ident,
            cliente=cliente,
            sucursal=sucursal,
            impuesto=impuesto,
            valor=valor,
            anulado=anulado,
        ))

    return mapping, cabeceras


def _generar_asientos_nca_desde_cabeceras(
    cabeceras: List[NotaCreditoCabecera],
    consecutivos_ya_procesados: set,
) -> List[dict]:
    """Construye asientos para NCA que NO vinieron en RadGrid.

    Cada NCA se compone de:
        Db 41750201 (DEVOLUCION EN VENTAS) por (Valor - Impuesto)
        Db 24080102 (IVA NC) por Impuesto
        Cr 11050500 o 130506 por Valor (según si es FE o no)

    Solo procesa NCA que no hayan sido procesadas ya en RadGrid.
    """
    filas: List[dict] = []
    comprobante = "21"

    for nc in cabeceras:
        if nc.anulado:
            continue
        # Solo procesar si es prefijo NCA y no está ya procesada
        if not nc.consecutivo.upper().startswith("NCA"):
            continue
        if nc.consecutivo in consecutivos_ya_procesados:
            continue

        documento = nc.consecutivo
        doc_ref = nc.factura_asociada or nc.consecutivo

        # Determinar si es FE (factura a empresa) por el prefijo de la asociada
        # o por el NIT del cliente
        es_factura_fe = (
            (nc.factura_asociada and nc.factura_asociada.upper().startswith("FE"))
            or (nc.identificacion and len(nc.identificacion) >= 9
                and nc.identificacion.startswith(("8", "9")))
        )

        if es_factura_fe:
            cuenta_contra = CUENTA_CLIENTES
            nit = nc.identificacion or NIT_GENERICO
        else:
            cuenta_contra = CUENTA_CAJA
            nit = NIT_GENERICO

        # Db DEVOLUCION (Valor - Impuesto) y Db IVA (Impuesto)
        valor_devol = _redondear(nc.valor - nc.impuesto)
        valor_iva = _redondear(nc.impuesto)
        valor_total = valor_devol + valor_iva

        if valor_iva > 0:
            filas.append(_crear_linea_plano(
                cuenta=CUENTA_IVA_NC,
                comprobante=comprobante,
                fecha=nc.fecha,
                documento=documento,
                doc_referencia=doc_ref,
                nit=nit,
                detalle=_detalle_para(CUENTA_IVA_NC),
                tr=TR_DEBITO,
                valor=valor_iva,
                base=valor_devol,
            ))

        if valor_devol > 0:
            filas.append(_crear_linea_plano(
                cuenta=CUENTA_DEVOLUCION,
                comprobante=comprobante,
                fecha=nc.fecha,
                documento=documento,
                doc_referencia=doc_ref,
                nit=nit,
                detalle=_detalle_para(CUENTA_DEVOLUCION),
                tr=TR_DEBITO,
                valor=valor_devol,
                base=0,
            ))

        if valor_total > 0:
            filas.append(_crear_linea_plano(
                cuenta=cuenta_contra,
                comprobante=comprobante,
                fecha=nc.fecha,
                documento=documento,
                doc_referencia=doc_ref,
                nit=nit,
                detalle=_detalle_para(cuenta_contra),
                tr=TR_CREDITO,
                valor=valor_total,
                base=0,
            ))

    return filas


# ============================================================
# Función principal
# ============================================================

def procesar_ventas_c13(
    archivo,
    archivo_notas=None,
) -> Tuple[pd.DataFrame, List[str], dict]:
    """Procesa el archivo RadGridExport y produce el plano contable.

    Args:
        archivo: file-like, bytes o path del Excel principal (RadGridExport).
        archivo_notas: opcional, archivo Documentos__3_.xlsx con la hoja
            'NOTAS CREDITO'. Si se pasa, enriquece el DOC REFERENCIA con
            la factura asociada de cada NC.

    Returns:
        (df_plano, log, resumen)
    """
    log: List[str] = []

    # Leer archivo
    log.append("📂 Leyendo archivo de facturación...")
    lineas = _leer_radgrid(archivo)
    log.append(f"  Total líneas leídas: {len(lineas)}")

    # Filtrar anuladas
    n_antes = len(lineas)
    lineas = [ln for ln in lineas if not ln.anulado]
    n_anuladas = n_antes - len(lineas)
    if n_anuladas > 0:
        log.append(f"  ⚠️ Filtradas {n_anuladas} líneas de documentos anulados (A=True).")

    # Mapeo de notas crédito y cabeceras (opcional)
    nc_map = {}
    nc_cabeceras: List[NotaCreditoCabecera] = []
    if archivo_notas:
        nc_map, nc_cabeceras = _leer_notas_credito(archivo_notas)
        if nc_map:
            log.append(f"  📋 Mapeo NC→factura cargado: {len(nc_map)} entradas.")
        if nc_cabeceras:
            log.append(f"  📋 Cabeceras NC cargadas: {len(nc_cabeceras)}.")

    # Agrupar por documento
    docs = _agrupar_por_documento(lineas)
    log.append(f"  📦 Documentos agrupados: {len(docs)}")

    # Estadísticas por tipo
    por_tipo: Dict[str, int] = {}
    desconocidos: List[str] = []
    todas_filas: List[dict] = []

    for doc_key in sorted(docs.keys()):
        doc = docs[doc_key]
        comp_tipo = _detectar_comprobante(doc.documento)
        if not comp_tipo:
            desconocidos.append(doc.documento)
            continue
        comprobante, tipo = comp_tipo
        por_tipo[tipo] = por_tipo.get(tipo, 0) + 1

        # Generar líneas según tipo
        if tipo in ("VTA_PVFE", "VTA_GBF"):
            filas_doc = _generar_asiento_venta_contado(doc, comprobante)
        elif tipo == "VTA_FE":
            filas_doc = _generar_asiento_venta_fe(doc, comprobante)
        elif tipo in ("NC_PVFE", "NC_GBF", "NC_DW", "NC_NCA"):
            filas_doc = _generar_asiento_nota_credito(doc, comprobante, tipo)
            # Aplicar mapeo NC→factura si está disponible
            if nc_map and doc.documento in nc_map:
                factura_asociada = nc_map[doc.documento]
                for f in filas_doc:
                    f["DOC REFERENCIA"] = factura_asociada
        else:
            filas_doc = []

        todas_filas.extend(filas_doc)

    # Reportar tipos detectados
    log.append("")
    log.append("📊 Documentos por tipo:")
    nombres_tipo = {p[2]: p[3] for p in MAPEO_PREFIJO_COMPROBANTE}
    for tipo, count in sorted(por_tipo.items()):
        nombre = nombres_tipo.get(tipo, tipo)
        comp = next((p[1] for p in MAPEO_PREFIJO_COMPROBANTE if p[2] == tipo), "?")
        log.append(f"   • {nombre} (Comp {comp}): {count} documentos")

    # Procesar NCA desde Documentos__3_ (no vienen en RadGrid)
    if nc_cabeceras:
        consecutivos_radgrid = set(docs.keys())
        filas_nca = _generar_asientos_nca_desde_cabeceras(nc_cabeceras, consecutivos_radgrid)
        if filas_nca:
            n_ncas = len(set(f["DOCUMENTO"] for f in filas_nca))
            log.append(f"   • Nota crédito NCA (Comp 21): {n_ncas} documentos (del archivo Documentos)")
            por_tipo["NC_NCA"] = n_ncas
            todas_filas.extend(filas_nca)

    if desconocidos:
        log.append(f"   ⚠️ {len(desconocidos)} documentos con prefijo desconocido: "
                   f"{desconocidos[:5]}{'...' if len(desconocidos) > 5 else ''}")

    # Construir DataFrame
    df_plano = pd.DataFrame(todas_filas, columns=COLUMNAS_PLANO)

    # Validación de cuadre por (comprobante, documento)
    log.append("")
    log.append("🧮 Verificación de cuadre por documento...")
    descuadres = []
    if len(df_plano) > 0:
        for (comp, doc_x), grupo in df_plano.groupby(["COMPROBANTE", "DOCUMENTO"]):
            db = int(grupo[grupo["TR"] == TR_DEBITO]["VALOR"].astype(int).sum())
            cr = int(grupo[grupo["TR"] == TR_CREDITO]["VALOR"].astype(int).sum())
            if db != cr:
                descuadres.append((comp, doc_x, db, cr, db - cr))

    if descuadres:
        log.append(f"   ⚠️ {len(descuadres)} documentos con descuadre:")
        for comp, doc_x, db, cr, dif in descuadres[:5]:
            log.append(f"      Comp {comp} Doc {doc_x}: Db=${db:,} Cr=${cr:,} dif=${dif:,}".replace(",","."))
    else:
        log.append("   ✅ Todos los documentos cuadran Db = Cr.")

    # Resumen general
    log.append("")
    log.append("=" * 60)
    log.append("📊 RESUMEN GENERAL")
    log.append("=" * 60)
    if len(df_plano) > 0:
        total_db = int(df_plano[df_plano["TR"] == TR_DEBITO]["VALOR"].astype(int).sum())
        total_cr = int(df_plano[df_plano["TR"] == TR_CREDITO]["VALOR"].astype(int).sum())
        log.append(f"   Total líneas: {len(df_plano)}")
        log.append(f"   Total Db:    $ {total_db:,}".replace(",", "."))
        log.append(f"   Total Cr:    $ {total_cr:,}".replace(",", "."))
        if total_db == total_cr:
            log.append("   ✅ Cuadre perfecto Db = Cr")
        else:
            log.append(f"   ❌ DESCUADRE: dif $ {total_db - total_cr:,}".replace(",", "."))

        # Resumen por comprobante
        log.append("")
        log.append("   Asientos por comprobante:")
        for comp, grupo in df_plano.groupby("COMPROBANTE"):
            db = int(grupo[grupo["TR"] == TR_DEBITO]["VALOR"].astype(int).sum())
            cr = int(grupo[grupo["TR"] == TR_CREDITO]["VALOR"].astype(int).sum())
            n_docs = grupo["DOCUMENTO"].nunique()
            log.append(f"      Comp {comp}: {n_docs} docs, {len(grupo)} líneas, Db=Cr=${db:,}".replace(",","."))

    resumen = {
        "total_lineas": len(df_plano),
        "total_documentos": len(docs),
        "anuladas_filtradas": n_anuladas,
        "por_tipo": por_tipo,
        "descuadres": len(descuadres),
        "comprobantes": sorted(df_plano["COMPROBANTE"].unique().tolist()) if len(df_plano) > 0 else [],
    }

    return df_plano, log, resumen


# ============================================================
# Exportar a TSV (mismo formato que los demás módulos)
# ============================================================

def dataframe_a_plano_tsv(
    df: pd.DataFrame,
    incluir_encabezado_excel: bool = True,
) -> bytes:
    """Convierte el DataFrame a bytes TSV compatible con el sistema contable."""
    df_out = df[COLUMNAS_PLANO].copy()
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str).str.replace("\t", " ", regex=False)
    tsv = df_out.to_csv(sep="\t", index=False, header=True, lineterminator="\r\n")
    if incluir_encabezado_excel:
        tsv = "sep=\t\r\n" + tsv
    return tsv.encode("utf-8")
