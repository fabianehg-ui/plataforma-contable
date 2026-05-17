"""
core/procesadores/reporte_comparativo_pos.py

Genera el reporte comparativo VENTAS POS REPORTADAS vs DIAN.

Compara, por sucursal:
  [A] POS reportado (del plano POS generado por el módulo Ingresos POS)
  [B] FE POS emitidas en DIAN (Token, prefijos_token de cada sucursal)
  [C] NC POS emitidas en DIAN (Token, prefijos_token_nc de cada sucursal)
  [D] NETO DIAN = B - C
  [E] DIFERENCIA = A - D

Si |E| > UMBRAL_ALERTA → ALERTA (no debería pasar, requiere revisión).
"""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook


# ============================================================
# Constantes
# ============================================================

# Umbral por encima del cual la diferencia se marca como ALERTA
UMBRAL_ALERTA_DEFAULT = 50_000  # $50.000

# Columnas esperadas en el plano POS (formato Silla3)
COL_PLANO_CUENTA = "CUENTA"
COL_PLANO_FECHA  = "FECHA"
COL_PLANO_VALOR  = "VALOR"
COL_PLANO_TR     = "TR"
COL_PLANO_CC     = "CENTRO DE COSTO"

# Cuentas del POS que representan ventas (no IVA, no INC, no caja)
# Estas son las cuentas "cta_base_v" del JSON datos_punto
CUENTAS_VENTA_POS = {"41401501", "41401503", "41200902", "41200906", "41200905"}


# ============================================================
# Funciones auxiliares
# ============================================================

def _normalizar_nit(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _parsear_fecha_universal(v):
    """Parsea cualquier formato común de fecha."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            pass
    return None


def _to_number(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        try:
            return float(str(v).replace(",", "").replace("$", "").strip())
        except Exception:
            return 0.0


# ============================================================
# Extracción de datos POS desde el plano
# ============================================================

def extraer_pos_desde_plano(
    df_plano: pd.DataFrame,
    cuentas_venta: set | None = None,
    anio: int | None = None,
    mes: int | None = None,
) -> dict:
    """
    Del plano POS generado, extrae el total reportado por CC y por fecha.

    El plano tiene líneas Db (caja) y Cr (ventas + INC).
    Para evitar duplicar, sumamos los CRÉDITOS de las cuentas de venta
    (no INC), que es exactamente lo facturable/comparable contra DIAN.

    Returns:
        dict {
            'por_cc':   { CC: {'total': float, 'docs': int, 'fechas': [date,...]} },
            'por_fecha_cc': { (fecha, CC): float },
            'total':    float,
        }
    """
    if cuentas_venta is None:
        cuentas_venta = CUENTAS_VENTA_POS

    if df_plano is None or len(df_plano) == 0:
        return {"por_cc": {}, "por_fecha_cc": {}, "total": 0.0}

    df = df_plano.copy()
    # Normalizar nombres de columnas: pueden venir como "CUENTA" o "Cuenta"
    df.columns = [str(c).strip().upper() for c in df.columns]

    # Detectar columnas alternativas
    col_cuenta = None
    col_fecha  = None
    col_valor  = None
    col_tr     = None
    col_cc     = None
    for c in df.columns:
        if c in ("CUENTA", "ACCOUNT"): col_cuenta = c
        elif c in ("FECHA", "FECHA(MM/DD/YYYY)", "DATE"): col_fecha = c
        elif c in ("VALOR", "VALUE", "AMOUNT"): col_valor = c
        elif c in ("TR", "TIPO", "TYPE"): col_tr = c
        elif c in ("CENTRO DE COSTO", "CC", "CENTRO COSTO"): col_cc = c

    if not all([col_cuenta, col_valor, col_tr, col_cc]):
        return {"por_cc": {}, "por_fecha_cc": {}, "total": 0.0,
                "error": "Faltan columnas del plano POS"}

    # Filtrar cuentas de venta y créditos (TR=2)
    df[col_cuenta] = df[col_cuenta].astype(str).str.strip()
    df[col_tr]     = df[col_tr].astype(str).str.strip()
    df[col_valor]  = df[col_valor].apply(_to_number)

    mask = df[col_cuenta].isin(cuentas_venta) & (df[col_tr] == "2")
    df_v = df[mask].copy()

    # Parsear fechas y filtrar por mes/año si se pidió
    if col_fecha:
        df_v["_fecha"] = df_v[col_fecha].apply(_parsear_fecha_universal)
        if anio is not None:
            df_v = df_v[df_v["_fecha"].apply(lambda f: f is not None and f.year == anio)]
        if mes is not None:
            df_v = df_v[df_v["_fecha"].apply(lambda f: f is not None and f.month == mes)]

    # Agrupar por CC
    df_v[col_cc] = df_v[col_cc].astype(str).str.strip()
    por_cc = {}
    for cc, grp in df_v.groupby(col_cc):
        por_cc[cc] = {
            "total":   round(grp[col_valor].sum(), 2),
            "docs":    len(grp),
            "fechas":  sorted({f for f in grp["_fecha"].dropna()}) if col_fecha else [],
        }

    # Por (fecha, CC)
    por_fecha_cc = {}
    if col_fecha:
        for (f, cc), grp in df_v.groupby(["_fecha", col_cc]):
            por_fecha_cc[(f, cc)] = round(grp[col_valor].sum(), 2)

    return {
        "por_cc": por_cc,
        "por_fecha_cc": por_fecha_cc,
        "total": round(df_v[col_valor].sum(), 2),
    }


# ============================================================
# Extracción de datos del Token DIAN
# ============================================================

def extraer_dian_desde_token(
    fuente_token,
    datos_punto: list,
    nit_empresa: str,
    anio: int,
    mes: int,
) -> dict:
    """
    Lee el Token DIAN y extrae:
      - FE POS DIAN agrupado por sucursal (mapeado vía prefijo_token)
      - NC POS DIAN agrupado por sucursal (mapeado vía prefijo_token_nc)

    Args:
        fuente_token: ruta/bytes del Excel del Token
        datos_punto: lista de sucursales (del datos_punto.json)
        nit_empresa: NIT del emisor a filtrar
        anio, mes: período

    Returns:
        dict {
            'fe_por_cc': { CC: {'total': float, 'docs': int, 'prefijo': str, 'nombre': str} },
            'nc_por_cc': { CC: {'total': float, 'docs': int, 'prefijo_nc': str} },
            'fe_total': float,
            'nc_total': float,
            'fe_huérfanas': { prefijo: {'total', 'docs'} },  # FE de prefijos no mapeados
            'nc_huérfanas': { prefijo: {'total', 'docs'} },
        }
    """
    # Cargar Token
    if hasattr(fuente_token, "read"):
        contenido = fuente_token.read()
        if hasattr(fuente_token, "seek"):
            fuente_token.seek(0)
        bio = io.BytesIO(contenido)
    elif isinstance(fuente_token, (bytes, bytearray)):
        bio = io.BytesIO(bytes(fuente_token))
    elif isinstance(fuente_token, str):
        bio = fuente_token
    else:
        raise TypeError(f"Tipo de fuente no soportado: {type(fuente_token)}")

    wb = load_workbook(bio, read_only=True, data_only=True)
    ws = wb.active

    headers = None
    filas = []
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(c).strip() if c else "" for c in row]
            continue
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        fila = {h: v for h, v in zip(headers, row)}
        filas.append(fila)
    wb.close()

    # Mapeo prefijo → CC
    pref_to_cc_fe = {}
    pref_to_cc_nc = {}
    cc_a_nombre   = {}
    for s in datos_punto:
        p = (s.get("prefijo_token") or "").strip().upper()
        if p:
            pref_to_cc_fe[p] = s["cc"]
            cc_a_nombre[s["cc"]] = s.get("nombre_reporte", s.get("sede", s["cc"]))
        pnc = (s.get("prefijo_token_nc") or "").strip().upper()
        if pnc:
            pref_to_cc_nc[pnc] = s["cc"]

    # Procesar
    fe_por_cc = {}
    nc_por_cc = {}
    fe_huerfanas = {}
    nc_huerfanas = {}

    for f in filas:
        nit_em = _normalizar_nit(f.get("NIT Emisor"))
        if nit_em != nit_empresa:
            continue

        fecha = _parsear_fecha_universal(f.get("Fecha Emisión"))
        if not fecha or fecha.year != anio or fecha.month != mes:
            continue

        tipo = str(f.get("Tipo de documento") or "").strip()
        pref = str(f.get("Prefijo") or "").strip().upper()
        total = _to_number(f.get("Total"))

        if tipo == "Factura electrónica":
            if pref in pref_to_cc_fe:
                cc = pref_to_cc_fe[pref]
                if cc not in fe_por_cc:
                    fe_por_cc[cc] = {"total": 0.0, "docs": 0, "prefijo": pref, "nombre": cc_a_nombre.get(cc, cc)}
                fe_por_cc[cc]["total"] += total
                fe_por_cc[cc]["docs"]  += 1
            else:
                # Prefijo no mapeado (STL, etc) — solo notamos
                if pref not in fe_huerfanas:
                    fe_huerfanas[pref] = {"total": 0.0, "docs": 0}
                fe_huerfanas[pref]["total"] += total
                fe_huerfanas[pref]["docs"]  += 1

        elif tipo == "Nota de crédito electrónica":
            if pref in pref_to_cc_nc:
                cc = pref_to_cc_nc[pref]
                if cc not in nc_por_cc:
                    nc_por_cc[cc] = {"total": 0.0, "docs": 0, "prefijo_nc": pref, "nombre": cc_a_nombre.get(cc, cc)}
                nc_por_cc[cc]["total"] += total
                nc_por_cc[cc]["docs"]  += 1
            else:
                if pref not in nc_huerfanas:
                    nc_huerfanas[pref] = {"total": 0.0, "docs": 0}
                nc_huerfanas[pref]["total"] += total
                nc_huerfanas[pref]["docs"]  += 1

    # Redondear
    for d in (fe_por_cc, nc_por_cc):
        for k in d:
            d[k]["total"] = round(d[k]["total"], 2)
    for d in (fe_huerfanas, nc_huerfanas):
        for k in d:
            d[k]["total"] = round(d[k]["total"], 2)

    return {
        "fe_por_cc":      fe_por_cc,
        "nc_por_cc":      nc_por_cc,
        "fe_huerfanas":   fe_huerfanas,
        "nc_huerfanas":   nc_huerfanas,
        "fe_total":       round(sum(v["total"] for v in fe_por_cc.values()), 2),
        "nc_total":       round(sum(v["total"] for v in nc_por_cc.values()), 2),
        "fe_total_huerfanas": round(sum(v["total"] for v in fe_huerfanas.values()), 2),
        "nc_total_huerfanas": round(sum(v["total"] for v in nc_huerfanas.values()), 2),
    }


# ============================================================
# Reporte comparativo (la función principal)
# ============================================================

def construir_reporte_comparativo(
    df_plano_pos: pd.DataFrame,
    fuente_token,
    datos_punto: list,
    nit_empresa: str,
    anio: int,
    mes: int,
    umbral_alerta: float = UMBRAL_ALERTA_DEFAULT,
) -> dict:
    """
    Genera el reporte comparativo completo.

    Args:
        df_plano_pos: DataFrame del plano POS generado por el módulo Ingresos POS.
        fuente_token: ruta/bytes del Excel del Token DIAN.
        datos_punto: lista de sucursales del datos_punto.json.
        nit_empresa: NIT de la empresa (emisor).
        anio, mes: período.
        umbral_alerta: monto en pesos para considerar la diferencia como alerta.

    Returns:
        dict con:
            'df_comparativo':  pd.DataFrame con una fila por sucursal y columnas
                               POS_REPORTADO, FE_DIAN, NC_DIAN, NETO_DIAN, DIFERENCIA, ALERTA
            'totales':         dict con sumas globales
            'alertas':         list de sucursales con diferencias > umbral
            'huérfanas':       prefijos no mapeados
    """
    # 1. POS reportado (del plano)
    pos = extraer_pos_desde_plano(df_plano_pos, anio=anio, mes=mes)

    # 2. DIAN (FE y NC del Token)
    dian = extraer_dian_desde_token(
        fuente_token, datos_punto, nit_empresa, anio, mes
    )

    # 3. Unir por CC
    todos_ccs = set(pos["por_cc"].keys()) | set(dian["fe_por_cc"].keys()) | set(dian["nc_por_cc"].keys())

    filas_reporte = []
    alertas = []
    cc_a_nombre = {}
    for s in datos_punto:
        cc_a_nombre[s["cc"]] = s.get("nombre_reporte", s.get("sede", s["cc"]))

    for cc in sorted(todos_ccs):
        pos_total = pos["por_cc"].get(cc, {}).get("total", 0.0)
        fe = dian["fe_por_cc"].get(cc, {})
        nc = dian["nc_por_cc"].get(cc, {})
        fe_total = fe.get("total", 0.0)
        nc_total = nc.get("total", 0.0)
        neto_dian = round(fe_total - nc_total, 2)
        diferencia = round(pos_total - neto_dian, 2)
        alerta = abs(diferencia) > umbral_alerta

        nombre = cc_a_nombre.get(cc, cc)
        if not nombre or nombre == cc:
            nombre = fe.get("nombre") or nc.get("nombre") or cc

        fila = {
            "CC":             cc,
            "SUCURSAL":       nombre,
            "POS_REPORTADO":  pos_total,
            "FE_DIAN":        fe_total,
            "NC_DIAN":        nc_total,
            "NETO_DIAN":      neto_dian,
            "DIFERENCIA":     diferencia,
            "FE_DOCS":        fe.get("docs", 0),
            "NC_DOCS":        nc.get("docs", 0),
            "ALERTA":         "🔴" if alerta else "✅",
        }
        filas_reporte.append(fila)

        if alerta:
            alertas.append({
                "cc":          cc,
                "sucursal":    nombre,
                "diferencia":  diferencia,
                "tipo":        "exceso_pos" if diferencia > 0 else "exceso_dian",
                "mensaje":     (
                    f"POS reportado (${pos_total:,.0f}) supera lo facturado en DIAN (${neto_dian:,.0f})"
                    if diferencia > 0 else
                    f"DIAN (${neto_dian:,.0f}) excede lo reportado POS (${pos_total:,.0f})"
                ),
            })

    df_rep = pd.DataFrame(filas_reporte)

    # Totales
    totales = {
        "pos_total":           pos["total"],
        "fe_total":            dian["fe_total"],
        "nc_total":            dian["nc_total"],
        "neto_dian_total":     round(dian["fe_total"] - dian["nc_total"], 2),
        "diferencia_total":    round(pos["total"] - (dian["fe_total"] - dian["nc_total"]), 2),
        "fe_huerfanas_total":  dian["fe_total_huerfanas"],
        "nc_huerfanas_total":  dian["nc_total_huerfanas"],
    }
    totales["alerta_total"] = abs(totales["diferencia_total"]) > umbral_alerta

    return {
        "df_comparativo":  df_rep,
        "totales":         totales,
        "alertas":         alertas,
        "fe_huerfanas":    dian["fe_huerfanas"],  # FE de prefijos no mapeados (STL, etc.)
        "nc_huerfanas":    dian["nc_huerfanas"],
        "umbral_alerta":   umbral_alerta,
        "periodo":         {"anio": anio, "mes": mes},
    }


# ============================================================
# Exportadores
# ============================================================

def reporte_a_xlsx(reporte: dict) -> bytes:
    """Exporta el reporte comparativo a Excel con varias hojas."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        # Hoja 1: Comparativo por sucursal
        reporte["df_comparativo"].to_excel(w, sheet_name="Comparativo", index=False)

        # Hoja 2: Totales y alerta
        totales = reporte["totales"]
        meta = reporte["periodo"]
        df_tot = pd.DataFrame([
            {"Concepto": "Período", "Valor": f"{meta['anio']}-{meta['mes']:02d}"},
            {"Concepto": "Umbral alerta", "Valor": f"${reporte['umbral_alerta']:,.0f}"},
            {"Concepto": "", "Valor": ""},
            {"Concepto": "[A] POS reportado total",   "Valor": totales["pos_total"]},
            {"Concepto": "[B] FE POS DIAN total",     "Valor": totales["fe_total"]},
            {"Concepto": "[C] NC POS DIAN total",     "Valor": totales["nc_total"]},
            {"Concepto": "[D] NETO DIAN = B - C",     "Valor": totales["neto_dian_total"]},
            {"Concepto": "[E] DIFERENCIA = A - D",    "Valor": totales["diferencia_total"]},
            {"Concepto": "¿ALERTA?",                  "Valor": "🔴 SÍ" if totales["alerta_total"] else "✅ NO"},
            {"Concepto": "", "Valor": ""},
            {"Concepto": "FE huérfanas (STL, etc.)",  "Valor": totales["fe_huerfanas_total"]},
            {"Concepto": "NC huérfanas (NC STL)",     "Valor": totales["nc_huerfanas_total"]},
        ])
        df_tot.to_excel(w, sheet_name="Totales", index=False)

        # Hoja 3: Alertas detalladas
        if reporte["alertas"]:
            pd.DataFrame(reporte["alertas"]).to_excel(w, sheet_name="Alertas", index=False)

        # Hoja 4: Prefijos huérfanos
        if reporte["fe_huerfanas"] or reporte["nc_huerfanas"]:
            df_huer = pd.DataFrame([
                {"Tipo": "FE", "Prefijo": p, "Docs": d["docs"], "Total": d["total"]}
                for p, d in reporte["fe_huerfanas"].items()
            ] + [
                {"Tipo": "NC", "Prefijo": p, "Docs": d["docs"], "Total": d["total"]}
                for p, d in reporte["nc_huerfanas"].items()
            ])
            df_huer.to_excel(w, sheet_name="Prefijos huérfanos", index=False)

    buf.seek(0)
    return buf.read()
