"""
core/procesadores/reporte_auditoria_xlsx.py

Genera el Excel de auditoría POS vs Token con formato profesional:
  - Cabecera con título, empresa, periodo
  - Tabla con colores por clase de sede
  - Subtotales destacados
  - Columna Efectividad con código de color (verde/amarillo/rojo)
  - Pie con resumen ejecutivo

EXPONE:
    generar_excel_auditoria(df_comp, ruta_salida, meta) -> bytes
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ─── Paletas de colores ──────────────────────────────────────

COLOR_HEADER_BG = "1F4E79"        # azul oscuro
COLOR_HEADER_FG = "FFFFFF"

COLOR_TITULO_BG = "203864"
COLOR_TITULO_FG = "FFFFFF"

COLOR_CLASE_SL_BG = "FFF2CC"      # amarillo suave SANTA LEÑA
COLOR_CLASE_RM_BG = "FCE4D6"      # naranja suave RESTAURANTE MILAGROS
COLOR_CLASE_STL_BG = "E2EFDA"     # verde suave STL

COLOR_SUBTOTAL_BG = "D9E1F2"      # azul claro
COLOR_GRAN_TOTAL_BG = "BDD7EE"    # azul medio
COLOR_GRAN_TOTAL_FG = "1F4E79"

COLOR_EFEC_OK = "C6EFCE"          # verde
COLOR_EFEC_OK_FG = "006100"
COLOR_EFEC_WARN = "FFEB9C"        # amarillo
COLOR_EFEC_WARN_FG = "9C5700"
COLOR_EFEC_BAD = "FFC7CE"         # rojo
COLOR_EFEC_BAD_FG = "9C0006"

COLOR_BORDER = "BFBFBF"


def _fill(color_hex: str) -> PatternFill:
    return PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")


def _border(thick: bool = False) -> Border:
    style = "medium" if thick else "thin"
    side = Side(border_style=style, color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _color_efectividad(pct: Optional[float]) -> tuple:
    """Devuelve (bg, fg) según rango de efectividad."""
    if pct is None:
        return ("F2F2F2", "595959")
    if pct >= 98 and pct <= 102:
        return (COLOR_EFEC_OK, COLOR_EFEC_OK_FG)
    if pct >= 95 and pct <= 105:
        return (COLOR_EFEC_WARN, COLOR_EFEC_WARN_FG)
    return (COLOR_EFEC_BAD, COLOR_EFEC_BAD_FG)


def generar_excel_auditoria(
    df_comp: pd.DataFrame,
    fecha_desde: str,
    fecha_hasta: str,
    empresa_nombre: str = "JIPER S.A.S",
    empresa_nit: str = "901038325",
) -> bytes:
    """
    Genera el archivo Excel y devuelve los bytes listos para descarga.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoría POS vs Token"

    # ─── 1. Cabecera principal ───────────────────────────────
    ws.merge_cells("A1:I1")
    c = ws.cell(row=1, column=1, value=f"📊 REPORTE DE AUDITORÍA — POS vs Token DIAN")
    c.font = Font(name="Calibri", size=16, bold=True, color=COLOR_TITULO_FG)
    c.fill = _fill(COLOR_TITULO_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    c = ws.cell(row=2, column=1, value=f"{empresa_nombre}  ·  NIT {empresa_nit}  ·  Periodo: {fecha_desde} → {fecha_hasta}")
    c.font = Font(name="Calibri", size=11, italic=True, color="595959")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:I3")
    c = ws.cell(row=3, column=1,
                value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    c.font = Font(name="Calibri", size=9, color="808080")
    c.alignment = Alignment(horizontal="center")

    # Fila 4 vacía
    fila_actual = 5

    # ─── 2. Encabezados de tabla ─────────────────────────────
    columnas = ["CC", "Sucursal", "Clase", "Base POS",
                "Base Token FE", "Base NC Token", "Token Neto",
                "Diferencia", "Efectividad %"]

    for i, col in enumerate(columnas, start=1):
        c = ws.cell(row=fila_actual, column=i, value=col)
        c.font = Font(name="Calibri", size=11, bold=True, color=COLOR_HEADER_FG)
        c.fill = _fill(COLOR_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
    ws.row_dimensions[fila_actual].height = 32

    fila_actual += 1

    # ─── 3. Filas de datos ───────────────────────────────────
    for _, row in df_comp.iterrows():
        es_subtotal = "SUBTOTAL" in str(row.get("Sucursal", "")).upper() or "──" in str(row.get("Sucursal", ""))
        es_gran_total = "GRAN TOTAL" in str(row.get("Sucursal", "")).upper() or "═══" in str(row.get("Sucursal", ""))

        clase = str(row.get("Clase", "")).strip().upper()
        if "SANTA LEÑA" in clase or "SANTA LENA" in clase:
            color_fila = COLOR_CLASE_SL_BG
        elif "MILAGROS" in clase:
            color_fila = COLOR_CLASE_RM_BG
        elif "STL" in clase:
            color_fila = COLOR_CLASE_STL_BG
        else:
            color_fila = "FFFFFF"

        if es_gran_total:
            color_fila = COLOR_GRAN_TOTAL_BG
        elif es_subtotal:
            color_fila = COLOR_SUBTOTAL_BG

        valores = [
            row.get("CC", ""),
            row.get("Sucursal", ""),
            row.get("Clase", ""),
            row.get("Base POS", ""),
            row.get("Base Token FE", ""),
            row.get("Base NC Token", ""),
            row.get("Token Neto", ""),
            row.get("Diferencia", ""),
            row.get("Efectividad %", None),
        ]

        for i, val in enumerate(valores, start=1):
            c = ws.cell(row=fila_actual, column=i, value=val if val != "" else None)

            # Formato numérico para columnas de plata
            if i in (4, 5, 6, 7, 8):
                c.number_format = '"$ "#,##0;[Red]"$ -"#,##0'
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif i == 9:  # Efectividad
                c.number_format = '0.00"%"'
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif i in (1, 3):
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")

            # Estilo de la fila
            c.fill = _fill(color_fila)

            if es_gran_total:
                c.font = Font(name="Calibri", size=12, bold=True, color=COLOR_GRAN_TOTAL_FG)
            elif es_subtotal:
                c.font = Font(name="Calibri", size=11, bold=True)
            else:
                c.font = Font(name="Calibri", size=10)

            c.border = _border(thick=es_gran_total)

        # Color especial en la celda de Efectividad
        ef_val = row.get("Efectividad %", None)
        if pd.notna(ef_val) and not es_subtotal and not es_gran_total:
            try:
                ef_float = float(ef_val)
                bg, fg = _color_efectividad(ef_float)
                cell_ef = ws.cell(row=fila_actual, column=9)
                cell_ef.fill = _fill(bg)
                cell_ef.font = Font(name="Calibri", size=10, bold=True, color=fg)
            except (ValueError, TypeError):
                pass

        ws.row_dimensions[fila_actual].height = 22 if es_gran_total or es_subtotal else 18
        fila_actual += 1

    # ─── 4. Resumen ejecutivo ────────────────────────────────
    fila_actual += 2
    ws.merge_cells(f"A{fila_actual}:I{fila_actual}")
    c = ws.cell(row=fila_actual, column=1, value="📋 LEYENDA DE EFECTIVIDAD")
    c.font = Font(name="Calibri", size=12, bold=True, color=COLOR_TITULO_FG)
    c.fill = _fill(COLOR_TITULO_BG)
    c.alignment = Alignment(horizontal="center")
    fila_actual += 1

    leyendas = [
        ("✅ Cuadre OK", "98% – 102%", COLOR_EFEC_OK, COLOR_EFEC_OK_FG),
        ("⚠️ Revisar", "95% – 98% ó 102% – 105%", COLOR_EFEC_WARN, COLOR_EFEC_WARN_FG),
        ("❌ Descuadre", "< 95% ó > 105%", COLOR_EFEC_BAD, COLOR_EFEC_BAD_FG),
    ]
    for desc, rango, bg, fg in leyendas:
        c = ws.cell(row=fila_actual, column=1, value=desc)
        c.font = Font(name="Calibri", size=10, bold=True, color=fg)
        c.fill = _fill(bg)
        c.alignment = Alignment(horizontal="center")
        c.border = _border()
        ws.merge_cells(start_row=fila_actual, start_column=2,
                       end_row=fila_actual, end_column=3)
        c2 = ws.cell(row=fila_actual, column=2, value=rango)
        c2.font = Font(name="Calibri", size=10)
        c2.alignment = Alignment(horizontal="left", indent=1)
        c2.border = _border()
        fila_actual += 1

    # ─── 5. Anchos de columna ────────────────────────────────
    anchos = {
        1: 10,   # CC
        2: 32,   # Sucursal
        3: 22,   # Clase
        4: 18,   # Base POS
        5: 18,   # Base Token FE
        6: 16,   # Base NC Token
        7: 18,   # Token Neto
        8: 16,   # Diferencia
        9: 14,   # Efectividad %
    }
    for col_num, ancho in anchos.items():
        ws.column_dimensions[get_column_letter(col_num)].width = ancho

    # Congelar paneles arriba de los datos
    ws.freeze_panes = "A6"

    # ─── 6. Exportar a bytes ─────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
