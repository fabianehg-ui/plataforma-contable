"""
Importador de balance de prueba en formato Siigo.

Reconoce el formato estándar de exportación "Balance de Prueba por Cuenta (Normal)"
de Siigo Contador, lee las cuentas PUC con sus saldos al cierre y mapea agregaciones
hacia las casillas del Formulario 110.

Formato esperado del XLSX:
    Hoja única, primeras filas con título y subtítulo, encabezados en fila 4:
        Cuenta | Equivalencia | Nombre | Saldo Anterior | Débitos | Créditos | Nuevo Saldo

Las cuentas tienen jerarquía PUC: 1, 11, 1105, 110505, 11050501. Solo se utilizan
los saldos a 4 dígitos para no doble-contar (Siigo presenta totales en cada nivel).
"""

from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, Iterable
import logging

log = logging.getLogger(__name__)


@dataclass
class CuentaPUC:
    """Representación de una cuenta del PUC con sus saldos."""
    codigo: str          # Ej: "1105", "5205"
    nombre: str          # Ej: "CAJA"
    saldo_anterior: float = 0.0
    debitos: float = 0.0
    creditos: float = 0.0
    saldo_nuevo: float = 0.0  # ← este es el que normalmente se usa

    @property
    def nivel(self) -> int:
        """Nivel jerárquico: 1=clase (1), 2=grupo (11), 3=mayor (1105), 4=auxiliar+."""
        n = len(self.codigo.strip())
        if n <= 1: return 1
        if n <= 2: return 2
        if n <= 4: return 3
        return 4

    @property
    def codigo_4digitos(self) -> str:
        """Truncado a 4 dígitos para agregaciones."""
        c = self.codigo.strip()
        return c[:4] if len(c) >= 4 else c.ljust(4, '0')


@dataclass
class BalanceSiigo:
    """Balance de Siigo parseado, indexado por código de cuenta."""
    razon_social: str = ""
    nit: str = ""
    periodo_desde: str = ""
    periodo_hasta: str = ""
    cuentas: dict[str, CuentaPUC] = field(default_factory=dict)

    def saldo_cuenta(self, codigo: str) -> float:
        """Retorna el saldo nuevo de una cuenta específica."""
        return self.cuentas.get(codigo, CuentaPUC(codigo=codigo, nombre="")).saldo_nuevo

    def saldo_grupo(self, prefijo: str) -> float:
        """
        Retorna el saldo agregado de todas las cuentas que empiezan por `prefijo`.
        Toma sólo las cuentas a 4 dígitos para evitar doble conteo.
        """
        return sum(
            c.saldo_nuevo for codigo, c in self.cuentas.items()
            if codigo.startswith(prefijo) and c.nivel == 3
        )

    def saldo_clase(self, prefijo: str) -> float:
        """
        Retorna el saldo de la cuenta de clase exacta (ej. '4' para ingresos).
        Si existe la cuenta resumen, la usa; si no, agrega los grupos.
        """
        cuenta_resumen = self.cuentas.get(prefijo)
        if cuenta_resumen and cuenta_resumen.nivel <= 2:
            return cuenta_resumen.saldo_nuevo
        # Fallback: sumar todas las cuentas de nivel 3 que empiecen por el prefijo
        return self.saldo_grupo(prefijo)


class ImportadorBalanceSiigo:
    """Lee y parsea un XLSX de Balance de Prueba exportado por Siigo Contador."""

    HEADER_ROW = 4
    COL_CUENTA = 0
    COL_EQUIV = 1
    COL_NOMBRE = 2
    COL_SALDO_ANT = 3
    COL_DEBITOS = 4
    COL_CREDITOS = 5
    COL_SALDO_NUEVO = 6

    def importar(self, ruta_xlsx: str | Path) -> BalanceSiigo:
        """Carga un XLSX de Siigo y devuelve un BalanceSiigo con las cuentas indexadas."""
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ImportError(
                "openpyxl es requerido para importar balances Siigo. "
                "Instale con: pip install openpyxl"
            ) from e

        ruta = Path(ruta_xlsx)
        if not ruta.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {ruta}")

        wb = load_workbook(ruta, data_only=True, read_only=True)
        ws = wb.active  # Siigo siempre exporta una sola hoja

        balance = BalanceSiigo()

        # Extraer metadata de las primeras filas (título)
        try:
            primera_fila = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            if primera_fila[0]:
                titulo = str(primera_fila[0])
                # Formato típico: "RAZÓN SOCIAL - NIT"
                if " - " in titulo:
                    partes = titulo.split(" - ", 1)
                    balance.razon_social = partes[0].strip()
                    balance.nit = partes[1].strip()
                else:
                    balance.razon_social = titulo.strip()
        except StopIteration:
            pass

        # Período (fila 3 típicamente: "Desde Ene-01-2025 Hasta Dic-31-2025")
        try:
            fila3 = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))
            if fila3 and fila3[0][0]:
                periodo = str(fila3[0][0])
                if "Desde" in periodo and "Hasta" in periodo:
                    desde = periodo.split("Desde")[1].split("Hasta")[0].strip()
                    hasta = periodo.split("Hasta")[1].strip()
                    balance.periodo_desde = desde
                    balance.periodo_hasta = hasta
        except (IndexError, StopIteration):
            pass

        # Leer cuentas desde la fila 5 (después del encabezado en fila 4)
        cuentas_leidas = 0
        for row in ws.iter_rows(min_row=self.HEADER_ROW + 1, values_only=True):
            cuenta = self._parsear_fila(row)
            if cuenta is None:
                continue
            balance.cuentas[cuenta.codigo] = cuenta
            cuentas_leidas += 1

        log.info(
            "Balance Siigo importado: %s — %d cuentas — período %s a %s",
            balance.razon_social, cuentas_leidas, balance.periodo_desde, balance.periodo_hasta
        )
        return balance

    def _parsear_fila(self, row: tuple) -> Optional[CuentaPUC]:
        """Convierte una fila del XLSX en una CuentaPUC, o None si la fila es inválida."""
        if not row or len(row) < 7:
            return None

        codigo_raw = row[self.COL_CUENTA]
        if codigo_raw is None:
            return None

        codigo = str(codigo_raw).strip()
        if not codigo or not codigo[0].isdigit():
            return None

        nombre = str(row[self.COL_NOMBRE] or "").strip()

        return CuentaPUC(
            codigo=codigo,
            nombre=nombre,
            saldo_anterior=self._a_float(row[self.COL_SALDO_ANT]),
            debitos=self._a_float(row[self.COL_DEBITOS]),
            creditos=self._a_float(row[self.COL_CREDITOS]),
            saldo_nuevo=self._a_float(row[self.COL_SALDO_NUEVO]),
        )

    @staticmethod
    def _a_float(v) -> float:
        """Convierte cualquier valor numérico a float, tolerando None y strings."""
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            # Filtrar valores residuales tipo 1e-7 que Siigo deja en cuentas a cero
            return 0.0 if abs(float(v)) < 0.005 else float(v)
        try:
            s = str(v).replace(",", "").replace(" ", "")
            f = float(s)
            return 0.0 if abs(f) < 0.005 else f
        except (ValueError, TypeError):
            return 0.0


# ============================================================
# MAPEO BALANCE → CASILLAS FORM 110
# ============================================================
# Estos mapeos se usan en conjunto con LiquidadorRenta para auto-cargar
# patrimonio, ingresos y gastos desde un BalanceSiigo importado.

def construir_diccionario_patrimonio(balance: BalanceSiigo) -> dict[str, float]:
    """
    Construye un dict {prefijo_puc: saldo} para patrimonio (clase 1 Activos, 2 Pasivos).
    Compatible con MAPEO_PUC_PATRIMONIO de liquidador.py.
    """
    return {
        # Efectivo y equivalentes (cas. 36)
        "11": balance.saldo_clase("11"),
        # Inversiones (cas. 37)
        "12": balance.saldo_clase("12"),
        # Cuentas por cobrar (cas. 38)
        "13": balance.saldo_clase("13"),
        # Inventarios (cas. 39)
        "14": balance.saldo_clase("14"),
        "15": balance.saldo_clase("15"),  # Inventarios servicios
        # Activos intangibles (cas. 40)
        "16": balance.saldo_clase("16"),
        # PP&E (cas. 42)
        "15": balance.saldo_clase("15"),
        # Otros activos (cas. 43)
        "17": balance.saldo_clase("17"),
        "18": balance.saldo_clase("18"),
        "19": balance.saldo_clase("19"),
        # Pasivos (cas. 45)
        "2": balance.saldo_clase("2"),
    }


def construir_diccionario_ingresos(balance: BalanceSiigo) -> dict[str, float]:
    """
    Construye dict de ingresos por grupo PUC.
    Cas. 47 (operacionales 41), 48 (financieros 4210), 57 (otros 42).
    """
    return {
        "41": balance.saldo_clase("41"),
        "4210": balance.saldo_grupo("4210"),  # Financieros
        "42": balance.saldo_clase("42"),
        # Devoluciones (clase 4 negativa)
        "417": balance.saldo_grupo("4175") + balance.saldo_grupo("4180"),
    }


def construir_diccionario_costos_gastos(balance: BalanceSiigo) -> dict[str, float]:
    """
    Costos (clase 6) y gastos (clase 5).
    Cas. 62 costos, 63 admin, 64 ventas, 65 financieros, 66 otros.
    """
    return {
        # Costos (cas. 62)
        "6": balance.saldo_clase("6"),
        "61": balance.saldo_clase("61"),
        "62": balance.saldo_clase("62"),
        # Gastos administración (cas. 63)
        "51": balance.saldo_clase("51"),
        # Gastos ventas (cas. 64)
        "52": balance.saldo_clase("52"),
        # Gastos financieros (cas. 65)
        "5305": balance.saldo_grupo("5305"),
        # GMF (parte del cas. 65 o de los gastos)
        "530525": balance.saldo_cuenta("530525"),
        # Otros gastos (cas. 66)
        "53": balance.saldo_clase("53"),
        # Provisión impuesto de renta (NO se carga al Form 110)
        "5405": balance.saldo_grupo("5405"),
    }
