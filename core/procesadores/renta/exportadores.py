"""
Generadores de entregables para la declaración de Renta.

Funciones públicas:
  - generar_excel_comparativo(liquidador, ruta_salida, valores_referencia=None)
      Crea un Excel con 6 hojas: Portada, Form 110 lado a lado, Conciliación,
      Detalle Retenciones, Datos PILA, Resumen Ejecutivo.

  - generar_dictamen_word(liquidador, ruta_salida)
      Crea un Word ejecutivo de ~8 páginas con el resultado de la liquidación,
      conciliación con base legal, retenciones, hallazgos y recomendaciones.

Ambas requieren un LiquidadorRenta ya ejecutado (con liquidar() llamado).
"""

from __future__ import annotations
from pathlib import Path
from typing import Optional, Any
import logging
import os
import subprocess
import json
import tempfile

log = logging.getLogger(__name__)


def generar_excel_comparativo(
    liquidador: Any,  # LiquidadorRenta - tipado lazy para evitar circular import
    ruta_salida: str | Path,
    valores_referencia: Optional[dict] = None,
    retenciones_detalle: Optional[list[dict]] = None,
) -> Path:
    """
    Genera un Excel comparativo profesional con el resultado de la liquidación.

    Args:
        liquidador: LiquidadorRenta ya ejecutado (con liquidar() llamado).
        ruta_salida: ruta donde guardar el .xlsx.
        valores_referencia: dict opcional {casilla: valor} con valores del
                            Excel modelo del contador para mostrar lado a lado.
                            Si es None, solo se muestran los valores calculados.
        retenciones_detalle: lista opcional de dicts con detalle de cada
                             certificado de retención. Cada item debe tener:
                             {nombre, nit, concepto, base, retencion_106, autoret_105}

    Returns:
        Path al archivo generado.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError as e:
        raise ImportError(
            "openpyxl es requerido. Instale con: pip install openpyxl"
        ) from e

    if liquidador.formulario is None:
        raise ValueError("El liquidador debe haber ejecutado .liquidar() antes")

    f110 = liquidador.formulario
    d = f110.to_dict()
    ruta = Path(ruta_salida)
    ruta.parent.mkdir(parents=True, exist_ok=True)

    # Estilos
    FONT_TITULO = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    FONT_SUBTITULO = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    FONT_SECCION = Font(name='Calibri', size=11, bold=True, color='1F4E78')
    FONT_NORMAL = Font(name='Calibri', size=10)
    FONT_NORMAL_BOLD = Font(name='Calibri', size=10, bold=True)
    FONT_NUM = Font(name='Calibri', size=10)
    FONT_TOTAL = Font(name='Calibri', size=10, bold=True)
    FONT_HEADER = Font(name='Calibri', size=10, bold=True, color='FFFFFF')

    FILL_TITULO = PatternFill('solid', start_color='1F4E78')
    FILL_SUBTITULO = PatternFill('solid', start_color='2E75B6')
    FILL_SECCION = PatternFill('solid', start_color='D9E1F2')
    FILL_OK = PatternFill('solid', start_color='C6EFCE')
    FILL_ERROR = PatternFill('solid', start_color='FFC7CE')
    FILL_HEADER = PatternFill('solid', start_color='305496')

    CENTER = Alignment(horizontal='center', vertical='center')
    LEFT = Alignment(horizontal='left', vertical='center', indent=1)
    RIGHT = Alignment(horizontal='right', vertical='center')
    THIN = Side(border_style='thin', color='BFBFBF')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    FORMAT_PESOS = '"$"#,##0;[Red]("$"#,##0);"-"'

    wb = Workbook()
    wb.remove(wb.active)

    # ----- HOJA 1: PORTADA -----
    ws = wb.create_sheet('Portada')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 35

    ws['B3'] = 'DECLARACIÓN DE RENTA — PERSONAS JURÍDICAS'
    ws['B3'].font = Font(name='Calibri', size=20, bold=True, color='1F4E78')
    ws['B5'] = liquidador.razon_social
    ws['B5'].font = Font(name='Calibri', size=16, bold=True)
    ws['B6'] = f'NIT {liquidador.nit}'
    ws['B6'].font = Font(name='Calibri', size=12)
    ws['B7'] = f'Año Gravable {liquidador.ano_gravable} — Formulario 110'
    ws['B7'].font = Font(name='Calibri', size=12, italic=True)

    ws['B10'] = 'COMPARATIVO PLATAFORMA' + (' vs EXCEL MODELO' if valores_referencia else '')
    ws['B10'].font = FONT_SECCION

    from datetime import datetime
    ws['B12'] = 'Fecha de generación:'
    ws['B12'].font = FONT_NORMAL_BOLD
    ws['C12'] = datetime.now().strftime('%Y-%m-%d %H:%M')

    # ----- HOJA 2: FORM 110 COMPARATIVO -----
    ws = wb.create_sheet('Form 110 Comparativo')
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 8

    titulo = 'FORMULARIO 110 — COMPARATIVO' if valores_referencia else 'FORMULARIO 110 — RESULTADO'
    ws.merge_cells('A1:F1')
    ws['A1'] = titulo
    ws['A1'].font = FONT_TITULO
    ws['A1'].fill = FILL_TITULO
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:F2')
    ws['A2'] = f'{liquidador.razon_social} — NIT {liquidador.nit} — AG {liquidador.ano_gravable}'
    ws['A2'].font = FONT_SUBTITULO
    ws['A2'].fill = FILL_SUBTITULO
    ws['A2'].alignment = CENTER
    ws.row_dimensions[2].height = 20

    encabezados = ['Casilla', 'Concepto', 'Plataforma']
    if valores_referencia:
        encabezados = ['Casilla', 'Concepto', 'Excel Modelo', 'Plataforma', 'Diferencia', 'OK']
    for col, val in enumerate(encabezados, 1):
        c = ws.cell(row=4, column=col, value=val)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER

    # Filas del Form 110 (mapeo casilla → valor)
    filas = _build_filas_form110(f110, d)

    row = 5
    for casilla, concepto, val_plat, tipo in filas:
        if tipo == "header":
            cols = 6 if valores_referencia else 3
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
            c = ws.cell(row=row, column=1, value=concepto)
            c.font = FONT_SECCION
            c.fill = FILL_SECCION
            c.alignment = LEFT
            c.border = BORDER
            row += 1
            continue

        es_total = (tipo == "total")
        c = ws.cell(row=row, column=1, value=casilla)
        c.font = FONT_TOTAL if es_total else FONT_NORMAL
        c.alignment = CENTER
        c.border = BORDER

        c = ws.cell(row=row, column=2, value=concepto)
        c.font = FONT_TOTAL if es_total else FONT_NORMAL
        c.alignment = LEFT
        c.border = BORDER

        if valores_referencia:
            val_excel = valores_referencia.get(casilla, 0)
            c = ws.cell(row=row, column=3, value=val_excel)
            c.font = FONT_TOTAL if es_total else FONT_NUM
            c.alignment = RIGHT
            c.number_format = FORMAT_PESOS
            c.border = BORDER

            c = ws.cell(row=row, column=4, value=int(round(val_plat)))
            c.font = FONT_TOTAL if es_total else FONT_NUM
            c.alignment = RIGHT
            c.number_format = FORMAT_PESOS
            c.border = BORDER

            diff = int(round(val_plat)) - val_excel
            c = ws.cell(row=row, column=5, value=diff)
            c.font = FONT_TOTAL if es_total else FONT_NUM
            c.alignment = RIGHT
            c.number_format = FORMAT_PESOS
            c.border = BORDER

            ok = "✓" if abs(diff) <= 1000 else "✗"
            c = ws.cell(row=row, column=6, value=ok)
            c.font = Font(name='Calibri', size=11, bold=True,
                          color='006100' if ok == "✓" else 'C00000')
            c.alignment = CENTER
            c.fill = FILL_OK if ok == "✓" else FILL_ERROR
            c.border = BORDER
        else:
            c = ws.cell(row=row, column=3, value=int(round(val_plat)))
            c.font = FONT_TOTAL if es_total else FONT_NUM
            c.alignment = RIGHT
            c.number_format = FORMAT_PESOS
            c.border = BORDER
        row += 1

    ws.freeze_panes = 'A5'

    # ----- HOJA 3: CONCILIACIÓN FISCAL -----
    if liquidador.resultado_conciliacion is not None:
        _agregar_hoja_conciliacion(wb, liquidador, FONT_TITULO, FONT_SUBTITULO,
                                    FONT_SECCION, FONT_NORMAL, FONT_NORMAL_BOLD,
                                    FONT_TOTAL, FONT_HEADER,
                                    FILL_TITULO, FILL_SUBTITULO, FILL_SECCION,
                                    FILL_HEADER, BORDER, FORMAT_PESOS, RIGHT, LEFT, CENTER)

    # ----- HOJA 4: DETALLE RETENCIONES (opcional) -----
    if retenciones_detalle:
        _agregar_hoja_retenciones(wb, retenciones_detalle, FONT_TITULO, FONT_SUBTITULO,
                                   FONT_HEADER, FONT_NORMAL, FONT_TOTAL,
                                   FILL_TITULO, FILL_SUBTITULO, FILL_HEADER,
                                   BORDER, FORMAT_PESOS, RIGHT, LEFT, CENTER)

    # ----- HOJA 5: RESUMEN EJECUTIVO -----
    _agregar_hoja_resumen(wb, liquidador, d, FONT_TITULO, FONT_SECCION,
                          FILL_TITULO, FILL_SECCION, FILL_OK, BORDER,
                          FORMAT_PESOS, RIGHT, CENTER)

    wb.save(ruta)
    log.info("Excel comparativo generado: %s", ruta)
    return ruta


def generar_dictamen_word(
    liquidador: Any,
    ruta_salida: str | Path,
    retenciones_detalle: Optional[list[dict]] = None,
    plazo_fecha_limite: Optional[str] = None,
) -> Path:
    """
    Genera el dictamen ejecutivo en Word (8 páginas aprox).

    Requiere que `node` esté instalado y la librería `docx` disponible globalmente.
    Si no están disponibles, lanza RuntimeError con instrucciones.

    Args:
        liquidador: LiquidadorRenta ya ejecutado.
        ruta_salida: ruta donde guardar el .docx.
        retenciones_detalle: detalle de retenciones (opcional).
        plazo_fecha_limite: fecha límite formateada (ej. "14 de mayo de 2026").

    Returns:
        Path al archivo generado.
    """
    if liquidador.formulario is None:
        raise ValueError("El liquidador debe haber ejecutado .liquidar() antes")

    # Verificar que node y docx-js estén disponibles
    try:
        result = subprocess.run(
            ['node', '-e', "require('docx'); console.log('OK')"],
            capture_output=True, text=True, timeout=10,
        )
        if result.returncode != 0 or 'OK' not in result.stdout:
            raise RuntimeError(
                "La librería 'docx' de Node.js no está disponible. "
                "Instale con: npm install -g docx"
            )
    except FileNotFoundError:
        raise RuntimeError(
            "Node.js no está instalado. Para generar el dictamen Word, instale Node y "
            "luego: npm install -g docx"
        )

    # Serializar datos para pasarlos al script Node
    f110 = liquidador.formulario
    payload = {
        "razon_social": liquidador.razon_social,
        "nit": liquidador.nit,
        "ano_gravable": liquidador.ano_gravable,
        "form110": f110.to_dict(),
        "conciliacion": _serializar_conciliacion(liquidador),
        "retenciones": retenciones_detalle or [],
        "plazo": plazo_fecha_limite or "(según último dígito del NIT)",
        "fecha_generacion": __import__('datetime').datetime.now().strftime('%Y-%m-%d'),
        "ruta_salida": str(ruta_salida),
    }

    # Encontrar la ruta del template Node embebido en este paquete
    template_dir = Path(__file__).parent / "templates"
    script_path = template_dir / "generar_dictamen.js"

    if not script_path.exists():
        raise FileNotFoundError(
            f"Template de dictamen no encontrado en {script_path}. "
            "Verifique que core/procesadores/renta/templates/generar_dictamen.js exista."
        )

    # Pasar el payload por stdin al script Node
    try:
        result = subprocess.run(
            ['node', str(script_path)],
            input=json.dumps(payload),
            capture_output=True, text=True, timeout=60,
            cwd=str(template_dir),
        )
        if result.returncode != 0:
            raise RuntimeError(f"Error generando dictamen: {result.stderr}")

        ruta = Path(ruta_salida)
        if not ruta.exists():
            raise RuntimeError(
                f"Script ejecutó pero no generó el archivo: {ruta}\n{result.stdout}"
            )
        log.info("Dictamen Word generado: %s", ruta)
        return ruta
    except subprocess.TimeoutExpired:
        raise RuntimeError("Timeout generando dictamen (>60s)")


# ============================================================
# HELPERS PRIVADOS
# ============================================================

def _build_filas_form110(f110, d: dict) -> list[tuple]:
    """Construye la lista de filas del Form 110 para la hoja comparativa."""
    return [
        (None, "DATOS INFORMATIVOS", None, "header"),
        ("33", "Total costos y gastos de nómina", d["total_costos_gastos_nomina"], None),
        ("34", "Aportes seguridad social", d["aportes_seguridad_social"], None),
        ("35", "Aportes SENA, ICBF, Cajas", d["aportes_sena_icbf_caja"], None),

        (None, "PATRIMONIO", None, "header"),
        ("36", "Efectivo y equivalentes", f110.patrimonio.efectivo_y_equivalentes, None),
        ("37", "Inversiones e instrumentos financieros", f110.patrimonio.inversiones_e_instrumentos, None),
        ("38", "Cuentas por cobrar", f110.patrimonio.cuentas_documentos_arrend_por_cobrar, None),
        ("39", "Inventarios", f110.patrimonio.inventarios, None),
        ("40", "Activos intangibles", f110.patrimonio.activos_intangibles, None),
        ("41", "Activos biológicos", f110.patrimonio.activos_biologicos, None),
        ("42", "PP&E y propiedades de inversión", f110.patrimonio.propiedades_planta_equipo, None),
        ("43", "Otros activos", f110.patrimonio.otros_activos, None),
        ("44", "TOTAL PATRIMONIO BRUTO", d["casilla_44_total_patrimonio_bruto"], "total"),
        ("45", "Pasivos", f110.patrimonio.pasivos, None),
        ("46", "TOTAL PATRIMONIO LÍQUIDO", d["casilla_46_total_patrimonio_liquido"], "total"),

        (None, "INGRESOS", None, "header"),
        ("47", "Ingresos brutos actividades ordinarias", f110.ingresos.ingresos_brutos_actividades_ordinarias, None),
        ("48", "Ingresos financieros", f110.ingresos.ingresos_financieros, None),
        ("57", "Otros ingresos", f110.ingresos.otros_ingresos, None),
        ("58", "TOTAL INGRESOS BRUTOS", d["casilla_58_total_ingresos_brutos"], "total"),
        ("59", "Devoluciones, rebajas y descuentos", f110.ingresos.devoluciones_rebajas_descuentos, None),
        ("60", "Ingresos no constitutivos de renta", f110.ingresos.ingresos_no_constitutivos_renta, None),
        ("61", "TOTAL INGRESOS NETOS", d["casilla_61_total_ingresos_netos"], "total"),

        (None, "COSTOS Y DEDUCCIONES", None, "header"),
        ("62", "Costos", f110.costos_y_deducciones.costos, None),
        ("63", "Gastos de administración", f110.costos_y_deducciones.gastos_administracion, None),
        ("64", "Gastos de distribución y ventas", f110.costos_y_deducciones.gastos_distribucion_ventas, None),
        ("65", "Gastos financieros", f110.costos_y_deducciones.gastos_financieros, None),
        ("66", "Otros gastos y deducciones", f110.costos_y_deducciones.otros_gastos_deducciones, None),
        ("67", "TOTAL COSTOS Y GASTOS DEDUCIBLES", d["casilla_67_total_costos_gastos_deducibles"], "total"),

        (None, "RENTA LÍQUIDA", None, "header"),
        ("72", "Renta líquida ordinaria del ejercicio", d["casilla_72_renta_liquida_ordinaria"], "total"),
        ("75", "Renta líquida", d["casilla_75_renta_liquida"], "total"),
        ("76", "Renta presuntiva", f110.renta.renta_presuntiva, None),
        ("77", "Renta exenta", f110.renta.renta_exenta, None),
        ("78", "Rentas gravables", f110.renta.rentas_gravables, None),
        ("79", "RENTA LÍQUIDA GRAVABLE", d["casilla_79_renta_liquida_gravable"], "total"),

        (None, "LIQUIDACIÓN PRIVADA", None, "header"),
        ("84", "Impuesto sobre rentas líquidas", f110.liquidacion.impuesto_rentas_liquidas, None),
        ("91", "Total impuesto rentas líquidas", d["casilla_91_total_impuesto_rentas_liquidas"], "total"),
        ("92", "Valor a adicionar (TMT)", f110.liquidacion.valor_a_adicionar_vaa, None),
        ("93", "Descuentos tributarios", f110.liquidacion.descuentos_tributarios, None),
        ("94", "Impuesto neto sin adicional", d["casilla_94_impuesto_neto_sin_adicional"], "total"),
        ("96", "Impuesto neto con adicional", d["casilla_96_impuesto_neto_con_adicional"], "total"),
        ("99", "TOTAL IMPUESTO A CARGO", d["casilla_99_total_impuesto_a_cargo"], "total"),
        ("104", "Saldo favor año anterior", f110.liquidacion.saldo_favor_ano_anterior, None),
        ("105", "Autorretenciones", f110.liquidacion.autorretenciones, None),
        ("106", "Otras retenciones", f110.liquidacion.otras_retenciones, None),
        ("107", "TOTAL RETENCIONES", d["casilla_107_total_retenciones"], "total"),
        ("108", "Anticipo año siguiente", f110.liquidacion.anticipo_renta_ano_siguiente, None),
        ("111", "Saldo a pagar por impuesto", d["casilla_111_saldo_pagar_impuesto"], "total"),
        ("112", "Sanciones", f110.liquidacion.sanciones, None),
        ("113", "TOTAL SALDO A PAGAR", d["casilla_113_total_saldo_a_pagar"], "total"),
        ("114", "TOTAL SALDO A FAVOR", d["casilla_114_total_saldo_a_favor"], "total"),
    ]


def _agregar_hoja_conciliacion(wb, liq, FONT_TITULO, FONT_SUBTITULO, FONT_SECCION,
                                 FONT_NORMAL, FONT_NORMAL_BOLD, FONT_TOTAL, FONT_HEADER,
                                 FILL_TITULO, FILL_SUBTITULO, FILL_SECCION, FILL_HEADER,
                                 BORDER, FORMAT_PESOS, RIGHT, LEFT, CENTER):
    """Agrega la hoja de Conciliación Fiscal al workbook."""
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.create_sheet('Conciliación Fiscal')
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 60

    ws.merge_cells('A1:D1')
    ws['A1'] = 'CONCILIACIÓN FISCAL'
    ws['A1'].font = FONT_TITULO
    ws['A1'].fill = FILL_TITULO
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 25

    row = 4
    if liq.utilidad_contable_antes_impuestos is not None:
        ws.cell(row=row, column=2, value='Utilidad contable antes de impuestos').font = FONT_NORMAL_BOLD
        c = ws.cell(row=row, column=3, value=liq.utilidad_contable_antes_impuestos)
        c.number_format = FORMAT_PESOS
        c.alignment = RIGHT
        c.font = FONT_NORMAL_BOLD
        row += 2

    res = liq.resultado_conciliacion
    if res:
        ws.cell(row=row, column=2, value='(+) AUMENTAN LA RENTA').font = FONT_SECCION
        ws.cell(row=row, column=2).fill = FILL_SECCION
        row += 1
        for p in res.partidas_aumentan:
            ws.cell(row=row, column=2, value=f"   {p.nombre}").font = FONT_NORMAL
            c = ws.cell(row=row, column=3, value=p.valor)
            c.number_format = FORMAT_PESOS
            c.alignment = RIGHT
            ws.cell(row=row, column=4, value=p.base_legal or '').font = FONT_NORMAL
            row += 1
        ws.cell(row=row, column=2, value='   TOTAL AUMENTOS').font = FONT_TOTAL
        c = ws.cell(row=row, column=3, value=res.total_aumentos)
        c.number_format = FORMAT_PESOS
        c.alignment = RIGHT
        c.font = FONT_TOTAL
        row += 2

        if res.partidas_disminuyen:
            ws.cell(row=row, column=2, value='(-) DISMINUYEN LA RENTA').font = FONT_SECCION
            ws.cell(row=row, column=2).fill = FILL_SECCION
            row += 1
            for p in res.partidas_disminuyen:
                ws.cell(row=row, column=2, value=f"   {p.nombre}").font = FONT_NORMAL
                c = ws.cell(row=row, column=3, value=p.valor)
                c.number_format = FORMAT_PESOS
                c.alignment = RIGHT
                ws.cell(row=row, column=4, value=p.base_legal or '').font = FONT_NORMAL
                row += 1
            row += 1

        c = ws.cell(row=row, column=2, value='= RENTA LÍQUIDA FISCAL DEL EJERCICIO')
        c.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', start_color='1F4E78')
        c = ws.cell(row=row, column=3, value=res.renta_liquida_fiscal)
        c.number_format = FORMAT_PESOS
        c.alignment = RIGHT
        c.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', start_color='1F4E78')


def _agregar_hoja_retenciones(wb, retenciones, FONT_TITULO, FONT_SUBTITULO, FONT_HEADER,
                                FONT_NORMAL, FONT_TOTAL, FILL_TITULO, FILL_SUBTITULO,
                                FILL_HEADER, BORDER, FORMAT_PESOS, RIGHT, LEFT, CENTER):
    """Agrega hoja con detalle de retenciones."""
    from openpyxl.styles import PatternFill, Font, Alignment

    ws = wb.create_sheet('Detalle Retenciones')
    for col, w in [('B', 5), ('C', 42), ('D', 14), ('E', 14), ('F', 16), ('G', 16), ('H', 16)]:
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:H1')
    ws['A1'] = 'DETALLE DE RETENCIONES'
    ws['A1'].font = FONT_TITULO
    ws['A1'].fill = FILL_TITULO
    ws['A1'].alignment = CENTER

    row = 3
    headers = ['', '#', 'Agente retenedor', 'NIT', 'Concepto', 'Base', 'Cas. 106', 'Cas. 105']
    for col, val in enumerate(headers, 1):
        if val:
            c = ws.cell(row=row, column=col, value=val)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.alignment = CENTER

    row = 4
    total_106 = 0
    total_105 = 0
    for i, ret in enumerate(retenciones, 1):
        ws.cell(row=row, column=2, value=i).alignment = CENTER
        ws.cell(row=row, column=3, value=ret.get('nombre', ''))
        ws.cell(row=row, column=4, value=ret.get('nit', ''))
        ws.cell(row=row, column=5, value=ret.get('concepto', ''))
        c = ws.cell(row=row, column=6, value=ret.get('base', 0))
        c.number_format = FORMAT_PESOS
        c.alignment = RIGHT
        c = ws.cell(row=row, column=7, value=ret.get('retencion_106', 0))
        c.number_format = FORMAT_PESOS
        c.alignment = RIGHT
        c = ws.cell(row=row, column=8, value=ret.get('autoret_105', 0))
        c.number_format = FORMAT_PESOS
        c.alignment = RIGHT
        total_106 += ret.get('retencion_106', 0)
        total_105 += ret.get('autoret_105', 0)
        row += 1


def _agregar_hoja_resumen(wb, liq, d, FONT_TITULO, FONT_SECCION, FILL_TITULO,
                           FILL_SECCION, FILL_OK, BORDER, FORMAT_PESOS, RIGHT, CENTER):
    """Hoja de resumen ejecutivo con KPIs."""
    from openpyxl.styles import Font

    ws = wb.create_sheet('Resumen Ejecutivo')
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25

    ws.merge_cells('A1:C1')
    ws['A1'] = 'RESUMEN EJECUTIVO'
    ws['A1'].font = FONT_TITULO
    ws['A1'].fill = FILL_TITULO
    ws['A1'].alignment = CENTER

    kpis = [
        ('Patrimonio bruto', d["casilla_44_total_patrimonio_bruto"]),
        ('Patrimonio líquido', d["casilla_46_total_patrimonio_liquido"]),
        ('Ingresos brutos', d["casilla_58_total_ingresos_brutos"]),
        ('Costos y gastos', d["casilla_67_total_costos_gastos_deducibles"]),
        ('Renta líquida gravable', d["casilla_79_renta_liquida_gravable"]),
        ('Total impuesto a cargo', d["casilla_99_total_impuesto_a_cargo"]),
        ('Total retenciones', d["casilla_107_total_retenciones"]),
        ('SALDO A PAGAR', d["casilla_113_total_saldo_a_pagar"]),
        ('SALDO A FAVOR', d["casilla_114_total_saldo_a_favor"]),
    ]
    row = 4
    for nombre, valor in kpis:
        ws.cell(row=row, column=2, value=nombre).font = Font(
            name='Calibri', size=11,
            bold='SALDO' in nombre,
        )
        c = ws.cell(row=row, column=3, value=valor)
        c.number_format = FORMAT_PESOS
        c.alignment = RIGHT
        if 'SALDO A FAVOR' in nombre and valor and valor > 0:
            c.fill = FILL_OK
            ws.cell(row=row, column=2).fill = FILL_OK
        row += 1


def _serializar_conciliacion(liq) -> dict:
    """Serializa la conciliación a un dict JSON-safe."""
    if liq.resultado_conciliacion is None:
        return {}
    res = liq.resultado_conciliacion
    return {
        "utilidad_contable_antes_impuestos": liq.utilidad_contable_antes_impuestos or 0,
        "partidas_aumentan": [
            {"codigo": p.codigo, "nombre": p.nombre, "valor": p.valor, "base_legal": p.base_legal}
            for p in res.partidas_aumentan
        ],
        "partidas_disminuyen": [
            {"codigo": p.codigo, "nombre": p.nombre, "valor": p.valor, "base_legal": p.base_legal}
            for p in res.partidas_disminuyen
        ],
        "total_aumentos": res.total_aumentos,
        "total_disminuciones": res.total_disminuciones,
        "renta_liquida_fiscal": res.renta_liquida_fiscal,
    }
