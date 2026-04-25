"""
Procesador de facturas electrónicas DIAN recibidas - versión web.

Adaptado del procesador_compras_dian.py del .exe v2.6.

Cambios vs versión escritorio:
    - Recibe file-like objects (BytesIO) en vez de rutas de archivo
    - Recibe un objeto Configuracion explícito en vez de usar módulo global
    - Retorna DataFrame + log de texto, igual que procesador_caja_menor

Lógica contable v2.6 (preservada 100%):
    1. IVA descontable automático: 24080308 (servicios) vs 24080201 (mercancía)
    2. IVA de NC compras DIAN: siempre 24080204
    3. Base gravable = Total − 16 columnas de impuestos/retenciones
    4. Retenciones del DIAN se ignoran (se calculan desde cero)
    5. Impuesto al consumo unificado → 511575
    6. ICUI → 511586 (línea aparte)
    7. INPP + IBUA → 519525 (línea aparte)
    8. NITs normalizados en todo el flujo
    9. Base mínima evaluada por factura (no acumulada)
   10. Pagos DIAN NO se generan aquí (módulo de Egresos DIAN aparte)

Comprobantes:
  - Facturas electrónicas:      Comprobante 3 (prefijo FDI)
  - Notas crédito electrónicas: Comprobante 7 (prefijo NDI)

Documento: AAAAMMNNN (año, mes, consecutivo 3 dígitos).

Fase 2: regla histórica del balance ACTIVA.
    Si el parámetro RESPETAR_HISTORICO_BALANCE = SI (default) y se pasa el
    diccionario `conocimiento` (cargado desde Supabase Storage), entonces:
      - Un NIT que está en el balance pero sin movimiento histórico en
        cuentas 2365 NO recibe retefuente, aunque supere los 10 UVT.
      - La cuenta de IVA del proveedor se toma del balance (24080201 vs
        24080308) cuando no hay regla explícita en DIAN_Proveedores.
      - La cuenta de gasto puede sugerirse del balance también.
"""
from __future__ import annotations
import io
import re
from datetime import datetime
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

from core.utils.configuracion_web import Configuracion
from core.utils.nits import normalizar_nit
from core.utils import conocimiento_balance as cb


# ============================================================
# Utilidades
# ============================================================

def _normalizar_texto(texto):
    if texto is None:
        return ""
    t = str(texto).strip().upper()
    for a, b in [("Á", "A"), ("É", "E"), ("Í", "I"), ("Ó", "O"),
                 ("Ú", "U"), ("Ñ", "N")]:
        t = t.replace(a, b)
    return " ".join(t.split())


def _formatear_fecha_para_plano(fecha_str):
    """Convierte fecha DIAN ('31-03-2026' o datetime) a 'MM/DD/YYYY'."""
    if fecha_str is None:
        return ""
    if isinstance(fecha_str, datetime):
        return fecha_str.strftime("%m/%d/%Y")
    s = str(fecha_str).strip().split()[0]
    m = re.match(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", s)
    if m:
        dia, mes, anio = m.groups()
        return f"{int(mes):02d}/{int(dia):02d}/{anio}"
    return s


def _parse_fecha(fecha_str):
    """Convierte la fecha DIAN a datetime para poder ordenar."""
    if fecha_str is None:
        return None
    if isinstance(fecha_str, datetime):
        return fecha_str
    s = str(fecha_str).strip().split()[0]
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _int_valor(v):
    if v is None:
        return 0
    try:
        return int(round(float(v)))
    except (ValueError, TypeError):
        return 0


# ============================================================
# Retenciones (lógica del módulo retenciones_compras del .exe)
# ============================================================

# Bases mínimas de UVT para ReteIVA 15% por concepto
UVT_MIN_RETEIVA = {
    "COMPRA_NATURAL":  10,
    "COMPRA_JURIDICA": 10,
    "HONORARIOS":      10,
    "COMISIONES":      10,
    "ARRENDAMIENTO":   10,
    "SERVICIO":         3,
}


def _retenciones_activas(cfg: Configuracion) -> bool:
    valor = str(cfg.parametro("APLICAR_RETENCIONES_COMPRAS", "SI")).strip().upper()
    return valor in ("SI", "SÍ", "YES", "TRUE", "1", "Y", "S")


def _uvt_vigente(cfg: Configuracion) -> int:
    try:
        return int(cfg.parametro("UVT_VIGENTE", "52374"))
    except (ValueError, TypeError):
        return 52374


def _detectar_tipo_especial(cfg: Configuracion, nombre_proveedor: str) -> str:
    """Detecta si el proveedor es AUTORRET, RST o vacío."""
    nombre_norm = _normalizar_texto(nombre_proveedor)
    for regla in cfg.dian_proveedores():
        palabra = _normalizar_texto(regla.get("palabra_clave", ""))
        if palabra and palabra in nombre_norm:
            return str(regla.get("tipo_especial", "")).strip().upper()
    return ""


def _es_rst(cfg: Configuracion, nombre_proveedor: str) -> bool:
    return _detectar_tipo_especial(cfg, nombre_proveedor) == "RST"


def _detectar_tipo_retencion(cfg: Configuracion, nombre_proveedor: str,
                             nit_proveedor: str) -> str:
    """Determina el tipo de retefuente que aplica al proveedor.

    Returns: tipo (COMPRA_NATURAL, COMPRA_JURIDICA, HONORARIOS, etc.) o "".
    """
    # 1. Buscar por palabra clave en DIAN_Proveedores
    nombre_norm = _normalizar_texto(nombre_proveedor)
    for regla in cfg.dian_proveedores():
        palabra = _normalizar_texto(regla.get("palabra_clave", ""))
        if palabra and palabra in nombre_norm:
            tipo = str(regla.get("tipo_retencion", "")).strip().upper()
            if tipo:
                return tipo
            break

    # 2. Regla automática por NIT (normalizado)
    nit = normalizar_nit(nit_proveedor)
    if not nit:
        return "COMPRA_JURIDICA"
    if len(nit) >= 9 and nit[0] in "89":
        return "COMPRA_JURIDICA"
    return "COMPRA_NATURAL"


def _respetar_historico(cfg: Configuracion) -> bool:
    valor = str(cfg.parametro("RESPETAR_HISTORICO_BALANCE", "SI")).strip().upper()
    return valor in ("SI", "SÍ", "YES", "TRUE", "1", "Y", "S")


def _calcular_retencion(
    cfg: Configuracion,
    base: int,
    tipo_retencion: str,
    nombre_proveedor: str,
    nit_proveedor: str = "",
    conocimiento: Optional[dict] = None,
) -> Tuple[int, str, str]:
    """Calcula retefuente. NO se practica si:
        - Proveedor es AUTORRET o RST
        - Base bajo el UVT mínimo
        - (Fase 2) NIT está en el balance pero sin retefuente histórica
          y RESPETAR_HISTORICO_BALANCE = SI

    Returns: (valor_retenido, cuenta_credito, tipo_aplicado).
    """
    if not _retenciones_activas(cfg):
        return 0, "", ""

    # AUTORRET o RST → no retefuente
    if nombre_proveedor:
        tipo_esp = _detectar_tipo_especial(cfg, nombre_proveedor)
        if tipo_esp in ("AUTORRET", "RST"):
            return 0, "", ""

    if not tipo_retencion:
        return 0, "", ""

    reglas = cfg.retenciones_compras()
    regla = reglas.get(tipo_retencion.upper())
    if not regla:
        return 0, "", ""

    base_int = int(base) if base else 0
    if base_int <= 0:
        return 0, "", ""

    uvt = _uvt_vigente(cfg)
    base_minima = int(regla["base_uvt"] * uvt)
    if base_int < base_minima:
        return 0, "", ""

    # Regla histórica del balance (v2.6): si el NIT existe en el balance
    # pero NUNCA se le ha retenido, no le practicamos retefuente esta vez.
    if conocimiento and nit_proveedor and _respetar_historico(cfg):
        info = cb.info_nit(conocimiento, nit_proveedor)
        if info is not None and not info.get("tiene_retefuente_historica"):
            return 0, "", ""

    tarifa = regla["tarifa"] / 100.0
    valor = int(round(base_int * tarifa))
    return valor, regla["cuenta"], tipo_retencion.upper()


def _calcular_reteiva(
    cfg: Configuracion,
    iva: int,
    tipo_retencion: str,
    nombre_proveedor: str,
) -> Tuple[int, str]:
    """Calcula ReteIVA 15% (cuenta 23670109) solo a proveedores RST."""
    if not _retenciones_activas(cfg):
        return 0, ""
    if not nombre_proveedor or not _es_rst(cfg, nombre_proveedor):
        return 0, ""

    iva_int = int(iva) if iva else 0
    if iva_int <= 0:
        return 0, ""

    tipo_upper = (tipo_retencion or "").upper()
    uvt_min = UVT_MIN_RETEIVA.get(tipo_upper, 10)
    uvt = _uvt_vigente(cfg)
    base_minima = uvt_min * uvt

    if iva_int < base_minima:
        return 0, ""

    valor = int(round(iva_int * 0.15))
    cuenta = cfg.parametro("CUENTA_RETEIVA_15", "23670109")
    return valor, cuenta


# ============================================================
# Clasificación por proveedor
# ============================================================

def _buscar_regla_proveedor(cfg: Configuracion, nombre_emisor: str):
    """Busca la regla de DIAN_Proveedores que coincide con el emisor."""
    if not nombre_emisor:
        return None
    nombre_norm = _normalizar_texto(nombre_emisor)
    for regla in cfg.dian_proveedores():
        palabra = _normalizar_texto(regla.get("palabra_clave", ""))
        if palabra and palabra in nombre_norm:
            return regla
    return None


def _clasificar_proveedor(
    cfg: Configuracion,
    nombre_emisor: str,
    nit_emisor: str = "",
    conocimiento: Optional[dict] = None,
) -> dict:
    """Determina cuenta de gasto y cuenta de pasivo para el proveedor.

    Orden de prioridad:
      1. Regla en DIAN_Proveedores (palabra clave) con cuentas definidas
      2. Cuenta histórica del balance (si conocimiento != None)
      3. Defaults (CUENTA_DIAN_COMPRA / CUENTA_DIAN_PROVEEDOR)
    """
    resultado = {
        "tipo": "mercancia",
        "cuenta_gasto": cfg.parametro("CUENTA_DIAN_COMPRA", "620505"),
        "cuenta_pasivo": cfg.parametro("CUENTA_DIAN_PROVEEDOR", "22050501"),
        "tipo_retencion": "",
        "tipo_especial": "",
        "origen": "default",
    }

    regla = _buscar_regla_proveedor(cfg, nombre_emisor)
    if regla:
        cg = (regla.get("cuenta_gasto") or "").strip()
        cp = (regla.get("cuenta_pasivo") or "").strip()
        resultado["tipo_retencion"] = str(regla.get("tipo_retencion", "")).strip().upper()
        resultado["tipo_especial"] = str(regla.get("tipo_especial", "")).strip().upper()
        if cg and cp:
            resultado["cuenta_gasto"] = cg
            resultado["cuenta_pasivo"] = cp
            resultado["tipo"] = "servicio"
            resultado["origen"] = "dian_proveedores"
            return resultado
        # Sin cuentas: solo marca tipo especial (caso LÁTTIMO RST)
        resultado["origen"] = "dian_proveedores_sin_cuentas"

    # Fase 2: si hay conocimiento del balance, usar la cuenta histórica
    if conocimiento and nit_emisor:
        cuenta_hist = cb.cuenta_gasto_sugerida(conocimiento, nit_emisor)
        if cuenta_hist:
            resultado["cuenta_gasto"] = cuenta_hist
            resultado["origen"] = "balance_historico"

    return resultado


def _cuenta_iva_factura(
    cfg: Configuracion,
    clasif: dict,
    nit_emisor: str = "",
    conocimiento: Optional[dict] = None,
) -> str:
    """Cuenta de IVA descontable según clasificación del proveedor (v2.6).

    Orden de prioridad:
      1. Regla en DIAN_Proveedores → 24080308 si HON/COM/ARR/AUTORRET, sino 24080201
      2. Si la regla NO existe pero hay balance: usar la cuenta IVA recurrente
         del NIT en el balance (Fase 2)
      3. Default: 24080201 (mercancía)
    """
    tipo_ret = clasif.get("tipo_retencion", "").upper()
    tipo_esp = clasif.get("tipo_especial", "").upper()
    origen = clasif.get("origen", "")

    # Si vino de DIAN_Proveedores, esa regla manda
    if origen == "dian_proveedores":
        if tipo_esp == "AUTORRET":
            return cfg.parametro("CUENTA_IVA_SERVICIOS", "24080308")
        if tipo_ret in ("HONORARIOS", "COMISIONES", "ARRENDAMIENTO"):
            return cfg.parametro("CUENTA_IVA_SERVICIOS", "24080308")
        return cfg.parametro("CUENTA_IVA_MERCANCIA", "24080201")

    # Aún si no hay cuentas definidas, AUTORRET / HON / COM / ARR mandan
    if tipo_esp == "AUTORRET":
        return cfg.parametro("CUENTA_IVA_SERVICIOS", "24080308")
    if tipo_ret in ("HONORARIOS", "COMISIONES", "ARRENDAMIENTO"):
        return cfg.parametro("CUENTA_IVA_SERVICIOS", "24080308")

    # Fase 2: si hay balance histórico para este NIT, usar su IVA recurrente
    if conocimiento and nit_emisor:
        iva_hist = cb.cuenta_iva_sugerida(conocimiento, nit_emisor)
        if iva_hist:
            return iva_hist

    return cfg.parametro("CUENTA_IVA_MERCANCIA", "24080201")


# ============================================================
# Lectura del archivo DIAN
# ============================================================

# Columnas que se restan del Total para obtener la base gravable
COLUMNAS_IMPUESTOS_BASE = [
    "IVA", "ICA", "IC", "INC", "Timbre", "INC Bolsas", "IN Carbono",
    "IN Combustibles", "IC Datos", "ICL", "INPP", "IBUA", "ICUI",
    "Rete IVA", "Rete Renta", "Rete ICA",
]

# Columnas que van a 511575 (Impuesto al consumo, SUMADAS)
COLUMNAS_IMP_CONSUMO_511575 = [
    "IC", "INC", "Timbre", "INC Bolsas", "IN Carbono",
    "IN Combustibles", "IC Datos", "ICL",
]

COLUMNA_ICUI = "ICUI"
COLUMNAS_INPP_IBUA = ["INPP", "IBUA"]


def _indice_columna(headers, nombre):
    """Encuentra índice 0-based de columna por nombre, tolerante a
    espacios/acentos/mayúsculas."""
    objetivo = _normalizar_texto(nombre)
    for i, h in enumerate(headers):
        if _normalizar_texto(h) == objetivo:
            return i
    objetivo_sin = objetivo.replace(" ", "")
    for i, h in enumerate(headers):
        if _normalizar_texto(h).replace(" ", "") == objetivo_sin:
            return i
    return -1


def _leer_documentos(archivo) -> list:
    """Lee el Excel DIAN desde file-like / bytes y retorna lista de dicts
    ordenados por fecha."""
    # Normalizar a BytesIO
    if hasattr(archivo, "read"):
        contenido = archivo.read()
        if hasattr(archivo, "seek"):
            archivo.seek(0)
        bio = io.BytesIO(contenido)
    elif isinstance(archivo, (bytes, bytearray)):
        bio = io.BytesIO(bytes(archivo))
    else:
        raise TypeError(f"Tipo de archivo no soportado: {type(archivo)}")

    wb = load_workbook(bio, data_only=True)
    ws = wb.active

    # Cabecera en fila 1
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        headers.append(str(v).strip() if v is not None else "")

    idx_fijos = {
        "tipo":           _indice_columna(headers, "Tipo de documento"),
        "folio":          _indice_columna(headers, "Folio"),
        "prefijo":        _indice_columna(headers, "Prefijo"),
        "fecha":          _indice_columna(headers, "Fecha Emisión"),
        "nit_emisor":     _indice_columna(headers, "NIT Emisor"),
        "nombre_emisor":  _indice_columna(headers, "Nombre Emisor"),
        "total":          _indice_columna(headers, "Total"),
        "estado":         _indice_columna(headers, "Estado"),
    }

    # Validación: si faltan columnas críticas, mensaje claro
    columnas_criticas = ["tipo", "fecha", "nit_emisor", "nombre_emisor", "total"]
    faltantes = [k for k in columnas_criticas if idx_fijos[k] < 0]
    if faltantes:
        raise ValueError(
            "El archivo DIAN no tiene las columnas esperadas. "
            f"Faltan: {', '.join(faltantes)}. "
            "Verifica que sea el reporte de Documentos Electrónicos del "
            "portal DIAN."
        )

    idx_impuestos = {col: _indice_columna(headers, col)
                     for col in COLUMNAS_IMPUESTOS_BASE}

    def _get(fila, i):
        return fila[i] if 0 <= i < len(fila) else None

    documentos = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        tipo = _get(fila, idx_fijos["tipo"])
        if not tipo:
            continue
        tipo_norm = str(tipo).strip().lower()

        # Nota débito: omitida (igual que en v2.6 .exe)
        if "debito" in tipo_norm or "débito" in tipo_norm:
            continue

        es_nc = "credito" in tipo_norm or "crédito" in tipo_norm

        doc = {
            "es_nc": es_nc,
            "tipo": str(tipo),
            "folio": str(_get(fila, idx_fijos["folio"]) or "").strip(),
            "prefijo": str(_get(fila, idx_fijos["prefijo"]) or "").strip(),
            "fecha_raw": _get(fila, idx_fijos["fecha"]),
            "fecha": _parse_fecha(_get(fila, idx_fijos["fecha"])),
            "nit_emisor": normalizar_nit(_get(fila, idx_fijos["nit_emisor"])),
            "nombre_emisor": str(_get(fila, idx_fijos["nombre_emisor"]) or "").strip(),
            "total": _int_valor(_get(fila, idx_fijos["total"])),
            "estado": str(_get(fila, idx_fijos["estado"]) or "").strip(),
            "impuestos": {},
        }

        for nombre_col, i in idx_impuestos.items():
            doc["impuestos"][nombre_col] = _int_valor(_get(fila, i))

        if doc["total"] == 0:
            continue
        documentos.append(doc)

    documentos.sort(key=lambda d: (d["fecha"] or datetime.min, d["folio"]))
    return documentos


# ============================================================
# Cálculo de impuestos por factura
# ============================================================

def _calcular_impuestos_factura(doc: dict) -> dict:
    """Calcula los componentes contables de una factura según v2.6."""
    imp = doc["impuestos"]
    iva = imp.get("IVA", 0)
    imp_consumo = sum(imp.get(col, 0) for col in COLUMNAS_IMP_CONSUMO_511575)
    icui = imp.get(COLUMNA_ICUI, 0)
    inpp_ibua = sum(imp.get(col, 0) for col in COLUMNAS_INPP_IBUA)

    # Base = Total - (suma de las 16 columnas)
    suma_descontable = sum(imp.get(col, 0) for col in COLUMNAS_IMPUESTOS_BASE)
    base = doc["total"] - suma_descontable

    return {
        "iva": iva,
        "imp_consumo_511575": imp_consumo,
        "icui_511586": icui,
        "inpp_ibua_519525": inpp_ibua,
        "base_gravable": base,
    }


def _generar_consecutivo(fecha, contadores) -> str:
    """Consecutivo AAAAMMNNN contando por año-mes."""
    if fecha is None:
        return ""
    clave = (fecha.year, fecha.month)
    contadores[clave] = contadores.get(clave, 0) + 1
    return f"{fecha.year:04d}{fecha.month:02d}{contadores[clave]:03d}"


# ============================================================
# Construcción de asientos
# ============================================================

def _construir_asiento_factura(
    cfg: Configuracion,
    doc: dict,
    consecutivo: str,
    conocimiento: Optional[dict] = None,
):
    """Construye las filas del asiento contable de una factura DIAN v2.6."""
    clasif = _clasificar_proveedor(
        cfg, doc["nombre_emisor"], doc["nit_emisor"], conocimiento,
    )
    impuestos = _calcular_impuestos_factura(doc)

    cuenta_compra = clasif["cuenta_gasto"]
    cuenta_proveedor = clasif["cuenta_pasivo"]
    cuenta_iva = _cuenta_iva_factura(
        cfg, clasif, doc["nit_emisor"], conocimiento,
    )
    cuenta_imp_consumo = cfg.parametro("CUENTA_IMP_CONSUMO", "511575")
    cuenta_icui = cfg.parametro("CUENTA_ICUI", "511586")
    cuenta_inpp_ibua = cfg.parametro("CUENTA_INPP_IBUA", "519525")

    comprobante = cfg.codigo_comprobante("FDI", "3")
    fecha_str = _formatear_fecha_para_plano(doc["fecha_raw"] or doc["fecha"])
    nombre_emisor = doc["nombre_emisor"][:80] if doc["nombre_emisor"] else "PROVEEDOR"

    base_gravable = impuestos["base_gravable"]
    iva = impuestos["iva"]
    imp_consumo = impuestos["imp_consumo_511575"]
    icui = impuestos["icui_511586"]
    inpp_ibua = impuestos["inpp_ibua_519525"]

    doc_referencia = (f"{doc['prefijo']}{doc['folio']}".strip()
                      if doc['prefijo'] else doc['folio'])

    base_fila = {
        "COMPROBANTE": comprobante,
        "FECHA": fecha_str,
        "DOCUMENTO": consecutivo,
        "DOCUMENTO REFERENCIA": doc_referencia,
        "NIT": doc["nit_emisor"],
        "CENTRO DE COSTOS": cfg.parametro("CENTRO_COSTOS_DEFAULT", "PRINCIPAL"),
    }

    filas = []

    # Db: compra/gasto (base gravable)
    if base_gravable != 0:
        filas.append({**base_fila,
            "CUENTA": cuenta_compra,
            "DETALLE": f"COMPRA FE {doc_referencia} - {nombre_emisor}",
            "TIPO DE TRANSACCION": "1",
            "VALOR": str(-base_gravable),
            "BASE DE IMPUESTOS": "0",
        })

    # Db: IVA descontable (cuenta según tipo)
    if iva != 0:
        filas.append({**base_fila,
            "CUENTA": cuenta_iva,
            "DETALLE": f"IVA FE {doc_referencia} - {nombre_emisor}",
            "TIPO DE TRANSACCION": "1",
            "VALOR": str(-iva),
            "BASE DE IMPUESTOS": str(base_gravable),
        })

    # Db: Impuesto al consumo unificado (511575)
    if imp_consumo != 0:
        filas.append({**base_fila,
            "CUENTA": cuenta_imp_consumo,
            "DETALLE": f"IMP CONSUMO FE {doc_referencia} - {nombre_emisor}",
            "TIPO DE TRANSACCION": "1",
            "VALOR": str(-imp_consumo),
            "BASE DE IMPUESTOS": "0",
        })

    # Db: ICUI (511586)
    if icui != 0:
        filas.append({**base_fila,
            "CUENTA": cuenta_icui,
            "DETALLE": f"ICUI FE {doc_referencia} - {nombre_emisor}",
            "TIPO DE TRANSACCION": "1",
            "VALOR": str(-icui),
            "BASE DE IMPUESTOS": "0",
        })

    # Db: INPP + IBUA (519525)
    if inpp_ibua != 0:
        filas.append({**base_fila,
            "CUENTA": cuenta_inpp_ibua,
            "DETALLE": f"INPP/IBUA FE {doc_referencia} - {nombre_emisor}",
            "TIPO DE TRANSACCION": "1",
            "VALOR": str(-inpp_ibua),
            "BASE DE IMPUESTOS": "0",
        })

    # Retefuente
    tipo_ret = _detectar_tipo_retencion(cfg, doc["nombre_emisor"], doc["nit_emisor"])
    valor_retenido, cuenta_retencion, tipo_aplicado = _calcular_retencion(
        cfg, base_gravable, tipo_ret, doc["nombre_emisor"],
        nit_proveedor=doc["nit_emisor"],
        conocimiento=conocimiento,
    )
    if valor_retenido > 0:
        filas.append({**base_fila,
            "CUENTA": cuenta_retencion,
            "DETALLE": f"RETEFUENTE {tipo_aplicado} FE {doc_referencia} - {nombre_emisor}",
            "TIPO DE TRANSACCION": "2",
            "VALOR": str(valor_retenido),
            "BASE DE IMPUESTOS": str(base_gravable),
        })

    # ReteIVA 15% si es RST
    valor_reteiva, cuenta_reteiva = _calcular_reteiva(
        cfg, iva, tipo_ret or tipo_aplicado, doc["nombre_emisor"],
    )
    if valor_reteiva > 0:
        filas.append({**base_fila,
            "CUENTA": cuenta_reteiva,
            "DETALLE": f"RETEIVA 15% RST FE {doc_referencia} - {nombre_emisor}",
            "TIPO DE TRANSACCION": "2",
            "VALOR": str(valor_reteiva),
            "BASE DE IMPUESTOS": str(iva),
        })

    # Cr: proveedor (neto a pagar).
    # Importante: NO usamos doc["total"] porque incluye las retenciones del
    # DIAN que estamos ignorando intencionalmente (v2.6: regla 6). El neto
    # correcto = lo que efectivamente debitamos − lo que retuvimos.
    suma_debitos = (
        base_gravable + iva + imp_consumo + icui + inpp_ibua
    )
    neto_proveedor = suma_debitos - valor_retenido - valor_reteiva
    filas.append({**base_fila,
        "CUENTA": cuenta_proveedor,
        "DETALLE": f"PROVEEDOR FE {doc_referencia} - {nombre_emisor}",
        "TIPO DE TRANSACCION": "2",
        "VALOR": str(neto_proveedor),
        "BASE DE IMPUESTOS": "0",
    })

    return filas, valor_retenido, valor_reteiva


def _construir_asiento_nc(
    cfg: Configuracion,
    doc: dict,
    consecutivo: str,
    conocimiento: Optional[dict] = None,
):
    """Construye asiento de NC DIAN. IVA NC → SIEMPRE 24080204 (v2.6)."""
    clasif = _clasificar_proveedor(
        cfg, doc["nombre_emisor"], doc["nit_emisor"], conocimiento,
    )
    impuestos = _calcular_impuestos_factura(doc)

    cuenta_reversa = clasif["cuenta_gasto"]
    cuenta_proveedor = clasif["cuenta_pasivo"]
    cuenta_nc_iva = cfg.parametro("CUENTA_DIAN_NC_IVA", "24080204")

    comprobante = cfg.codigo_comprobante("NDI", "7")
    fecha_str = _formatear_fecha_para_plano(doc["fecha_raw"] or doc["fecha"])
    nombre_emisor = doc["nombre_emisor"][:80] if doc["nombre_emisor"] else "PROVEEDOR"

    iva = impuestos["iva"]
    base_gravable = impuestos["base_gravable"]
    imp_consumo = impuestos["imp_consumo_511575"]
    icui = impuestos["icui_511586"]
    inpp_ibua = impuestos["inpp_ibua_519525"]

    doc_referencia = doc["folio"]

    base_fila = {
        "COMPROBANTE": comprobante,
        "FECHA": fecha_str,
        "DOCUMENTO": consecutivo,
        "DOCUMENTO REFERENCIA": doc_referencia,
        "NIT": doc["nit_emisor"],
        "CENTRO DE COSTOS": cfg.parametro("CENTRO_COSTOS_DEFAULT", "PRINCIPAL"),
    }

    filas = []

    # Suma de "lo que se reversa" = lo que se acreditó originalmente
    suma_reversa = base_gravable + iva + imp_consumo + icui + inpp_ibua

    # Db: proveedor (cancela pasivo por la suma contable, no por el total DIAN)
    filas.append({**base_fila,
        "CUENTA": cuenta_proveedor,
        "DETALLE": f"NC {doc_referencia} PROVEEDOR - {nombre_emisor}",
        "TIPO DE TRANSACCION": "1",
        "VALOR": str(-suma_reversa),
        "BASE DE IMPUESTOS": "0",
    })

    # Cr: reversa compra/gasto
    if base_gravable != 0:
        filas.append({**base_fila,
            "CUENTA": cuenta_reversa,
            "DETALLE": f"NC {doc_referencia} REVERSA - {nombre_emisor}",
            "TIPO DE TRANSACCION": "2",
            "VALOR": str(base_gravable),
            "BASE DE IMPUESTOS": "0",
        })

    # Cr: IVA reversado (cuenta específica NC)
    if iva != 0:
        filas.append({**base_fila,
            "CUENTA": cuenta_nc_iva,
            "DETALLE": f"NC {doc_referencia} IVA - {nombre_emisor}",
            "TIPO DE TRANSACCION": "2",
            "VALOR": str(iva),
            "BASE DE IMPUESTOS": str(base_gravable),
        })

    # Reversa de impuestos especiales
    if imp_consumo != 0:
        filas.append({**base_fila,
            "CUENTA": cfg.parametro("CUENTA_IMP_CONSUMO", "511575"),
            "DETALLE": f"NC {doc_referencia} IMP CONSUMO - {nombre_emisor}",
            "TIPO DE TRANSACCION": "2",
            "VALOR": str(imp_consumo),
            "BASE DE IMPUESTOS": "0",
        })
    if icui != 0:
        filas.append({**base_fila,
            "CUENTA": cfg.parametro("CUENTA_ICUI", "511586"),
            "DETALLE": f"NC {doc_referencia} ICUI - {nombre_emisor}",
            "TIPO DE TRANSACCION": "2",
            "VALOR": str(icui),
            "BASE DE IMPUESTOS": "0",
        })
    if inpp_ibua != 0:
        filas.append({**base_fila,
            "CUENTA": cfg.parametro("CUENTA_INPP_IBUA", "519525"),
            "DETALLE": f"NC {doc_referencia} INPP/IBUA - {nombre_emisor}",
            "TIPO DE TRANSACCION": "2",
            "VALOR": str(inpp_ibua),
            "BASE DE IMPUESTOS": "0",
        })

    return filas


def _detectar_proveedores_nuevos(
    cfg: Configuracion,
    documentos: list,
    conocimiento: Optional[dict] = None,
) -> list:
    """Lista proveedores que NO están en DIAN_Proveedores ni en el balance.

    Un proveedor se considera "nuevo" si:
      - No coincide con ninguna palabra clave de DIAN_Proveedores, Y
      - Su NIT no aparece en el balance histórico (cuando hay balance).
    """
    proveedores = {}
    for doc in documentos:
        nit = doc["nit_emisor"]
        if not nit:
            continue
        if nit not in proveedores:
            proveedores[nit] = {
                "nit": nit,
                "nombre": doc["nombre_emisor"],
                "num_facturas": 0,
                "total_acumulado": 0,
            }
        proveedores[nit]["num_facturas"] += 1
        proveedores[nit]["total_acumulado"] += doc["total"]

    nuevos = []
    for nit, info in proveedores.items():
        # Si tiene regla por palabra clave → no es nuevo
        if _buscar_regla_proveedor(cfg, info["nombre"]):
            continue
        # Si está en el balance → tampoco es nuevo (hay info histórica)
        if conocimiento and cb.es_nit_conocido(conocimiento, nit):
            continue
        info["tipo_retencion_sugerido"] = (
            "COMPRA_JURIDICA" if (nit and nit[0] in "89") else "COMPRA_NATURAL"
        )
        nuevos.append(info)
    return nuevos


# ============================================================
# Función principal (firma estilo Caja Menor)
# ============================================================

# Orden de columnas del plano (igual que Caja Menor)
COLUMNAS_PLANO = [
    "CUENTA", "COMPROBANTE", "FECHA", "DOCUMENTO", "DOCUMENTO REFERENCIA",
    "NIT", "DETALLE", "TIPO DE TRANSACCION", "VALOR", "BASE DE IMPUESTOS",
    "CENTRO DE COSTOS",
]


def procesar_compras_dian(
    archivo_dian,
    cfg: Configuracion,
    conocimiento: Optional[dict] = None,
) -> Tuple[pd.DataFrame, List[str], List[dict]]:
    """Procesa el archivo DIAN y genera el DataFrame plano.

    Args:
        archivo_dian: file-like o bytes del Excel del portal DIAN.
        cfg: Configuracion de la empresa.
        conocimiento: (opcional) dict cargado desde Supabase Storage con
            la información histórica del balance de prueba. Si se pasa,
            activa la regla de respeto al histórico v2.6:
              - NIT en balance sin retefuente histórica → no retefuente.
              - Cuenta de gasto / IVA sugerida desde el balance.

    Returns:
        (df_plano, log_mensajes, proveedores_nuevos)
        df_plano: DataFrame con las columnas del plano contable.
        log_mensajes: lista de strings para mostrar al usuario.
        proveedores_nuevos: lista de dicts con info de proveedores no
            registrados en DIAN_Proveedores ni en el balance.
    """
    log: List[str] = []

    documentos = _leer_documentos(archivo_dian)
    if not documentos:
        log.append("⚠️ No se detectaron documentos en el archivo DIAN.")
        df_vacio = pd.DataFrame(columns=COLUMNAS_PLANO)
        return df_vacio, log, []

    # Indicar al usuario si se activó la regla histórica
    if conocimiento:
        n_nits_balance = len(conocimiento.get("nits", {}))
        if _respetar_historico(cfg):
            log.append(
                f"ℹ️ Balance histórico cargado: {n_nits_balance} NITs. "
                "Regla histórica de retefuente ACTIVA."
            )
        else:
            log.append(
                f"ℹ️ Balance histórico cargado: {n_nits_balance} NITs "
                "(solo informativo, regla histórica desactivada)."
            )

    proveedores_nuevos = _detectar_proveedores_nuevos(
        cfg, documentos, conocimiento,
    )

    contadores = {}
    filas_salida = []
    n_facturas = 0
    n_nc = 0
    total_retenido = 0
    total_reteiva = 0

    for doc in documentos:
        consecutivo = _generar_consecutivo(doc["fecha"], contadores)
        if not consecutivo:
            continue

        if doc["es_nc"]:
            filas = _construir_asiento_nc(cfg, doc, consecutivo, conocimiento)
            n_nc += 1
        else:
            filas, val_ret, val_reteiva = _construir_asiento_factura(
                cfg, doc, consecutivo, conocimiento,
            )
            n_facturas += 1
            total_retenido += val_ret
            total_reteiva += val_reteiva

        filas_salida.extend(filas)

    log.append(
        f"Documentos DIAN: {n_facturas} facturas + {n_nc} notas crédito "
        f"= {n_facturas + n_nc} total"
    )
    if total_retenido > 0:
        log.append(f"Retefuente aplicada:  $ {total_retenido:,.0f}".replace(",", "."))
    if total_reteiva > 0:
        log.append(f"ReteIVA 15% (RST):    $ {total_reteiva:,.0f}".replace(",", "."))
    if proveedores_nuevos:
        log.append(
            f"⚠️ Proveedores nuevos detectados: {len(proveedores_nuevos)} "
            "(se procesaron con cuentas por defecto; ajústalos en mapeos.xlsx)"
        )

    df_salida = pd.DataFrame(filas_salida, columns=COLUMNAS_PLANO)
    return df_salida, log, proveedores_nuevos


# ============================================================
# Generar TSV (idéntico al de Caja Menor)
# ============================================================

def dataframe_a_plano_tsv(df: pd.DataFrame, incluir_encabezado_excel: bool = True) -> bytes:
    """Convierte el DataFrame plano a bytes TSV para descarga.

    Args:
        incluir_encabezado_excel: si True, agrega 'sep=\\t' al inicio para que
                                   Excel abra el archivo correctamente.
    """
    buf = io.StringIO()
    if incluir_encabezado_excel:
        buf.write("sep=\t\n")
    df.to_csv(buf, sep="\t", index=False)
    return buf.getvalue().encode("utf-8-sig")
