"""
core/procesadores/parser_token_dian.py

Lee el reporte del Token DIAN (un Excel descargado del portal de la DIAN
con todas las facturas electrónicas emitidas/recibidas por el NIT) y
produce un dataframe agregado por (fecha, prefijo) listo para comparar
contra el plano contable del POS.

Reglas de filtrado aplicadas:
  1. NIT Emisor = NIT de la empresa (solo lo que la empresa EMITIÓ).
  2. Tipo de documento = "Factura electrónica" (descarta application
     response, nómina, DSE, notas crédito — esas tienen flujo propio).
  3. Prefijos del maestro: solo se aceptan los prefijos registrados en
     `datos_punto.json` (campo `prefijo_token` de cada sucursal). Los
     prefijos extras se reportan como "ruido" sin matar el proceso.
  4. Prefijos OMITIDOS por configuración: por defecto se omite "STL"
     porque sus facturas tienen tarifa IVA variable y se procesan por
     otro flujo (Henko).

Reglas de desglose IVA/INC:
  El Token solo trae el TOTAL con impuestos incluidos (sin desglose
  base/IVA). Asumimos tarifa fija 8% INC (Santa Leña / Milagros) y
  reconstruimos:
      base_teorica = total / 1.08
      inc_teorico  = base_teorica * 0.08
  Si (base + inc) ≈ total → desglose correcto.
  Si la diferencia es ~10% del total → es propina (manejo interno),
  se omite.
  Si la diferencia es otra → marcar como "Revisar manualmente".

Funciones públicas:
    parsear_token_dian(fuente, nit_empresa, sucursales,
                       prefijos_omitidos=("STL",))
        → dict con:
            'agregado_fecha_prefijo': DataFrame con
                fecha, prefijo, sucursal_cc, sucursal_nombre,
                docs, total_bruto, base_teorica, inc_teorico,
                propina_estimada, estado_desglose
            'prefijos_no_mapeados': dict prefijo→count (no están en
                el maestro)
            'prefijos_omitidos': dict prefijo→count (excluidos por
                config como STL)
            'descartados_por_tipo': dict tipo_doc→count
            'total_filas_leidas': int
            'log': list[str] (mensajes informativos)
"""

from __future__ import annotations

import io
from collections import Counter, defaultdict
from datetime import date, datetime
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


# ============================================================
# Constantes
# ============================================================

NIT_JIPER = "901038325"  # default; el caller debe pasar el NIT real

# Columnas esperadas del Token (orden fijo del export DIAN)
COL_TIPO_DOC       = "Tipo de documento"
COL_FOLIO          = "Folio"
COL_PREFIJO        = "Prefijo"
COL_FECHA_EMISION  = "Fecha Emisión"
COL_NIT_EMISOR     = "NIT Emisor"
COL_NIT_RECEPTOR   = "NIT Receptor"
COL_IVA            = "IVA"
COL_INC            = "IC"
COL_TOTAL          = "Total"

TIPO_FACTURA_E     = "Factura electrónica"
TIPO_NOTA_CREDITO  = "Nota de crédito electrónica"

# Tarifa INC vigente Santa Leña / Milagros (puede pasarse por parámetro)
TARIFA_INC_DEFAULT = 0.08

# Tolerancia para considerar (base+inc ≈ total)
TOLERANCIA_REDONDEO_PESOS = 5

# Margen para detectar propina (~10% del total)
PROPINA_MIN = 0.085
PROPINA_MAX = 0.115


# ============================================================
# Helpers
# ============================================================

def _abrir_excel(fuente):
    """Acepta ruta, bytes o file-like y devuelve un openpyxl Workbook."""
    if hasattr(fuente, "read"):
        contenido = fuente.read()
        if hasattr(fuente, "seek"):
            fuente.seek(0)
        bio = io.BytesIO(contenido)
    elif isinstance(fuente, (bytes, bytearray)):
        bio = io.BytesIO(bytes(fuente))
    elif isinstance(fuente, str):
        bio = fuente
    else:
        raise TypeError(f"Tipo de fuente no soportado: {type(fuente)}")
    return load_workbook(bio, read_only=True, data_only=True)


def _parsear_fecha_token(v) -> Optional[date]:
    """Parsea la fecha del Token. Formato típico: '01-03-2026' (DD-MM-YYYY)."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _to_number(v) -> float:
    """Convierte a número, devuelve 0.0 si no se puede."""
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _normalizar_prefijo(p) -> str:
    """Devuelve el prefijo en mayúsculas o cadena vacía."""
    if p is None:
        return ""
    s = str(p).strip().upper()
    return s


def _desglosar_total(total: float, tarifa_inc: float = TARIFA_INC_DEFAULT) -> dict:
    """
    Toma el total bruto y devuelve el desglose teórico.

    Estados posibles:
      - 'correcto': base + inc ≈ total (sin propina)
      - 'con_propina': la diferencia es ~10% del total
      - 'revisar': la diferencia no encaja en propina ni en redondeo
      - 'vacio': total es 0 o negativo
    """
    if total <= 0:
        return {
            "base_teorica":     0.0,
            "inc_teorico":      0.0,
            "propina_estimada": 0.0,
            "estado_desglose":  "vacio",
        }

    base = round(total / (1 + tarifa_inc))
    inc = round(base * tarifa_inc)
    suma = base + inc
    diferencia = total - suma

    if abs(diferencia) <= TOLERANCIA_REDONDEO_PESOS:
        return {
            "base_teorica":     float(base),
            "inc_teorico":      float(inc),
            "propina_estimada": 0.0,
            "estado_desglose":  "correcto",
        }

    # ¿La diferencia es propina (~10% sobre la base)?
    if base > 0:
        ratio = diferencia / base
        if PROPINA_MIN <= ratio <= PROPINA_MAX:
            return {
                "base_teorica":     float(base),
                "inc_teorico":      float(inc),
                "propina_estimada": float(round(diferencia)),
                "estado_desglose":  "con_propina",
            }

    return {
        "base_teorica":     float(base),
        "inc_teorico":      float(inc),
        "propina_estimada": float(round(diferencia)),
        "estado_desglose":  "revisar",
    }


# ============================================================
# Función principal
# ============================================================

def parsear_token_dian(
    fuente,
    nit_empresa: str,
    sucursales: List,
    prefijos_omitidos: Tuple[str, ...] = ("STL",),
    tarifa_inc: float = TARIFA_INC_DEFAULT,
) -> dict:
    """
    Lee el Excel del Token DIAN y devuelve un agregado por (fecha, prefijo).

    Args:
        fuente: ruta, bytes o file-like del Excel del Token.
        nit_empresa: NIT de la empresa (solo se procesa lo que ELLA emitió).
        sucursales: lista de Sucursal del maestro (con campo prefijo_token).
        prefijos_omitidos: tupla de prefijos a omitir (default: STL).
        tarifa_inc: tarifa INC para el desglose (default: 0.08 = 8%).

    Returns:
        dict con el resultado del procesamiento (ver docstring del módulo).
    """
    nit_empresa = str(nit_empresa).strip()
    prefijos_omitidos = tuple(p.strip().upper() for p in prefijos_omitidos)

    # Indexar el maestro por prefijo
    sucs_por_prefijo: dict = {}
    for s in sucursales:
        pref = (getattr(s, "prefijo_token", "") or "").strip().upper()
        if pref:
            sucs_por_prefijo[pref] = s

    log: List[str] = []
    log.append(f"📋 Maestro: {len(sucs_por_prefijo)} sucursales con prefijo_token registrado.")
    if prefijos_omitidos:
        log.append(f"🚫 Prefijos omitidos por configuración: {', '.join(prefijos_omitidos)}")

    wb = _abrir_excel(fuente)
    ws = wb.active
    log.append(f"📂 Hoja leída: '{ws.title}' ({ws.max_row:,} filas)")

    # Localizar encabezado
    encabezado_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    encabezado = next(encabezado_iter, None)
    if not encabezado:
        raise ValueError("El Excel del Token está vacío o sin encabezado.")
    encabezado = [str(c).strip() if c is not None else "" for c in encabezado]

    def idx_col(nombre: str) -> int:
        try:
            return encabezado.index(nombre)
        except ValueError:
            raise ValueError(
                f"Columna '{nombre}' no encontrada en el Token. "
                f"Encabezado leído: {encabezado}"
            )

    i_tipo     = idx_col(COL_TIPO_DOC)
    i_folio    = idx_col(COL_FOLIO)
    i_prefijo  = idx_col(COL_PREFIJO)
    i_fecha    = idx_col(COL_FECHA_EMISION)
    i_emisor   = idx_col(COL_NIT_EMISOR)
    i_iva      = idx_col(COL_IVA)
    i_inc      = idx_col(COL_INC)
    i_total    = idx_col(COL_TOTAL)

    # Acumuladores
    # Clave: (fecha_date, prefijo_str) → datos agregados
    agregado: dict = defaultdict(lambda: {
        "docs": 0,
        "total_bruto": 0.0,
        "iva_reportado": 0.0,
        "inc_reportado": 0.0,
    })
    prefijos_no_mapeados: Counter = Counter()
    prefijos_omitidos_count: Counter = Counter()
    descartados_por_tipo: Counter = Counter()
    descartados_otro_emisor = 0
    descartados_fecha_invalida = 0
    total_filas = 0

    # Iterar las filas
    for fila in ws.iter_rows(min_row=2, values_only=True):
        total_filas += 1
        if not fila or fila[i_tipo] is None:
            continue

        tipo_doc = str(fila[i_tipo]).strip()

        # Filtro 1: solo factura electrónica
        if tipo_doc != TIPO_FACTURA_E:
            descartados_por_tipo[tipo_doc] += 1
            continue

        # Filtro 2: solo lo que LA EMPRESA emitió
        nit_emi = str(fila[i_emisor] or "").strip()
        if nit_emi != nit_empresa:
            descartados_otro_emisor += 1
            continue

        # Filtro 3: prefijo presente
        prefijo = _normalizar_prefijo(fila[i_prefijo])
        if not prefijo:
            # Sin prefijo, no podemos asignarlo a una sucursal
            prefijos_no_mapeados["(sin prefijo)"] += 1
            continue

        # Filtro 4: omitidos por configuración (STL)
        if prefijo in prefijos_omitidos:
            prefijos_omitidos_count[prefijo] += 1
            continue

        # Filtro 5: prefijo debe estar en el maestro
        if prefijo not in sucs_por_prefijo:
            prefijos_no_mapeados[prefijo] += 1
            continue

        # Filtro 6: fecha válida
        f = _parsear_fecha_token(fila[i_fecha])
        if not f:
            descartados_fecha_invalida += 1
            continue

        total = _to_number(fila[i_total])
        if total <= 0:
            # Documentos con total 0 o negativos no se agregan a ventas
            continue

        clave = (f, prefijo)
        agregado[clave]["docs"] += 1
        agregado[clave]["total_bruto"] += total
        agregado[clave]["iva_reportado"] += _to_number(fila[i_iva])
        agregado[clave]["inc_reportado"] += _to_number(fila[i_inc])

    # Construir DataFrame final con desglose teórico
    filas_df = []
    for (fecha, prefijo), data in sorted(agregado.items()):
        suc = sucs_por_prefijo[prefijo]
        desglose = _desglosar_total(data["total_bruto"], tarifa_inc=tarifa_inc)
        filas_df.append({
            "fecha":            fecha,
            "prefijo":          prefijo,
            "sucursal_cc":      suc.cc,
            "sucursal_nombre":  suc.nombre_reporte,
            "clase":            suc.clase,
            "docs":             data["docs"],
            "total_bruto":      float(round(data["total_bruto"])),
            "base_teorica":     desglose["base_teorica"],
            "inc_teorico":      desglose["inc_teorico"],
            "propina_estimada": desglose["propina_estimada"],
            "estado_desglose":  desglose["estado_desglose"],
            "iva_reportado":    float(round(data["iva_reportado"])),
            "inc_reportado":    float(round(data["inc_reportado"])),
        })

    df_agregado = pd.DataFrame(filas_df, columns=[
        "fecha", "prefijo", "sucursal_cc", "sucursal_nombre", "clase",
        "docs", "total_bruto", "base_teorica", "inc_teorico",
        "propina_estimada", "estado_desglose",
        "iva_reportado", "inc_reportado",
    ])

    # Mensajes de log
    log.append(f"📊 Filas leídas: {total_filas:,}")
    log.append(f"   Descartadas por tipo de documento: {sum(descartados_por_tipo.values()):,}")
    log.append(f"   Descartadas por otro emisor: {descartados_otro_emisor:,}")
    log.append(f"   Descartadas por fecha inválida: {descartados_fecha_invalida:,}")
    if prefijos_omitidos_count:
        log.append(f"   Omitidas por prefijo configurado: "
                   f"{sum(prefijos_omitidos_count.values()):,} "
                   f"(detalle: {dict(prefijos_omitidos_count)})")
    if prefijos_no_mapeados:
        log.append(f"   ⚠️ Prefijos sin mapeo: {sum(prefijos_no_mapeados.values()):,}")

    log.append(f"✅ Filas agregadas (fecha × prefijo): {len(df_agregado):,}")

    if len(df_agregado) > 0:
        total_token = int(df_agregado["total_bruto"].sum())
        log.append(f"💰 Total ventas Token: ${total_token:,}".replace(",", "."))

        # Estados de desglose
        estados = df_agregado["estado_desglose"].value_counts().to_dict()
        log.append(f"🔍 Estados de desglose: {estados}")

    return {
        "agregado_fecha_prefijo":  df_agregado,
        "prefijos_no_mapeados":    dict(prefijos_no_mapeados),
        "prefijos_omitidos":       dict(prefijos_omitidos_count),
        "descartados_por_tipo":    dict(descartados_por_tipo),
        "descartados_otro_emisor": descartados_otro_emisor,
        "total_filas_leidas":      total_filas,
        "log":                     log,
    }


# ============================================================
# Función auxiliar: agregar el plano POS por (fecha, CC) para comparar
# ============================================================

def agregar_plano_pos_por_fecha_cc(df_plano: pd.DataFrame) -> pd.DataFrame:
    """
    Toma el dataframe del plano contable POS (devuelto por procesar_pos)
    y lo agrega por (fecha, centro de costo) para poder compararlo
    contra el Token.

    Para cada (fecha, CC) suma el valor en TR=2 (créditos) de las cuentas
    de venta — base + INC + bolsas — que es lo que debería corresponder
    al total bruto del Token.

    Args:
        df_plano: DataFrame del plano POS con columnas:
            CUENTA, COMPROBANTE, FECHA, DOCUMENTO, DOC REFERENCIA,
            NIT, DETALLE, TR, VALOR, BASE, CENTRO DE COSTO

    Returns:
        DataFrame con: fecha (date), sucursal_cc, total_pos
    """
    if df_plano is None or len(df_plano) == 0:
        return pd.DataFrame(columns=["fecha", "sucursal_cc", "total_pos"])

    df = df_plano.copy()

    # Normalizar tipos
    df["TR"] = df["TR"].astype(str)
    df["VALOR"] = pd.to_numeric(df["VALOR"], errors="coerce").fillna(0)
    df["CENTRO DE COSTO"] = df["CENTRO DE COSTO"].astype(str)

    # Convertir FECHA (formato "MM/DD/YYYY") a date
    def _to_date(s):
        try:
            return datetime.strptime(str(s).strip(), "%m/%d/%Y").date()
        except (ValueError, TypeError):
            return None

    df["_fecha_d"] = df["FECHA"].apply(_to_date)

    # Filtrar: solo créditos (TR=2). La caja (Db) duplica el monto, no la queremos.
    # Pero la caja TR=1 = base + INC + bolsas + IVA bolsas — es el TOTAL bruto neto.
    # Más limpio: tomar la fila de caja (Db), que es el "total final" de la sucursal-día.
    # En el procesador POS la línea de caja siempre es TR=1, y su VALOR es el bruto.
    df_caja = df[df["TR"] == "1"].copy()

    # Identificar la fila de caja: es la que tiene cuenta_caja específica (11050xxx)
    # y agrupa todos los créditos del día. Usamos directamente la suma de créditos
    # de las cuentas de venta (TR=2), que es equivalente y más a prueba de errores
    # si se cambian las cuentas.
    df_cr = df[df["TR"] == "2"].copy()

    agg = (
        df_cr
        .groupby(["_fecha_d", "CENTRO DE COSTO"])["VALOR"]
        .sum()
        .reset_index()
        .rename(columns={
            "_fecha_d": "fecha",
            "CENTRO DE COSTO": "sucursal_cc",
            "VALOR": "total_pos",
        })
    )
    agg["total_pos"] = agg["total_pos"].astype(float)
    agg = agg.dropna(subset=["fecha"]).reset_index(drop=True)
    return agg


# ============================================================
# Parser de Notas Crédito (función paralela a parsear_token_dian)
# ============================================================

def parsear_notas_credito_token(
    fuente,
    nit_empresa: str,
    sucursales: List,
    tarifa_inc: float = TARIFA_INC_DEFAULT,
) -> dict:
    """
    Parsea las Notas Crédito Electrónicas del Token DIAN.

    Las NC tienen prefijos propios (NCVI, NCLU, NCT, NCFA, etc.) que NO
    coinciden con los prefijos de venta. En `datos_punto.json` cada sucursal
    tiene un campo `prefijo_token_nc` que mapea el prefijo NC al CC.

    Reglas:
      - Solo `Nota de crédito electrónica`.
      - Solo cuando JIPER es emisor (NIT Emisor = nit_empresa).
      - Solo prefijos NC mapeados en el maestro (`prefijo_token_nc`).
      - Las NC SIN prefijo se omiten siempre (NO se contabilizan automáticamente):
          * Si folio empieza por "NC2" → motivo "STL (flujo Henko)".
          * Otro folio → motivo "interna (subida manual)".
        En ambos casos quedan en `nc_omitidas` para que el contador las
        suba a la contabilidad por separado.
      - Igual que las facturas: el INC del Token siempre viene en 0 para
        las NC, así que recalculamos: base = total/1.08, inc = base*0.08.

    Args:
        fuente: ruta, bytes o file-like del Excel del Token DIAN.
        nit_empresa: NIT de la empresa (solo lo que ella emitió).
        sucursales: lista de Sucursal del maestro.
        tarifa_inc: tarifa INC para el desglose (default 0.08 = 8%).

    Returns:
        dict con:
            'detalle_nc': DataFrame con una fila por cada NC POS, con
                columnas: fecha, prefijo, folio, sucursal_cc,
                sucursal_nombre, clase, nit_receptor, cliente,
                total_bruto, base_teorica, inc_teorico.
            'nc_omitidas': DataFrame con las NC sin prefijo (STL e internas).
                Columnas: fecha, folio, nit_receptor, cliente, total_bruto,
                motivo. NO se contabilizan; quedan para referencia.
            'nc_stl': alias de nc_omitidas filtrado por motivo STL
                (mantenido por compatibilidad con código anterior).
            'nc_no_mapeadas': dict prefijo→list[dict] con NCs cuyo prefijo
                NO está en el maestro (acción requerida del contador).
            'total_filas_leidas': int
            'log': list[str]
    """
    nit_empresa = str(nit_empresa).strip()

    # Indexar el maestro por prefijo_token_nc
    sucs_por_prefijo_nc: dict = {}
    for s in sucursales:
        pref_nc = (getattr(s, "prefijo_token_nc", "") or "").strip().upper()
        if pref_nc:
            sucs_por_prefijo_nc[pref_nc] = s

    log: List[str] = []
    log.append(f"📋 Maestro: {len(sucs_por_prefijo_nc)} sucursales con prefijo_token_nc registrado.")

    wb = _abrir_excel(fuente)
    ws = wb.active

    # Localizar encabezado (mismo que en parsear_token_dian)
    encabezado_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    encabezado = next(encabezado_iter, None)
    if not encabezado:
        raise ValueError("El Excel del Token está vacío o sin encabezado.")
    encabezado = [str(c).strip() if c is not None else "" for c in encabezado]

    def idx_col(nombre: str) -> int:
        try:
            return encabezado.index(nombre)
        except ValueError:
            raise ValueError(f"Columna '{nombre}' no encontrada en el Token.")

    i_tipo     = idx_col(COL_TIPO_DOC)
    i_folio    = idx_col(COL_FOLIO)
    i_prefijo  = idx_col(COL_PREFIJO)
    i_fecha    = idx_col(COL_FECHA_EMISION)
    i_emisor   = idx_col(COL_NIT_EMISOR)
    i_receptor = idx_col(COL_NIT_RECEPTOR)
    # Columna 12 = "Nombre Receptor"
    i_nombre_rec = 12
    i_total    = idx_col(COL_TOTAL)

    filas_nc_pos = []
    filas_nc_omitidas = []      # STL + sin prefijo (subidas manualmente por el contador)
    nc_no_mapeadas: dict = {}   # solo prefijos NC desconocidos (errores reales)
    total_filas = 0

    for fila in ws.iter_rows(min_row=2, values_only=True):
        total_filas += 1
        if not fila or fila[i_tipo] is None:
            continue

        tipo_doc = str(fila[i_tipo]).strip()
        if tipo_doc != TIPO_NOTA_CREDITO:
            continue

        nit_emi = str(fila[i_emisor] or "").strip()
        if nit_emi != nit_empresa:
            continue

        prefijo = _normalizar_prefijo(fila[i_prefijo])
        folio = str(fila[i_folio] or "").strip()
        fecha = _parsear_fecha_token(fila[i_fecha])
        total = _to_number(fila[i_total])
        nit_rec = str(fila[i_receptor] or "").strip()
        cliente = ""
        if i_nombre_rec < len(fila) and fila[i_nombre_rec]:
            cliente = str(fila[i_nombre_rec]).strip()

        if not fecha or total <= 0:
            continue

        # NC SIN PREFIJO: se omiten del flujo automático.
        # Pueden ser de dos tipos:
        #   - STL/Henko (folio "NC2xxx") → se procesarán con módulo STL futuro.
        #   - Internas (folio distinto)  → el contador las sube manualmente.
        # En ambos casos quedan reportadas para referencia, no se contabilizan.
        if not prefijo:
            if folio.upper().startswith("NC2"):
                motivo = "STL (flujo Henko)"
            else:
                motivo = "interna (subida manual)"
            filas_nc_omitidas.append({
                "fecha":         fecha,
                "folio":         folio,
                "nit_receptor":  nit_rec,
                "cliente":       cliente,
                "total_bruto":   float(round(total)),
                "motivo":        motivo,
            })
            continue

        # NC POS con prefijo desconocido (error real: alguien emitió con prefijo
        # nuevo que no está en el maestro → hay que ajustar el maestro)
        if prefijo not in sucs_por_prefijo_nc:
            nc_no_mapeadas.setdefault(prefijo, []).append({
                "fecha": fecha, "folio": folio, "total": total, "cliente": cliente,
            })
            continue

        suc = sucs_por_prefijo_nc[prefijo]
        desglose = _desglosar_total(total, tarifa_inc=tarifa_inc)

        filas_nc_pos.append({
            "fecha":           fecha,
            "prefijo":         prefijo,
            "folio":           folio,
            "sucursal_cc":     suc.cc,
            "sucursal_nombre": suc.nombre_reporte,
            "clase":           suc.clase,
            "nit_receptor":    nit_rec,
            "cliente":         cliente,
            "total_bruto":     float(round(total)),
            "base_teorica":    desglose["base_teorica"],
            "inc_teorico":     desglose["inc_teorico"],
            "estado_desglose": desglose["estado_desglose"],
        })

    # DataFrames
    df_pos = pd.DataFrame(filas_nc_pos, columns=[
        "fecha", "prefijo", "folio", "sucursal_cc", "sucursal_nombre",
        "clase", "nit_receptor", "cliente",
        "total_bruto", "base_teorica", "inc_teorico", "estado_desglose",
    ])
    df_omitidas = pd.DataFrame(filas_nc_omitidas, columns=[
        "fecha", "folio", "nit_receptor", "cliente", "total_bruto", "motivo",
    ])

    log.append(f"📊 Notas Crédito procesadas:")
    log.append(f"   - NC POS contabilizables: {len(df_pos):,}")
    if len(df_pos):
        log.append(f"     · Total: ${df_pos['total_bruto'].sum():,.0f}".replace(",", "."))
    log.append(f"   - NC omitidas (sin prefijo): {len(df_omitidas):,}")
    if len(df_omitidas):
        log.append(f"     · Total: ${df_omitidas['total_bruto'].sum():,.0f}".replace(",", "."))
        # Desglosar por motivo
        for motivo, sub in df_omitidas.groupby("motivo"):
            log.append(f"     · {motivo}: {len(sub)} doc(s), ${sub['total_bruto'].sum():,.0f}".replace(",", "."))
    if nc_no_mapeadas:
        cant_no_map = sum(len(v) for v in nc_no_mapeadas.values())
        log.append(
            f"   - ⚠️ NC con prefijo no mapeado: {cant_no_map:,} "
            f"({list(nc_no_mapeadas.keys())})"
        )

    return {
        "detalle_nc":      df_pos,
        "nc_omitidas":     df_omitidas,
        "nc_stl":          df_omitidas[df_omitidas["motivo"].str.startswith("STL")] if len(df_omitidas) else df_omitidas,  # compat
        "nc_no_mapeadas":  nc_no_mapeadas,
        "total_filas_leidas": total_filas,
        "log":             log,
    }


# ============================================================
# Generación de líneas contables para NC POS
# ============================================================

def generar_lineas_nc_pos(
    df_nc: pd.DataFrame,
    sucursales: List,
    comprobante_default: str = "498",
) -> pd.DataFrame:
    """
    Genera las líneas del plano contable a partir de las NC POS.

    Asiento por cada NC:
        Db cta_devoluciones (41754001) ← base_teorica
        Db cta_ico (24800505)          ← inc_teorico
        Cr cuenta_caja sucursal        ← total_bruto

    Args:
        df_nc: DataFrame de `parsear_notas_credito_token()['detalle_nc']`.
        sucursales: lista de Sucursal (para tomar las cuentas).
        comprobante_default: código de comprobante para las NC (default '498',
            distinto del POS que es '497').

    Returns:
        DataFrame con las mismas 11 columnas que el plano POS:
            CUENTA, COMPROBANTE, FECHA, DOCUMENTO, DOC REFERENCIA, NIT,
            DETALLE, TR, VALOR, BASE, CENTRO DE COSTO.
    """
    suc_por_cc = {s.cc: s for s in sucursales}

    filas = []
    for _, r in df_nc.iterrows():
        cc = str(r["sucursal_cc"])
        suc = suc_por_cc.get(cc)
        if not suc:
            continue
        if not r["fecha"] or r["total_bruto"] <= 0:
            continue

        fecha_str = (
            r["fecha"].strftime("%m/%d/%Y")
            if hasattr(r["fecha"], "strftime")
            else str(r["fecha"])
        )
        base = int(round(r["base_teorica"]))
        inc = int(round(r["inc_teorico"]))
        total = int(round(r["total_bruto"]))

        # Ajuste de redondeo: si base + inc != total, ajustar la base
        suma = base + inc
        if suma != total:
            base += (total - suma)

        documento = f"{r['prefijo']}{r['folio']}"[:14]  # prefijo ya empieza por NC; acotar a 14 chars
        nit_rec = (r.get("nit_receptor") or "222222222").strip() or "222222222"
        detalle = f"NC POS {suc.nombre_reporte}"

        cta_dev = suc.cta_devoluciones or "41754001"

        # Db devoluciones (base)
        if base > 0:
            filas.append({
                "CUENTA":          cta_dev,
                "COMPROBANTE":     comprobante_default,
                "FECHA":           fecha_str,
                "DOCUMENTO":       documento,
                "DOC REFERENCIA":  documento,
                "NIT":             nit_rec,
                "DETALLE":         detalle,
                "TR":              "1",
                "VALOR":           base,
                "BASE":            base,
                "CENTRO DE COSTO": cc,
            })
        # Db INC
        if inc > 0:
            filas.append({
                "CUENTA":          suc.cta_ico,
                "COMPROBANTE":     comprobante_default,
                "FECHA":           fecha_str,
                "DOCUMENTO":       documento,
                "DOC REFERENCIA":  documento,
                "NIT":             nit_rec,
                "DETALLE":         detalle,
                "TR":              "1",
                "VALOR":           inc,
                "BASE":            base,
                "CENTRO DE COSTO": cc,
            })
        # Cr caja (total)
        filas.append({
            "CUENTA":          suc.cuenta_caja,
            "COMPROBANTE":     comprobante_default,
            "FECHA":           fecha_str,
            "DOCUMENTO":       documento,
            "DOC REFERENCIA":  documento,
            "NIT":             nit_rec,
            "DETALLE":         detalle,
            "TR":              "2",
            "VALOR":           total,
            "BASE":            0,
            "CENTRO DE COSTO": cc,
        })

    df_lineas = pd.DataFrame(filas, columns=[
        "CUENTA", "COMPROBANTE", "FECHA", "DOCUMENTO", "DOC REFERENCIA",
        "NIT", "DETALLE", "TR", "VALOR", "BASE", "CENTRO DE COSTO",
    ])
    return df_lineas
