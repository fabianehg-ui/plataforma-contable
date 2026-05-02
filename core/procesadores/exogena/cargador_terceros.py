"""
Cargador de terceros revisados desde Excel a la base de datos.

Usa el clasificador_nits para aplicar reglas oficiales DIAN antes de cargar.

Uso desde el módulo de exógena:
    from cargador_terceros import cargar_terceros_desde_excel
    
    resultado = cargar_terceros_desde_excel(
        archivo_xlsx='terceros_revisados.xlsx',
        empresa_id=1,
        supabase_client=supabase
    )
"""

from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
import re

import openpyxl

try:
    from .clasificador_nits import reclasificar_tercero, calc_dv
except ImportError:
    from clasificador_nits import reclasificar_tercero, calc_dv


@dataclass
class ResultadoCarga:
    total_filas: int = 0
    insertados: int = 0
    actualizados: int = 0
    descartados: int = 0
    errores: list[str] = None

    def __post_init__(self):
        if self.errores is None:
            self.errores = []


def _clean(s, max_len=None):
    if s is None:
        return ''
    s = re.sub(r'\s+', ' ', str(s).strip())
    return s[:max_len] if max_len else s


def parsear_excel_terceros(archivo: str | Path, aplicar_clasificador: bool = True) -> list[dict]:
    """Lee el Excel revisable y devuelve lista de dicts listos para BD.
    
    Si aplicar_clasificador=True, pasa cada tercero por reclasificar_tercero()
    antes de devolverlo, lo que aplica reglas oficiales DIAN y detecta NITs
    con DV pegado.
    """
    wb = openpyxl.load_workbook(archivo, data_only=True, read_only=True)
    ws = wb['Terceros'] if 'Terceros' in wb.sheetnames else wb.active

    # Detectar si tiene la columna 'NIT Original' (Excel reclasificado v2)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []
    
    header = [str(h or '').strip().lower() for h in rows[0]]
    tiene_nit_original = 'nit original' in header

    terceros = []
    for row in rows[1:]:
        if not row[0]:
            continue
        nit = re.sub(r'[^\d]', '', str(row[0]))
        if len(nit) < 7 or len(nit) > 11:
            continue

        # Layout del Excel revisable v2:
        # 0:NIT 1:NIT Original 2:DV 3:TipoDoc 4:TipoPersona 5:RazónSocial
        # 6:PrimerApellido 7:SegundoApellido 8:PrimerNombre 9:OtrosNombres
        # 10:Dirección 11:CodDpto 12:CodMcp 13:CodPaís 14:Email 15:CIIU
        # 16:ReglaAplicada 17:Sugerencias
        # Si no tiene NIT Original (v1), restar 1 a los índices >=2
        offset = 0 if tiene_nit_original else -1
        
        def col(i):
            idx = i if i < 2 else i + offset
            return row[idx] if idx < len(row) else None
        
        tipo_persona = _clean(col(4)).lower() or 'natural'

        t = {
            'nit': nit,
            'dv': calc_dv(nit) if nit.isdigit() else None,
            'nit_original': _clean(col(1), 20) if tiene_nit_original else '',
            'tipo_documento': int(col(3)) if col(3) else (31 if tipo_persona == 'juridica' else 13),
            'tipo_persona': tipo_persona,
            'razon_social': _clean(col(5), 450),
            'primer_apellido': _clean(col(6), 60),
            'segundo_apellido': _clean(col(7), 60),
            'primer_nombre': _clean(col(8), 60),
            'otros_nombres': _clean(col(9), 60),
            'direccion': _clean(col(10), 200),
            'codigo_dpto': _clean(col(11), 2),
            'codigo_municipio': _clean(col(12), 3),
            'codigo_pais': _clean(col(13), 3) or '169',
            'email': _clean(col(14), 100),
            'actividad_ciiu': _clean(col(15), 4),
            'regla_clasificacion': _clean(col(16)) if tiene_nit_original else '',
            'sugerencias': _clean(col(17)) if tiene_nit_original else '',
        }
        
        # Aplicar clasificador si se solicita (recalcula tipo, detecta DV pegado, etc.)
        if aplicar_clasificador:
            reclasificar_tercero(t)
        
        terceros.append(t)

    wb.close()
    return terceros


def cargar_terceros_desde_excel(
    archivo_xlsx: str | Path,
    empresa_id: int,
    supabase_client,
    sobreescribir: bool = True,
) -> ResultadoCarga:
    """
    Carga terceros revisados a la tabla exogena_terceros.
    
    Args:
        archivo_xlsx: ruta al Excel revisable (con la hoja 'Terceros')
        empresa_id: ID de la empresa en la tabla empresas
        supabase_client: cliente de Supabase ya inicializado
        sobreescribir: si True, hace upsert; si False, salta los existentes
    """
    res = ResultadoCarga()
    try:
        terceros = parsear_excel_terceros(archivo_xlsx)
    except Exception as e:
        res.errores.append(f"Error leyendo Excel: {e}")
        return res

    res.total_filas = len(terceros)

    for t in terceros:
        t['empresa_id'] = empresa_id
        try:
            if sobreescribir:
                # Upsert por (empresa_id, nit)
                resp = supabase_client.table('exogena_terceros').upsert(
                    t, on_conflict='empresa_id,nit'
                ).execute()
                if resp.data:
                    res.insertados += 1
            else:
                # Insertar sin sobreescribir
                resp = supabase_client.table('exogena_terceros').insert(t).execute()
                if resp.data:
                    res.insertados += 1
        except Exception as e:
            msg = str(e)
            if 'duplicate' in msg.lower() or 'unique' in msg.lower():
                res.descartados += 1
            else:
                res.errores.append(f"NIT {t['nit']}: {e}")

    return res


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print("Uso: python cargador_terceros.py <archivo.xlsx>")
        sys.exit(1)
    
    archivo = sys.argv[1]
    print(f"Parseando {archivo}...")
    terceros = parsear_excel_terceros(archivo)
    print(f"\n{len(terceros)} terceros parseados correctamente")
    
    naturales = sum(1 for t in terceros if t['tipo_persona'] == 'natural')
    juridicas = sum(1 for t in terceros if t['tipo_persona'] == 'juridica')
    print(f"  Naturales: {naturales}")
    print(f"  Jurídicas: {juridicas}")
    
    print("\nEjemplos:")
    for t in terceros[:3]:
        nombre = t['razon_social'] or f"{t['primer_nombre']} {t['primer_apellido']}".strip()
        print(f"  NIT {t['nit']}-{t['dv']} ({t['tipo_persona']}): {nombre}")
