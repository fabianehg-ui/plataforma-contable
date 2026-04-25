"""
Módulo de conocimiento del balance de prueba - versión web.

Adaptado de conocimiento_balance.py del .exe v2.6.

Cambios vs versión escritorio:
    - Recibe file-like / bytes en vez de rutas de archivo
    - Persiste en Supabase Storage (bucket empresas-config)
    - No usa caché global; cada empresa carga el suyo cuando lo necesita

Lo que hace:
    Lee un balance de prueba (Excel exportado del sistema contable) y
    extrae información histórica POR NIT que sirve al procesador de
    Compras DIAN para tomar mejores decisiones contables:

    1. NITs con cuenta de gasto histórica (grupo 5 y 6, débitos):
       → Sugiere cuenta de gasto al encontrar un proveedor "conocido".

    2. NITs con retefuente histórica (grupo 2365, créditos):
       → Si un NIT NUNCA se le ha retenido en el balance, NO se le practica
         retefuente aunque supere los 10 UVT (regla de Casa UnoTres).
         Esta es la "regla histórica del balance" que el procesador
         consulta cuando RESPETAR_HISTORICO_BALANCE = SI.

    3. NITs con IVA descontable recurrente (24080201 / 24080308):
       → Para determinar la cuenta de IVA por proveedor.

    El resultado se guarda como JSON en Supabase Storage:
        bucket "empresas-config"
        path   "{empresa_id}/conocimiento_balance.json"
"""
from __future__ import annotations
import io
import json
import re
from collections import defaultdict
from datetime import datetime
from typing import Optional, Tuple

from openpyxl import load_workbook

from core.utils.nits import normalizar_nit


# ============================================================
# Constantes (idénticas al .exe v2.6)
# ============================================================

CUENTAS_IVA_DESCONTABLE = ("24080201", "24080308")
PREFIJO_RETEFUENTE = "2365"
PREFIJOS_GASTO = ("5", "6")

# NITs genéricos del sistema (no se incluyen en el conocimiento)
NITS_GENERICOS = {"222222222", "22222222222", "999", "99"}

# Bucket en Supabase Storage donde se guarda todo lo de configuración
BUCKET_CONFIG = "empresas-config"
NOMBRE_ARCHIVO_CONOCIMIENTO = "conocimiento_balance.json"


# ============================================================
# Helpers privados
# ============================================================

def _es_nit_generico(nit_normalizado: str) -> bool:
    """True si el NIT es uno de los genéricos del sistema."""
    if not nit_normalizado:
        return True
    if nit_normalizado in NITS_GENERICOS:
        return True
    # Cadenas como '2222...' o '9999...' (NITs inválidos / cierre contable)
    if re.fullmatch(r"2+", nit_normalizado):
        return True
    if re.fullmatch(r"9+", nit_normalizado):
        return True
    return False


# ============================================================
# Procesamiento del balance
# ============================================================

def procesar_balance(archivo) -> dict:
    """Lee el Excel del balance y extrae información por NIT.

    Args:
        archivo: file-like (BytesIO, UploadedFile de Streamlit) o bytes.

    Returns:
        dict con la estructura:
        {
            "fecha_balance": "...",
            "empresa": "...",
            "fecha_procesamiento": "YYYY-MM-DDTHH:MM:SS",
            "nits": {
                "900473959": {
                    "nombre": "PROVEEDOR XYZ",
                    "cuentas_gasto": ["620505", "519525"],
                    "cuenta_iva_recurrente": "24080201",
                    "tiene_retefuente_historica": True,
                    "cuenta_retefuente_usada": "23654005"
                },
                ...
            },
            "estadisticas": {
                "total_nits": 42,
                "con_retefuente_historica": 18,
                "con_iva_mercancia": 23,
                "con_iva_servicios": 9
            }
        }
    """
    # Normalizar a bytes para load_workbook
    if hasattr(archivo, "read"):
        contenido = archivo.read()
        if hasattr(archivo, "seek"):
            archivo.seek(0)
        bio = io.BytesIO(contenido)
    elif isinstance(archivo, (bytes, bytearray)):
        bio = io.BytesIO(bytes(archivo))
    else:
        raise TypeError(f"Tipo de archivo no soportado: {type(archivo)}")

    wb = load_workbook(bio, data_only=True, read_only=True)
    ws = wb.active

    # Las primeras 2 filas suelen tener nombre de empresa y fecha del balance
    empresa = ""
    fecha_balance = ""
    for i, fila in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 and fila and fila[0]:
            empresa = str(fila[0]).strip()
        if i == 1 and fila and fila[0]:
            fecha_balance = str(fila[0]).strip()
        if i >= 2:
            break

    # Acumuladores por NIT
    nit_cuentas_gasto = defaultdict(lambda: defaultdict(float))
    nit_iva = defaultdict(lambda: defaultdict(float))
    nit_retefuente = defaultdict(lambda: defaultdict(float))
    nit_nombres = {}

    # Leer desde la fila 4 (después de encabezado + título de columnas)
    for fila in ws.iter_rows(min_row=4, values_only=True):
        if not fila or len(fila) < 8:
            continue
        cuenta, equiv, nombre_cta, nit, nombre_nit, saldo_ant, debitos, creditos = fila[:8]

        if not cuenta or not nit:
            continue

        c = str(cuenta).strip()
        nit_norm = normalizar_nit(nit)

        if not nit_norm or _es_nit_generico(nit_norm):
            continue

        try:
            deb = float(debitos or 0)
            cre = float(creditos or 0)
        except (ValueError, TypeError):
            continue

        if nombre_nit and nit_norm not in nit_nombres:
            nit_nombres[nit_norm] = str(nombre_nit).strip()

        # Cuentas de gasto / compra (grupo 5 o 6, débitos)
        if c.startswith(PREFIJOS_GASTO) and deb > 0:
            nit_cuentas_gasto[nit_norm][c] += deb

        # IVA descontable (débitos)
        if c in CUENTAS_IVA_DESCONTABLE and deb > 0:
            nit_iva[nit_norm][c] += deb

        # Retefuente histórica (créditos en cuentas 2365)
        if c.startswith(PREFIJO_RETEFUENTE) and cre > 0:
            nit_retefuente[nit_norm][c] += cre

    # Consolidar resultados por NIT
    todos_nits = (
        set(nit_cuentas_gasto.keys())
        | set(nit_iva.keys())
        | set(nit_retefuente.keys())
    )

    resultado_nits = {}
    for nit in sorted(todos_nits):
        cuentas_gasto = list(nit_cuentas_gasto.get(nit, {}).keys())
        # Cuenta más usada primero
        cuentas_gasto.sort(key=lambda c: -nit_cuentas_gasto[nit][c])

        ivas = nit_iva.get(nit, {})
        cuenta_iva_recurrente = ""
        if ivas:
            cuenta_iva_recurrente = max(ivas.items(), key=lambda x: x[1])[0]

        retfs = nit_retefuente.get(nit, {})
        tiene_retefuente = bool(retfs)
        cuenta_ret_usada = ""
        if retfs:
            cuenta_ret_usada = max(retfs.items(), key=lambda x: x[1])[0]

        resultado_nits[nit] = {
            "nombre": nit_nombres.get(nit, ""),
            "cuentas_gasto": cuentas_gasto,
            "cuenta_iva_recurrente": cuenta_iva_recurrente,
            "tiene_retefuente_historica": tiene_retefuente,
            "cuenta_retefuente_usada": cuenta_ret_usada,
        }

    # Estadísticas
    total_nits = len(resultado_nits)
    con_retefuente = sum(
        1 for v in resultado_nits.values() if v["tiene_retefuente_historica"]
    )
    con_iva_merc = sum(
        1 for v in resultado_nits.values() if v["cuenta_iva_recurrente"] == "24080201"
    )
    con_iva_serv = sum(
        1 for v in resultado_nits.values() if v["cuenta_iva_recurrente"] == "24080308"
    )

    return {
        "fecha_balance": fecha_balance,
        "empresa": empresa,
        "fecha_procesamiento": datetime.now().isoformat(timespec="seconds"),
        "nits": resultado_nits,
        "estadisticas": {
            "total_nits": total_nits,
            "con_retefuente_historica": con_retefuente,
            "con_iva_mercancia": con_iva_merc,
            "con_iva_servicios": con_iva_serv,
        },
    }


# ============================================================
# Persistencia en Supabase Storage
# ============================================================

def guardar_conocimiento_en_storage(empresa_id: str, conocimiento: dict) -> bool:
    """Sube el JSON de conocimiento al bucket Supabase Storage de la empresa.

    Args:
        empresa_id: UUID de la empresa (de la tabla 'empresas').
        conocimiento: dict producido por procesar_balance().

    Returns:
        True si se subió correctamente, False si falló.

    Raises:
        Exception: si el bucket no existe o las credenciales fallan.
    """
    from db.supabase_client import get_supabase_admin, get_supabase

    sb = get_supabase_admin() or get_supabase()

    contenido = json.dumps(conocimiento, indent=2, ensure_ascii=False).encode("utf-8")
    ruta = f"{empresa_id}/{NOMBRE_ARCHIVO_CONOCIMIENTO}"

    try:
        # upsert para sobrescribir si ya existe
        sb.storage.from_(BUCKET_CONFIG).upload(
            path=ruta,
            file=contenido,
            file_options={
                "content-type": "application/json",
                "upsert": "true",
            },
        )
        return True
    except Exception:
        # Algunos clientes no soportan upsert en upload; intentar update
        try:
            sb.storage.from_(BUCKET_CONFIG).update(
                path=ruta,
                file=contenido,
                file_options={"content-type": "application/json"},
            )
            return True
        except Exception:
            raise


def cargar_conocimiento_desde_storage(empresa_id: str) -> Optional[dict]:
    """Descarga el conocimiento_balance.json de la empresa desde Storage.

    Args:
        empresa_id: UUID de la empresa.

    Returns:
        dict con el conocimiento, o None si la empresa todavía no ha
        cargado un balance.
    """
    from db.supabase_client import get_supabase

    sb = get_supabase()
    ruta = f"{empresa_id}/{NOMBRE_ARCHIVO_CONOCIMIENTO}"

    try:
        resp = sb.storage.from_(BUCKET_CONFIG).download(ruta)
    except Exception:
        return None

    if not resp:
        return None

    try:
        # resp puede ser bytes o tener .content; manejar ambos
        contenido = resp if isinstance(resp, (bytes, bytearray)) else getattr(resp, "content", None)
        if contenido is None:
            return None
        return json.loads(contenido.decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError, AttributeError):
        return None


def eliminar_conocimiento_de_storage(empresa_id: str) -> bool:
    """Elimina el archivo del Storage. Útil para "empezar de cero"."""
    from db.supabase_client import get_supabase_admin, get_supabase

    sb = get_supabase_admin() or get_supabase()
    ruta = f"{empresa_id}/{NOMBRE_ARCHIVO_CONOCIMIENTO}"
    try:
        sb.storage.from_(BUCKET_CONFIG).remove([ruta])
        return True
    except Exception:
        return False


# ============================================================
# API de consulta (sin estado global; recibe el dict explícito)
# ============================================================

def info_nit(conocimiento: Optional[dict], nit) -> Optional[dict]:
    """Retorna el dict de info histórica de un NIT, o None."""
    if not conocimiento:
        return None
    nit_norm = normalizar_nit(nit)
    if not nit_norm:
        return None
    return conocimiento.get("nits", {}).get(nit_norm)


def tiene_retefuente_historica(conocimiento: Optional[dict], nit) -> bool:
    """True si el NIT tiene movimiento histórico en cuentas 2365."""
    info = info_nit(conocimiento, nit)
    return bool(info and info.get("tiene_retefuente_historica"))


def cuenta_iva_sugerida(conocimiento: Optional[dict], nit) -> str:
    """Cuenta de IVA descontable más usada con este NIT en el balance."""
    info = info_nit(conocimiento, nit)
    return info["cuenta_iva_recurrente"] if info else ""


def cuenta_gasto_sugerida(conocimiento: Optional[dict], nit) -> str:
    """Cuenta de gasto/compra más usada con este NIT."""
    info = info_nit(conocimiento, nit)
    if not info or not info.get("cuentas_gasto"):
        return ""
    return info["cuentas_gasto"][0]


def es_nit_conocido(conocimiento: Optional[dict], nit) -> bool:
    """True si el NIT aparece en el balance con alguna cuenta."""
    return info_nit(conocimiento, nit) is not None
