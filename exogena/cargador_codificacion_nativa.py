"""
Cargador del archivo de Codificación de Formatos del software contable.

El archivo nativo tiene esta estructura:
    Código Formato | Nombre Formato | Cuenta Inicial | Cuenta Final | Concepto | Tipo Contrato | Valor | Registros

Donde:
    - Código Formato: '001001', '001003', etc → se normaliza a '1001', '1003'
    - Cuenta Inicial/Final: rango de cuentas con padding inconsistente (espacios,
      ceros). Se compara con padding a 10 dígitos (0s al final del inicial,
      9s al final del final).
    - Concepto: viene como '5007  Compra de activos movibles' → se separa
      el código (5007) de la descripción.

Una misma cuenta puede ser referenciada por múltiples reglas (ambigüedad
intencional para que el contador decida según el NIT/contexto).
"""

from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
import re

import openpyxl


@dataclass
class ReglaMapeoNativo:
    formato_dian: str
    concepto_dian: int
    cuenta_inicial: str
    cuenta_final: str
    descripcion_concepto: str
    tipo_contrato: str = ''
    valor_aplicable: int = 1
    fila_origen: int = 0


@dataclass
class ResultadoCargaMapeo:
    total_filas: int = 0
    reglas_validas: int = 0
    reglas_descartadas: int = 0
    formatos_detectados: dict = field(default_factory=dict)
    errores: list = field(default_factory=list)
    reglas: list = field(default_factory=list)


# Regex para extraer "5007  Descripción..." → ('5007', 'Descripción...')
RE_CONCEPTO = re.compile(r'^\s*(\d+)\s+(.*?)\s*$')


def _normalizar_codigo_formato(cod_raw: str) -> str:
    """'001001' → '1001'."""
    cod = str(cod_raw).strip().lstrip('0')
    return cod if cod else '0'


def _limpiar_cuenta(c) -> str:
    """Quita espacios y ceros a la izquierda de un código de cuenta.
    Acepta tanto strings como números."""
    if c is None:
        return ''
    s = str(c).strip()
    return s


def cargar_codificacion_nativa(archivo) -> ResultadoCargaMapeo:
    """Lee el Excel de codificación nativa y devuelve la lista de reglas.

    Acepta:
      - str / Path: ruta del archivo en disco
      - UploadedFile / BytesIO / file-like: objeto de archivo en memoria
        (típico cuando viene de Streamlit st.file_uploader)

    El parser asume:
      - Hoja 'Datos' (o primera hoja)
      - Filas 1-3 son encabezados
      - Fila con '#N/A' marca el final
    """
    res = ResultadoCargaMapeo()

    # Si es ruta (str/Path), validar que existe. Si es file-like, dejarlo pasar.
    if isinstance(archivo, (str, Path)):
        archivo = Path(archivo)
        if not archivo.exists():
            res.errores.append(f"Archivo no existe: {archivo}")
            return res

    wb = openpyxl.load_workbook(archivo, data_only=True, read_only=True)
    ws = wb['Datos'] if 'Datos' in wb.sheetnames else wb.active

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= 3:
            continue
        if not row[0]:
            continue
        cod_raw = str(row[0]).strip()
        if cod_raw in ('#N/A', ''):
            continue

        res.total_filas += 1

        formato = _normalizar_codigo_formato(cod_raw)
        nombre_formato = str(row[1]).strip() if row[1] else ''
        cuenta_ini = _limpiar_cuenta(row[2])
        cuenta_fin = _limpiar_cuenta(row[3])
        concepto_raw = str(row[4]).strip() if row[4] else ''
        tipo_contrato = str(row[5]).strip() if row[5] else ''

        # Parsear concepto: "5007  Compra de activos movibles" → (5007, "Compra...")
        m = RE_CONCEPTO.match(concepto_raw)
        if not m:
            res.reglas_descartadas += 1
            res.errores.append(
                f"Fila {i}: concepto sin formato esperado: {concepto_raw!r}"
            )
            continue

        try:
            concepto = int(m.group(1))
        except ValueError:
            res.reglas_descartadas += 1
            continue

        descripcion = m.group(2).strip()

        if not cuenta_ini or not cuenta_fin:
            res.reglas_descartadas += 1
            res.errores.append(f"Fila {i}: rango de cuentas vacío")
            continue

        try:
            valor = int(row[6]) if row[6] is not None else 1
        except (ValueError, TypeError):
            valor = 1

        regla = ReglaMapeoNativo(
            formato_dian=formato,
            concepto_dian=concepto,
            cuenta_inicial=cuenta_ini,
            cuenta_final=cuenta_fin,
            descripcion_concepto=descripcion,
            tipo_contrato=tipo_contrato,
            valor_aplicable=valor,
            fila_origen=i,
        )
        res.reglas.append(regla)
        res.reglas_validas += 1
        res.formatos_detectados[formato] = res.formatos_detectados.get(formato, 0) + 1

    wb.close()
    return res


def cuenta_en_rango(cuenta: str, inicial: str, final: str) -> bool:
    """Determina si una cuenta cae dentro del rango [inicial, final].

    Lógica: padding a 10 caracteres con '0' en el inicial (mínimo) y '9'
    en el final (máximo), comparación lexicográfica como strings.
    Esto maneja correctamente rangos donde inicial y final tienen distinta
    longitud (ej. '18' → '18600511').
    """
    c = str(cuenta).strip().ljust(10, '0')
    ini = str(inicial).strip().ljust(10, '0')
    fin = str(final).strip().ljust(10, '9')
    return ini <= c <= fin


def cargar_a_supabase(
    archivo_xlsx,
    empresa_id,
    año_gravable: int,
    supabase_client,
    reemplazar_existente: bool = True,
) -> ResultadoCargaMapeo:
    """Carga el archivo nativo a la tabla exogena_mapeo_empresa de Supabase.

    archivo_xlsx puede ser ruta (str/Path) o file-like (UploadedFile de Streamlit).
    empresa_id puede ser int o str (UUID).
    """
    res = cargar_codificacion_nativa(archivo_xlsx)
    if not res.reglas:
        return res

    if reemplazar_existente:
        supabase_client.table('exogena_mapeo_empresa').delete().eq(
            'empresa_id', empresa_id
        ).eq('año_gravable', año_gravable).execute()

    registros = [
        {
            'empresa_id': empresa_id,
            'año_gravable': año_gravable,
            'formato_dian': r.formato_dian,
            'concepto_dian': r.concepto_dian,
            'cuenta_inicial': r.cuenta_inicial,
            'cuenta_final': r.cuenta_final,
            'descripcion_concepto': r.descripcion_concepto,
            'tipo_contrato': r.tipo_contrato,
            'valor_aplicable': r.valor_aplicable,
            'fila_origen': r.fila_origen,
        }
        for r in res.reglas
    ]

    # Insertar en lotes de 100
    LOTE = 100
    for i in range(0, len(registros), LOTE):
        chunk = registros[i:i + LOTE]
        supabase_client.table('exogena_mapeo_empresa').insert(chunk).execute()

    return res


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print("Uso: python cargador_codificacion_nativa.py <archivo.xlsx>")
        sys.exit(1)

    archivo = sys.argv[1]
    print(f"Parseando {archivo}...\n")
    res = cargar_codificacion_nativa(archivo)

    print(f"Total filas analizadas: {res.total_filas}")
    print(f"Reglas válidas:          {res.reglas_validas}")
    print(f"Reglas descartadas:      {res.reglas_descartadas}")
    print(f"\nFormatos detectados:")
    for fmt, n in sorted(res.formatos_detectados.items()):
        print(f"  {fmt}: {n} reglas")

    if res.errores:
        print(f"\nErrores ({len(res.errores)}):")
        for e in res.errores[:5]:
            print(f"  - {e}")

    print(f"\nMuestra de reglas:")
    for r in res.reglas[:5]:
        print(
            f"  fmt {r.formato_dian} | cpt {r.concepto_dian} | "
            f"[{r.cuenta_inicial} → {r.cuenta_final}] | {r.descripcion_concepto[:50]}"
        )
