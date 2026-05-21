"""
Generador de Excel editable para conciliación tributaria.

Crea un archivo Excel descargable con:
    - Hoja "Resumen": dictamen por tipo de conciliación
    - Hoja "Seguridad Social": cuentas + saldos + columnas para empleador/trabajador
    - Hoja "GMF": cuentas + total + cálculo automático del 50%
    - Hoja "Otras": cuentas adicionales para conciliación manual
    - Hoja "Movimientos detallados": detalle por NIT (read-only)

El usuario puede:
    - Editar los valores de empleador/trabajador
    - Agregar notas y soportes
    - Re-cargar el Excel modificado al sistema
"""

from __future__ import annotations
from datetime import datetime
from io import BytesIO
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Estilos
HEADER_FILL = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
SUBHEADER_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
EDITABLE_FILL = PatternFill(start_color="FFF9E5", end_color="FFF9E5", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="ABEBC6", end_color="ABEBC6", fill_type="solid")
WARNING_FILL = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")

THIN = Side(border_style="thin", color="999999")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def _aplicar_header(ws, row, cols_data, freeze=True):
    """Aplica formato a la fila de encabezados."""
    for col, value in enumerate(cols_data, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    if freeze:
        ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _set_col_widths(ws, widths):
    """Establece anchos de columnas."""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generar_excel_conciliacion(
    dictamen,
    cabecera_empresa: dict,
    año_gravable: int,
) -> BytesIO:
    """Genera el Excel editable de conciliación.
    
    Args:
        dictamen: DictamenConciliacion con cuentas detectadas
        cabecera_empresa: dict con keys 'razon_social', 'nit'
        año_gravable: int
    
    Returns:
        BytesIO con el archivo Excel listo para descargar
    """
    wb = openpyxl.Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # ============================================================
    # HOJA 1: Resumen
    # ============================================================
    ws = wb.create_sheet("Resumen")
    
    ws['A1'] = "CONCILIACIÓN TRIBUTARIA - INFORMACIÓN EXÓGENA"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    ws['A3'] = "Empresa:"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = cabecera_empresa.get('razon_social', '')
    
    ws['A4'] = "NIT:"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = cabecera_empresa.get('nit', '')
    
    ws['A5'] = "Año gravable:"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = año_gravable
    
    ws['A6'] = "Generado:"
    ws['A6'].font = Font(bold=True)
    ws['B6'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    
    # Tabla resumen
    _aplicar_header(ws, 9, [
        "Tipo de Conciliación", "Cuentas", "Saldo Total", "Observación",
    ], freeze=False)
    
    fila = 10
    ws.cell(row=fila, column=1, value="Seguridad Social (5010, 5011, 5012)")
    ws.cell(row=fila, column=2, value=len(dictamen.cuentas_seguridad_social))
    ws.cell(row=fila, column=3, value=float(dictamen.total_seguridad_social))
    ws.cell(row=fila, column=4, value="Separar aporte empleador (deducible) vs trabajador (no deducible)")
    fila += 1
    
    ws.cell(row=fila, column=1, value="GMF (5045)")
    ws.cell(row=fila, column=2, value=len(dictamen.cuentas_gmf))
    ws.cell(row=fila, column=3, value=float(dictamen.total_gmf))
    ws.cell(row=fila, column=4, value="Solo 50% es deducible (ET art. 115)")
    fila += 1
    
    ws.cell(row=fila, column=1, value="Otras (manual)")
    ws.cell(row=fila, column=2, value=len(dictamen.cuentas_otras))
    ws.cell(row=fila, column=3, value=float(sum(c.saldo_total for c in dictamen.cuentas_otras)))
    ws.cell(row=fila, column=4, value="Definir manualmente")
    fila += 1
    
    # Total
    ws.cell(row=fila, column=1, value="TOTAL GENERAL").font = Font(bold=True)
    ws.cell(row=fila, column=3, value=float(dictamen.total_general)).font = Font(bold=True)
    for col in range(1, 5):
        ws.cell(row=fila, column=col).fill = TOTAL_FILL
    
    # Formatear columna de valores como moneda
    for r in range(10, fila + 1):
        ws.cell(row=r, column=3).number_format = '"$"#,##0.00'
    
    _set_col_widths(ws, [40, 12, 20, 60])
    
    # Instrucciones
    fila += 3
    ws.cell(row=fila, column=1, value="📋 INSTRUCCIONES").font = Font(bold=True, size=12)
    fila += 1
    instrucciones = [
        "1. Ve a la hoja 'Seguridad Social' y completa los valores de aporte EMPLEADOR vs TRABAJADOR.",
        "2. Ve a la hoja 'GMF' y verifica/edita el total certificado por banco.",
        "3. Ve a la hoja 'Otras' para agregar conciliaciones manuales.",
        "4. Las celdas amarillas son editables.",
        "5. Las celdas verdes son cálculos automáticos.",
        "6. Una vez completado, sube este Excel de vuelta al sistema en el tab Conciliación.",
    ]
    for inst in instrucciones:
        ws.cell(row=fila, column=1, value=inst)
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
        fila += 1
    
    # ============================================================
    # HOJA 2: Seguridad Social
    # ============================================================
    ws = wb.create_sheet("Seguridad Social")
    
    ws['A1'] = "CONCILIACIÓN: APORTES SEGURIDAD SOCIAL (PILA)"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    ws['A2'] = ("Para cada cuenta, indica cuánto del saldo corresponde al APORTE EMPLEADOR "
                "(deducible) vs APORTE TRABAJADOR (no deducible). El aporte trabajador NO se reporta.")
    ws.merge_cells('A2:H2')
    ws['A2'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    headers = [
        "Cuenta", "Nombre Cuenta", "Concepto", "Tipo",
        "Saldo Balance", "Aporte Empleador", "Aporte Trabajador", "Verificación",
    ]
    _aplicar_header(ws, 4, headers)
    
    fila = 5
    for c in dictamen.cuentas_seguridad_social:
        ws.cell(row=fila, column=1, value=c.codigo_cuenta)
        ws.cell(row=fila, column=2, value=c.nombre_cuenta)
        ws.cell(row=fila, column=3, value=c.concepto_sugerido)
        ws.cell(row=fila, column=4, value=c.descripcion_concepto)
        ws.cell(row=fila, column=5, value=float(c.saldo_total)).number_format = '"$"#,##0.00'
        
        # Celdas editables (amarillas) - empleador y trabajador
        cell_emp = ws.cell(row=fila, column=6, value=0)
        cell_emp.fill = EDITABLE_FILL
        cell_emp.number_format = '"$"#,##0.00'
        
        cell_trab = ws.cell(row=fila, column=7, value=0)
        cell_trab.fill = EDITABLE_FILL
        cell_trab.number_format = '"$"#,##0.00'
        
        # Si es ARL/CCF/SENA/ICBF, sugerir 100% empleador
        if c.concepto_sugerido in (5012, 5013, 5014, 5015):
            cell_emp.value = float(c.saldo_total)
            cell_trab.value = 0
        
        # Verificación: empleador + trabajador debe igualar saldo
        cell_check = ws.cell(row=fila, column=8,
                              value=f'=IF(F{fila}+G{fila}=E{fila},"✓ OK","⚠ Diferencia: "&TEXT(E{fila}-F{fila}-G{fila},"$#,##0"))')
        cell_check.fill = TOTAL_FILL
        
        fila += 1
    
    # Totales
    if dictamen.cuentas_seguridad_social:
        fila_total = fila
        ws.cell(row=fila_total, column=2, value="TOTAL").font = Font(bold=True)
        ws.cell(row=fila_total, column=5,
                value=f'=SUM(E5:E{fila-1})').number_format = '"$"#,##0.00'
        ws.cell(row=fila_total, column=6,
                value=f'=SUM(F5:F{fila-1})').number_format = '"$"#,##0.00'
        ws.cell(row=fila_total, column=7,
                value=f'=SUM(G5:G{fila-1})').number_format = '"$"#,##0.00'
        for col in range(1, 9):
            ws.cell(row=fila_total, column=col).fill = TOTAL_FILL
            ws.cell(row=fila_total, column=col).font = Font(bold=True)
    
    _set_col_widths(ws, [12, 30, 10, 18, 18, 18, 18, 25])
    
    # ============================================================
    # HOJA 3: GMF
    # ============================================================
    ws = wb.create_sheet("GMF")
    
    ws['A1'] = "CONCILIACIÓN: GMF (Gravamen a los Movimientos Financieros)"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    
    ws['A2'] = ("Verifica el total certificado por cada banco. El sistema calcula automáticamente "
                "el 50% deducible (ET art. 115).")
    ws.merge_cells('A2:G2')
    ws['A2'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    headers = [
        "Cuenta", "Nombre Cuenta", "Saldo Balance",
        "Banco / NIT", "Total Certificado", "Deducible 50%", "No Deducible 50%",
    ]
    _aplicar_header(ws, 4, headers)
    
    fila = 5
    for c in dictamen.cuentas_gmf:
        ws.cell(row=fila, column=1, value=c.codigo_cuenta)
        ws.cell(row=fila, column=2, value=c.nombre_cuenta)
        ws.cell(row=fila, column=3, value=float(c.saldo_total)).number_format = '"$"#,##0.00'
        
        # Sugerir el primer NIT como banco (editable)
        primer_nit = c.nits_principales[0] if c.nits_principales else ''
        cell_banco = ws.cell(row=fila, column=4, value=primer_nit)
        cell_banco.fill = EDITABLE_FILL
        
        # Total certificado (editable, default = saldo balance)
        cell_total = ws.cell(row=fila, column=5, value=float(c.saldo_total))
        cell_total.fill = EDITABLE_FILL
        cell_total.number_format = '"$"#,##0.00'
        
        # 50% deducible (auto)
        ws.cell(row=fila, column=6, value=f'=E{fila}*0.5').number_format = '"$"#,##0.00'
        ws.cell(row=fila, column=6).fill = TOTAL_FILL
        
        # 50% no deducible (auto)
        ws.cell(row=fila, column=7, value=f'=E{fila}*0.5').number_format = '"$"#,##0.00'
        ws.cell(row=fila, column=7).fill = TOTAL_FILL
        
        fila += 1
    
    if dictamen.cuentas_gmf:
        fila_total = fila
        ws.cell(row=fila_total, column=2, value="TOTAL").font = Font(bold=True)
        for col_idx in [3, 5, 6, 7]:
            col_letter = get_column_letter(col_idx)
            ws.cell(row=fila_total, column=col_idx,
                    value=f'=SUM({col_letter}5:{col_letter}{fila-1})').number_format = '"$"#,##0.00'
        for col in range(1, 8):
            ws.cell(row=fila_total, column=col).fill = TOTAL_FILL
            ws.cell(row=fila_total, column=col).font = Font(bold=True)
    
    _set_col_widths(ws, [12, 30, 18, 22, 18, 18, 18])
    
    # ============================================================
    # HOJA 4: Otras
    # ============================================================
    ws = wb.create_sheet("Otras")
    
    ws['A1'] = "CONCILIACIONES ADICIONALES"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    
    ws['A2'] = ("Agrega aquí cuentas que requieran ajuste manual: gastos no deducibles, "
                "amortizaciones, valores no reportables, etc.")
    ws.merge_cells('A2:G2')
    ws.row_dimensions[2].height = 30
    
    headers = [
        "Cuenta", "Nombre Cuenta", "Saldo Balance", "Valor Deducible", 
        "Valor No Deducible", "Motivo del Ajuste", "Soporte/Notas",
    ]
    _aplicar_header(ws, 4, headers)
    
    fila = 5
    for c in dictamen.cuentas_otras:
        ws.cell(row=fila, column=1, value=c.codigo_cuenta)
        ws.cell(row=fila, column=2, value=c.nombre_cuenta)
        ws.cell(row=fila, column=3, value=float(c.saldo_total)).number_format = '"$"#,##0.00'
        # Editables
        for col in range(4, 8):
            ws.cell(row=fila, column=col).fill = EDITABLE_FILL
        ws.cell(row=fila, column=4).number_format = '"$"#,##0.00'
        ws.cell(row=fila, column=5).number_format = '"$"#,##0.00'
        fila += 1
    
    # Agregar 10 filas vacías editables para nuevas conciliaciones
    for _ in range(10):
        for col in range(1, 8):
            ws.cell(row=fila, column=col).fill = EDITABLE_FILL
        ws.cell(row=fila, column=3).number_format = '"$"#,##0.00'
        ws.cell(row=fila, column=4).number_format = '"$"#,##0.00'
        ws.cell(row=fila, column=5).number_format = '"$"#,##0.00'
        fila += 1
    
    _set_col_widths(ws, [12, 30, 18, 18, 18, 30, 30])
    
    # ============================================================
    # Guardar a BytesIO
    # ============================================================
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


if __name__ == '__main__':
    # Test
    import sys
    sys.path.insert(0, '.')
    from conciliacion import construir_dictamen
    
    movs = [
        {'codigo_cuenta': '510303', 'nombre_cuenta': 'SALUD EPS',
         'nit': '901589040', 'debitos': 4000000, 'creditos': 0,
         'saldo_final': 4000000, 'formato_dian': '1001', 'concepto_dian': 5010},
        {'codigo_cuenta': '510306', 'nombre_cuenta': 'PENSION',
         'nit': '901589040', 'debitos': 6000000, 'creditos': 0,
         'saldo_final': 6000000, 'formato_dian': '1001', 'concepto_dian': 5011},
        {'codigo_cuenta': '530525', 'nombre_cuenta': 'GMF BANCOLOMBIA',
         'nit': '890903938', 'debitos': 1500000, 'creditos': 0,
         'saldo_final': 1500000, 'formato_dian': '1001', 'concepto_dian': 5045},
    ]
    
    d = construir_dictamen(movs)
    excel = generar_excel_conciliacion(d, {'razon_social': 'TEST SAS', 'nit': '900000000'}, 2025)
    
    with open('/tmp/test_conciliacion.xlsx', 'wb') as f:
        f.write(excel.getvalue())
    print(f"Excel generado: /tmp/test_conciliacion.xlsx ({len(excel.getvalue()):,} bytes)")
