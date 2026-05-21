"""
Cargador de terceros desde el archivo nativo del software contable o revisable.

Detecta automáticamente el formato del archivo y aplica el parser correcto.

Formato nativo del software contable (Rutas del Mar):
    [0] NIT  [1] Tipo  [2] Nombre  [3] Direccion  [4] Ciudad  [5] Telefono
    [6] Municipio (DANE 5 dig)  [7] Activo  [8] Tiene RUT  [9] Pais
    [10] Primer Nombre  [11] Segundo Nombre  [12] Primer Apellido
    [13] Segundo Apellido  [14] Email  [15] Celular  [16] Plazo
    [17] Actividad CIIU  [18] Indicativo  [19] Naturaleza

Formato revisable (Excel limpio generado por este sistema):
    [0] NIT  [1] NIT Original  [2] DV  [3] Tipo Doc  [4] Tipo Persona
    [5] Razón Social  [6] Primer Apellido ... etc

El parser:
    - Limpia espacios y caracteres extra
    - Aplica clasificador_nits (rangos oficiales DIAN)
    - Detecta NITs con DV pegado y los corrige automáticamente
    - Marca filas dudosas para revisión manual
"""

from __future__ import annotations
from pathlib import Path
import re

import openpyxl

try:
    from .clasificador_nits import reclasificar_tercero, calc_dv
except ImportError:
    from clasificador_nits import reclasificar_tercero, calc_dv


def _clean(s, max_len=None):
    """Limpia un string: quita espacios extra y trunca."""
    if s is None:
        return ''
    s = re.sub(r'\s+', ' ', str(s).strip())
    return s[:max_len] if max_len else s


def _clean_nit(s):
    """Extrae solo dígitos de un campo NIT (quita puntos, guiones, espacios, DV)."""
    if s is None:
        return None
    s = str(s).strip()
    # Si tiene formato '901.589.040-1', quitar el guión y el dígito final (DV)
    if '-' in s:
        s = s.split('-')[0]
    # Quitar todo lo que no sea dígito
    s = re.sub(r'[^\d]', '', s)
    return s if s else None


def _es_archivo_nativo_software(rows):
    """Detecta si es el formato nativo del software contable."""
    if not rows:
        return False
    primera = rows[0]
    if not primera or len(primera) < 5:
        return False
    headers = [str(c or '').strip().lower() for c in primera[:5]]
    return (headers[0] == 'nit' and headers[2] == 'nombre')


def _es_archivo_revisable(rows):
    """Detecta si es el formato del Excel revisable que generamos."""
    if not rows:
        return False
    primera = rows[0]
    if not primera:
        return False
    headers = [str(c or '').strip().lower() for c in primera[:6]]
    return ('nit original' in headers) or ('tipo doc' in headers)


def parsear_archivo_nativo_software(rows: list, aplicar_clasificador: bool = True) -> list[dict]:
    """Parsea el formato del software contable (con datos sucios)."""
    terceros = []
    nits_vistos = set()

    for row in rows[1:]:
        if not row or not row[0]:
            continue
        nit = _clean_nit(row[0])
        if not nit or len(nit) < 7 or len(nit) > 11:
            continue
        if nit in nits_vistos:
            continue
        nits_vistos.add(nit)

        nombre_raw = _clean(row[2] if len(row) > 2 else '', 200)
        direccion = _clean(row[3] if len(row) > 3 else '', 200)

        # Municipio (código DANE 5 dígitos: 2 dpto + 3 mcp)
        cod_mun = _clean(row[6] if len(row) > 6 else '')
        cod_dpto = ''
        cod_mcp = ''
        if cod_mun and cod_mun.isdigit() and len(cod_mun) == 5:
            cod_dpto = cod_mun[:2]
            cod_mcp = cod_mun[2:]

        # País
        cod_pais = _clean(row[9] if len(row) > 9 else '')
        if cod_pais and cod_pais.isdigit():
            cod_pais = cod_pais.zfill(3)
        else:
            cod_pais = '169'

        primer_nombre = _clean(row[10] if len(row) > 10 else '', 60)
        segundo_nombre = _clean(row[11] if len(row) > 11 else '', 60)
        primer_apellido = _clean(row[12] if len(row) > 12 else '', 60)
        segundo_apellido = _clean(row[13] if len(row) > 13 else '', 60)

        email = _clean(row[14] if len(row) > 14 else '', 100)
        if email and '@' not in email:
            email = ''

        ciiu_raw = _clean(row[17] if len(row) > 17 else '')
        actividad_ciiu = ciiu_raw if ciiu_raw.isdigit() and len(ciiu_raw) == 4 else ''

        t = {
            'nit': nit,
            'dv': calc_dv(nit) if nit.isdigit() else None,
            'tipo_documento': 31,
            'tipo_persona': 'desconocido',
            'razon_social': nombre_raw,
            'primer_apellido': primer_apellido,
            'segundo_apellido': segundo_apellido,
            'primer_nombre': primer_nombre,
            'otros_nombres': segundo_nombre,
            'direccion': direccion,
            'codigo_dpto': cod_dpto,
            'codigo_municipio': cod_mcp,
            'codigo_pais': cod_pais,
            'email': email,
            'actividad_ciiu': actividad_ciiu,
        }

        if aplicar_clasificador:
            reclasificar_tercero(t)

        terceros.append(t)

    return terceros


def parsear_archivo_revisable(rows: list, aplicar_clasificador: bool = True) -> list[dict]:
    """Parsea el formato del Excel revisable (limpio)."""
    if not rows:
        return []

    header = [str(h or '').strip().lower() for h in rows[0]]
    tiene_nit_original = 'nit original' in header

    terceros = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        nit = _clean_nit(row[0])
        if not nit or len(nit) < 7 or len(nit) > 11:
            continue

        offset = 0 if tiene_nit_original else -1

        def col(i):
            idx = i if i < 2 else i + offset
            return row[idx] if idx < len(row) else None

        tipo_persona = _clean(col(4)).lower() or 'natural'

        t = {
            'nit': nit,
            'dv': calc_dv(nit) if nit.isdigit() else None,
            'nit_original': _clean(col(1), 20) if tiene_nit_original else '',
            'tipo_documento': int(col(3)) if col(3) and str(col(3)).strip().isdigit() else (31 if tipo_persona == 'juridica' else 13),
            'tipo_persona': tipo_persona,
            'razon_social': _clean(col(5), 450),
            'primer_apellido': _clean(col(6), 60),
            'segundo_apellido': _clean(col(7), 60),
            'primer_nombre': _clean(col(8), 60),
            'otros_nombres': _clean(col(9), 60),
            'direccion': _clean(col(10), 200),
            'codigo_dpto': _clean(col(11), 2),
            'codigo_municipio': _clean(col(12), 3),
            'codigo_pais': _clean(col(13), 3) or '169',
            'email': _clean(col(14), 100),
            'actividad_ciiu': _clean(col(15), 4),
            'regla_clasificacion': _clean(col(16)) if tiene_nit_original else '',
            'sugerencias': _clean(col(17)) if tiene_nit_original else '',
        }

        if aplicar_clasificador:
            reclasificar_tercero(t)

        terceros.append(t)

    return terceros


def parsear_excel_terceros(archivo, aplicar_clasificador: bool = True) -> list[dict]:
    """Lee un Excel de terceros (nativo o revisable) y devuelve lista de dicts.

    Detecta automáticamente el formato:
        - Nativo del software contable (con datos sucios)
        - Revisable (limpio, generado por este sistema)

    Acepta:
        - str / Path: ruta del archivo
        - UploadedFile / file-like: típico de Streamlit st.file_uploader
    """
    # Si es ruta, validar que existe; si es file-like, dejarlo pasar
    if isinstance(archivo, (str, Path)):
        archivo = Path(archivo)
        if not archivo.exists():
            raise FileNotFoundError(f"Archivo no existe: {archivo}")

    wb = openpyxl.load_workbook(archivo, data_only=True, read_only=True)
    ws = wb['Terceros'] if 'Terceros' in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    if _es_archivo_nativo_software(rows):
        return parsear_archivo_nativo_software(rows, aplicar_clasificador)
    elif _es_archivo_revisable(rows):
        return parsear_archivo_revisable(rows, aplicar_clasificador)
    else:
        # Asumir nativo si tiene NIT en columna 0 sin importar nombre del header
        return parsear_archivo_nativo_software(rows, aplicar_clasificador)


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print("Uso: python cargador_terceros.py <archivo.xlsx>")
        sys.exit(1)

    print(f"Parseando {sys.argv[1]}...")
    terceros = parsear_excel_terceros(sys.argv[1])

    from collections import Counter
    tipos = Counter(t['tipo_persona'] for t in terceros)

    print(f"\n{len(terceros)} terceros parseados")
    for k, v in tipos.items():
        print(f"  {k}: {v}")

    n_revisar = sum(1 for t in terceros if t.get('requiere_revision'))
    print(f"  Requieren revisión: {n_revisar}")

    print("\nEjemplos:")
    for t in terceros[:5]:
        nombre = t.get('razon_social') or f"{t.get('primer_nombre','')} {t.get('primer_apellido','')}"
        mcp = f"{t.get('codigo_dpto','')}-{t.get('codigo_municipio','')}"
        print(f"  NIT {t['nit']}-{t.get('dv','')} ({t['tipo_persona']}) mcp={mcp}: {nombre.strip()[:60]}")
