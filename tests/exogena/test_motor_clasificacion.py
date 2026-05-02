"""Test integral del motor de clasificación con datos reales de Rutas del Mar."""
from pathlib import Path
import sys

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from core.exogena.motor_clasificacion import (
    MotorClasificacion, Movimiento, ReglaCapa1, ReglaCapa2, ReglaCapa3,
    cuenta_en_rango,
)
from core.exogena.cargador_codificacion_nativa import cargar_codificacion_nativa

import openpyxl


# ================================================================
# Tests unitarios de la lógica de rangos
# ================================================================
def test_cuenta_en_rango():
    """La función de rango maneja correctamente longitudes distintas."""
    # Rango raíz que cubre toda una cuenta
    assert cuenta_en_rango('18', '18', '18600511')
    assert cuenta_en_rango('1810', '18', '18600511')
    assert cuenta_en_rango('18600511', '18', '18600511')
    assert not cuenta_en_rango('18600512', '18', '18600511')
    assert not cuenta_en_rango('19', '18', '18600511')
    assert not cuenta_en_rango('17', '18', '18600511')

    # Rango exacto subcuenta
    assert cuenta_en_rango('236540', '236540', '23654099')
    assert cuenta_en_rango('23654099', '236540', '23654099')
    assert not cuenta_en_rango('23654100', '236540', '23654099')

    # Rango cuenta única
    assert cuenta_en_rango('15200504', '15200504', '15200504')
    assert not cuenta_en_rango('15200505', '15200504', '15200504')
    print("  ✓ cuenta_en_rango: 11 casos correctos")


# ================================================================
# Tests del motor
# ================================================================
def test_motor_capa3_tiene_prioridad():
    """Capa 3 (manual) gana sobre capas 1 y 2."""
    motor = MotorClasificacion(
        reglas_capa1=[
            ReglaCapa1(codigo_cuenta='1405', formato_dian='1001', concepto_dian=5007),
        ],
        reglas_capa2=[
            ReglaCapa2(formato_dian='1001', concepto_dian=5008,
                       cuenta_inicial='1405', cuenta_final='14059999'),
        ],
        reglas_capa3=[
            ReglaCapa3(codigo_cuenta='1405', nit='900123456',
                       formato_dian='1001', concepto_dian=5099),
        ],
    )
    mov = Movimiento(codigo_cuenta='1405', nit='900123456', debitos=1000)
    resultado = motor.clasificar_movimiento(mov)
    assert len(resultado) == 1
    assert resultado[0].capa_resolucion == 'mapeo_manual'
    assert resultado[0].concepto_dian == 5099
    print("  ✓ Capa 3 tiene prioridad sobre capas 1 y 2")


def test_motor_capa2_ambiguedad():
    """Capa 2 con múltiples reglas marca requiere_revision."""
    motor = MotorClasificacion(
        reglas_capa1=[],
        reglas_capa2=[
            ReglaCapa2(formato_dian='1001', concepto_dian=5004,
                       cuenta_inicial='236540', cuenta_final='23654099',
                       descripcion_concepto='Servicios'),
            ReglaCapa2(formato_dian='1001', concepto_dian=5007,
                       cuenta_inicial='236540', cuenta_final='23654099',
                       descripcion_concepto='Compras'),
            ReglaCapa2(formato_dian='1001', concepto_dian=5008,
                       cuenta_inicial='236540', cuenta_final='23654099',
                       descripcion_concepto='Activos fijos'),
        ],
    )
    mov = Movimiento(codigo_cuenta='236540', nit='900123', debitos=5000)
    resultado = motor.clasificar_movimiento(mov)
    assert len(resultado) == 3
    assert all(c.requiere_revision for c in resultado)
    assert {c.concepto_dian for c in resultado} == {5004, 5007, 5008}
    print(f"  ✓ Capa 2 ambigua genera {len(resultado)} clasificaciones, todas requieren revisión")


def test_motor_fallback_capa1():
    """Si capa 2 y 3 no resuelven, cae a capa 1."""
    motor = MotorClasificacion(
        reglas_capa1=[
            ReglaCapa1(codigo_cuenta='1110', formato_dian='1012', concepto_dian=1110,
                       nombre_cuenta='Bancos'),
        ],
        reglas_capa2=[],  # vacía
    )
    mov = Movimiento(codigo_cuenta='111005', saldo_final=50000)
    # Debe encontrar por prefijo 4d ('1110')
    resultado = motor.clasificar_movimiento(mov)
    assert len(resultado) == 1
    assert resultado[0].capa_resolucion == 'puc_generico'
    assert resultado[0].formato_dian == '1012'
    print("  ✓ Fallback a capa 1 con búsqueda por prefijo")


def test_motor_sin_resolver():
    """Cuenta sin match en ninguna capa marca sin_resolver + requiere_revision."""
    motor = MotorClasificacion(reglas_capa1=[], reglas_capa2=[])
    mov = Movimiento(codigo_cuenta='999999', debitos=100)
    resultado = motor.clasificar_movimiento(mov)
    assert len(resultado) == 1
    assert resultado[0].capa_resolucion == 'sin_resolver'
    assert resultado[0].requiere_revision
    print("  ✓ Movimiento sin regla queda como sin_resolver+revisar")


# ================================================================
# Test integral con datos reales: Rutas del Mar
# ================================================================
def test_integracion_rutas_del_mar():
    """Aplicar el motor completo al balance real de Rutas del Mar."""
    base_dir = Path(__file__).parent.parent.parent

    # 1. Cargar reglas capa 2 desde archivo nativo
    archivo_nativo = '/mnt/user-data/uploads/Codificación_Formatos__Dic-31-20255220269145.xlsx'
    res_nativa = cargar_codificacion_nativa(archivo_nativo)
    reglas_capa2 = [
        ReglaCapa2(
            formato_dian=r.formato_dian,
            concepto_dian=r.concepto_dian,
            cuenta_inicial=r.cuenta_inicial,
            cuenta_final=r.cuenta_final,
            descripcion_concepto=r.descripcion_concepto,
            fila_origen=r.fila_origen,
        )
        for r in res_nativa.reglas
    ]
    print(f"  · {len(reglas_capa2)} reglas nativas cargadas")

    # 2. Cargar PUC genérico como capa 1
    wb = openpyxl.load_workbook(
        '/mnt/user-data/uploads/PUC_-_Exo_gena_Actualizado_1_.xlsx',
        data_only=True, read_only=True
    )
    reglas_capa1 = []
    for sheet in ['puc', 'pasivo']:
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1: continue
            codigo, desc, formato, concepto = row[0], row[1], row[2], row[3]
            if not codigo or not formato: continue
            f = str(formato).strip()
            if not f.isdigit(): continue
            try:
                cpt = int(concepto) if concepto else None
            except (ValueError, TypeError):
                cpt = None
            reglas_capa1.append(ReglaCapa1(
                codigo_cuenta=str(codigo).strip(),
                formato_dian=f,
                concepto_dian=cpt,
                nombre_cuenta=str(desc or '').strip(),
            ))
    wb.close()
    print(f"  · {len(reglas_capa1)} reglas PUC genérico cargadas")

    # 3. Cargar movimientos del balance
    wb = openpyxl.load_workbook(
        '/mnt/user-data/uploads/BALANCE_PARA_MAPEO_MEDIOS_RUTAS_DE_MAR.xlsx',
        data_only=True, read_only=True
    )
    ws = wb['Datos']
    movimientos = []
    
    def safe_float(v):
        try:
            return float(v) if v is not None and str(v).strip() not in ('#N/A', '') else 0.0
        except (ValueError, TypeError):
            return 0.0
    
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= 4: continue
        codigo, _, nombre, nit, nom_nit, _, db, cr, ns = row[:9]
        if not codigo or not nit or not str(nit).strip(): continue
        if str(codigo).strip() == '#N/A': continue
        movimientos.append(Movimiento(
            codigo_cuenta=str(codigo).strip(),
            nit=str(nit).strip().replace('.', '').replace('-', '').replace(' ', ''),
            debitos=safe_float(db),
            creditos=safe_float(cr),
            saldo_final=safe_float(ns),
            nombre_cuenta=str(nombre or '').strip(),
            nombre_tercero=str(nom_nit or '').strip(),
        ))
    wb.close()
    print(f"  · {len(movimientos)} movimientos por tercero del balance")

    # 4. Clasificar
    motor = MotorClasificacion(reglas_capa1=reglas_capa1, reglas_capa2=reglas_capa2)
    resultado = motor.clasificar_balance(movimientos)

    print(f"\n  Resultado:")
    e = resultado.estadisticas
    print(f"    Movimientos: {e['total_movimientos']}")
    print(f"    Clasificaciones generadas: {e['total_clasificaciones']}")
    print(f"      Capa 2 únicos:    {e['capa2_unico']:>4} ({e['capa2_unico']/e['total_movimientos']*100:.1f}%)")
    print(f"      Capa 2 ambiguos:  {e['capa2_ambiguo']:>4} ({e['capa2_ambiguo']/e['total_movimientos']*100:.1f}%)")
    print(f"      Capa 1 (PUC):     {e['capa1']:>4} ({e['capa1']/e['total_movimientos']*100:.1f}%)")
    print(f"      Sin resolver:     {e['sin_resolver']:>4} ({e['sin_resolver']/e['total_movimientos']*100:.1f}%)")
    print(f"    Requieren revisión: {e['requieren_revision']}")

    # Validaciones
    assert e['total_movimientos'] >= 700
    assert e['capa2_unico'] + e['capa2_ambiguo'] >= 400, "Capa 2 debe resolver mayoría"
    assert e['sin_resolver'] < 150, "Sin resolver debe ser <20%"
    print("  ✓ Test integral pasó las validaciones")

    # Mostrar distribución por formato
    from collections import Counter
    fmt_count = Counter(c.formato_dian for c in resultado.movimientos
                        if c.formato_dian and not c.requiere_revision)
    print(f"\n  Distribución por formato (sólo clasificados firmes):")
    for f, n in sorted(fmt_count.items()):
        print(f"    {f}: {n}")


if __name__ == '__main__':
    print("\n=== Tests del motor de clasificación ===\n")
    print("Lógica de rangos:")
    test_cuenta_en_rango()

    print("\nMotor unitario:")
    test_motor_capa3_tiene_prioridad()
    test_motor_capa2_ambiguedad()
    test_motor_fallback_capa1()
    test_motor_sin_resolver()

    print("\nTest integral con datos reales (Rutas del Mar):")
    test_integracion_rutas_del_mar()

    print("\n✅ Todos los tests pasaron")
