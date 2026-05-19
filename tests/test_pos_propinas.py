"""
Tests del manejo de propinas en procesador_pos.

Casos clave:
  1. Sin propinas (caso clásico): asiento de 3 líneas, Db=Cr.
  2. Con propinas: asiento de 4 líneas, Db=Cr, propina = final - inc/0.08 - inc.
  3. Tolerancia: si la diferencia es <= TOLERANCIA_PROPINA_PESOS no se trata como propina.
  4. Sin INC: se mantiene la lógica clásica (cuadre forzado).
"""
import io
import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core.procesadores.procesador_pos import (
    procesar_pos,
    Sucursal,
    TARIFA_INC_POS,
    CUENTA_PROPINAS_DEFAULT,
)


# ============================================================
# Helpers para construir Excels in-memory
# ============================================================

def _xlsx_test(filas_chili: list[dict]) -> bytes:
    """Crea un xlsx mínimo con DATOS PUNTO + REPORTE CHILI.

    Cada fila de `filas_chili` debe tener: fecha, sucursal, neto, impuestos, final.
    """
    wb = Workbook()
    # Hoja 1: DATOS PUNTO
    ws = wb.active
    ws.title = "DATOS PUNTO"
    ws.append([
        "COMPROBANTE", "NOMBRE EN REPORTE", "SEDE", "CC",
        "CUENTA DE CAJA", "CTA BASE V", "CTA ICO", "CLASE DE SEDE",
        "CTA PROPINAS",
    ])
    ws.append([
        "497", "Indiana", "Indiana", "001101",
        "11050510", "41401501", "24800505", "SANTA LEÑA",
        "28150505",
    ])
    ws.append([
        "497", "Oviedo", "Oviedo", "001102",
        "11050511", "41401501", "24800505", "SANTA LEÑA",
        "28150505",
    ])

    # Hoja 2: REPORTE CHILI
    ws2 = wb.create_sheet("REPORTE CHILI")
    # Por cada sucursal pongo un sub-reporte
    fila_actual = 1
    sucursales_unicas = {f["sucursal"] for f in filas_chili}
    for nombre_suc in sucursales_unicas:
        ws2.cell(row=fila_actual, column=1, value=nombre_suc)
        fila_actual += 1
        # Encabezado del subreporte (sin columna SUCURSAL para forzar default)
        ws2.cell(row=fila_actual, column=1, value="FECHA")
        ws2.cell(row=fila_actual, column=2, value="TOTAL NETO")
        ws2.cell(row=fila_actual, column=3, value="IMPUESTOS")
        ws2.cell(row=fila_actual, column=4, value="TOTAL FINAL")
        fila_actual += 1
        for f in filas_chili:
            if f["sucursal"] != nombre_suc:
                continue
            ws2.cell(row=fila_actual, column=1, value=f["fecha"])
            ws2.cell(row=fila_actual, column=2, value=f["neto"])
            ws2.cell(row=fila_actual, column=3, value=f["impuestos"])
            ws2.cell(row=fila_actual, column=4, value=f["final"])
            fila_actual += 1
        fila_actual += 1  # separador

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _check_cuadre(df: pd.DataFrame) -> tuple[int, int]:
    """Devuelve (total_db, total_cr) del DataFrame del plano."""
    db = int(df[df["TR"] == "1"]["VALOR"].astype(int).sum())
    cr = int(df[df["TR"] == "2"]["VALOR"].astype(int).sum())
    return db, cr


# ============================================================
# Tests
# ============================================================

class TestPropinasPOS:

    def test_sin_propina_cuadre_clasico(self):
        """Caso clásico: neto=100.000, INC=8.000, final=108.000.
        No hay propina → 3 líneas, Db=Cr=108.000."""
        from datetime import date
        xlsx = _xlsx_test([
            {"fecha": date(2026, 3, 1), "sucursal": "Indiana",
             "neto": 100000, "impuestos": 8000, "final": 108000},
        ])
        df, log, sucs_no = procesar_pos(xlsx)

        assert len(df) == 3, f"Esperaba 3 líneas, obtuve {len(df)}"
        db, cr = _check_cuadre(df)
        assert db == cr == 108000

        # No debe haber línea de propinas
        assert (df["CUENTA"] == CUENTA_PROPINAS_DEFAULT).sum() == 0

    def test_con_propina_4_lineas_cuadrado(self):
        """Caso con propina: final=120.000, INC=8.000
        → base_inc = 8.000/0.08 = 100.000
        → propina = 120.000 - 100.000 - 8.000 = 12.000
        Plano debe tener 4 líneas y Db=Cr."""
        from datetime import date
        # neto reportado puede no encajar (POS a veces incluye propina en final)
        xlsx = _xlsx_test([
            {"fecha": date(2026, 3, 1), "sucursal": "Indiana",
             "neto": 100000, "impuestos": 8000, "final": 120000},
        ])
        df, log, sucs_no = procesar_pos(xlsx)

        assert len(df) == 4, f"Esperaba 4 líneas, obtuve {len(df)}"
        db, cr = _check_cuadre(df)
        assert db == cr == 120000, f"Descuadre: Db={db} Cr={cr}"

        # Propina = 12.000 exactos
        propina_row = df[df["CUENTA"] == CUENTA_PROPINAS_DEFAULT]
        assert len(propina_row) == 1
        assert int(propina_row["VALOR"].iloc[0]) == 12000

        # Base venta = 100.000 (= INC/0.08)
        base_row = df[df["CUENTA"] == "41401501"]
        assert int(base_row["VALOR"].iloc[0]) == 100000

        # INC = 8.000
        inc_row = df[df["CUENTA"] == "24800505"]
        assert int(inc_row["VALOR"].iloc[0]) == 8000

    def test_tolerancia_no_propina(self):
        """Si la diferencia es 1 peso (redondeo), NO se trata como propina."""
        from datetime import date
        # neto=100.000 + inc=8.000 = 108.000, pero final=108.001 (1 peso de redondeo)
        xlsx = _xlsx_test([
            {"fecha": date(2026, 3, 1), "sucursal": "Indiana",
             "neto": 100000, "impuestos": 8000, "final": 108001},
        ])
        df, log, sucs_no = procesar_pos(xlsx)

        # No debe haber línea de propinas
        assert (df["CUENTA"] == CUENTA_PROPINAS_DEFAULT).sum() == 0
        # Debe seguir cuadrando
        db, cr = _check_cuadre(df)
        assert db == cr == 108001

    def test_sin_inc_logica_clasica(self):
        """Sin INC reportado: lógica clásica (cuadre forzado), 3 líneas máx."""
        from datetime import date
        xlsx = _xlsx_test([
            {"fecha": date(2026, 3, 1), "sucursal": "Indiana",
             "neto": 100000, "impuestos": 0, "final": 100000},
        ])
        df, log, sucs_no = procesar_pos(xlsx)

        # Solo 2 líneas: Db caja + Cr venta (sin IC, sin propina)
        assert (df["CUENTA"] == CUENTA_PROPINAS_DEFAULT).sum() == 0
        db, cr = _check_cuadre(df)
        assert db == cr == 100000

    def test_propina_multiples_sucursales_dias(self):
        """Múltiples sucursales/días con propinas — todos cuadran."""
        from datetime import date
        xlsx = _xlsx_test([
            {"fecha": date(2026, 3, 1), "sucursal": "Indiana",
             "neto": 50000, "impuestos": 4000, "final": 60000},   # propina 6.000
            {"fecha": date(2026, 3, 2), "sucursal": "Indiana",
             "neto": 75000, "impuestos": 6000, "final": 81000},   # sin propina (cuadra exacto)
            {"fecha": date(2026, 3, 1), "sucursal": "Oviedo",
             "neto": 200000, "impuestos": 16000, "final": 240000}, # propina 24.000
        ])
        df, log, sucs_no = procesar_pos(xlsx)

        # Cuadre global
        db, cr = _check_cuadre(df)
        assert db == cr, f"Descuadre global: Db={db} Cr={cr}"

        # Total Db = 60.000 + 81.000 + 240.000 = 381.000
        assert db == 381000

        # 2 líneas de propinas (la del 2/3 cuadra exacto, no genera propina)
        propina_rows = df[df["CUENTA"] == CUENTA_PROPINAS_DEFAULT]
        assert len(propina_rows) == 2
        assert sorted(propina_rows["VALOR"].astype(int).tolist()) == [6000, 24000]

    def test_centro_de_costo_propina_es_de_sucursal(self):
        """La línea de propina debe llevar el CC de la sucursal."""
        from datetime import date
        xlsx = _xlsx_test([
            {"fecha": date(2026, 3, 1), "sucursal": "Indiana",
             "neto": 100000, "impuestos": 8000, "final": 120000},
        ])
        df, log, sucs_no = procesar_pos(xlsx)

        propina_row = df[df["CUENTA"] == CUENTA_PROPINAS_DEFAULT]
        assert propina_row["CENTRO DE COSTO"].iloc[0] == "001101"  # Indiana
